"""Build a persistent index of the user's manual FTM XLSX archive.

The user maintains a master workbook (`FTM_2026_NC Estates (N).xlsx` in
their Downloads folder) with one tab per ISO week. Each tab contains
their manually-pulled probate cases with verified case numbers + PR
mailings. As newspaper-published notices (which lack case numbers) flow
into our weekly scraper, we want to automatically match them against
this manual archive — the case was almost certainly filed weeks earlier
and the user already pulled it.

This script:
  1. Finds the latest FTM XLSX in `C:\\Users\\omark\\Downloads\\`
  2. Walks every "Week N" tab, extracts rows with a non-blank Case No.
  3. Builds a JSON index keyed by (county, normalized-decedent-tokens)
  4. Writes to `output/.manual_archive_index.json` for use by
     `fix_addresses_and_prep.py:backfill_from_manual_archive`.

Re-run any time you update the manual workbook.

Usage:
    python build_manual_archive_index.py
    python build_manual_archive_index.py --xlsx "C:\\path\\to\\file.xlsx"
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S"
)
logger = logging.getLogger("manual_index")


INDEX_PATH = Path("output") / ".manual_archive_index.json"
DOWNLOADS_DIR = Path(r"C:\Users\omark\Downloads")
XLSX_GLOB = "FTM*NC*Estates*.xlsx"


CSV_GLOB = "FTM*NC*Estates*Week*.csv"


def find_latest_xlsx() -> Path | None:
    """Latest master workbook in Downloads, or None. The user's workflow moved
    to per-week CSV exports (see find_weekly_csvs), so a master XLSX is now
    optional — returning None instead of raising keeps the pipeline from
    dumping a FileNotFoundError traceback into the daily log."""
    candidates = sorted(DOWNLOADS_DIR.glob(XLSX_GLOB),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def _week_num(text: str) -> int | None:
    """Extract the ISO week number from a tab name ('Week 27') or filename
    ('FTM_2026_NC Estates - Week 27 (1).csv')."""
    m = re.search(r"week\s*0*(\d+)", text or "", re.I)
    return int(m.group(1)) if m else None


def find_weekly_csvs() -> list[Path]:
    """Per-week manual CSV exports in Downloads, one Path per ISO week (the
    most recently modified file wins when a week has multiple '(N)' copies).

    The user exports one CSV per week (e.g. 'FTM_2026_NC Estates - Week 27
    (1).csv') rather than maintaining a single master workbook, so these are
    the primary archive source now."""
    by_week: dict[int, Path] = {}
    for p in DOWNLOADS_DIR.glob(CSV_GLOB):
        wk = _week_num(p.name)
        if wk is None:
            continue
        if wk not in by_week or p.stat().st_mtime > by_week[wk].stat().st_mtime:
            by_week[wk] = p
    return [by_week[w] for w in sorted(by_week)]


def name_token_key(name: str) -> str:
    """Order-independent name key — covers 'Last, First Middle' / 'First
    Middle Last' / 'AKA <alt>' variations by sorting normalized tokens.
    """
    s = (name or "").upper()
    s = re.sub(r"\bAKA\b.*", "", s)
    s = re.sub(r"\b(JR|SR|II|III|IV|MR|MRS|MS|DR)\.?\b", "", s)
    tokens = sorted(t for t in re.findall(r"[A-Z]+", s) if len(t) >= 3)
    return " ".join(tokens)


def cell(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _ingest_table(
    headers: list[str],
    rows: list[tuple],
    week_label: str,
    index: dict[str, dict],
    counters: dict[str, int],
) -> int:
    """Fold one table (xlsx tab or csv) into the shared index. Returns the
    number of NEW unique (county, decedent) entries added. First occurrence
    across all tables wins (tables are ingested oldest-week-first)."""
    try:
        col_county = headers.index("County")
        col_case = headers.index("Case No.")
        col_dec = headers.index("Deceased Owner")
    except ValueError:
        logger.warning("Skipping %s (missing required column)", week_label)
        return 0

    def maybe(name: str) -> int | None:
        return headers.index(name) if name in headers else None
    optional = {
        "first_name": maybe("First Name"), "last_name": maybe("Last Name"),
        "mailing_address": maybe("Mailing Address"), "mailing_city": maybe("Mailing City"),
        "mailing_state": maybe("Mailing State"), "mailing_zip": maybe("Mailing Zip"),
        "parcel_id": maybe("Parcel ID"),
        "property_address": maybe("Property Address"), "property_city": maybe("Property City"),
        "property_zip": maybe("Property Zip"),
    }

    added = 0
    for r in rows:
        county = cell(r[col_county]) if col_county < len(r) else ""
        case_no = cell(r[col_case]) if col_case < len(r) else ""
        dec = cell(r[col_dec]) if col_dec < len(r) else ""
        if not dec:
            continue
        if not case_no:
            counters["skipped_no_case"] += 1
            continue
        key = f"{county.upper()}||{name_token_key(dec)}"
        if key in index:
            continue  # first occurrence (oldest week) wins
        entry = {"case_no": case_no, "week": week_label,
                 "deceased_owner": dec, "county": county}
        for name, idx in optional.items():
            if idx is not None and idx < len(r):
                val = cell(r[idx])
                if val:
                    entry[name] = val
        index[key] = entry
        added += 1
    return added


def _load_xlsx_tables(xlsx_path: Path) -> list[tuple[int, str, list[str], list[tuple]]]:
    """Return (week_num, label, headers, rows) for each 'Week N' tab."""
    wb = load_workbook(xlsx_path, data_only=True)
    tables = []
    for tab in (s for s in wb.sheetnames if s.lower().startswith("week ")):
        ws = wb[tab]
        headers = [cell(c.value) for c in ws[1]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wk = _week_num(tab)
        tables.append((wk if wk is not None else 9999, f"{tab} (xlsx)", headers, rows))
    return tables


def _load_csv_table(csv_path: Path) -> tuple[int, str, list[str], list[tuple]]:
    """Return (week_num, label, headers, rows) for one per-week CSV."""
    import csv as _csv
    with csv_path.open(encoding="utf-8-sig", newline="") as f:
        reader = list(_csv.reader(f))
    headers = [cell(h) for h in reader[0]] if reader else []
    rows = [tuple(r) for r in reader[1:]]
    wk = _week_num(csv_path.name)
    label = f"Week {wk}" if wk is not None else csv_path.stem
    return (wk if wk is not None else 9999, f"{label} (csv)", headers, rows)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=None, help="Path to a manual XLSX file")
    args = ap.parse_args()

    # Gather every source table: master XLSX tabs (if any) + per-week CSVs.
    tables: list[tuple[int, str, list[str], list[tuple]]] = []
    sources: list[str] = []

    xlsx_path = Path(args.xlsx) if args.xlsx else find_latest_xlsx()
    if xlsx_path:
        logger.info("Reading manual XLSX: %s", xlsx_path)
        tables.extend(_load_xlsx_tables(xlsx_path))
        sources.append(str(xlsx_path))

    for csv_path in find_weekly_csvs():
        logger.info("Reading manual CSV: %s", csv_path.name)
        tables.append(_load_csv_table(csv_path))
        sources.append(str(csv_path))

    if not tables:
        logger.warning("No manual archive found in %s (looked for %s and %s). "
                       "Writing empty index.", DOWNLOADS_DIR, XLSX_GLOB, CSV_GLOB)

    # Oldest week first so first-occurrence-wins keeps the earliest filing.
    tables.sort(key=lambda t: t[0])

    index: dict[str, dict] = {}  # key: "COUNTY||sorted-name-tokens"
    counters = {"skipped_no_case": 0}
    by_label_counts: list[tuple[str, int]] = []
    for _wk, label, headers, rows in tables:
        added = _ingest_table(headers, rows, label, index, counters)
        by_label_counts.append((label, added))

    INDEX_PATH.parent.mkdir(parents=True, exist_ok=True)
    with INDEX_PATH.open("w", encoding="utf-8") as f:
        json.dump({"_sources": sources, "_entries": index}, f,
                  separators=(",", ":"), ensure_ascii=False)

    logger.info("Wrote index: %s", INDEX_PATH)
    logger.info("  Total unique (county, decedent) entries: %d", len(index))
    logger.info("  Skipped rows without Case No.: %d", counters["skipped_no_case"])
    logger.info("Per-source additions (new unique decedents):")
    for label, n in by_label_counts:
        logger.info("  %s: %d", label, n)


if __name__ == "__main__":
    main()
