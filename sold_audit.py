"""Recently-sold loss audit — parcel-level ground truth from county GIS.

Reads every week tab of the latest consolidated FTM workbook, dedupes
parcels, queries each county's GIS BY PARCEL ID (exact match — no address
fuzz), and reports any probate property whose GIS sale date falls on/after
--since (default 2026-06-01). Independent of DataSift entirely.

A parcel is also flagged when its GIS reading moved forward since the last run,
which is the only way some counties are visible at all — see
CHANGE_MAX_AGE_DAYS below.

A parcel is ALSO flagged when the sale falls on/after the estate's file date
(a transfer after the estate opened is a settlement or a sale no matter how
old the window is — that is how a January sale on a Week-50 case stayed in
the mail lane until August 2026).

Usage:
    python sold_audit.py                       # since 2026-06-01
    python sold_audit.py --since 2026-01-01    # any 2026 sale
    python sold_audit.py --since-days 90       # rolling window (nightly uses this)
    python sold_audit.py --since-days 90 --crm-legacy
        # + every "Courthouse Data" record in DataSift that is NOT in the
        #   workbook (Jul-2025..May-2026 vintages, ~2,000 parcels). Self-
        #   throttled to once per --crm-every days (default 7) because it is
        #   ~2,700 record reads + ~2,000 GIS calls. --crm-force ignores the
        #   throttle. Needs a DataSift token (same path push_sold_tags uses).
Output: output/sold_audit_<today>.csv + console summary (Case No. included).
Feeds push_sold_tags.py, which tags the real sales "Sold" in DataSift.
"""
from __future__ import annotations

import argparse
import csv
import glob
import os
import re
import sys
import time
from datetime import date, timedelta

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "src"))

import requests  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from nc_gis_lookup import (  # noqa: E402
    _ARCGIS_CONFIG, _ARCGIS_HEADERS, _catawba_parcel_report,
    _parse_arcgis_sale_date, _parse_arcgis_sale_price, _polaris3g_get_page,
)

ARC_COUNTIES = {"rowan", "cabarrus", "gaston", "iredell", "lincoln"}


def load_rows() -> list[dict]:
    books = sorted(glob.glob("output/FTM_*_NC_Estates_throughWeek*.xlsx"),
                   key=os.path.getmtime)
    if not books:
        raise SystemExit("No consolidated workbook found in output/")
    book = books[-1]
    print(f"Workbook: {book}")
    wb = load_workbook(book, read_only=True, data_only=True)
    rows: list[dict] = []
    for ws in wb.worksheets:
        header: dict[str, int] = {}
        for r_i, row in enumerate(ws.iter_rows(values_only=True)):
            vals = ["" if v is None else str(v).strip() for v in row]
            if not header:
                for c_i, v in enumerate(vals):
                    lv = v.lower()
                    if "county" == lv:
                        header["county"] = c_i
                    elif "case" in lv and "no" in lv:
                        header["case"] = c_i
                    elif "parcel" in lv:
                        header["parcel"] = c_i
                    elif lv.startswith("property address"):
                        header["address"] = c_i
                    elif "deceased" in lv or lv == "decedent":
                        header["decedent"] = c_i
                    elif lv.startswith("personal representative"):
                        header["pr"] = c_i
                    elif lv == "file date":
                        header["filed"] = c_i
                if header and "parcel" not in header:
                    header = {}
                continue
            def g(key):
                i = header.get(key)
                return vals[i] if i is not None and i < len(vals) else ""
            county = g("county").lower()
            parcel = g("parcel")
            if not county or not parcel:
                continue
            rows.append({
                "week": ws.title,
                "county": county,
                "case": g("case"),
                "decedent": g("decedent"),
                "pr": g("pr"),
                "filed": g("filed"),
                "address": g("address"),
                "parcel": parcel,
            })
    print(f"Rows with parcels: {len(rows)}")
    return rows


_ENTITY_MARKERS = (
    "LLC", "L L C", "INC", "CORP", "PROPERTIES", "HOLDINGS", "HOMES",
    "INVESTMENTS", "CAPITAL", "VENTURES", "REALTY", "REI", "GROUP",
    "PARTNERS", "ENTERPRISES", "SOLUTIONS", "TRUST CO",
)


def _surnames(name: str) -> set[str]:
    return {t.strip(",.").upper() for t in name.replace(",", " ").split()
            if len(t.strip(",.")) > 2}


def classify(row: dict) -> str:
    """HEIR TRANSFER / INVESTOR PURCHASE / MARKET SALE / UNCLEAR TRANSFER."""
    owner = (row.get("owner") or "").upper()
    fam = _surnames(row.get("decedent", "")) | _surnames(row.get("pr", ""))
    if fam & _surnames(owner):
        return "HEIR TRANSFER"
    if any(m in owner for m in _ENTITY_MARKERS):
        return "INVESTOR PURCHASE"
    price = row.get("sale_price")
    try:
        pricef = float(price) if price not in (None, "") else 0.0
    except (TypeError, ValueError):
        pricef = 0.0
    if pricef >= 5000:
        return "MARKET SALE"
    return "UNCLEAR TRANSFER"


def update_competitor_log(flagged: list[dict]) -> None:
    """Append investor/market sales to a durable competitor log (dedup by case)."""
    path = "output/competitor_log.csv"
    existing: set[str] = set()
    if os.path.exists(path):
        with open(path, encoding="utf-8-sig") as f:
            existing = {r["case"] for r in csv.DictReader(f)}
    new = [r for r in flagged
           if r["class"] in ("INVESTOR PURCHASE", "MARKET SALE")
           and r["case"] not in existing]
    if not new:
        print("Competitor log: no new entries")
        return
    write_header = not os.path.exists(path)
    with open(path, "a", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=[
            "found", "case", "county", "decedent", "filed", "address",
            "sale_date", "sale_price", "buyer", "class", "week"],
            extrasaction="ignore")
        if write_header:
            w.writeheader()
        for r in new:
            w.writerow({**r, "found": date.today().isoformat(),
                        "buyer": r.get("owner", "")})
    print(f"Competitor log: +{len(new)} entries -> {path}")


# A date window alone can't see a sale in every county. Measured 2026-08-23
# across two snapshots 22 days apart:
#   Lincoln  57/57  and Rowan 116/116 dated sales read "Jan 1" — those counties
#                   publish a sale YEAR, not a date, so a "last 90 days" window
#                   only sees them in Jan-Mar.
#   Iredell         posts 16-61+ days after the sale (Mecklenburg: ~10 days).
# So the sweep also fires when a parcel's GIS reading CHANGED since the last
# run — that catches a year flip (2025-01-01 -> 2026-01-01) and a slow county
# whenever it finally posts, with no date window involved.
CHANGE_MAX_AGE_DAYS = 730  # a year-only county's "Jan 1" can already be ~600 days old


def load_prior_readings() -> dict[tuple, str]:
    """(county, parcel) -> sale_date from the most recent PREVIOUS audit CSV.

    Empty on the very first run, which disables change detection for that run —
    without a baseline every parcel would look "changed" and flag at once.
    """
    # Compare BASENAMES: glob returns Windows backslash paths
    # ("output\\sold_audit_....csv") which never equal a forward-slash literal,
    # so today's own file survived the filter and became its own baseline —
    # zero changes detected, Lincoln and Rowan silently back to blind.
    today = f"sold_audit_{date.today().isoformat()}.csv"
    files = sorted(f for f in glob.glob("output/sold_audit_*.csv")
                   if os.path.basename(f) != today)
    if not files:
        print("No prior audit CSV — change detection off this run (building baseline)")
        return {}
    print(f"Change baseline: {files[-1]}")
    out: dict[tuple, str] = {}
    with open(files[-1], encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            sd = (r.get("sale_date") or "")[:10]
            out[(r.get("county", ""), r.get("parcel", ""))] = sd
    return out


def changed_since_last_run(prior: dict[tuple, str], county: str, pid: str,
                           sale_date: str | None) -> bool:
    """True when this parcel's sale record moved FORWARD to a recent date.

    Guarded three ways so a GIS correction to an old deed doesn't read as a
    sale: there must be a prior reading, the new date must be strictly NEWER
    than the old one, and it must fall inside CHANGE_MAX_AGE_DAYS.
    """
    key = (county, pid)
    if key not in prior or not sale_date:
        return False
    was = prior[key]
    if not was or sale_date <= was:
        return False
    try:
        sold_on = date.fromisoformat(sale_date[:10])
    except ValueError:
        return False
    return 0 <= (date.today() - sold_on).days <= CHANGE_MAX_AGE_DAYS


def norm_pid(pid: str) -> str:
    """Strip float artifacts ('4673892477.00000000') and whitespace."""
    pid = pid.strip()
    if re.fullmatch(r"\d+\.0+", pid):
        pid = pid.split(".")[0]
    return pid


# Candidate (field, mode) query attempts per county, in order. Older workbook
# weeks recorded the ALTERNATE id type for Gaston (PIN, pre-2026-07-09) and
# Lincoln (PIN, pre-2026-07-23); Cabarrus workbook holds the 10-digit PIN
# while the GIS canonical field is 14-digit PIN14 (prefix match); Iredell
# PINs are sometimes stored dashed.
_FIELD_ATTEMPTS: dict[str, list[tuple[str, str]]] = {
    "rowan":    [("PARCEL_ID", "exact")],
    "cabarrus": [("PIN14", "exact"), ("PIN14", "like")],
    "gaston":   [("PID", "exact"), ("PIN", "exact")],
    "iredell":  [("PIN", "exact"), ("PIN", "like"), ("PIN", "dashed")],
    "lincoln":  [("PARCELID", "exact"), ("PIN", "exact")],
}


# Cabarrus was 0-for-98 on sale dates until 2026-08-30: the `Parcels` layer
# nc_gis_lookup uses for name search carries owner + mailing but NO sale
# fields, so every Cabarrus parcel read "found, never sold". The OpenData
# Tax_Parcels layer publishes SaleYear / SaleMonth / SalePrice / DeedBook /
# DeedPage on the same PIN14 key (verified current: it showed the Jan-2026
# estate deed on 5607 Dorchester Ave that Parcels could not). Sale-date
# resolution is month-level there (we synthesize the 1st).
_SALE_LAYER: dict[str, str] = {
    "cabarrus": ("https://location.cabarruscounty.us/arcgisservices/rest/"
                 "services/OpenData/Tax_Parcels/MapServer/1"),
}


def q_arcgis(county: str, pid: str) -> dict | None:
    cfg = _ARCGIS_CONFIG[county]
    url = _SALE_LAYER.get(county, cfg["url"])
    pid = norm_pid(pid)
    pid_esc = pid.replace("'", "''")
    attempts = _FIELD_ATTEMPTS.get(county, [(cfg["parcel_field"], "exact")])
    for field, mode in attempts:
        if mode == "exact":
            where = f"{field}='{pid_esc}'"
        elif mode == "like":
            where = f"{field} LIKE '{pid_esc}%'"
        elif mode == "dashed" and len(pid) == 10:
            dashed = f"{pid[:4]}-{pid[4:6]}-{pid[6:]}"
            where = f"{field}='{dashed}'"
        else:
            continue
        try:
            r = requests.get(f"{url}/query", params={
                "where": where,
                "outFields": "*", "returnGeometry": "false", "f": "json",
            }, headers=_ARCGIS_HEADERS, timeout=30)
            data = r.json() or {}
            if data.get("error"):
                continue  # invalid field on this layer — try next
            feats = data.get("features", [])
            if not feats:
                continue
            attrs = feats[0].get("attributes", {})
            owner = " / ".join(
                str(attrs.get(f) or "").strip()
                for f in cfg.get("owner_fields", []) if attrs.get(f))
            return {
                "sale_date": _parse_arcgis_sale_date(attrs),
                "sale_price": _parse_arcgis_sale_price(attrs),
                "owner": owner,
            }
        except Exception as e:  # noqa: BLE001
            print(f"  ! {county} {pid} ({field} {mode}): {e}")
    return None


def q_catawba(pid: str) -> dict | None:
    try:
        rec = _catawba_parcel_report(pid)
        if not rec:
            return None
        sd = str(rec.get("sale_date") or "")[:10] or None
        if sd and "/" in sd:
            m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", sd)
            if m:
                sd = f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
        price = rec.get("sale_amount")
        try:
            price = float(price) if price not in (None, "") else None
        except (TypeError, ValueError):
            price = None
        owner = str(rec.get("owner") or rec.get("owner_name") or "").strip()
        return {"sale_date": sd, "sale_price": price, "owner": owner}
    except Exception as e:  # noqa: BLE001
        print(f"  ! catawba {pid}: {e}")
        return None


_meck_session = None
_meck_supported = True


def q_meck(pid: str) -> dict | None:
    """polaris3g bolt parcel lookup — the param is `pid` (verified 2026-08-01;
    record includes sale_date, sale_price, sale_qualified, owner, situs)."""
    global _meck_session
    if _meck_session is None:
        _meck_session = requests.Session()
    try:
        rows = _polaris3g_get_page(_meck_session, {"pid": norm_pid(pid)})
        if not rows:
            return None
        rec = rows[0]
        sd = str(rec.get("sale_date") or "")[:10] or None
        if sd and "/" in sd:
            m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", sd)
            if m:
                sd = f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
        owners = rec.get("owner")
        if isinstance(owners, list):
            parts = []
            for o in owners[:3]:
                if isinstance(o, dict):
                    nm = (o.get("fullname")
                          or " ".join(str(o.get(k) or "") for k in
                                      ("firstname", "firstName",
                                       "lastname", "lastName")).strip())
                    if nm:
                        parts.append(str(nm))
                elif o:
                    parts.append(str(o))
            owner = " / ".join(parts)[:80]
        elif isinstance(owners, dict):
            owner = str(owners.get("fullname") or owners)[:80]
        else:
            owner = str(owners or "")[:80]
        price = rec.get("sale_price")
        try:
            price = float(price) if price not in (None, "") else None
        except (TypeError, ValueError):
            price = None
        return {"sale_date": sd, "sale_price": price, "owner": owner}
    except Exception as e:  # noqa: BLE001
        print(f"  ! mecklenburg {pid}: {e}")
        return None


def _iso(d: str) -> str:
    """'2026-07-15', '2026-07-15 00:00:00', '07/15/2026' -> '2026-07-15' ('' if unreadable)."""
    d = (d or "").strip()
    if not d:
        return ""
    if len(d) >= 10 and d[4] == "-" and d[7] == "-":
        return d[:10]
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", d)
    if m:
        return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return ""


# Sale-date resolution per county. Lincoln and Rowan publish only a YEAR
# (read as Jan 1), Cabarrus a year+month (read as the 1st), so "on/after the
# file date" has to be compared at that resolution or a Rowan estate filed in
# May whose house sold in June reads 2026-01-01 < 2026-05-xx and never flags.
YEAR_ONLY = {"lincoln", "rowan"}
MONTH_ONLY = {"cabarrus"}


def _post_filing(county: str, sale_date: str | None, filed: str) -> bool:
    """True when the GIS sale lands on/after the estate's file date, compared
    at the county's date resolution. classify() then decides whether it is a
    settlement to family (report-only) or a sale (suppress)."""
    if not sale_date or not filed:
        return False
    if county in YEAR_ONLY:
        return sale_date[:4] >= filed[:4]
    if county in MONTH_ONLY:
        return sale_date[:7] >= filed[:7]
    return sale_date >= filed


CRM_CACHE = "output/.sold_crm_parcels.json"
CRM_STAMP = "output/.sold_crm_legacy_last_run"
CRM_VINTAGE_SLACK_DAYS = 45   # a sale can record a few weeks before the list was pulled
_SUPPORTED = ARC_COUNTIES | {"catawba", "mecklenburg"}


def load_crm_rows(skip_keys: set[tuple], every_days: int, force: bool) -> list[dict]:
    """Courthouse Data records in DataSift that the workbook does not cover.

    The sweep read only the CURRENT-year consolidated workbook (Week 24+), so
    ~2,000 courthouse records uploaded Jul-2025..May-2026 had no GIS check at
    all and were protected solely by DataSift's own `last_sold` field, which
    stopped refreshing in July 2026. This pulls every record tagged
    "Courthouse Data", keeps the ones with a county + parcel that the workbook
    does not already sweep, and shapes them like workbook rows. Because the
    case file date is not on the record, the floor is the upload vintage
    ("List Purchased County MM/YYYY") minus CRM_VINTAGE_SLACK_DAYS.

    Record reads are cached 7 days (CRM_CACHE); the whole pass is throttled to
    once per `every_days` (CRM_STAMP) unless `force`. Sold-tagged records are
    skipped — nothing left to suppress. Any failure returns [] so the workbook
    sweep still runs.
    """
    if not force and os.path.exists(CRM_STAMP):
        try:
            last = date.fromisoformat(open(CRM_STAMP, encoding="utf-8").read().strip())
            if (date.today() - last).days < every_days:
                print(f"CRM legacy sweep: last ran {last}, next due in "
                      f"{every_days - (date.today() - last).days} day(s) — skipped")
                return []
        except ValueError:
            pass
    import json
    recs: list[dict] | None = None
    if os.path.exists(CRM_CACHE):
        try:
            blob = json.load(open(CRM_CACHE, encoding="utf-8"))
            if (date.today() - date.fromisoformat(blob["fetched"])).days < 7:
                recs = blob["rows"]
                print(f"CRM legacy: {len(recs)} records from cache ({blob['fetched']})")
        except (ValueError, KeyError, OSError):
            recs = None
    if recs is None:
        try:
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            from push_sold_tags import headers, tag_map
            h = headers()
            tags = tag_map(h)
            ch = tags.get("courthouse data")
            if not ch:
                print("CRM legacy: no 'Courthouse Data' tag on the account — skipped")
                return []
            api = "https://apiv2.reisift.io/api/internal/property/"
            listing: list[dict] = []
            off = 0
            while True:
                r = requests.post(api, headers={**h, "x-http-method-override": "GET"},
                                  json={"limit": 200, "offset": off,
                                        "query": {"must": {"any_tags": [ch]}}}, timeout=120)
                r.raise_for_status()
                page = r.json().get("results") or r.json().get("data") or []
                listing += page
                if len(page) < 200:
                    break
                off += 200
            print(f"CRM legacy: {len(listing)} Courthouse Data records, reading details...")
            recs = []
            for i, row in enumerate(listing, 1):
                try:
                    d = requests.get(f"{api}{row['uuid']}/", headers=h, timeout=60).json()
                except Exception as e:  # noqa: BLE001
                    print(f"  ! detail {row.get('uuid')}: {e}")
                    continue
                a = d.get("address") or {}
                o = d.get("owner") or {}
                tg = [t.get("title") if isinstance(t, dict) else t for t in (d.get("tags") or [])]
                recs.append({
                    "uuid": d.get("uuid"), "county": (a.get("county") or "").lower(),
                    "street": a.get("street") or "", "city": a.get("city") or "",
                    "parcel": str(d.get("parcel_id") or "").strip(),
                    "owner": " ".join(x for x in (o.get("first_name"), o.get("last_name")) if x),
                    "tags": [t for t in tg if t],
                    "last_sold": d.get("last_sold") or "",
                })
                if i % 250 == 0:
                    print(f"  ...{i}/{len(listing)}")
            json.dump({"fetched": date.today().isoformat(), "rows": recs},
                      open(CRM_CACHE, "w", encoding="utf-8"))
        except Exception as e:  # noqa: BLE001
            print(f"CRM legacy: DataSift read failed ({e}) — workbook sweep only")
            return []
    out: list[dict] = []
    skipped = {"no parcel": 0, "in workbook": 0, "sold-tagged": 0, "county": 0}
    for rec in recs:
        county, pid = rec["county"], norm_pid(rec["parcel"])
        if not pid:
            skipped["no parcel"] += 1
            continue
        if county not in _SUPPORTED:
            skipped["county"] += 1
            continue
        if (county, pid) in skip_keys:
            skipped["in workbook"] += 1
            continue
        if any(t.lower() == "sold" for t in rec["tags"]):
            skipped["sold-tagged"] += 1
            continue
        vintage = ""
        for t in rec["tags"]:
            m = re.match(r"List Purchased County (\d{2})/(\d{4})", t)
            if m:
                vintage = f"{m.group(2)}-{m.group(1)}"
        floor = ""
        if vintage:
            floor = (date.fromisoformat(vintage + "-01")
                     - timedelta(days=CRM_VINTAGE_SLACK_DAYS)).isoformat()
        out.append({
            "week": f"CRM {vintage or 'unknown vintage'}", "county": county, "case": "",
            "decedent": "", "pr": rec["owner"], "filed": floor,
            "address": rec["street"], "parcel": pid, "uuid": rec["uuid"],
        })
    print(f"CRM legacy: {len(out)} parcels to sweep "
          f"(skipped {', '.join(f'{v} {k}' for k, v in skipped.items())})")
    open(CRM_STAMP, "w", encoding="utf-8").write(date.today().isoformat())
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--since", default="2026-06-01")
    ap.add_argument("--since-days", type=int,
                    help="rolling window instead of a fixed --since date; what "
                         "the nightly run uses so the cutoff can't go stale")
    ap.add_argument("--crm-legacy", action="store_true",
                    help="also sweep Courthouse Data records in DataSift that "
                         "the workbook does not cover (throttled, see --crm-every)")
    ap.add_argument("--crm-every", type=int, default=7,
                    help="days between --crm-legacy passes (default 7)")
    ap.add_argument("--crm-force", action="store_true",
                    help="run --crm-legacy now even if it ran recently")
    args = ap.parse_args()
    if args.since_days:
        args.since = (date.today() - timedelta(days=args.since_days)).isoformat()
        print(f"rolling window: sales on/after {args.since}")

    prior = load_prior_readings()

    rows = load_rows()
    seen: dict[tuple, dict] = {}
    for r in rows:
        key = (r["county"], r["parcel"])
        seen.setdefault(key, r)  # keep first (oldest week) occurrence
    print(f"Unique (county, parcel): {len(seen)}")
    if args.crm_legacy:
        for r in load_crm_rows(set(seen), args.crm_every, args.crm_force):
            seen.setdefault((r["county"], r["parcel"]), r)
        print(f"Unique (county, parcel) incl. CRM legacy: {len(seen)}")

    results = []
    stats: dict[str, list[int]] = {}
    for i, ((county, pid), r) in enumerate(sorted(seen.items()), 1):
        if county in ARC_COUNTIES:
            gis = q_arcgis(county, pid)
        elif county == "catawba":
            gis = q_catawba(pid)
        elif county == "mecklenburg":
            gis = q_meck(pid)
        else:
            gis = None
        st = stats.setdefault(county, [0, 0, 0])  # [queried, found, flagged]
        st[0] += 1
        flag = ""
        if gis:
            st[1] += 1
            sd = gis.get("sale_date")
            filed = _iso(r.get("filed", ""))
            if sd and sd >= args.since:
                flag = f"SOLD since {args.since}"
            elif _post_filing(county, sd, filed):
                # Older than the window but on/after the estate opened: the
                # window never sees it, and the first run's baseline already
                # held the post-sale reading so change detection is blind too.
                flag = f"SOLD after filing {filed}"
            elif changed_since_last_run(prior, county, pid, sd):
                flag = f"SOLD - GIS record changed (was {prior[(county, pid)]})"
            if flag:
                st[2] += 1
        results.append({**r, **(gis or {}), "flag": flag})
        if i % 50 == 0:
            print(f"  ...{i}/{len(seen)}")
        time.sleep(0.3)

    out_path = f"output/sold_audit_{date.today().isoformat()}.csv"
    results.sort(key=lambda x: (x["flag"] == "", x["county"]))
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=[
            # "pr" is NOT cosmetic: push_sold_tags.py re-derives classify()
            # from this CSV, and classify() matches the new owner's surname
            # against decedent OR PR. Without pr the two disagree — an estate
            # deeded to a married daughter who is the PR reads HEIR TRANSFER
            # here and MARKET SALE there, which would suppress a hot lead.
            "flag", "county", "case", "decedent", "pr", "parcel", "address",
            "sale_date", "sale_price", "owner", "week", "filed", "uuid"],
            extrasaction="ignore")
        w.writeheader()
        w.writerows(results)
    print(f"\nWrote {out_path}")

    print("\n=== Coverage (queried / found in GIS / flagged sold) ===")
    for county, (q, f_, fl) in sorted(stats.items()):
        print(f"  {county:12s} {q:4d} / {f_:4d} / {fl}")
    flagged = [r for r in results if r["flag"]]
    for r in flagged:
        r["class"] = classify(r)
    print(f"\n=== FLAGGED: {len(flagged)} probate properties transferred since {args.since} ===")
    for r in flagged:
        print(f"  {r['case'] or 'NO CASE#'} | {r['county'].title():12s} | "
              f"{r['decedent'][:26]:26s} | {r['address'][:32]:32s} | "
              f"{r.get('sale_date')} | {str(r.get('sale_price') or ''):>9s} | {r['class']}")

    update_competitor_log(flagged)

    # Heir-transfer retarget CSV — new owner already holds clear title.
    heirs = [r for r in flagged if r["class"] == "HEIR TRANSFER"]
    if heirs:
        hpath = f"output/heir_transfers_{date.today().isoformat()}.csv"
        with open(hpath, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=[
                "case", "county", "decedent", "pr", "filed", "address",
                "sale_date", "owner", "week"], extrasaction="ignore")
            w.writeheader()
            w.writerows(heirs)
        print(f"Heir-transfer retarget list: {len(heirs)} -> {hpath}")


if __name__ == "__main__":
    main()
