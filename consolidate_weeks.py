"""Consolidate weekly NC Estates DataSift CSVs into one multi-tab workbook.

Mirrors the user's manual FTM-style workbook: one XLSX with a tab per
week (Week 20, Week 21, etc.), each tab styled identically (dark green
header, yellow banding, single-line rows, County dropdown, frozen
header). Tabs are sorted by ISO week ascending so the oldest week is
the leftmost tab and the newest is rightmost.

The workbook filename is `FTM_2026_NC_Estates_throughWeekN.xlsx` —
overwritten with each run so there's always one canonical workbook.

Usage:
    # Auto-pick the latest DataSift CSV per ISO week from output/
    python consolidate_weeks.py

    # Specify exact files (in any order)
    python consolidate_weeks.py \
        output/nc_estates_ftm_X_week20_datasift.csv \
        output/nc_estates_ftm_Y_week21_datasift.csv
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / "src"))

from nc_ftm_writer import FTM_COLUMNS, HIDDEN_FROM_WORKBOOK, NC_COUNTY_COLORS, NC_COUNTY_OPTIONS, workbook_header  # noqa: E402

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("consolidate")


# Same styling constants used in nc_ftm_writer / reenrich_ftm_executors
_HEADER_FILL = "1B5E20"
_BAND_FILL = "FFFDE7"
_DEFAULT_ROW_HEIGHT = 16
_COL_WIDTHS = {
    "File Date": 11, "County": 14, "Case No.": 18, "Deceased Owner": 32,
    "Personal Representative": 25, "First Name": 16, "Last Name": 18,
    "Mailing Address": 28, "Mailing City": 16, "Mailing State": 7, "Mailing Zip": 8,
    "Parcel ID": 16, "Property Address": 28, "Property City": 16,
    "Property State": 8, "Property Zip": 8, "Property use": 14,
    "Property Value": 14,
    "Notes": 40, "Beneficiaries": 80, "Phone 1": 14, "Phone 1 Tier": 12, "Tags": 26, "List": 10,
    "DM Name": 22, "DM Relationship": 14, "DM Phone": 14, "DM Phone Tier": 13, "DM Email": 26,
    "DM 2 Name": 22, "DM 2 Relationship": 14,
    "DM 3 Name": 22, "DM 3 Relationship": 14,
}


def load_csv(path: Path) -> list[dict]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    # Back-compat: pre-rename CSVs used "Executor Full Name".
    for r in rows:
        if "Executor Full Name" in r and "Personal Representative" not in r:
            r["Personal Representative"] = r.pop("Executor Full Name")
    return rows


def week_from_csv(rows: list[dict]) -> tuple[int, int]:
    """Pull the ISO (year, week) from the most common Tags value.

    Tag pattern: 'NC Estates Week N YYYY'. Falls back to (year, week) of
    the most common File Date if the tag is missing or unparseable.
    """
    tag_pat = re.compile(r"NC Estates Week (\d+) (\d{4})", re.IGNORECASE)
    tag_counts: Counter[tuple[int, int]] = Counter()
    for r in rows:
        m = tag_pat.search(r.get("Tags") or "")
        if m:
            tag_counts[(int(m.group(2)), int(m.group(1)))] += 1
    if tag_counts:
        return tag_counts.most_common(1)[0][0]
    # Fallback: derive from File Date
    date_counts: Counter[tuple[int, int]] = Counter()
    for r in rows:
        d = (r.get("File Date") or "").strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                dt = datetime.strptime(d, fmt)
                yw = dt.isocalendar()
                date_counts[(yw.year, yw.week)] += 1
                break
            except ValueError:
                continue
    if date_counts:
        return date_counts.most_common(1)[0][0]
    return (datetime.now().year, datetime.now().isocalendar().week)


def auto_pick_weekly_files() -> dict[tuple[int, int], Path]:
    """Pick the most recent CSV per (year, week) from output/.

    Considers THREE naming patterns (higher priority wins per week):
      *_weekN_datasift.csv                 (Step 4 output)
      *_weekN_<ts>_ecourts_backfilled.csv  (post-Step-5 eCourts name-search
                                            backfill — has additional Case
                                            No. fills that the datasift CSV
                                            doesn't have yet)
      *_weekN_dm_enriched.csv              (post-deep-prospect — adds DM
                                            Name/Phone/Email for no-contact
                                            rows; built FROM whichever of the
                                            above consolidate would have picked)

    The stage chain (datasift -> backfilled -> dm_enriched) is a superset ONLY
    within a single scrape run — each stage is built from the prior one of the
    SAME run. Across runs it is NOT: a fresh scrape's raw datasift can contain
    cases a stale-but-heavily-enriched file from an earlier run never saw
    (that's how an 11-row 6/29 dm_enriched shadowed a 17-row 6/30 datasift).
    But the nightly build ALSO re-buckets past weeks from accumulated raw data,
    producing a fresh-but-raw datasift with the SAME rows as an already-enriched
    prior file — there we must keep the enriched one.

    So per week: take the most-enriched file (newest among those) as the
    baseline — the canonical output of the latest COMPLETED run, carrying its
    DM enrichment. Override it ONLY when a strictly-newer scrape produced MORE
    rows (genuinely new cases in the in-progress week) — not when an older run
    merely carried a since-removed row. This keeps a fresh 17-row current week
    fresh while protecting a past week's 55-row enriched file from the nightly
    raw re-bucketing.
    """
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).parent))
    from iso_week_archive import get_archived_weeks
    archived = get_archived_weeks()

    def _row_count(fp: Path) -> int:
        try:
            with open(fp, newline="", encoding="utf-8-sig") as f:
                return max(0, sum(1 for _ in csv.reader(f)) - 1)  # minus header
        except OSError:
            return 0

    # Gather every candidate per week as (rows, priority, ts, path).
    # Priority 0 = datasift, 1 = ecourts_backfilled, 2 = dm_enriched.
    cands: dict[tuple[int, int], list[tuple[int, int, str, Path]]] = {}
    for pattern, priority in [
        ("nc_estates_ftm_*_week*_datasift.csv", 0),
        ("nc_estates_ftm_*_week*_ecourts_backfilled.csv", 1),
        ("nc_estates_ftm_*_week*_dm_enriched.csv", 2),
    ]:
        for fp in sorted(Path("output").glob(pattern)):
            m = re.search(r"_week(\d+)", fp.name)
            if not m:
                continue
            wk = int(m.group(1))
            if wk in archived:
                continue  # Archived weeks are excluded from the workbook
            ym = re.search(r"_(\d{4})-\d{2}-\d{2}_", fp.name)
            year = int(ym.group(1)) if ym else datetime.now().year
            # First timestamp in the name is the originating scrape's stamp,
            # carried through every enrichment stage.
            ts_m = re.search(r"(\d{4}-\d{2}-\d{2}_\d{6})", fp.name)
            ts = ts_m.group(1) if ts_m else ""
            cands.setdefault((year, wk), []).append((_row_count(fp), priority, ts, fp))

    picked: dict[tuple[int, int], Path] = {}
    for key, lst in cands.items():
        # Baseline = most-enriched file, newest among those. It's the canonical
        # output of the latest completed run and carries its DM enrichment.
        baseline = max(lst, key=lambda c: (c[1], c[2]))
        # Override only when a STRICTLY NEWER scrape yielded MORE rows — real new
        # cases, not an older run carrying a since-removed row. Most rows wins,
        # newest breaks ties.
        fuller = [c for c in lst if c[0] > baseline[0] and c[2] > baseline[2]]
        chosen = max(fuller, key=lambda c: (c[0], c[2])) if fuller else baseline
        picked[key] = chosen[3]
    return picked


def add_tab(wb, title: str, rows: list[dict]) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet(title=title[:31])  # XLSX tab name max 31 chars

    # Sort rows by County then Case No. so each county groups together
    rows = sorted(rows, key=lambda r: ((r.get("County") or "ZZZ"), r.get("Case No.", "")))

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=_HEADER_FILL, end_color=_HEADER_FILL, fill_type="solid")
    for c_idx, col_name in enumerate(FTM_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c_idx, value=workbook_header(col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # Per-county color tint (replaces old yellow-alternating band)
    county_fills = {
        c: PatternFill(start_color=h, end_color=h, fill_type="solid")
        for c, h in NC_COUNTY_COLORS.items()
    }
    multiline_cols = {"Notes", "Beneficiaries"}
    band_fill = PatternFill(start_color=_BAND_FILL, end_color=_BAND_FILL, fill_type="solid")
    for r_idx, r in enumerate(rows, start=2):
        # Per-county tint applied ONLY to the County column cell — per Oren
        # 2026-06-26, full-row tinting was visually noisy. Mirrors
        # nc_ftm_writer.write_ftm_xlsx so the consolidated workbook matches
        # the weekly file's look.
        row_fill = county_fills.get(r.get("County", ""))
        # Alternating band: odd data rows get a pale tint for readability;
        # skip the County column so its per-county color stands out.
        is_banded = (r_idx - 2) % 2 == 1
        for c_idx, col_name in enumerate(FTM_COLUMNS, start=1):
            val = r.get(col_name, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            # Multi-line columns (Notes, Beneficiaries) keep their newlines and
            # wrap vertically — one parcel/heir per line — per Oren. Matches
            # nc_ftm_writer.write_ftm_xlsx so weekly + consolidated look alike.
            if col_name in multiline_cols:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if col_name == "County" and row_fill:
                cell.fill = row_fill
            elif is_banded and col_name != "County":
                cell.fill = band_fill
        ws.row_dimensions[r_idx].height = _DEFAULT_ROW_HEIGHT

    county_col_idx = FTM_COLUMNS.index("County") + 1
    county_col_letter = get_column_letter(county_col_idx)
    county_formula = '"' + ",".join(NC_COUNTY_OPTIONS) + '"'
    dv = DataValidation(type="list", formula1=county_formula, allow_blank=True)
    dv.error = "Pick one of the 7 NC counties"
    dv.errorTitle = "Invalid county"
    dv.prompt = "Select a NC county"
    dv.promptTitle = "County"
    ws.add_data_validation(dv)
    dv.add(f"{county_col_letter}2:{county_col_letter}1048576")

    for c_idx, col_name in enumerate(FTM_COLUMNS, start=1):
        dim = ws.column_dimensions[get_column_letter(c_idx)]
        dim.width = _COL_WIDTHS.get(col_name, 14)
        if col_name in HIDDEN_FROM_WORKBOOK:
            dim.hidden = True
    ws.freeze_panes = "A2"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("csv_files", nargs="*",
                    help="DataSift CSV files (in any order). Omit to auto-pick "
                         "the latest per ISO week from output/.")
    ap.add_argument("--output", default=None,
                    help="Output XLSX path (default: FTM_<year>_NC_Estates_throughWeek<N>.xlsx)")
    args = ap.parse_args()

    if args.csv_files:
        files = [Path(p) for p in args.csv_files]
        per_week = {}
        for fp in files:
            rows = load_csv(fp)
            yr, wk = week_from_csv(rows)
            per_week[(yr, wk)] = (fp, rows)
    else:
        auto = auto_pick_weekly_files()
        per_week = {}
        for key, fp in auto.items():
            per_week[key] = (fp, load_csv(fp))

    if not per_week:
        logger.error("No DataSift CSV files found. Pass paths as args or run "
                     "prepare_for_datasift.py first.")
        sys.exit(1)

    logger.info("Consolidating %d weekly file(s):", len(per_week))
    for (yr, wk), (fp, rows) in sorted(per_week.items()):
        logger.info("  Week %d %d: %s (%d rows)", wk, yr, fp.name, len(rows))

    from openpyxl import Workbook
    wb = Workbook()
    # Workbook starts with a default empty "Sheet" — remove it
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Add tabs sorted by (year, week) ascending — oldest left, newest right
    for (yr, wk), (_fp, rows) in sorted(per_week.items()):
        title = f"Week {wk} {yr}"
        add_tab(wb, title, rows)
        logger.info("Added tab '%s' with %d rows", title, len(rows))

    # Default output filename: through the latest week
    latest_yr, latest_wk = sorted(per_week.keys())[-1]
    if args.output:
        out_path = Path(args.output)
    else:
        out_path = Path("output") / f"FTM_{latest_yr}_NC_Estates_throughWeek{latest_wk}.xlsx"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(out_path)
        logger.info("Wrote consolidated workbook: %s", out_path)
    except PermissionError:
        # The workbook is open in Excel (Windows locks it), so the nightly
        # rebuild can't overwrite it. Silently failing means Oren keeps looking
        # at a STALE workbook without knowing it (2026-07-17: a whole night's
        # audit fixes never reached the file he had open). Write a timestamped
        # fallback so the fresh data is never lost, and shout about it.
        alt = out_path.with_name(out_path.stem + "_LOCKED_REOPEN" + out_path.suffix)
        wb.save(alt)
        logger.warning(
            "Consolidated workbook %s is OPEN IN EXCEL — could not overwrite. "
            "Wrote fresh copy to %s instead. CLOSE the workbook and rename this "
            "file over it (or re-run consolidate_weeks.py) to get tonight's data.",
            out_path.name, alt.name)
        print(f"\n{'*' * 70}\n*** WORKBOOK LOCKED: {out_path.name} is open in Excel.\n"
              f"*** Tonight's data was saved to {alt.name} instead.\n"
              f"*** Close Excel and re-run consolidate_weeks.py to refresh the "
              f"real workbook.\n{'*' * 70}")


if __name__ == "__main__":
    main()
