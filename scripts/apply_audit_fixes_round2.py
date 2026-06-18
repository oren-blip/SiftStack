"""Re-apply audit fixes that polish/consolidate regressed.

Idempotent — checks current parcel before overwriting.

  - Kinney 26E000557-480: 145 Carriage (Jr's) -> 0 Carriage SR HEIRS vacant
  - Keller 26E000601-790: 0 Mt Hope (vacant) -> 3820 Mt Hope (SFR)
  - Mauser 26E000641-170: add a Note listing all 9 HEIRS parcels (polish picks
    one but doesn't know which is best — user picks during audit)
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook


WORKBOOK = Path("output/FTM_2026_NC_Estates_throughWeek23.xlsx")
TAB = "Week 23 2026"

MAUSER_PARCEL_NOTE = (
    "9 HEIRS parcels in Catawba GIS: "
    "1020 5TH ST NE Hickory (1.45ac), "
    "930 5TH ST NE Hickory (2.24ac), "
    "5TH ST NE Hickory (4.77ac), "
    "1341 9TH AVE NE Hickory (1.06ac), "
    "3679 ROCKY FORD RD Newton (481.39ac), "
    "plus 4 vacant lots. "
    "Pick best during audit."
)


def main() -> int:
    wb = load_workbook(WORKBOOK)
    ws = wb[TAB]
    header = [str(c.value or "").strip() for c in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(header)}

    def find_row(case_no: str) -> int | None:
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, col["Case No."]).value or "").strip() == case_no:
                return r
        return None

    def set_field(row: int, field: str, value):
        if field in col:
            ws.cell(row, col[field]).value = value

    # --- Kinney: swap to SR HEIRS vacant ---
    r = find_row("26E000557-480")
    if r:
        pid = str(ws.cell(r, col["Parcel ID"]).value or "")
        if "4712038255" not in pid:
            set_field(r, "Parcel ID", "4712038255.000")
            set_field(r, "Property Address", "0 CARRIAGE RD")
            set_field(r, "Property City", "Statesville")
            set_field(r, "Property State", "NC")
            set_field(r, "Property Zip", "28677")
            set_field(r, "Property use", "Vacant Land")
            set_field(r, "Property Value", 15190)
            print(f"Kinney row {r}: swapped {pid!r} -> 4712038255.000 (SR HEIRS vacant)")
        else:
            print(f"Kinney row {r}: already SR HEIRS (no change)")

    # --- Keller: swap to SFR ---
    r = find_row("26E000601-790")
    if r:
        pid = str(ws.cell(r, col["Parcel ID"]).value or "").strip()
        if "422 010" not in pid and "422010" not in pid.replace(" ", ""):
            set_field(r, "Parcel ID", "422 010")
            set_field(r, "Property Address", "3820 MT HOPE CHURCH RD")
            set_field(r, "Property City", "Salisbury")
            set_field(r, "Property State", "NC")
            set_field(r, "Property Zip", "28146")
            set_field(r, "Property use", "SFR")
            set_field(r, "Property Value", 158536)
            print(f"Keller row {r}: swapped {pid!r} -> 422 010 (3820 Mt Hope Church Rd SFR)")
        else:
            print(f"Keller row {r}: already SFR (no change)")

    # --- Mauser: add multi-parcel note (polish can't choose between 9) ---
    r = find_row("26E000641-170")
    if r and "Notes" in col:
        existing = str(ws.cell(r, col["Notes"]).value or "")
        if "HEIRS parcels" not in existing:
            new_note = (existing + " | " + MAUSER_PARCEL_NOTE).strip(" |")
            ws.cell(r, col["Notes"]).value = new_note
            print(f"Mauser row {r}: added 9-parcel note")
        else:
            print(f"Mauser row {r}: note already present")

    wb.save(WORKBOOK)
    print(f"\nSaved {WORKBOOK}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
