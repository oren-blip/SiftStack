"""Compare Oren's manual XLSX (with one tab per ISO week) against the
auto-generated multi-week workbook. Prints per-week + per-county
breakdowns and per-case diffs.

Usage:
    python scripts/diff_manual_xlsx_vs_workbook.py \\
        --manual "C:/Users/omark/Downloads/FTM_2026_NC Estates (2).xlsx" \\
        --weeks 22 23
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


def load_xlsx_tab(path: Path, tab: str) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    if tab not in wb.sheetnames:
        return []
    ws = wb[tab]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c or "").strip() for c in rows[0]]
    # Some tabs use "Estates Case No." — normalize to "Case No."
    header = ["Case No." if h == "Estates Case No." else h for h in header]
    out = []
    for r in rows[1:]:
        d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        cn = str(d.get("Case No.", "") or "").strip()
        if cn:
            out.append(d)
    return out


def key(row: dict) -> tuple[str, str]:
    return (
        str(row.get("County", "") or "").strip().title(),
        str(row.get("Case No.", "") or "").strip().upper(),
    )


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--manual", required=True)
    ap.add_argument("--workbook", default="output/FTM_2026_NC_Estates_throughWeek23.xlsx")
    ap.add_argument("--weeks", nargs="+", default=["22", "23"])
    args = ap.parse_args()

    manual_path = Path(args.manual)
    wb_path = Path(args.workbook)

    for wk in args.weeks:
        manual = load_xlsx_tab(manual_path, f"Week {wk}")
        wb_rows = load_xlsx_tab(wb_path, f"Week {wk} 2026")
        manual_map = {key(r): r for r in manual}
        wb_map = {key(r): r for r in wb_rows}

        in_both = sorted(set(manual_map) & set(wb_map))
        manual_only = sorted(set(manual_map) - set(wb_map))
        wb_only = sorted(set(wb_map) - set(manual_map))

        print(f"\n{'=' * 75}")
        print(f"WEEK {wk}: manual={len(manual)} | workbook={len(wb_rows)}")
        print(f"{'=' * 75}")
        print(f"  In both: {len(in_both)} | Manual-only: {len(manual_only)} | Workbook-only: {len(wb_only)}")

        # Per-county breakdown
        all_counties = sorted({k[0] for k in set(manual_map) | set(wb_map)})
        print(f"\n  {'County':14} {'Manual':>7} {'Wkbk':>7} {'Both':>5} {'M-only':>7} {'W-only':>7}")
        for c in all_counties:
            m = sum(1 for k in manual_map if k[0] == c)
            w = sum(1 for k in wb_map if k[0] == c)
            b = sum(1 for k in in_both if k[0] == c)
            mo = sum(1 for k in manual_only if k[0] == c)
            wo = sum(1 for k in wb_only if k[0] == c)
            print(f"  {c:14} {m:>7} {w:>7} {b:>5} {mo:>7} {wo:>7}")

        if manual_only:
            print(f"\n  --- MANUAL ONLY (scraper missed these) ---")
            for k in manual_only:
                r = manual_map[k]
                print(f"    {k[1]:18} {k[0]:12} | {r.get('Deceased Owner','')}")

        if wb_only:
            print(f"\n  --- WORKBOOK ONLY (extras) ---")
            for k in wb_only:
                r = wb_map[k]
                print(f"    {k[1]:18} {k[0]:12} | {r.get('Deceased Owner',''):30} PR: {r.get('Personal Representative','')}")

        # Property/parcel disagreements on matched rows
        disagreements = 0
        for k in in_both:
            m = manual_map[k]
            w = wb_map[k]
            diffs = []
            for field in ("Deceased Owner", "Personal Representative", "Property Address", "Parcel ID"):
                mv = str(m.get(field, "") or "").strip().lower()
                wv = str(w.get(field, "") or "").strip().lower()
                if mv and wv and mv != wv:
                    diffs.append(f"{field}: M={m.get(field)!r} | W={w.get(field)!r}")
            if diffs:
                disagreements += 1
                if disagreements <= 8:
                    print(f"\n  --- {k[1]} {k[0]} disagrees: ---")
                    for d in diffs:
                        print(f"      {d[:160]}")
        if disagreements > 8:
            print(f"\n  ... and {disagreements - 8} more rows with field disagreements")

    return 0


if __name__ == "__main__":
    sys.exit(main())
