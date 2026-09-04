"""Daily NC pipeline report.

Runs as the final step of the 6 PM daily build. Reads the day's outputs,
builds a plain-text report, writes it to output/daily_report_YYYY-MM-DD.txt,
and optionally emails it via Gmail/Workspace SMTP.

Configured via .env (already loaded by the pipeline):
  SMTP_USER       — sender Gmail/Workspace address (e.g. oren@remedihomesolutions.com)
  SMTP_PASS       — Google app password (NOT your regular password)
  REPORT_TO       — comma-separated recipient list
  SMTP_HOST       — defaults to smtp.gmail.com
  SMTP_PORT       — defaults to 587 (TLS)

To run by hand:
    .venv\\Scripts\\python.exe scripts\\daily_report.py
    .venv\\Scripts\\python.exe scripts\\daily_report.py --no-email
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import smtplib
import subprocess
import sys
from collections import Counter
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from pathlib import Path

from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


OUTPUT = Path("output")
LOGS = Path("logs")
# Running deep-prospecting ledger (repo root, committed). Every DP run —
# whether it produced a research doc or just an address/phone fix — appends a
# row here; the report's DEEP PROSPECTING section reads it (Oren, 2026-08-12).
DP_LOG = Path("dp_log.csv")


def read_dp_log() -> list[dict]:
    if not DP_LOG.exists():
        return []
    try:
        with DP_LOG.open(newline="", encoding="utf-8-sig") as f:
            return [r for r in csv.DictReader(f) if (r.get("Case No.") or "").strip()]
    except OSError:
        return []


# ── Discovery ────────────────────────────────────────────────────────


def find_latest_workbook() -> Path | None:
    candidates = sorted(OUTPUT.glob("FTM_*_NC_Estates_throughWeek*.xlsx"))
    return candidates[-1] if candidates else None


def find_latest_datasift(week_n: int) -> Path | None:
    files = sorted(OUTPUT.glob(f"nc_estates_ftm_*_week{week_n}_datasift.csv"))
    return files[-1] if files else None


def find_yesterdays_datasift(week_n: int, today_path: Path) -> Path | None:
    """Datasift from a prior run — used to compute 'new today'."""
    files = sorted(OUTPUT.glob(f"nc_estates_ftm_*_week{week_n}_datasift.csv"))
    older = [f for f in files if f != today_path]
    return older[-1] if older else None


def find_latest_heir_transfer(week_n: int) -> Path | None:
    files = sorted(OUTPUT.glob(f"heir_transfer_review_week{week_n}_*.xlsx"))
    return files[-1] if files else None


# ── Parse outputs ────────────────────────────────────────────────────


def read_workbook_cases(wb_path: Path) -> tuple[int, list[dict]]:
    wb = load_workbook(wb_path, read_only=True)
    rows: list[dict] = []
    week_n = 0
    for sn in wb.sheetnames:
        m = re.search(r"Week (\d+)", sn)
        if m:
            week_n = max(week_n, int(m.group(1)))
        ws = wb[sn]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        header = [str(c or "").strip() for c in all_rows[0]]
        for r in all_rows[1:]:
            d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
            if (d.get("Case No.") or "").strip():
                d["_tab"] = sn
                rows.append(d)
    return week_n, rows


def read_csv_cases(path: Path) -> list[dict]:
    if not path or not path.exists():
        return []
    with path.open(encoding="utf-8-sig") as f:
        return [r for r in csv.DictReader(f) if (r.get("Case No.") or "").strip()]


def read_heir_transfer(path: Path | None) -> list[dict]:
    if not path or not path.exists():
        return []
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c or "").strip() for c in rows[0]]
    out = []
    for r in rows[1:]:
        d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        if d.get("Case No."):
            out.append(d)
    return out


_DROP_PATTERNS = {
    "no_parcel":     re.compile(r"Dropped \(no parcel\): (\d+)"),
    "over_500k":     re.compile(r"Dropped >\$500K: (\d+)"),
    "heir_occupied": re.compile(r"Dropped heir-occupied: (\d+)"),
    "commercial":    re.compile(r"Dropped commercial: (\d+)"),
    "archive_dupe":  re.compile(r"Dropped archive-duplicate rows: (\d+)"),
    "non_pending":   re.compile(r"Dropped non-Pending: (\d+)"),
    # Real events print "REPICK {county}/{decedent}: ..." — require the slash so
    # the report's own "Smart-picker REPICK events: N" line (the report text is
    # echoed into nc_daily_run.log by the nightly bat) doesn't count itself.
    "repick":        re.compile(r"REPICK \S+/"),
    "refound":       re.compile(r"Re-found correct parcels: (\d+)"),
    "under_min_value": re.compile(r"Dropped under-min-value: (\d+)"),
}


def parse_polish_log(log_path: Path, since: datetime | None = None) -> dict:
    """Parse polish log for drop counts + matcher events. Robust to UTF-8/UTF-16."""
    out = {k: 0 for k in _DROP_PATTERNS}
    if not log_path.exists():
        return out
    # Try UTF-8 first, fall back to UTF-16
    text = ""
    for enc in ("utf-8", "utf-16-le", "utf-16"):
        try:
            text = log_path.read_text(encoding=enc, errors="ignore")
            if text:
                break
        except (OSError, UnicodeDecodeError):
            continue
    if since is not None:
        # Keep only lines from today's run boundary onward
        cutoff = since.strftime("%H:%M")
        marker = f"=== Daily run started"
        if marker in text:
            text = text[text.rindex(marker):]
    for key, pat in _DROP_PATTERNS.items():
        matches = pat.findall(text)
        if key == "repick":
            out[key] = len(matches)
            continue
        # Sum all matches — polish processes multiple weeks in one run, each
        # producing its own "Dropped X: N" line. Total across all weeks.
        total = 0
        for m in matches:
            try:
                total += int(m)
            except ValueError:
                total += 1
        out[key] = total
    return out


def pipeline_runtime_min(log_path: Path) -> float | None:
    """Extract today's daily run wall-clock from the log's start/end markers."""
    if not log_path.exists():
        return None
    text = ""
    for enc in ("utf-8", "utf-16-le", "utf-16"):
        try:
            text = log_path.read_text(encoding=enc, errors="ignore")
            if text:
                break
        except (OSError, UnicodeDecodeError):
            continue
    starts = re.findall(r"Daily run started.*?(\d{2}):(\d{2}):(\d{2})", text)
    # Only accept a done marker AFTER the last start — a deadline-killed run
    # never prints one, and pairing it with an older run's end fabricates a
    # plausible-looking but bogus runtime.
    last_start = text.rfind("Daily run started")
    ends = re.findall(r"Daily run done.*?(\d{2}):(\d{2}):(\d{2})",
                      text[last_start:]) if last_start >= 0 else []
    if not starts or not ends:
        return None
    h1, m1, s1 = map(int, starts[-1])
    h2, m2, s2 = map(int, ends[-1])
    return ((h2 * 3600 + m2 * 60 + s2) - (h1 * 3600 + m1 * 60 + s1)) / 60.0


# ── Tonight's run — summary (Oren, 2026-08-31: "send a summary like this
# with the daily report") ─────────────────────────────────────────────
# Reads tonight's slice of logs/nc_daily_run.log and turns the step-by-step
# chatter into the same shape as a hand-written run recap: headline, per-week
# polish, DataSift upload, deep prospecting, sold sweep, cost, and a
# NEEDS A HAND list. Everything is best-effort — a missing line just drops
# its sentence, never the section.


def _read_log_text(log_path: Path) -> str:
    if not log_path.exists():
        return ""
    for enc in ("utf-8", "utf-16-le", "utf-16"):
        try:
            text = log_path.read_text(encoding=enc, errors="ignore")
            if text:
                return text
        except (OSError, UnicodeDecodeError):
            continue
    return ""


def tonight_log_slice(log_path: Path) -> str:
    """Text from the last '=== Daily run started' marker onward, cut before
    any echoed daily report (the bat tees this script's own output into the
    same log, so a manual re-run must not read last night's report as data)."""
    text = _read_log_text(log_path)
    marker = "=== Daily run started"
    if marker not in text:
        return ""
    return text[text.rindex(marker):]


_TN_WEEK_HDR = re.compile(r"=== WEEK(\d+): (\S+) ===")


def parse_tonight(text: str, today: date | None = None) -> dict:
    """Pull the numbers a human would quote from tonight's log slice."""
    today = today or date.today()
    t: dict = {"ok": bool(text)}
    if not text:
        return t
    # A finished run prints its done stamp AFTER this report is echoed into
    # the log; grab it before cutting the echo away (manual re-runs only —
    # inside the nightly the run is still going when this parses).
    m = re.search(r"Daily run done \w+ \d{2}/\d{2}/\d{4} (\d{2}):(\d{2}):(\d{2})", text)
    t["done_hms"] = tuple(int(x) for x in m.groups()) if m else None
    echo = text.find("NC Probate Pipeline")
    if echo > 0:
        text = text[:echo]

    m = re.search(r"Daily run started \w+ (\d{2})/(\d{2})/(\d{4}) (\d{2}):(\d{2}):(\d{2})", text)
    if m:
        mo, dd, yy, hh, mi, ss = map(int, m.groups())
        t["start"] = datetime(yy, mo, dd, hh, mi, ss)
    t["budget_hit"] = "*** BUDGET" in text
    t["tracebacks"] = len(re.findall(r"^Traceback \(most recent call last\)", text, re.M))

    m = re.search(r"Summary: (\d+) pass, (\d+) fail, (\d+) skip", text)
    if m:
        t["smoke"] = tuple(int(x) for x in m.groups())

    # Raw scrape rows landed tonight (files stamped with today's date).
    raw_today: Counter = Counter()
    for d, wk, n in re.findall(
            r"nc_estates_ftm_(\d{4}-\d{2}-\d{2})_\d+\.csv -> week (\d+) \d{4} \((\d+) rows\)", text):
        if d == today.isoformat():
            raw_today[int(wk)] += int(n)
    t["raw_today"] = dict(raw_today)

    # Per-week polish blocks.
    weeks: dict[int, dict] = {}
    hdrs = list(_TN_WEEK_HDR.finditer(text))
    for i, h in enumerate(hdrs):
        wk = int(h.group(1))
        blk = text[h.end(): hdrs[i + 1].start() if i + 1 < len(hdrs) else len(text)]
        # The polish output ends where the next pipeline step starts.
        for stop in ("LLM USAGE this process", "[budget]"):
            j = blk.find(stop)
            if j > 0:
                blk = blk[:j]
        w: dict = {}
        m = re.search(r"Loaded (\d+) rows", blk)
        w["loaded"] = int(m.group(1)) if m else 0
        w["pr_backfilled"] = len(re.findall(r"^\s+PR backfill ", blk, re.M))
        w["dm_mailing_found"] = len(re.findall(r"DM-MAILING FOUND", blk))
        m = re.search(r"parties backfill: skipped (\d+) case", blk)
        w["parties_skipped"] = int(m.group(1)) if m else 0
        m = re.search(r"parties backfill: (\d+) rows need the court; doing (\d+) this run", blk)
        w["parties_pending"] = (int(m.group(1)), int(m.group(2))) if m else None
        m = re.search(r"Rows in: (\d+)\s+Dropped \(no parcel\): (\d+)\s+DM-promoted: (\d+)"
                      r"\s+Beneficiary-promoted: (\d+)\s+Generic Heirs-of: (\d+).*?Out: (\d+)", blk)
        if m:
            (w["step4_in"], w["no_parcel"], w["dm_promoted"], w["ben_promoted"],
             w["heirs_of"], w["step4_out"]) = (int(x) for x in m.groups())
        w["heir_occupied"] = sum(int(x) for x in re.findall(r"Dropped heir-occupied: (\d+)", blk))
        m = re.search(r"Dropped via Zillow status: (\d+)", blk)
        w["zillow_drops"] = int(m.group(1)) if m else 0
        w["zillow_lines"] = [
            (cty, dec.strip(), addr, why) for cty, dec, addr, why in
            re.findall(r"ZILLOW DROP (\w+)/(.+?): '(.*?)'\s*\W*\s*(\w+)", blk)]
        m = re.search(r"Same-property rows merged: (\d+)", blk)
        w["spouse_merged"] = int(m.group(1)) if m else 0
        m = re.search(r"Rows held from upload: (\d+)", blk)
        w["held"] = int(m.group(1)) if m else 0
        w["held_lines"] = re.findall(r"OCCUPIED-HOLD (\w+)/(\S+): '(.+?)' at the home", blk)
        m = re.search(r"_upload\.csv\s+\((\d+) rows", blk)
        w["upload_ready"] = int(m.group(1)) if m else None
        m = re.search(r"Rows updated from fetched case-doc cache: (\d+)", blk)
        w["casedoc_updates"] = int(m.group(1)) if m else 0
        weeks[wk] = w
    t["weeks"] = weeks

    # DataSift upload (auto_upload_netnew + upload_netnew_datasift).
    up: dict = {"uploaded": {}, "skipped": []}
    for wk, n in re.findall(r"auto-upload: week (\d+): uploading (\d+) row", text):
        up["uploaded"][int(wk)] = int(n)
    for wk in re.findall(r"auto-upload: week (\d+): \S+ has 0 rows -- skip", text):
        up["skipped"].append(int(wk))
    up["committed"] = "Upload committed (wizard closed)" in text
    up["skip_trace"] = "Skip trace started" in text
    up["tiers"] = "Phone tags uploaded" in text
    up["touches"] = "Text touches uploaded" in text
    up["tagpush"] = len(re.findall(r"Upload the file: tagpush_", text))
    up["disabled"] = "NC_AUTO_UPLOAD=0" in text
    t["upload"] = up

    # Deep prospecting (nc_deep_prospect + tracers).
    dp: dict = {}
    m = re.search(r"Tracerfy batch complete: (\d+)/(\d+) matched, (\d+) phones, (\d+) emails, \$([\d.]+)", text)
    if m:
        dp["tracerfy"] = (int(m.group(1)), int(m.group(2)), int(m.group(3)),
                          int(m.group(4)), float(m.group(5)))
    m = re.search(r"Enformion filled phones for (\d+) row\(s\).*?billed \$([\d.]+)", text)
    if m:
        dp["enformion"] = (int(m.group(1)), float(m.group(2)))
    dp["enformion_rejected"] = len(re.findall(r"Enformion match rejected", text))
    m = re.search(r"Trestle scoring (\d+) unique phones across (\d+) records \(~\$([\d.]+)\)", text)
    if m:
        dp["trestle"] = (int(m.group(1)), int(m.group(2)), float(m.group(3)))
    m = re.search(r"Deep prospecting complete: (\d+) DM names found, (\d+) phones traced", text)
    if m:
        dp["complete"] = (int(m.group(1)), int(m.group(2)))
    t["dp"] = dp

    # LLM spend — the line prints twice per process (stdout + logger), so
    # dedupe consecutive identical readings before summing.
    llm_calls = 0
    llm_cost = 0.0
    last = None
    for calls, cost in re.findall(r"LLM USAGE this process: (\d+) Anthropic calls, est \$([\d.]+)", text):
        if (calls, cost) == last:
            continue
        last = (calls, cost)
        llm_calls += int(calls)
        llm_cost += float(cost)
    t["llm"] = (llm_calls, llm_cost)

    # Sold sweep (sold_audit + push_sold_tags).
    sold: dict = {}
    m = re.search(r"Rows with parcels: (\d+)", text)
    if m:
        sold["parcels"] = int(m.group(1))
    m = re.search(r"Unique \(county, parcel\) incl\. CRM legacy: (\d+)", text)
    if m:
        sold["checked"] = int(m.group(1))
    m = re.search(r"=== FLAGGED: (\d+) probate properties transferred since (\S+)", text)
    if m:
        sold["flagged"] = int(m.group(1))
        sold["since"] = m.group(2)
    m = re.search(r"(\d+) flagged sales: (\d+) to suppress, (\d+) report-only", text)
    if m:
        sold["to_suppress"] = int(m.group(2))
        sold["report_only"] = int(m.group(3))
    sold["tagged"] = re.findall(r"TAGGED \(verified\)\s+(\S+) \| (.+)", text)
    sold["not_in_crm"] = len(re.findall(r"^\s+NOT IN CRM", text, re.M))
    sold["failed"] = re.findall(r"^\s+FAILED\s+(\S+) \| (.+)", text, re.M)
    sold["would_tag"] = len(re.findall(r"^\s+would tag\s", text, re.M))
    m = re.search(r"CRM legacy sweep: (.+)", text)
    sold["legacy"] = m.group(1).strip() if m else ""
    sold["ran"] = bool(sold.get("parcels") or "No flagged sales" in text)
    sold["disabled"] = "NC_SOLD_SWEEP=0" in text
    t["sold"] = sold

    # Post-upload sweeps.
    m = re.search(r"Tier backfill sweep failed \((.*)\)", text)
    t["tier_fail"] = ("DataSift dropped the connection" if m and "ConnectionReset" in m.group(1)
                      else (m.group(1)[:80] if m else ""))
    m = re.search(r"UNTIERED: (\d+) phone\(s\) on (\d+) record\(s\)", text)
    if m:
        t["untiered"] = (int(m.group(1)), int(m.group(2)))
    m = re.search(r"Text touches: (\d+) record\(s\) OK, (\d+) MISSING, (\d+) DRIFT", text)
    if m:
        t["touch_audit"] = tuple(int(x) for x in m.groups())

    # Owner/mailing drift (audit_owner_mailing_drift).
    t["drift"] = re.findall(
        r"^\s*(?:\d\d:\d\d:\d\d\s+)?(\S+): '(.+?)' -> '(.+?)', mailing still '(.+?)'", text, re.M)

    # Warnings — grouped so the report says "Lincoln GIS x2", not 2 raw lines.
    warn: Counter = Counter()
    for ln in re.findall(r"^.*\[WARNING\].*$", text, re.M):
        if "Tier backfill sweep failed" in ln:
            continue  # surfaced as its own NEEDS A HAND item
        if "ArcGIS error at" in ln:
            mm = re.search(r"https?://[^/]*?(\w+)county", ln, re.I)
            warn[f"{(mm.group(1).title() if mm else 'a county')} GIS query error"] += 1
        elif "Connection aborted" in ln or "ConnectionReset" in ln:
            mm = re.search(r"https?://[^/]*?(\w+)county", ln, re.I)
            who = (mm.group(1).title() + " GIS" if mm
                   else ("DataSift" if "netnew_upload" in ln else "a remote host"))
            warn[f"{who} dropped the connection"] += 1
        elif "No manual archive found" in ln:
            continue  # expected since manual pulls stopped (Week 32)
        else:
            mm = re.search(r"\[WARNING\]\s*(?:\S+:\s*)?(.{0,70})", ln)
            warn[(mm.group(1).strip() if mm else ln[:70])] += 1
    t["warnings"] = warn
    return t


def _money(x: float) -> str:
    return f"${x:.2f}"


def _tonight_costs(t: dict) -> tuple[float, list[str]]:
    """Tonight's paid-service burn as (total, ['Tracerfy $0.10', ...]).
    Shared by the text summary and the HTML email so the two never disagree."""
    dp = t.get("dp") or {}
    costs: list[str] = []
    total = 0.0
    if "tracerfy" in dp:
        costs.append(f"Tracerfy {_money(dp['tracerfy'][4])}")
        total += dp["tracerfy"][4]
    if "enformion" in dp:
        costs.append(f"Enformion {_money(dp['enformion'][1])}")
        total += dp["enformion"][1]
    if "trestle" in dp:
        costs.append(f"Trestle ~{_money(dp['trestle'][2])}")
        total += dp["trestle"][2]
    calls, llm = t.get("llm", (0, 0.0))
    if calls:
        costs.append(f"LLM {_money(llm)} / {calls} calls")
        total += llm
    return total, costs


def tonight_summary_section(t: dict, condensed: bool = False) -> tuple[list[str], list[str]]:
    """Render TONIGHT'S RUN — SUMMARY. Returns (lines, needs_a_hand) so the
    caller can also fold the hand-items into WHAT TO WORK ON NEXT."""
    lines: list[str] = []
    hand: list[str] = []
    if not t.get("ok"):
        return lines, hand
    lines.append("TONIGHT'S RUN — SUMMARY")

    # Headline: when it ran, how long, whether it finished clean.
    start = t.get("start")
    if start:
        end = datetime.now()
        if t.get("done_hms"):
            hh, mi, ss = t["done_hms"]
            end = start.replace(hour=hh, minute=mi, second=ss)
            if end < start:
                end += timedelta(days=1)
        mins = int((end - start).total_seconds() // 60)
        span = f"{mins // 60}h {mins % 60:02d}m" if mins >= 60 else f"{mins} min"
        if t.get("budget_hit"):
            state = "*** KILLED BY THE TIME BUDGET — back half did not run ***"
        elif t.get("tracebacks"):
            state = f"{t['tracebacks']} step(s) crashed (see log)"
        else:
            state = "finished clean"
        lines.append(f"  Ran {start:%H:%M} -> {end:%H:%M} ({span}), {state}.")
        if t.get("budget_hit"):
            hand.append("Tonight's run hit the 4.5h time budget — the back half (upload/DP/sold sweep) "
                        "did not run; grep '*** BUDGET' in logs/nc_daily_run.log")
    raw = t.get("raw_today") or {}
    smoke = t.get("smoke")
    bits = []
    if raw:
        bits.append(f"Scraped {sum(raw.values())} raw row(s) today (" +
                    ", ".join(f"wk{w}: {n}" for w, n in sorted(raw.items())) + ")")
    elif start:
        bits.append("Scrape landed 0 new rows today")
    if smoke:
        p, f, _ = smoke
        bits.append(f"GIS smoke test {p}/{p + f} pass" + (" — SEE OUTAGE BELOW" if f else ""))
    if bits:
        lines.append("  " + "; ".join(bits) + ".")
    lines.append("")

    # Per-week polish, newest week first.
    weeks = t.get("weeks") or {}
    up = t.get("upload") or {}
    for i, wk in enumerate(sorted(weeks, reverse=True)):
        w = weeks[wk]
        label = "new" if i == 0 else "re-polish"
        ready = w.get("upload_ready")
        if ready is None:
            ready = max(0, w.get("step4_out", 0) - w.get("zillow_drops", 0) - w.get("held", 0))
        kept = ready + w.get("held", 0)
        head = f"  Week {wk} ({label}):"
        parts = [f"{w.get('loaded', 0)} in -> {kept} kept"
                 + (f" ({ready} mail-ready, {w['held']} held)" if w.get("held") else "")]
        n_up = up.get("uploaded", {}).get(wk)
        if n_up:
            parts.append(f"{n_up} uploaded to DataSift tonight")
        elif wk in up.get("skipped", []):
            parts.append("0 net-new to upload (all already in DataSift)")
        lines.append(f"{head:23}{', '.join(parts)}.")
        detail = []
        if w.get("pr_backfilled"):
            detail.append(f"{w['pr_backfilled']} PRs from the court")
        if w.get("dm_mailing_found"):
            detail.append(f"{w['dm_mailing_found']} contact addresses")
        if w.get("dm_promoted"):
            detail.append(f"{w['dm_promoted']} heirs promoted to contact")
        if w.get("heirs_of"):
            detail.append(f'{w["heirs_of"]} still "Heirs of"')
        if detail:
            lines.append(f"{'':23}found: " + ", ".join(detail))
        drops = []
        if w.get("no_parcel"):
            drops.append(f"{w['no_parcel']} no parcel")
        if w.get("heir_occupied"):
            drops.append(f"{w['heir_occupied']} heir-occupied")
        if w.get("zillow_drops"):
            drops.append(f"{w['zillow_drops']} Zillow (listed/sold)")
        if w.get("spouse_merged"):
            drops.append(f"{w['spouse_merged']} spouse pair merged")
        if w.get("held"):
            drops.append(f"{w['held']} occupied-hold (workbook only, not uploaded)")
        if drops:
            lines.append(f"{'':23}dropped: " + ", ".join(drops))
        if w.get("parties_skipped"):
            lines.append(f"{'':23}court hasn't indexed Parties for {w['parties_skipped']} "
                         "case(s) yet — PR/heirs retry tomorrow")
        elif w.get("parties_pending"):
            need, did = w["parties_pending"]
            lines.append(f"{'':23}Parties backfill: {did} of {need} pending done tonight, rest tomorrow")
        if not condensed:
            for cty, dec, addr, why in w.get("zillow_lines", []):
                lines.append(f"{'':25}Zillow drop  {cty:11} {dec[:26]:26} {addr[:28]:28} {why}")
            for cty, case, who in w.get("held_lines", []):
                lines.append(f"{'':25}Held         {case:16} {cty:11} '{who}' lives at the property")
    if weeks:
        lines.append("")

    # DataSift.
    if up.get("disabled"):
        lines.append("  DataSift:     auto-upload OFF (NC_AUTO_UPLOAD=0) — nothing pushed tonight.")
    elif up.get("uploaded"):
        n = sum(up["uploaded"].values())
        steps = [f"uploaded {n} record(s)"]
        steps.append("skip trace started" if up.get("skip_trace") else "skip trace NOT started")
        steps.append("dial tiers tagged" if up.get("tiers") else "dial tiers NOT tagged")
        steps.append("text touches written" if up.get("touches") else "text touches NOT written")
        lines.append("  DataSift:     " + ", ".join(steps) + ".")
        if not up.get("skip_trace"):
            hand.append(f"DataSift skip trace did not start on tonight's {n} upload(s) — "
                        "trace them by hand (Send To -> Skip Trace on tag 'NC Upload "
                        f"{date.today().isoformat()}')")
    elif weeks:
        lines.append("  DataSift:     nothing net-new to upload tonight.")

    # Deep prospecting.
    dp = t.get("dp") or {}
    if dp:
        parts = []
        if "tracerfy" in dp:
            hit, tot, ph, em, cost = dp["tracerfy"]
            parts.append(f"Tracerfy {hit}/{tot} matched ({ph} phones) {_money(cost)}")
        if "enformion" in dp:
            n, cost = dp["enformion"]
            rej = dp.get("enformion_rejected", 0)
            parts.append(f"Enformion filled {n}"
                         + (f" (rejected {rej} wrong-town match(es))" if rej else "")
                         + f" {_money(cost)}")
        if "trestle" in dp:
            ph, rec, cost = dp["trestle"]
            parts.append(f"Trestle scored {ph} phones ~{_money(cost)}")
        if parts:
            lines.append("  Deep prosp.:  " + "; ".join(parts) + ".")

    # Sold sweep.
    sold = t.get("sold") or {}
    if sold.get("disabled"):
        lines.append("  Sold sweep:   OFF (NC_SOLD_SWEEP=0).")
    elif sold.get("ran"):
        checked = sold.get("checked") or sold.get("parcels") or 0
        flagged = sold.get("flagged", 0)
        sup = sold.get("to_suppress", 0)
        rep = sold.get("report_only", 0)
        n_tag = len(sold.get("tagged", []))
        n_fail = len(sold.get("failed", []))
        line = (f"  Sold sweep:   {checked} parcels checked -> {flagged} transfers since "
                f"{sold.get('since', '?')}: ")
        if flagged:
            line += (f"{sup} real sale(s) ({n_tag} tagged Sold, {sold.get('not_in_crm', 0)} not in CRM"
                     + (f", {n_fail} FAILED" if n_fail else "")
                     + f"), {rep} heir transfer(s) kept as leads.")
        else:
            line += "none."
        lines.append(line)
        if sold.get("would_tag"):
            lines.append(f"{'':16}(dry run — {sold['would_tag']} would have been tagged; nothing written)")
        if sold.get("tagged"):
            show = sold["tagged"] if not condensed else sold["tagged"][:6]
            more = len(sold["tagged"]) - len(show)
            lines.append(f"{'':16}tagged: " + "; ".join(f"{c} {a.strip()[:28]}" for c, a in show)
                         + (f"; +{more} more" if more else ""))
        if n_fail:
            hand.append(f"Sold tag FAILED on {n_fail} record(s): " +
                        ", ".join(c for c, _ in sold["failed"][:5]) + " — tag them by hand")
        if sold.get("legacy") and "skipped" not in sold["legacy"]:
            lines.append(f"{'':16}pre-Week-24 CRM records were swept too (weekly pass)")
    elif weeks:
        lines.append("  Sold sweep:   did not run tonight.")

    # Cost.
    total, costs = _tonight_costs(t)
    if costs:
        lines.append(f"  Cost tonight: ~{_money(total)} ({', '.join(costs)}).")
    lines.append("")

    # Needs a hand — things only a person can close.
    if t.get("tier_fail"):
        hand.append(f"Nightly dial-tier sweep died mid-run ({t['tier_fail']}) — the 07:00 "
                    "'SiftStack Tier Sweep' task will cover it; check logs/trestle_sweep.log "
                    "tomorrow, or run  python trestle_api_backfill.py --apply --max-cost 2 --headless")
    for case, was, now, mailing in t.get("drift", []):
        hand.append(f"{case} was renamed '{was}' -> '{now}' but still MAILS {mailing} "
                    "(the old heir's address) — fix the mailing address in DataSift "
                    "(detail: output/owner_mailing_drift.csv)")
    warn = t.get("warnings") or Counter()
    if warn:
        top = "; ".join(f"{k} x{n}" for k, n in warn.most_common(4))
        lines.append(f"  Blips: {sum(warn.values())} warning(s) — {top}. One-offs unless "
                     "the same one shows up again tomorrow.")
    if hand:
        lines.append("  NEEDS A HAND")
        for h in hand:
            lines.append(f"  >> {h}")
    lines.append("")
    return lines, hand


# ── Render ───────────────────────────────────────────────────────────


def _row_key(r: dict) -> tuple[str, str]:
    return ((r.get("County") or "").strip(), (r.get("Case No.") or "").strip().upper())


def compute_diffs(today_rows: list[dict], yesterday_rows: list[dict]) -> dict:
    """Compare two polished CSVs (same week, different runs).

    Returns dict with: new, dropped, parcel_changed, reason_today.
    - new:            rows in today but not yesterday
    - dropped:        rows in yesterday but not today (regression signal —
                      polish dropped something it had before)
    - parcel_changed: rows present in both but with a different Parcel ID
    - reason_today:   Counter of Match Reason tags from today's run
    """
    by_yest = {_row_key(r): r for r in yesterday_rows}
    by_today = {_row_key(r): r for r in today_rows}

    new = [r for k, r in by_today.items() if k not in by_yest]
    dropped = [r for k, r in by_yest.items() if k not in by_today]

    parcel_changed: list[dict] = []
    for k, r_today in by_today.items():
        r_yest = by_yest.get(k)
        if r_yest is None:
            continue
        p_y = (r_yest.get("Parcel ID") or "").strip()
        p_t = (r_today.get("Parcel ID") or "").strip()
        if p_y and p_t and p_y != p_t:
            parcel_changed.append({
                "case_no": r_today.get("Case No.", ""),
                "county":  r_today.get("County", ""),
                "decedent": r_today.get("Deceased Owner", ""),
                "parcel_yesterday": p_y,
                "parcel_today":     p_t,
                "addr_yesterday":   (r_yest.get("Property Address") or "").strip(),
                "addr_today":       (r_today.get("Property Address") or "").strip(),
            })

    reason_today: Counter[str] = Counter()
    for r in today_rows:
        raw = (r.get("Match Reason") or "").strip()
        if not raw:
            reason_today["(scrape direct)"] += 1
            continue
        for tag in (t.strip() for t in raw.split("|") if t.strip()):
            reason_today[tag] += 1

    return {
        "new": new,
        "dropped": dropped,
        "parcel_changed": parcel_changed,
        "reason_today": reason_today,
    }


# One-line plain-English description per Match Reason tag, shown as a glossary
# under the distribution and appended inline to each line. Dynamic-suffix tags
# (lot-cluster-N, low-confidence-parcel(NN%)) are matched by prefix in
# _describe_reason(). Keep in sync with tag_reason() calls in fix_addresses_and_prep.py.
REASON_GLOSSARY: dict[str, str] = {
    "(scrape direct)": "Parcel + PR came straight from the court record — no guessing",
    "default-sfr": "Property type unknown — labeled Single-Family (safe default)",
    "addr-corrected": "Property street address fixed/standardized from county records",
    "centroid-geocode": "No exact street number — map pin placed at the parcel's center",
    "name-research": "Property found by searching county records for the deceased's name",
    "name-nearmiss-addr-corroborated": "Name wasn't exact but the address confirmed it — treated as solid",
    "decedent-address": "Property matched using the deceased's own address from the court file",
    "app-realestate-parcel": "Parcel read from the real-estate section of the court Application PDF",
    "collapsed-same-property": "Duplicate rows for the same property merged into one",
    "condo-overruled-by-zillow": "Zillow says it's a house, not a condo — kept",
    "lincoln-ahdesc-structure": "Lincoln County structure type read from the GIS description field",
    # --- flagged: the matcher made a call but isn't fully sure (eyeball) ---
    "verify-middle-initial": "Matched via middle initial to separate same-named people — double-check",
    "verify-name-nomiddle": "Name matched but deed had no middle name to confirm — eyeball",
    "verify-name-ambiguous": "Several people could fit the name — best guess, eyeball",
    "verify-swap-lowconf": "Parcel swapped in with low confidence — eyeball",
    "verify-swap-ambiguous": "Parcel swap was ambiguous — eyeball",
    "low-confidence-parcel": "Property match scored below the 'strong' bar — eyeball",
    "addr-uncorroborated": "Address applied but not independently confirmed — eyeball",
    "verify-occupied-confirmed": "Court-confirmed home the surviving spouse occupies — kept (locked)",
    # --- multi-parcel / land ---
    "lot-cluster": "Estate with adjacent lots on one street — mobile-home-on-land play",
    "swap-on-dq": "Main property disqualified (e.g. heir lives there) — mailed a different estate parcel",
    "swap-on-over-cap": "Main house over your $500K cap — points to an under-cap estate parcel",
    "subdivide-exempt": "SFR/MH on >2 acres — gets the $1M cap (land is the play)",
    "tiny-vacant-lot": "Small vacant lot (low standalone value)",
    # --- where the contact / PR came from ---
    "heirs-of-fallback": "No usable PR/beneficiary — contact defaults to 'Heirs of [Deceased]' (deep-prospecting)",
    "coowner-dm": "Contact taken from a co-owner on the deed",
    "coowner-address": "Mailing address taken from a co-owner on the deed",
    "dm-promoted-pr": "No court executor — the obituary-found living heir (a real person) was promoted to mail target",
    "beneficiary-promoted-pr": "No court executor + no obit heir — a named beneficiary (real person w/ address) was promoted to mail target",
    "beneficiary-address": "Mailing address taken from a named beneficiary",
    "pr-backfill-parties": "PR filled from the eCourts Parties list",
    "pr-from-app-over-obit": "PR taken from the court Application, overriding a weaker obituary guess",
    "pr-from-app-override": "PR taken from the court Application, overriding another source",
    "pr-people-search": "PR's address found via free people-search",
    "pr-tracerfy": "PR's contact found via Tracerfy skip trace",
    "tracerfy": "Contact found via Tracerfy skip trace",
    "mailing-from-property": "PR mailing filled from the property address (so mail still lands)",
    "mailing-from-prior-main": "PR mailing kept at the estate home the main parcel was swapped away from",
    "mailing-addr-split": "Mailing address split out of a combined field",
    "second-pass-obit-full": "Heir/PR found on a second obituary pass (full-name match)",
    "second-pass-obit-name-only": "Heir/PR found on a second obituary pass (name-only match)",
    # --- case data / values ---
    "landportal-revalued": "Vacant lot's value filled from the LandPortal data source",
    "late-doc-apply": "A court document arrived late — its info was applied",
    "small-estate-disposed-recent": "Recently-filed small-estate/disposed case (you work these)",
    "small-estate-address": "Property address from a small-estate filing",
    "heir-transferred-deed": "Deed already moved to a next-gen heir — still a lead, flagged",
    "dq-recently-sold": "Flagged — the property sold recently",
    "rowan-condo": "Rowan County condo (flagged — condos are normally dropped)",
    "audit-repick": "Nightly audit re-picked a better parcel for this case",
    "audit-blanked": "Nightly audit removed a parcel that no longer passes the matcher",
    "pdf-phone": "Phone number pulled from the court PDF",
    "pdf-phone-verified": "Phone number pulled from the court PDF and cross-checked",
}

# Tags meaning "matched, but not fully sure" — summed into the 'Cases to eyeball'
# headline. Matched by prefix so dynamic-suffix variants are included.
_EYEBALL_PREFIXES = (
    "verify-middle-initial", "verify-name-nomiddle", "verify-name-ambiguous",
    "verify-swap-lowconf", "verify-swap-ambiguous",
    "low-confidence-parcel", "addr-uncorroborated",
)


def _describe_reason(tag: str) -> str:
    """Plain-English one-liner for a Match Reason tag (handles dynamic suffixes)."""
    if tag in REASON_GLOSSARY:
        return REASON_GLOSSARY[tag]
    for key, desc in REASON_GLOSSARY.items():
        if tag.startswith(key):
            return desc
    return ""


def _is_eyeball_reason(tag: str) -> bool:
    """True if the tag means the matcher made an uncertain call worth a glance."""
    return any(tag.startswith(p) for p in _EYEBALL_PREFIXES)


def _tab_week(r: dict) -> int:
    """ISO week number of the workbook tab a row came from (0 if unknown)."""
    m = re.search(r"Week (\d+)", str(r.get("_tab") or ""))
    return int(m.group(1)) if m else 0


def _is_heirs_of(r: dict) -> bool:
    return (r.get("Personal Representative") or "").strip().lower().startswith("heirs of")


def _has_named_contact(r: dict) -> bool:
    name = ((r.get("First Name") or "") + (r.get("Last Name") or "")).strip() \
        or (r.get("Personal Representative") or "").strip()
    return bool(name) and not _is_heirs_of(r)


def _has_phone(r: dict) -> bool:
    return bool((r.get("Phone 1") or "").strip() or (r.get("DM Phone") or "").strip())


def _week_pull_dates(week_n: int, year: int) -> str:
    """Mon-Fri date span of an ISO week — when that week's cases were pulled."""
    try:
        mon = date.fromisocalendar(year, week_n, 1)
        fri = date.fromisocalendar(year, week_n, 5)
        return f"{mon.strftime('%b %d')}-{fri.strftime('%b %d')}"
    except ValueError:
        return "?"


def eyeball_cases(rows: list[dict]) -> list[tuple[dict, list[str]]]:
    """Rows whose Match Reason carries at least one uncertain (eyeball) tag."""
    out: list[tuple[dict, list[str]]] = []
    for r in rows:
        tags = [t.strip() for t in (r.get("Match Reason") or "").split("|") if t.strip()]
        eye = [t for t in tags if _is_eyeball_reason(t)]
        if eye:
            out.append((r, eye))
    return out


def _why(tag: str) -> str:
    """Glossary description with the trailing '— eyeball'-style nag stripped
    (the section these lines appear in already says they need a look)."""
    desc = _describe_reason(tag) or tag
    head, sep, tail = desc.rpartition("—")
    if sep and ("eyeball" in tail or "double-check" in tail):
        return head.strip()
    return desc


def get_stale_docs() -> list[dict]:
    """High-priority cases whose court docs still aren't scanned (21+ days)."""
    try:
        sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
        import case_doc_queue as _cdq  # noqa: E402
        from ecourts_scraper import cases_needing_docs as _cnd  # noqa: E402
        return _cdq.long_pending_high_priority(set(_cnd()), min_age_days=21)
    except Exception:
        return []


def get_late_doc_updates() -> list[dict]:
    """Archived-week rows that just got a real named contact (last 7 days)."""
    try:
        import json as _json
        entries = _json.loads((OUTPUT / ".late_doc_updates.json")
                              .read_text(encoding="utf-8")).get("entries", [])
        cut = (datetime.now() - timedelta(days=7)).isoformat(timespec="seconds")
        return [e for e in entries if e.get("found_iso", "") >= cut]
    except Exception:
        return []


def _cases_preview(rows: list[dict], key: str, n: int = 3) -> str:
    """First n case numbers as a comma string — Oren looks cases up by number."""
    cases = [c for c in ((r.get(key) or "").strip() for r in rows) if c][:n]
    out = ", ".join(cases)
    if len(rows) > n:
        out += f" (+{len(rows) - n} more)"
    return out


def tracerfy_budget_line() -> str:
    """This week's Tracerfy burn vs cap, or '' if unavailable/unfunded."""
    try:
        sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
        from tracerfy_budget import budget_summary  # noqa: E402
        return budget_summary() or ""
    except Exception:
        return ""


def pr_push_queue_line() -> str:
    """Cases the nightly corrected in the WORKBOOK that still await their
    DataSift push, or '' when the queue is clear.

    fix_addresses_and_prep queues a case whenever the court's Parties list
    replaces an invented/placeholder PR on an already-uploaded row (Headen
    26E002921-590 sat 16 days with a phantom co-owner at the decedent's
    property because nothing surfaced this). The consumer is manual:
    python pr_upgrade_step.py --queued  (pushes name + mailing together).
    """
    try:
        cases = [ln.strip() for ln in
                 (OUTPUT / "pr_push_queue.txt").read_text(encoding="utf-8")
                 .splitlines() if ln.strip()]
    except OSError:
        return ""
    # Never nag about a case the user KILLED. A killed case is gone from the
    # weekly CSVs, so pr_upgrade_step can never cover it and it would sit here
    # forever (Vandall 26E000533-540 did, from 8/31). Read-only here — the
    # actual prune happens in pr_upgrade_step.load_queue().
    try:
        sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
        from fix_addresses_and_prep import _load_manual_drops
        killed = set(_load_manual_drops())
        cases = [c for c in cases if c.upper() not in killed]
    except Exception:  # noqa: BLE001
        pass
    if not cases:
        return ""
    preview = ", ".join(cases[:5]) + ("..." if len(cases) > 5 else "")
    return (f"PR PUSH QUEUE: {len(cases)} corrected case(s) not yet in DataSift "
            f"— run  python pr_upgrade_step.py --queued  ({preview})")


def kpi_totals(today: date) -> tuple[Counter, Counter, str] | None:
    """(this-week totals, last-week totals, latest ledger day) from
    output/kpi_daily_ledger.csv (maintained by scripts/kpi_refresh.py, which
    pulls DataSift call activity nightly). None when the ledger is missing
    or has no dials this week/last week — quiet weeks add no section."""
    ledger = OUTPUT / "kpi_daily_ledger.csv"
    if not ledger.exists():
        return None
    try:
        with open(ledger, encoding="utf-8", newline="") as fh:
            rows = {r["day"]: r for r in csv.DictReader(fh)}
    except Exception:
        return None

    def _tot(day_from: date, day_to: date) -> Counter:
        t = Counter()
        for d, r in rows.items():
            if day_from.isoformat() <= d <= day_to.isoformat():
                for f in ("dials", "answered", "conversations", "correct_numbers",
                          "leads", "not_interested", "talk_seconds", "sms_sent"):
                    t[f] += int(r.get(f) or 0)
        return t

    monday = today - timedelta(days=today.weekday())
    wk = _tot(monday, today)
    lw = _tot(monday - timedelta(days=7), monday - timedelta(days=1))
    if not wk["dials"] and not lw["dials"]:
        return None
    return wk, lw, (max(rows) if rows else "")


def kpi_section(today: date, condensed: bool = False) -> list[str]:
    """Short phone-KPI text block; [] when kpi_totals has nothing to show."""
    totals = kpi_totals(today)
    if not totals:
        return []
    wk, lw, stale = totals

    def _line(label: str, t: dict) -> str:
        ans = f"{t['answered'] / t['dials'] * 100:.0f}%" if t["dials"] else "-"
        mins = t["talk_seconds"] // 60
        return (f"  {label:13} {t['dials']} dials, {t['answered']} answered ({ans}), "
                f"{t['conversations']} conversations, {t['correct_numbers']} correct #s, "
                f"{t['leads']} leads, {mins}m talk")

    out = ["PHONES THIS WEEK"]
    out.append(_line("Mon-today:", wk))
    out.append(_line("Last week:", lw))
    if wk["sms_sent"] or lw["sms_sent"]:
        out.append(f"  Texts sent:   {wk['sms_sent']} this week, {lw['sms_sent']} last week")
    if stale and (today - date.fromisoformat(stale)).days > 2:
        out.append(f"  (numbers only current through {stale} — KPI refresh hasn't run since)")
    out.append("")
    return out


def summarize_week_improvements(monday: date) -> list[str]:
    """Plain-English bullets describing the week's pipeline improvements.

    Reads this week's commit subjects + bodies and asks Haiku to group them
    into a few non-technical bullets ("what changed and why it matters").
    Disk-cached on the commit-set hash, so it costs at most ONE small call
    per night (standing directive: keep LLM cost very low). Returns [] when
    git/LLM is unavailable — callers fall back to raw commit subjects.
    """
    try:
        r = subprocess.run(
            ["git", "log", f"--since={monday.isoformat()} 00:00", "--no-merges",
             "--pretty=%h%x1f%s%x1f%b%x1e"],
            capture_output=True, text=True, timeout=15,
            cwd=str(Path(__file__).parent.parent))
        raw = r.stdout if r.returncode == 0 else ""
    except Exception:
        raw = ""
    commits: list[tuple[str, str, str]] = []
    for chunk in raw.split("\x1e"):
        parts = chunk.strip("\n").split("\x1f")
        if len(parts) >= 2 and parts[0].strip():
            body = parts[2].strip() if len(parts) > 2 else ""
            commits.append((parts[0].strip(), parts[1].strip(), body))
    if not commits:
        return []

    import hashlib
    import json as _json
    key = hashlib.sha1("|".join(h for h, _, _ in commits).encode()).hexdigest()
    cache_path = OUTPUT / ".report_changes_cache.json"
    try:
        cached = _json.loads(cache_path.read_text(encoding="utf-8"))
        if cached.get("key") == key and cached.get("bullets"):
            return list(cached["bullets"])
    except Exception:
        pass

    commit_lines = []
    for _h, subj, body in commits[:30]:
        first_para = body.split("\n\n")[0].replace("\n", " ").strip()[:400]
        commit_lines.append(f"- {subj}" + (f" :: {first_para}" if first_para else ""))
    prompt = (
        "You write the \"what improved this week\" section of a nightly status "
        "email for Oren, a solo real-estate investor whose automated pipeline "
        "finds probate leads and preps them for direct mail and cold calls.\n\n"
        "This week's code changes (subject :: details):\n"
        + "\n".join(commit_lines)
        + "\n\nRewrite as 3-7 short bullets. Group related changes into one "
        "bullet. Each bullet says what changed and why it matters to his "
        "leads, mail, or calls — in plain English a non-programmer gets. "
        "No file names, no jargon, no commit hashes. Most impactful first.\n\n"
        'Return JSON only: {"bullets": ["...", "..."]}'
    )
    bullets: list[str] = []
    try:
        sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
        import llm_client  # noqa: E402
        resp = llm_client.chat_json(prompt, max_tokens=700)
        if resp and isinstance(resp.get("bullets"), list):
            bullets = [str(b).strip() for b in resp["bullets"] if str(b).strip()][:8]
    except Exception:
        bullets = []
    if not bullets:
        return []
    try:
        cache_path.write_text(_json.dumps({"key": key, "bullets": bullets}),
                              encoding="utf-8")
    except OSError:
        pass
    return bullets


def git_fixes_since(day: date) -> list[str]:
    """Commit subjects since `day` — the week's pipeline work, for the
    'what got done' section. 'Ddd<TAB>subject' lines; [] if git unavailable."""
    try:
        r = subprocess.run(
            ["git", "log", f"--since={day.isoformat()} 00:00", "--no-merges",
             "--pretty=%cd\t%s", "--date=format:%a"],
            capture_output=True, text=True, timeout=15,
            cwd=str(Path(__file__).parent.parent))
        if r.returncode != 0:
            return []
        return [ln.strip() for ln in r.stdout.splitlines() if ln.strip()]
    except Exception:
        return []


def datasift_upload_status(rows: list[dict]) -> tuple[int, int]:
    """(already uploaded, awaiting upload) for this week's rows, from the
    NETNEW upload ledger. (-1, -1) if the ledger can't be read."""
    try:
        sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
        from nc_datasift_export import load_upload_ledger  # noqa: E402
        ledger = load_upload_ledger()
    except Exception:
        return -1, -1
    in_ds = sum(1 for r in rows
                if (r.get("Case No.") or "").strip().upper() in ledger)
    return in_ds, len(rows) - in_ds


def compute_model(
    workbook_rows: list[dict],
    week_n: int,
    today_csv_rows: list[dict],
    yesterday_csv_rows: list[dict],
    tonight: dict | None = None,
) -> dict:
    """Everything the report derives from tonight's outputs, computed once.

    Both renderers read from this — render_report (text) and
    render_html_email — so the email can never disagree with the attached
    report about a count."""
    today = date.today()
    diffs = compute_diffs(today_csv_rows, yesterday_csv_rows)
    # Weekly lens (Oren, 2026-08-14): the current week's polished CSV IS this
    # week's inventory — lead with it, demote day-to-day churn and plumbing.
    week_rows = today_csv_rows

    wk_counts = Counter(_tab_week(r) for r in workbook_rows)
    wk_counts.pop(0, None)
    prev_weeks = sorted(w for w in wk_counts if w < week_n)
    prev_week = prev_weeks[-1] if prev_weeks else None
    prev_total = wk_counts.get(prev_week, 0) if prev_week is not None else 0

    ready = [r for r in week_rows
             if _has_named_contact(r) and (r.get("Property Address") or "").strip()]
    with_phone = [r for r in week_rows if _has_phone(r)]
    heirs_of = [r for r in week_rows if _is_heirs_of(r)]

    # The FULL deep-prospecting queue — every "Heirs of" row still in the
    # workbook, all weeks, not just this week's slice (Oren, 2026-08-19:
    # "how many are in the DP queue total and what are the dates they were
    # pulled? Need to know this in future reports").
    # Archived week tabs are frozen, so a case that was since researched never
    # leaves the workbook — the queue must subtract finished work itself
    # (Oren, 2026-08-20): skip cases the DP ledger marks resolved, and rows
    # where research already filled a DM Name (a named contact exists even
    # though the PR column still says "Heirs of").
    dp_rows = read_dp_log()
    dp_done_cases = {
        (r.get("Case No.") or "").strip().upper() for r in dp_rows
        if (r.get("Outcome") or "").strip() in ("resolved", "resolved-rejected")
    }
    all_heirs = [r for r in workbook_rows if _is_heirs_of(r)]
    dp_queue = [
        r for r in all_heirs
        if (r.get("Case No.") or "").strip().upper() not in dp_done_cases
        and not str(r.get("DM Name") or "").strip()
    ]
    dp_already_done = len(all_heirs) - len(dp_queue)
    dp_wk_counts: Counter = Counter()
    for r in dp_queue:
        mm = re.search(r"Week (\d+)\s*(\d{4})?", str(r.get("_tab") or ""))
        wk = int(mm.group(1)) if mm else 0
        yr = int(mm.group(2)) if mm and mm.group(2) else today.year
        dp_wk_counts[(yr, wk)] += 1
    dp_by_week = [(yr, wk, dp_wk_counts[(yr, wk)])
                  for (yr, wk) in sorted(dp_wk_counts, reverse=True)]

    eyeballs = eyeball_cases(week_rows)
    in_ds, awaiting = datasift_upload_status(week_rows)
    open_dp = [r for r in dp_rows if (r.get("Outcome") or "").strip() == "open"]
    stale_docs = get_stale_docs()
    late_docs = get_late_doc_updates()

    # What to work on next — the priority queue, ranked. Hand-entry work that
    # nothing else surfaces comes first; volume queues last. Case numbers up
    # front (Oren looks cases up by number). The hand items don't depend on
    # the condensed flag, so one call covers both renderers.
    _, tn_hand = tonight_summary_section(tonight or {}, condensed=True)
    prios: list[str] = []
    if late_docs:
        prios.append("Type these into DataSift by hand — archived cases just got a "
                     f"named contact: {_cases_preview(late_docs, 'case_number')}")
    # Unscanned court docs are NOT a priority item — Oren doesn't do courthouse
    # trips (2026-08-19). They stay in the full report as an FYI list only.
    prios.extend(tn_hand)
    if eyeballs:
        prios.append(f"Double-check {len(eyeballs)} uncertain matches before "
                     "mailing (worst listed below)")
    if open_dp:
        prios.append(f"Finish {len(open_dp)} open deep-prospecting case(s): "
                     f"{_cases_preview(open_dp, 'Case No.')}")
    if in_ds >= 0 and awaiting > 0:
        prios.append(f"{awaiting} polished rows are waiting for the next DataSift upload "
                     "(nothing uploads by itself — they go in when the upload script is run)")
    if heirs_of:
        prios.append(f'{len(heirs_of)} new "Heirs of" rows this week — DP queue is '
                     f"{len(dp_queue)} total (breakdown below)")

    county_this = Counter((r.get("County") or "").strip() for r in week_rows)
    county_prev = Counter((r.get("County") or "").strip() for r in workbook_rows
                          if prev_week is not None and _tab_week(r) == prev_week)

    # A county GIS that went down during the polish (see the comment where
    # this is rendered — rows recover on the next clean run).
    gis_outage: dict = {}
    try:
        import json as _json
        gis_outage = _json.loads((OUTPUT / ".gis_outage_last_run.json")
                                 .read_text(encoding="utf-8"))
    except Exception:
        gis_outage = {}

    return {
        "diffs": diffs,
        "new_today": diffs["new"],
        "week_rows": week_rows,
        "wk_counts": wk_counts,
        "prev_week": prev_week,
        "prev_total": prev_total,
        "ready": ready,
        "with_phone": with_phone,
        "heirs_of": heirs_of,
        "dp_rows": dp_rows,
        "dp_queue": dp_queue,
        "dp_already_done": dp_already_done,
        "dp_by_week": dp_by_week,
        "eyeballs": eyeballs,
        "in_ds": in_ds,
        "awaiting": awaiting,
        "open_dp": open_dp,
        "stale_docs": stale_docs,
        "late_docs": late_docs,
        "prios": prios,
        "county_this": county_this,
        "county_prev": county_prev,
        "gis_outage": gis_outage,
    }


def render_report(
    workbook_path: Path | None,
    workbook_rows: list[dict],
    week_n: int,
    today_csv_rows: list[dict],
    yesterday_csv_rows: list[dict],
    heir_transfer_rows: list[dict],
    polish_stats: dict,
    runtime_min: float | None,
    include_heir_transfer: bool = False,
    condensed: bool = False,
    improvements: list[str] | None = None,
    tonight: dict | None = None,
    model: dict | None = None,
) -> str:
    """Render the report. condensed=True builds the short email body (the
    most important items only, with a pointer to the attached full report);
    condensed=False builds the full .txt with every fine-print section."""
    today = date.today()
    today_str = today.strftime("%a %b %d, %Y")
    wd = today.weekday()
    day_note = f"court day {wd + 1} of 5" if wd < 5 else "weekend catch-up run"
    monday = today - timedelta(days=wd)
    m = model or compute_model(workbook_rows, week_n, today_csv_rows,
                               yesterday_csv_rows, tonight)
    diffs = m["diffs"]
    new_today = m["new_today"]
    week_rows = m["week_rows"]
    this_week_total = len(week_rows)
    wk_counts = m["wk_counts"]
    prev_week = m["prev_week"]
    prev_total = m["prev_total"]
    ready = m["ready"]
    with_phone = m["with_phone"]
    heirs_of = m["heirs_of"]
    dp_rows = m["dp_rows"]
    dp_queue = m["dp_queue"]
    dp_already_done = m["dp_already_done"]
    eyeballs = m["eyeballs"]
    in_ds, awaiting = m["in_ds"], m["awaiting"]

    lines: list[str] = []
    lines.append(f"NC Probate Pipeline — Week {week_n}")
    lines.append(f"  {today_str} — {day_note}")
    lines.append("=" * 60)
    lines.append("")

    lines.append("THIS WEEK AT A GLANCE")
    cmp_note = f"    (Week {prev_week} finished at {prev_total})" if prev_week else ""
    lines.append(f"  Cases this week:      {this_week_total}{cmp_note}")
    lines.append(f"  Added today:          {len(new_today)}")
    lines.append(f"  Ready to work:        {len(ready)}    have a property + a named person to contact")
    # Phones mostly arrive AFTER upload (DataSift skip trace) — only show the
    # in-CSV count when deep prospecting actually put numbers on rows, so a 0
    # here doesn't read as "no phones anywhere".
    if with_phone:
        lines.append(f"  Have a phone number:  {len(with_phone)}    (found by the pipeline itself — court docs / DP; "
                     "skip-trace phones live in DataSift and aren't counted here)")
    lines.append(f"  \"Heirs of\" rows:      {len(heirs_of)}    new this week — DP queue is "
                 f"{len(dp_queue)} total (see DEEP PROSPECTING QUEUE below)")
    lines.append(f"  Check these:          {len(eyeballs)}    uncertain matches — listed below")
    if in_ds >= 0:
        lines.append(f"  In DataSift already:  {in_ds}    ({awaiting} waiting for the next upload)")
    lines.append("")
    county_this = m["county_this"]
    county_prev = m["county_prev"]
    lines.append("  By county:           this wk   last wk")
    for c, _n in county_this.most_common():
        label = c if c else "(blank)"
        lines.append(f"    {label:16} {county_this.get(c, 0):>7}   {county_prev.get(c, 0):>7}")
    for c in sorted(set(county_prev) - set(county_this)):
        label = c if c else "(blank)"
        lines.append(f"    {label:16} {0:>7}   {county_prev[c]:>7}")
    lines.append("")

    # Tonight's run in plain words (Oren, 2026-08-31) — the same recap a
    # person would write after reading the log, in both email and full report.
    tn_lines, _tn_hand = tonight_summary_section(tonight or {}, condensed=condensed)
    lines.extend(tn_lines)

    lines.extend(kpi_section(today, condensed=condensed))

    # A county GIS that went down during the polish. Its rows lost their parcel
    # to a dead server (not to "owns nothing") and were dropped at Step 4 — so
    # tonight's sheet is short for that county. The next clean run restores them,
    # and auto_archive_weeks.py won't freeze the week until then; this is just so
    # a short county count doesn't read as a quiet week.
    _o = m["gis_outage"]
    if _o.get("counties"):
        lines.append("*** COUNTY GIS OUTAGE THIS RUN ***")
        lines.append(f"  Down: {', '.join(_o['counties'])}")
        lines.append(f"  Rows for these counties are INCOMPLETE in week(s) "
                     f"{_o.get('weeks', [])} — they will be recovered on the "
                     f"next clean run, and the week won't be archived until then.")
        lines.append("")

    # What to work on next — the priority queue, ranked (built in
    # compute_model so the HTML email shows the identical list).
    open_dp = m["open_dp"]
    stale_docs = m["stale_docs"]
    late_docs = m["late_docs"]
    prios = m["prios"]
    lines.append("WHAT TO WORK ON NEXT")
    if not prios:
        lines.append("  (nothing urgent — pipeline is clean)")
    for i, p in enumerate(prios, 1):
        lines.append(f"  {i}. {p}")
    lines.append("")

    # What got done — the week's progress in plain terms: cases in, contacts
    # found, deep-prospecting worked by hand, pipeline fixes shipped (git).
    monday_iso = monday.isoformat()
    dp_resolved_wk = [r for r in dp_rows
                      if (r.get("Date") or "") >= monday_iso
                      and (r.get("Outcome") or "").strip() in ("resolved", "partial")]
    fixes = git_fixes_since(monday)
    lines.append(f"WHAT GOT DONE THIS WEEK (since Mon {monday.strftime('%b %d')})")
    lines.append(f"  {this_week_total} new estate cases found and matched to property")
    phone_note = (f"; {len(with_phone)} came with a phone the pipeline found itself "
                  "(court docs / DP — the rest get phones from DataSift skip trace "
                  "after upload)") if with_phone else ""
    lines.append(f"  {len(ready)} are mail-ready with a named contact{phone_note}")
    if dp_resolved_wk:
        if condensed:
            lines.append(f"  {len(dp_resolved_wk)} deep-prospecting cases worked by "
                         "hand (case detail in the attached report)")
        else:
            lines.append(f"  {len(dp_resolved_wk)} deep-prospecting cases worked by hand:")
            for r in dp_resolved_wk:
                note = (r.get("Notes") or "").strip().replace("\n", " ")
                lines.append(f"    {r.get('Case No.',''):16}  {(r.get('Decedent') or '')[:24]:24}  {note[:64]}")
    # Improvements: plain-English grouped bullets (LLM, cached) when available;
    # raw commit subjects otherwise. The full report also keeps the commit list
    # underneath as the audit trail.
    import textwrap
    if improvements:
        lines.append("  Pipeline improvements this week — what changed and why it matters:")
        for b in improvements:
            for j, seg in enumerate(textwrap.wrap(b, width=86)):
                lines.append(f"    {'*' if j == 0 else ' '} {seg}")
    if fixes and (not condensed or not improvements):
        n_show = 6 if condensed else 15
        lines.append(f"  {len(fixes)} commit(s) shipped:")
        for fx in fixes[:n_show]:
            day_abbr, _, subj = fx.partition("\t")
            lines.append(f"    {day_abbr:4} {subj[:90]}")
        if len(fixes) > n_show:
            lines.append(f"    ... and {len(fixes) - n_show} more")
    lines.append("")

    # New cases
    lines.append("NEW CASES TODAY")
    n_new_cap = 15 if condensed else 50
    if not new_today:
        lines.append("  (none — workbook unchanged)")
    else:
        for r in new_today[:n_new_cap]:
            cn = (r.get("Case No.") or "").strip()
            cty = (r.get("County") or "").strip()
            dec = (r.get("Deceased Owner") or "").strip()
            prop = (r.get("Property Address") or "(no parcel)").strip()
            lines.append(f"  {cn:18}  {cty:12}  {dec[:32]:32}  ->  {prop}")
        if len(new_today) > n_new_cap:
            lines.append(f"  ... and {len(new_today) - n_new_cap} more")
    lines.append("")

    # Uncertain matches — the actual rows (with Case No.), not just tag counts.
    # Condensed: one line per case (why inline, address left to the workbook).
    lines.append("CHECK THESE — UNCERTAIN MATCHES (this week)")
    if not eyeballs:
        lines.append("  (none — every match this week is solid)")
    elif condensed:
        lines.append("  (the matcher made a judgment call — worth a look before mailing)")
        for r, tags in eyeballs[:8]:
            cn = (r.get("Case No.") or "").strip()
            cty = (r.get("County") or "").strip()
            dec = (r.get("Deceased Owner") or "").strip()
            lines.append(f"  >> {cn:18}  {cty:12}  {dec[:24]:24}  {_why(tags[0])}")
        if len(eyeballs) > 8:
            lines.append(f"  ... and {len(eyeballs) - 8} more — full list in the attached report")
    else:
        lines.append("  (the matcher made a judgment call — worth a look before mailing)")
        for r, tags in eyeballs[:20]:
            cn = (r.get("Case No.") or "").strip()
            cty = (r.get("County") or "").strip()
            dec = (r.get("Deceased Owner") or "").strip()
            prop = (r.get("Property Address") or "(no parcel)").strip()
            lines.append(f"  >> {cn:18}  {cty:12}  {dec[:28]:28}  {prop}")
            for t in tags:
                lines.append(f"       why: {_why(t)}")
        if len(eyeballs) > 20:
            lines.append(f"  ... and {len(eyeballs) - 20} more — sort the workbook's Match Reason column")
    lines.append("")

    # Deep-prospecting queue — every "Heirs of" row across ALL workbook weeks,
    # grouped by the week it was pulled (with the pull dates). Condensed email
    # shows the total + by-week breakdown; the full report also lists every case.
    lines.append('DEEP PROSPECTING QUEUE — ALL "HEIRS OF" ROWS (no named contact yet)')
    if not dp_queue:
        lines.append("  (empty)" + (f" — {dp_already_done} already researched"
                                    if dp_already_done else ""))
    else:
        by_wk: dict[tuple[int, int], list[dict]] = {}
        for r in dp_queue:
            m = re.search(r"Week (\d+)\s*(\d{4})?", str(r.get("_tab") or ""))
            wk = int(m.group(1)) if m else 0
            yr = int(m.group(2)) if m and m.group(2) else today.year
            by_wk.setdefault((yr, wk), []).append(r)
        done_note = (f"  (another {dp_already_done} already researched — "
                     "not counted)") if dp_already_done else ""
        lines.append(f"  Total: {len(dp_queue)} cases, pulled across "
                     f"{len(by_wk)} week(s){done_note}")
        for (yr, wk) in sorted(by_wk, reverse=True):
            span = _week_pull_dates(wk, yr)
            lines.append(f"    Week {wk:<3} pulled {span:14} {len(by_wk[(yr, wk)]):>4} case(s)")
        if not condensed:
            lines.append("  Every case (newest first):")
            for (yr, wk) in sorted(by_wk, reverse=True):
                for r in by_wk[(yr, wk)]:
                    cn = (r.get("Case No.") or "").strip()
                    cty = (r.get("County") or "").strip()
                    dec = (r.get("Deceased Owner") or "").strip()
                    filed = str(r.get("File Date") or "").strip()
                    lines.append(f"    {cn:18}  {cty:12}  {dec[:30]:30}  filed {filed}")
    lines.append("")

    # Deep prospecting still open — the hand-work queue. The condensed email
    # already surfaces these in WHAT TO WORK ON NEXT; detail is full-report only.
    if not condensed:
        lines.append("DEEP PROSPECTING — STILL OPEN")
        if not open_dp:
            lines.append("  (nothing open)")
        else:
            for r in open_dp:
                lines.append(f"  {r.get('Case No.',''):16}  {(r.get('Decedent') or '')[:30]:30}  "
                             f"{(r.get('Notes') or '')[:60]}")
        if dp_rows:
            n_docs = len(list((OUTPUT / "reports").glob("DP_*.md")))
            by_outcome = Counter((r.get("Outcome") or "?").strip() for r in dp_rows)
            outcome_str = ", ".join(f"{n} {k}" for k, n in by_outcome.most_common())
            lines.append(f"  (ledger: {len(dp_rows)} runs, {n_docs} research docs — {outcome_str})")
        lines.append("")

    # Documents the court still hasn't scanned for high-priority leads. We retry
    # nightly for months, but if the Application/Will never appears the only way
    # to the executor name is a manual courthouse pull — surface the oldest so
    # Oren can grab them by hand.
    if stale_docs and not condensed:
        lines.append("DOCS THE COURT STILL HASN'T SCANNED (FYI — 21+ days; only worth "
                     "pulling if you happen to be at the courthouse)")
        for e in stale_docs[:15]:
            lines.append(f"  {e.get('case_number',''):18} {e.get('county',''):12} "
                         f"{e['age_days']:>3}d  needs {','.join(e.get('needs', []))}")
        if len(stale_docs) > 15:
            lines.append(f"  ... and {len(stale_docs) - 15} more")
        lines.append("")

    # Wills/Applications that landed for weeks already archived. apply_late_docs.py
    # writes the name onto the archived sheet, but an archived week is not in the
    # workbook — this report is the ONLY place Oren would ever learn a dead
    # "Heirs of" row turned into a real named contact. Nothing pushes these to
    # DataSift, so they need typing in by hand; keep the case no. front and centre.
    if late_docs and not condensed:
        lines.append("LATE DOCS — ARCHIVED WEEKS JUST GOT A NAME (add to DataSift by hand)")
        for e in late_docs:
            _rel = f" ({e['relationship']})" if e.get("relationship") else ""
            lines.append(f"  {e.get('case_number',''):18} {e.get('county',''):12} "
                         f"wk{e.get('week','?')}  {e.get('was','')} -> {e.get('now','')}{_rel}")
        lines.append("")

    # Day-to-day churn — one compact section (weekly lens). Dropped rows are
    # still the regression signal; parcel changes are usually a smart-picker
    # repick or a new sibling parcel winning main.
    dropped = diffs["dropped"]
    pc = diffs["parcel_changed"]
    n_drop_cap = 5 if condensed else 15
    n_pc_cap = 3 if condensed else 10
    lines.append("CHANGES SINCE YESTERDAY'S RUN")
    if not dropped and not pc:
        lines.append("  (steady — nothing dropped, no parcels changed)")
    if dropped:
        lines.append(f"  Dropped ({len(dropped)}) — in yesterday's sheet, gone tonight:")
        for r in dropped[:n_drop_cap]:
            cn = (r.get("Case No.") or "").strip()
            cty = (r.get("County") or "").strip()
            dec = (r.get("Deceased Owner") or "").strip()
            prop = (r.get("Property Address") or "(no parcel)").strip()
            lines.append(f"    {cn:18}  {cty:12}  {dec[:30]:30}  {prop}")
        if len(dropped) > n_drop_cap:
            lines.append(f"    ... and {len(dropped) - n_drop_cap} more")
    if pc:
        lines.append(f"  Parcel changed ({len(pc)}) — same case, different parcel picked:")
        for e in pc[:n_pc_cap]:
            lines.append(f"    {e['case_no']:18}  {e['county']:12}  {e['decedent'][:30]:30}")
            lines.append(f"      was: {e['parcel_yesterday']:18}  {e['addr_yesterday']}")
            lines.append(f"      now: {e['parcel_today']:18}  {e['addr_today']}")
        if len(pc) > n_pc_cap:
            lines.append(f"    ... and {len(pc) - n_pc_cap} more")
    lines.append("")

    # Week-over-week pace from the workbook tabs (current week still filling).
    lines.append("WEEK-OVER-WEEK PACE")
    recent_weeks = sorted(wk_counts)[-6:]
    max_n = max((wk_counts[w] for w in recent_weeks), default=1) or 1
    for w in recent_weeks:
        n = wk_counts[w]
        bar = "#" * max(1, round(n / max_n * 28))
        tail = "   <- in progress" if w == week_n else ""
        lines.append(f"  Week {w:<3} {n:>4}  {bar}{tail}")
    lines.append("")

    # Condensed email stops here: fold runtime, drop total, and Tracerfy burn
    # into a two-line footer that points at the attached full report.
    ts_line = tracerfy_budget_line()
    q_line = pr_push_queue_line()
    if q_line:
        lines.append(q_line)
        lines.append("")
    if condensed:
        total_drops = sum(polish_stats.get(k, 0) for k in (
            "over_500k", "under_min_value", "heir_occupied", "commercial",
            "no_parcel", "archive_dupe", "non_pending"))
        bits = []
        if runtime_min is not None:
            bits.append(f"runtime {runtime_min:.0f} min")
        bits.append(f"{total_drops} rows dropped by filters tonight")
        bits.append(f"workbook total {len(workbook_rows)} cases")
        lines.append("FINE PRINT")
        lines.append("  " + " / ".join(bits))
        if ts_line:
            lines.append(f"  {ts_line}")
        lines.append("  Full detail — match reasons, drop breakdown, DP ledger — "
                     "is in the attached report file")
        lines.append("")
        lines.append("=" * 60)
        lines.append(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        return "\n".join(lines)

    # Tracerfy spend (when funded). One-line summary of this week's burn
    # vs cap — useful for catching runaway costs before they hit the cap.
    if ts_line:
        lines.append("TRACERFY BUDGET")
        lines.append(f"  {ts_line}")
        lines.append("")

    # Heir-transfer — off by default (Oren, 2026-07-18). The review XLSX still
    # generates in output/; it's just no longer summarized here or emailed.
    # Re-enable with `daily_report.py --include-heir-transfer`.
    if include_heir_transfer:
        lines.append("HEIR-TRANSFER FLAGS (review file)")
        if not heir_transfer_rows:
            lines.append("  (none flagged)")
        else:
            # Count unique decedents
            unique_decedents = {(r.get("County"), r.get("Case No.")) for r in heir_transfer_rows}
            lines.append(f"  {len(unique_decedents)} decedent(s) with {len(heir_transfer_rows)} candidate parcel(s) flagged")
            # Group by case; show first 5
            by_case: dict[tuple, list[dict]] = {}
            for r in heir_transfer_rows:
                key = (r.get("County"), r.get("Case No."), r.get("Deceased Owner"))
                by_case.setdefault(key, []).append(r)
            for (cty, cn, dec), cands in list(by_case.items())[:5]:
                lines.append(f"  {cn}  {cty}  {dec}  ({len(cands)} candidate parcels)")
            if len(by_case) > 5:
                lines.append(f"  ... and {len(by_case) - 5} more decedents")
        lines.append("")

    # Fine print — plumbing numbers, scoped to tonight's run (parse_polish_log
    # is called with since=now so these aren't the all-history sums anymore).
    lines.append("TONIGHT'S RUN — FINE PRINT")
    lines.append(f"  Workbook:          {workbook_path.name if workbook_path else '(missing!)'}")
    lines.append(f"  All weeks total:   {len(workbook_rows)} cases")
    if runtime_min is not None:
        lines.append(f"  Runtime:           {runtime_min:.0f} min")
    drop_labels = [
        ("over_500k",       "Dropped — over $500K (buy-box cap)"),
        ("under_min_value", "Dropped — under $10K (scrap parcel)"),
        ("heir_occupied",   "Dropped — heir-occupied (PR mailing = property)"),
        ("commercial",      "Dropped — commercial/industrial"),
        ("no_parcel",       "Dropped — no parcel found in county GIS"),
        ("archive_dupe",    "Dropped — cross-week archive duplicate"),
        ("non_pending",     "Dropped — case no longer Pending"),
    ]
    for key, label in drop_labels:
        n = polish_stats.get(key, 0)
        if n:
            lines.append(f"  {label:48} {n}")
    if polish_stats.get("repick"):
        lines.append(f"  Smart-picker REPICK events:                      {polish_stats['repick']}")
    if polish_stats.get("refound"):
        lines.append(f"  Step 0.5 re-found correct parcels:               {polish_stats['refound']}")
    lines.append("")

    # Match Reason distribution — how each kept row got to its current state.
    # (scrape direct) = parcel + PR from court directly (high confidence,
    # no polish-step fallback fired). Other tags name the fallback step.
    reason_today = diffs["reason_today"]
    lines.append("HOW THIS WEEK'S ROWS WERE MATCHED (fine print)")
    lines.append("  (how each kept case got its property + contact — '>>' = worth an eyeball)")
    if not reason_today:
        lines.append("  (no rows in today's CSV)")
    else:
        for tag, n in reason_today.most_common():
            flag = ">>" if _is_eyeball_reason(tag) else "  "
            desc = _describe_reason(tag)
            lines.append(f"  {flag} {tag:30} {n:>4}   {desc}")
    lines.append("")

    lines.append("=" * 60)
    lines.append(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    return "\n".join(lines)


# ── HTML email rendering ─────────────────────────────────────────────
# The condensed plain-text report stays the text/plain fallback and the full
# .txt rides along as the attachment; this builds the HTML part the inbox
# actually shows. It renders prose + small tables straight from
# compute_model() — the same numbers as the text report, none of the
# monospace ledger look (Oren, 2026-09-01: "easier to read, less of a
# ledger"). Email-safe: inline styles only, table layout, system fonts.

_GREEN = "#1b5e20"
_F = ("font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,"
      "Helvetica,Arial,sans-serif;")
_MONO = ("font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,"
         "'Liberation Mono',monospace;font-size:12px;")


def _esc(s) -> str:
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _p(inner: str, *, color: str = "#374151", size: str = "13.5px") -> str:
    """A paragraph. `inner` is already-escaped/marked-up HTML."""
    return (f'<p style="{_F}font-size:{size};line-height:1.6;color:{color};'
            f'margin:0 0 8px;">{inner}</p>')


def _case(cn) -> str:
    """Case numbers stay monospace — Oren looks cases up by number."""
    return f'<span style="{_MONO}white-space:nowrap;">{_esc(cn)}</span>'


def _h(title: str) -> str:
    return (f'<div style="{_F}font-size:11px;font-weight:700;letter-spacing:.6px;'
            f'text-transform:uppercase;color:{_GREEN};margin:18px 0 7px;">'
            f'{_esc(title)}</div>')


def _chip(inner: str) -> str:
    return (f'<span style="display:inline-block;background:#eef1f4;color:#374151;'
            f'border-radius:12px;padding:3px 10px;margin:0 6px 6px 0;'
            f'{_F}font-size:12px;">{inner}</span>')


def _tbl(headers: list[str], rows: list[list[str]], note: str = "") -> str:
    """A small data table. Cell values are already-escaped/marked-up HTML."""
    th = "".join(
        f'<th align="left" style="{_F}font-size:10.5px;font-weight:700;color:#6b7280;'
        f'text-transform:uppercase;letter-spacing:.4px;padding:6px 10px;'
        f'border-bottom:1px solid #e4e8eb;background:#f6f7f9;">{_esc(h)}</th>'
        for h in headers)
    body = []
    for i, cells in enumerate(rows):
        bg = "background:#fbfcfd;" if i % 2 else "background:#ffffff;"
        tds = "".join(
            f'<td valign="top" style="{_F}font-size:13px;color:#374151;line-height:1.45;'
            f'padding:6px 10px;{bg}">{c}</td>' for c in cells)
        body.append(f"<tr>{tds}</tr>")
    note_html = (f'<div style="{_F}font-size:11.5px;color:#8a919a;margin:4px 2px 0;">'
                 f'{_esc(note)}</div>') if note else ""
    return (f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0" '
            f'style="border:1px solid #e7eaee;border-radius:8px;border-collapse:separate;'
            f'border-spacing:0;overflow:hidden;">'
            f"<tr>{th}</tr>" + "".join(body) + "</table>" + note_html)


def _stat_tile(number, label, accent_bg, accent_fg) -> str:
    return (
        f'<td style="width:25%;padding:6px;" valign="top">'
        f'<div style="background:{accent_bg};border-radius:8px;padding:14px 12px;text-align:center;">'
        f'<div style="font-size:30px;font-weight:800;color:{accent_fg};line-height:1;">{number}</div>'
        f'<div style="font-size:10.5px;font-weight:600;letter-spacing:.6px;text-transform:uppercase;'
        f'color:#6b7280;margin-top:6px;">{label}</div>'
        f'</div></td>'
    )


def _tonight_html(t: dict) -> str:
    """Tonight's run as short prose — headline sentence, then money + blips."""
    if not t.get("ok"):
        return ""
    sent: list[str] = []
    start = t.get("start")
    if start:
        end = datetime.now()
        if t.get("done_hms"):
            hh, mi, ss = t["done_hms"]
            end = start.replace(hour=hh, minute=mi, second=ss)
            if end < start:
                end += timedelta(days=1)
        mins = int((end - start).total_seconds() // 60)
        span = f"{mins // 60}h {mins % 60:02d}m" if mins >= 60 else f"{mins} min"
        if t.get("budget_hit"):
            state = ('<b style="color:#b91c1c;">was killed by the 4.5-hour time '
                     'budget</b> — the back half (upload / DP / sold sweep) did not run')
        elif t.get("tracebacks"):
            state = (f'<b style="color:#b91c1c;">crashed in {t["tracebacks"]} '
                     'step(s)</b> (see the log)')
        else:
            state = "<b>finished clean</b>"
        sent.append(f"The run {state} in {span} ({start:%H:%M}&ndash;{end:%H:%M}).")
    raw = t.get("raw_today") or {}
    if raw:
        sent.append(f"It scraped <b>{sum(raw.values())}</b> raw row(s) today.")
    elif start:
        sent.append("The scrape landed 0 new rows today.")
    up = t.get("upload") or {}
    if up.get("disabled"):
        sent.append("Auto-upload was OFF (NC_AUTO_UPLOAD=0) — nothing was pushed to DataSift.")
    elif up.get("uploaded"):
        n = sum(up["uploaded"].values())
        after = [
            "skip trace started" if up.get("skip_trace")
            else '<b style="color:#b45309;">skip trace NOT started</b>',
            "dial tiers tagged" if up.get("tiers") else "dial tiers NOT tagged",
            "text touches written" if up.get("touches") else "text touches NOT written",
        ]
        sent.append(f"<b>{n}</b> record(s) went to DataSift ({'; '.join(after)}).")
    elif t.get("weeks"):
        sent.append("Nothing net-new to upload to DataSift tonight.")
    sold = t.get("sold") or {}
    if sold.get("ran"):
        checked = sold.get("checked") or sold.get("parcels") or 0
        flagged = sold.get("flagged", 0)
        if flagged:
            n_tag = len(sold.get("tagged", []))
            rep = sold.get("report_only", 0)
            sent.append(f"The sold sweep checked {checked} parcels and flagged "
                        f"<b>{flagged}</b> transfer(s) — {n_tag} tagged Sold, "
                        f"{rep} heir transfer(s) kept as leads.")
        else:
            sent.append(f"The sold sweep checked {checked} parcels — no new transfers.")
    out = _p(" ".join(sent)) if sent else ""
    tail: list[str] = []
    total, costs = _tonight_costs(t)
    if costs:
        tail.append(f"Tonight cost about <b>{_money(total)}</b> ({_esc(', '.join(costs))}).")
    warn = t.get("warnings") or Counter()
    if warn:
        top = "; ".join(f"{k} x{n}" for k, n in warn.most_common(3))
        tail.append(f"{sum(warn.values())} warning blip(s) — {_esc(top)} — one-offs "
                    "unless they repeat tomorrow.")
    if tail:
        out += _p(" ".join(tail), color="#8a919a", size="12.5px")
    return out


def render_html_email(m: dict, *, tonight: dict, week_n: int,
                      improvements: list[str] | None,
                      runtime_min: float | None,
                      polish_stats: dict,
                      workbook_total: int) -> str:
    d = date.today()
    wd = d.weekday()
    day_note = f"court day {wd + 1} of 5" if wd < 5 else "weekend catch-up run"
    date_str = d.strftime("%A, %B ") + str(d.day) + d.strftime(", %Y")

    # Stat tiles — eyeball tile turns amber only when there's something to check.
    eyeball_n = len(m["eyeballs"])
    eye_bg, eye_fg = (("#fff8e1", "#b45309") if eyeball_n else ("#eef1f4", "#374151"))
    tiles = (
        _stat_tile(len(m["week_rows"]), "This week", "#eef1f4", "#111827")
        + _stat_tile(len(m["new_today"]), "New today", "#e7f4ea", _GREEN)
        + _stat_tile(len(m["ready"]), "Ready to work", "#e7f4ea", _GREEN)
        + _stat_tile(eyeball_n, "To check", eye_bg, eye_fg)
    )

    chips = ""
    for name, n in m["county_this"].most_common():
        label = name if name else "(blank)"
        was = m["county_prev"].get(name, 0)
        chips += _chip(f'{_esc(label)}&nbsp;<b>{n}</b>'
                       f'<span style="color:#9aa1a9;">&nbsp;&middot;&nbsp;was {was}</span>')

    parts: list[str] = []

    outage = m.get("gis_outage") or {}
    if outage.get("counties"):
        parts.append(
            f'<div style="{_F}background:#fef2f2;border:1px solid #fecaca;border-radius:8px;'
            f'padding:10px 14px;margin:14px 0 4px;color:#b91c1c;font-size:13px;font-weight:600;">'
            f'County GIS outage this run: {_esc(", ".join(outage["counties"]))} — rows for '
            f'these counties are incomplete and will be recovered on the next clean run.</div>')

    tn = _tonight_html(tonight or {})
    if tn:
        parts.append(_h("Tonight's run"))
        parts.append(tn)

    parts.append(_h("What to work on next"))
    if m["prios"]:
        items = "".join(
            f'<li style="{_F}font-size:13px;color:#513c06;line-height:1.55;'
            f'margin:0 0 7px;">{_esc(p)}</li>' for p in m["prios"])
        parts.append(
            f'<div style="background:#fffbeb;border:1px solid #fde68a;border-radius:8px;'
            f'padding:10px 14px 4px;"><ol style="margin:0;padding:0 0 0 18px;">{items}</ol></div>')
    else:
        parts.append(_p("Nothing urgent — the pipeline is clean.", color=_GREEN))

    parts.append(_h(f"New cases today ({len(m['new_today'])})"))
    if not m["new_today"]:
        parts.append(_p("None — the workbook is unchanged from yesterday.", color="#8a919a"))
    else:
        cap = 12
        rows = [[
            _case((r.get("Case No.") or "").strip()),
            _esc((r.get("County") or "").strip()),
            _esc((r.get("Deceased Owner") or "").strip()),
            _esc((r.get("Property Address") or "(no parcel)").strip()),
        ] for r in m["new_today"][:cap]]
        note = (f"...and {len(m['new_today']) - cap} more in the attached report"
                if len(m["new_today"]) > cap else "")
        parts.append(_tbl(["Case No.", "County", "Deceased", "Property"], rows, note))

    if m["eyeballs"]:
        parts.append(_h(f"Check before mailing ({len(m['eyeballs'])})"))
        parts.append(_p("The matcher made a judgment call on these — worth a look "
                        "before they get mail.", color="#8a919a", size="12.5px"))
        rows = [[
            _case((r.get("Case No.") or "").strip()),
            _esc((r.get("County") or "").strip()),
            _esc((r.get("Deceased Owner") or "").strip()),
            _esc(_why(tags[0])),
        ] for r, tags in m["eyeballs"][:8]]
        note = (f"...and {len(m['eyeballs']) - 8} more in the attached report"
                if len(m["eyeballs"]) > 8 else "")
        parts.append(_tbl(["Case No.", "County", "Deceased", "Why"], rows, note))

    parts.append(_h("Deep prospecting queue"))
    if not m["dp_queue"]:
        done = m["dp_already_done"]
        parts.append(_p("Empty" + (f" — {done} already researched." if done else "."),
                        color="#8a919a"))
    else:
        parts.append(_p(f'<b>{len(m["dp_queue"])}</b> &ldquo;Heirs of&rdquo; cases still '
                        f'need a named contact, pulled across {len(m["dp_by_week"])} '
                        f'week(s). The full case list is in the attached report.'))
        wk_chips = "".join(
            _chip(f'Wk {wk} <b>{n}</b><span style="color:#9aa1a9;">&nbsp;'
                  f'{_esc(_week_pull_dates(wk, yr))}</span>')
            for yr, wk, n in m["dp_by_week"][:8])
        more_wks = len(m["dp_by_week"]) - 8
        if more_wks > 0:
            wk_chips += (f'<span style="{_F}font-size:12px;color:#9aa1a9;">'
                         f'+{more_wks} more week(s)</span>')
        parts.append(f"<div>{wk_chips}</div>")

    kpi = kpi_totals(d)
    if kpi:
        wk_t, lw_t, stale = kpi

        def _kpirow(label: str, t: Counter) -> list[str]:
            ans = f"{t['answered'] / t['dials'] * 100:.0f}%" if t["dials"] else "-"
            return [_esc(label), str(t["dials"]), f"{t['answered']} ({ans})",
                    str(t["conversations"]), str(t["correct_numbers"]),
                    str(t["leads"]), f"{t['talk_seconds'] // 60}m"]

        note = ""
        if stale and (d - date.fromisoformat(stale)).days > 2:
            note = f"Numbers only current through {stale} — the KPI refresh hasn't run since."
        parts.append(_h("Phones this week"))
        parts.append(_tbl(["", "Dials", "Answered", "Convos", "Correct #s", "Leads", "Talk"],
                          [_kpirow("Mon-today", wk_t), _kpirow("Last week", lw_t)], note))
        if wk_t["sms_sent"] or lw_t["sms_sent"]:
            parts.append(_p(f"Texts sent: {wk_t['sms_sent']} this week, "
                            f"{lw_t['sms_sent']} last week.", color="#8a919a", size="12.5px"))

    parts.append(_h("This week so far"))
    phone_bit = (f"; {len(m['with_phone'])} already have a phone the pipeline found itself"
                 if m["with_phone"] else "")
    parts.append(_p(f"<b>{len(m['week_rows'])}</b> new estate cases matched to property, "
                    f"<b>{len(m['ready'])}</b> mail-ready with a named contact{phone_bit}."))
    if improvements:
        items = "".join(
            f'<li style="{_F}font-size:13px;color:#374151;line-height:1.55;'
            f'margin:0 0 6px;">{_esc(b)}</li>' for b in improvements)
        parts.append(_p("Pipeline improvements shipped this week:",
                        color="#6b7280", size="12.5px"))
        parts.append(f'<ul style="margin:0 0 8px;padding:0 0 0 18px;">{items}</ul>')

    dropped, pc = m["diffs"]["dropped"], m["diffs"]["parcel_changed"]
    if dropped or pc:
        bits = []
        if dropped:
            bits.append(f"<b>{len(dropped)}</b> row(s) dropped since yesterday "
                        f"({_esc(_cases_preview(dropped, 'Case No.'))})")
        if pc:
            bits.append(f"<b>{len(pc)}</b> case(s) had their parcel re-picked")
        parts.append(_h("Changes since yesterday"))
        parts.append(_p("; ".join(bits) + " — detail in the attached report."))

    q_line = pr_push_queue_line()
    if q_line:
        parts.append(
            f'<div style="{_F}background:#fffbeb;border:1px solid #fde68a;border-radius:8px;'
            f'padding:10px 14px;margin:14px 0 4px;color:#7a5300;font-size:13px;">'
            f'{_esc(q_line)}</div>')

    total_drops = sum(polish_stats.get(k, 0) for k in (
        "over_500k", "under_min_value", "heir_occupied", "commercial",
        "no_parcel", "archive_dupe", "non_pending"))
    fine = []
    if runtime_min is not None:
        fine.append(f"runtime {runtime_min:.0f} min")
    fine.append(f"{total_drops} rows dropped by filters tonight")
    fine.append(f"workbook total {workbook_total} cases")
    ts_line = tracerfy_budget_line()
    if ts_line:
        fine.append(ts_line)
    parts.append(
        f'<div style="{_F}font-size:11.5px;color:#9aa1a9;margin:18px 0 0;'
        f'border-top:1px solid #e4e8eb;padding-top:10px;">'
        + _esc(" / ".join(fine))
        + "<br>Full detail — match reasons, drop breakdown, DP ledger — is in the "
          "attached report file.</div>")

    body_html = "\n".join(parts)

    return f"""\
<!doctype html><html><body style="margin:0;padding:0;background:#eef0f2;">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#eef0f2;padding:24px 12px;">
<tr><td align="center">
<table role="presentation" width="700" cellpadding="0" cellspacing="0" style="max-width:700px;width:100%;background:#ffffff;border:1px solid #e2e5e9;border-radius:12px;overflow:hidden;{_F}">
  <tr><td style="background:{_GREEN};padding:22px 24px;">
    <div style="color:#ffffff;font-size:19px;font-weight:800;letter-spacing:.2px;">NC Probate Pipeline &middot; Week {week_n}</div>
    <div style="color:#bfe3c5;font-size:13px;margin-top:3px;">{date_str} &middot; {day_note}</div>
  </td></tr>
  <tr><td style="padding:14px 18px 4px;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>{tiles}</tr></table>
  </td></tr>
  <tr><td style="padding:6px 24px 0;">
    <div style="font-size:11px;font-weight:600;letter-spacing:.5px;text-transform:uppercase;color:#6b7280;margin-bottom:8px;">By county &mdash; this week vs last</div>
    <div>{chips}</div>
  </td></tr>
  <tr><td style="padding:0 24px 8px;">
    {body_html}
  </td></tr>
  <tr><td style="padding:12px 24px 22px;color:#9aa1a9;font-size:11px;">
    Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} &middot; SiftStack
  </td></tr>
</table>
</td></tr></table>
</body></html>"""


# ── Email ────────────────────────────────────────────────────────────


def _send_email_resend(
    subject: str, body: str, attachments: list[Path] | None = None,
    html: str | None = None,
) -> tuple[bool, str]:
    """Send via Resend HTTP API. Used when RESEND_API_KEY is in .env."""
    import base64
    import requests

    api_key = os.environ["RESEND_API_KEY"]
    sender = os.environ.get("REPORT_FROM", "reports@remedihomesolutions.com")
    recipients_raw = os.environ.get("REPORT_TO", "")
    recipients = [r.strip() for r in recipients_raw.split(",") if r.strip()]
    if not recipients:
        return False, "REPORT_TO not set in .env"

    payload: dict = {
        "from": sender,
        "to": recipients,
        "subject": subject,
        "text": body,   # plain-text fallback for clients that don't render HTML
    }
    if html:
        payload["html"] = html

    if attachments:
        attached = []
        for att in attachments:
            if not att.exists():
                continue
            attached.append({
                "filename": att.name,
                "content": base64.b64encode(att.read_bytes()).decode("ascii"),
            })
        if attached:
            payload["attachments"] = attached

    try:
        r = requests.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=30,
        )
    except requests.RequestException as e:
        return False, f"Resend HTTP error: {type(e).__name__}: {e}"
    if r.status_code >= 300:
        return False, f"Resend rejected (HTTP {r.status_code}): {r.text[:300]}"
    return True, f"Sent via Resend to {', '.join(recipients)} (id={r.json().get('id', '?')})"


def _send_email_smtp(
    subject: str, body: str, attachments: list[Path] | None = None,
    html: str | None = None,
) -> tuple[bool, str]:
    """Fallback: send via Gmail/Workspace SMTP using app password."""
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")
    report_to = os.environ.get("REPORT_TO") or smtp_user
    smtp_host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))

    if not smtp_user or not smtp_pass:
        return False, "Neither RESEND_API_KEY nor SMTP_USER/SMTP_PASS set in .env — skipping email."

    msg = EmailMessage()
    msg["From"] = smtp_user
    msg["To"] = report_to
    msg["Subject"] = subject
    msg.set_content(body)   # plain-text part
    if html:
        # multipart/alternative — clients pick HTML, fall back to text
        msg.add_alternative(html, subtype="html")

    for att in attachments or []:
        if not att.exists():
            continue
        ctype = "application/octet-stream"
        if att.suffix.lower() == ".xlsx":
            ctype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        maintype, subtype = ctype.split("/", 1)
        with att.open("rb") as fp:
            msg.add_attachment(fp.read(), maintype=maintype, subtype=subtype, filename=att.name)

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as s:
            s.starttls()
            s.login(smtp_user, smtp_pass)
            s.send_message(msg)
    except smtplib.SMTPAuthenticationError as e:
        return False, f"Gmail rejected the app password: {e}. Re-generate at https://myaccount.google.com/apppasswords"
    except Exception as e:
        return False, f"Email send failed: {type(e).__name__}: {e}"
    return True, f"Sent via SMTP to {report_to}"


def send_email(subject: str, body: str, attachments: list[Path] | None = None,
               html: str | None = None) -> tuple[bool, str]:
    """Send via Resend if RESEND_API_KEY is set, else fall back to SMTP."""
    if os.environ.get("RESEND_API_KEY"):
        return _send_email_resend(subject, body, attachments=attachments, html=html)
    return _send_email_smtp(subject, body, attachments=attachments, html=html)


# ── Main ─────────────────────────────────────────────────────────────


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-email", action="store_true", help="Skip the email send (file output only)")
    ap.add_argument("--attach-workbook", action="store_true", help="Attach the FTM workbook to the email")
    ap.add_argument("--include-heir-transfer", action="store_true",
                    help="Re-enable the heir-transfer review section + XLSX attachment "
                         "(off by default per Oren; the XLSX still generates in output/)")
    args = ap.parse_args(argv)

    # Top up the phone-KPI ledger (call activity from DataSift) so the PHONES
    # THIS WEEK section is current. Best-effort: a dead token, DataSift outage,
    # or timeout just leaves the section on yesterday's numbers. KPI_EMAIL=0
    # skips the refresh entirely.
    if os.environ.get("KPI_EMAIL", "1") != "0":
        try:
            subprocess.run(
                [sys.executable, str(Path(__file__).parent / "kpi_refresh.py")],
                capture_output=True, timeout=900,
                cwd=str(Path(__file__).parent.parent))
        except Exception:
            pass

    wb_path = find_latest_workbook()
    if not wb_path:
        print("ERROR: no FTM workbook found in output/", file=sys.stderr)
        return 1
    week_n, workbook_rows = read_workbook_cases(wb_path)

    today_datasift = find_latest_datasift(week_n)
    today_rows = read_csv_cases(today_datasift) if today_datasift else []
    yesterday_datasift = find_yesterdays_datasift(week_n, today_datasift) if today_datasift else None
    yesterday_rows = read_csv_cases(yesterday_datasift) if yesterday_datasift else []

    heir_path = find_latest_heir_transfer(week_n)
    heir_rows = read_heir_transfer(heir_path)

    polish_log = LOGS / "nc_daily_run.log"
    # since=now truncates the log at the last "=== Daily run started" marker so
    # drop counts reflect tonight's run, not the whole log history.
    polish_stats = parse_polish_log(polish_log, since=datetime.now())
    runtime = pipeline_runtime_min(polish_log)
    tonight = parse_tonight(tonight_log_slice(polish_log))
    model = compute_model(workbook_rows, week_n, today_rows, yesterday_rows, tonight)

    monday = date.today() - timedelta(days=date.today().weekday())
    improvements = summarize_week_improvements(monday)

    common = dict(
        workbook_path=wb_path,
        workbook_rows=workbook_rows,
        week_n=week_n,
        today_csv_rows=today_rows,
        yesterday_csv_rows=yesterday_rows,
        heir_transfer_rows=heir_rows,
        polish_stats=polish_stats,
        runtime_min=runtime,
        include_heir_transfer=args.include_heir_transfer,
        improvements=improvements,
        tonight=tonight,
        model=model,
    )
    # Full report → attached .txt (every fine-print section, uncapped lists);
    # condensed report → the email body itself (most important items only).
    report = render_report(**common)
    email_body = render_report(**common, condensed=True)

    out_path = OUTPUT / f"daily_report_{date.today().strftime('%Y-%m-%d')}.txt"
    out_path.write_text(report, encoding="utf-8")
    email_path = OUTPUT / f"daily_report_{date.today().strftime('%Y-%m-%d')}_email.txt"
    email_path.write_text(email_body, encoding="utf-8")
    print(f"Wrote {out_path}")
    print(f"Wrote {email_path}")
    print()
    print(report)

    if args.no_email:
        return 0

    attachments = [out_path]
    if heir_path and args.include_heir_transfer:
        attachments.append(heir_path)
    if args.attach_workbook:
        attachments.append(wb_path)

    new_count = len(model["new_today"])
    subject = (f"NC Probate Wk {week_n} — {new_count} new today, "
               f"{len(today_rows)} this week")

    # Build the HTML email from the same model as the text report. Best-effort:
    # if the HTML render ever breaks, the email still goes out text-only.
    try:
        html = render_html_email(
            model, tonight=tonight, week_n=week_n, improvements=improvements,
            runtime_min=runtime, polish_stats=polish_stats,
            workbook_total=len(workbook_rows),
        )
    except Exception as e:
        print(f"HTML email render failed ({type(e).__name__}: {e}) — "
              "sending text-only", file=sys.stderr)
        html = None

    ok, msg = send_email(subject, email_body, attachments=attachments, html=html)
    print(f"\nEmail: {msg}")
    return 0 if ok else 0  # File was written; email is best-effort


if __name__ == "__main__":
    sys.exit(main())
