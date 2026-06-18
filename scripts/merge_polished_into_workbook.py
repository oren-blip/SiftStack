"""Append new cases from today's polished CSV into the Week 23 workbook tab.

Strategy:
  - Read today's polished CSV (48 rows)
  - For each case NOT already in the workbook, append it as a new row
  - Preserve existing workbook rows AS-IS (Kinney SR-HEIRS swap, any other edits)
  - Match formatting from existing rows (copy styles from row above)
  - Sort by County then Case No. after appending
"""
from __future__ import annotations
import csv
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


POLISHED_CSV = Path("output/nc_estates_ftm_2026-06-03_202412_week23_datasift.csv")
WB_PATH = Path("output/FTM_2026_NC_Estates_throughWeek23.xlsx")
TAB = "Week 23 2026"


def main() -> int:
    if not POLISHED_CSV.exists():
        print(f"Polished CSV not found: {POLISHED_CSV}")
        return 1
    if not WB_PATH.exists():
        print(f"Workbook not found: {WB_PATH}")
        return 1

    # Load polished CSV
    with open(POLISHED_CSV, encoding="utf-8") as f:
        csv_rows = list(csv.DictReader(f))
    csv_cases = {r.get("Case No.", "").strip().upper() for r in csv_rows}
    print(f"Polished CSV: {len(csv_rows)} rows")

    # Load workbook
    wb = load_workbook(WB_PATH)
    ws = wb[TAB]
    header = [str(c.value or "").strip() for c in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(header)}
    print(f"Workbook headers: {len(header)} cols")

    # Existing case nos
    existing = {}
    last_data_row = 1
    for r in range(2, ws.max_row + 1):
        cn_cell = ws.cell(r, col["Case No."]).value
        if cn_cell:
            existing[str(cn_cell).strip().upper()] = r
            last_data_row = r
    print(f"Workbook existing rows: {len(existing)}")

    # Identify new cases
    new_cases = [c for c in csv_cases if c not in existing]
    print(f"New cases to append: {len(new_cases)}")

    # Template row (copy style from existing last data row)
    template_row = last_data_row

    appended = 0
    for cn in new_cases:
        src = next(r for r in csv_rows if r.get("Case No.", "").strip().upper() == cn)
        last_data_row += 1
        new_idx = last_data_row
        # Copy styles from template row
        for c in range(1, ws.max_column + 1):
            src_cell = ws.cell(template_row, c)
            tgt_cell = ws.cell(new_idx, c)
            if src_cell.has_style:
                tgt_cell.font = copy(src_cell.font)
                tgt_cell.fill = copy(src_cell.fill)
                tgt_cell.border = copy(src_cell.border)
                tgt_cell.alignment = copy(src_cell.alignment)
                tgt_cell.number_format = src_cell.number_format
        # Populate values
        for h, cidx in col.items():
            # Strip BOM-prefixed File Date if present
            v = src.get(h, "")
            if not v and h == "File Date":
                v = src.get("﻿File Date", "")
            ws.cell(new_idx, cidx).value = v
        appended += 1

    print(f"Appended {appended} new rows. Final row: {last_data_row}")

    # Sort: read all data rows, sort by County then Case No., rewrite values
    all_data = []
    for r in range(2, last_data_row + 1):
        row_vals = {h: ws.cell(r, col[h]).value for h in col}
        all_data.append(row_vals)
    all_data.sort(key=lambda d: (
        str(d.get("County", "") or ""),
        str(d.get("Case No.", "") or ""),
    ))
    # Re-write values (styles already set per-row; we keep them)
    for i, row_vals in enumerate(all_data, start=2):
        for h, cidx in col.items():
            ws.cell(i, cidx).value = row_vals.get(h)

    wb.save(WB_PATH)
    print(f"Saved {WB_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
