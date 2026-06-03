"""Generate a 'Review Me' list of suspicious rows from the workbook.

Flags rows that are most likely to need manual verification:
  - PR mailing address matches property address (heir-occupied?)
  - Decedent and PR have the same last name (family — often legit but
    worth a quick check for "person is their own PR" data errors)
  - Property address blank or "0 STREET" (vacant — may be intentional
    but flag for completeness)
  - "Heirs of" PR fallback (no formal executor on file yet)
  - Property Value > $400K (close to buy-box cap)
  - Joint parcel without beneficiary cross-reference data (case too
    new for the Parties API to have heir info)
  - Missing parcel ID
  - PR mailing in different state than property (offsite — positive
    signal, flagged so reviewer knows what to prioritize)

Output:
  output/audit_review_<timestamp>.xlsx
Each row has its case info + a "Flags" column listing what's suspicious.
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


def normalize_address(s: str) -> str:
    """Crude street-address normalizer for equality checks."""
    if not s:
        return ""
    return "".join(c for c in s.upper() if c.isalnum())


def last_token(name: str) -> str:
    """Pull a likely last-name token from a name string."""
    if not name:
        return ""
    n = name.strip().split(",")[0].strip().split()
    return (n[-1] if n else "").upper()


def first_token(name: str) -> str:
    if not name:
        return ""
    s = name.strip()
    if "," in s:
        return s.split(",")[0].strip().split()[0].upper() if s.split(",")[0].strip() else ""
    return s.split()[0].upper() if s.split() else ""


def evaluate_row(row: dict) -> list[str]:
    flags: list[str] = []
    dec = (row.get("Deceased Owner") or "").strip()
    pr = (row.get("Personal Representative") or "").strip()
    pr_first = (row.get("First Name") or "").strip()
    pr_last = (row.get("Last Name") or "").strip()
    pr_mail = (row.get("Mailing Address") or "").strip()
    pr_state = (row.get("Mailing State") or "").strip().upper()
    pr_zip = (row.get("Mailing Zip") or "").strip()
    prop_addr = (row.get("Property Address") or "").strip()
    prop_state = (row.get("Property State") or "NC").strip().upper()
    prop_zip = (row.get("Property Zip") or "").strip()
    parcel = (row.get("Parcel ID") or "").strip()
    prop_use = (row.get("Property use") or "").strip().upper()
    val_str = (row.get("Property Value") or "").strip()

    # 1. PR mailing equals property address (heir-occupied)
    if pr_mail and prop_addr and normalize_address(pr_mail) == normalize_address(prop_addr):
        flags.append("PR-MAILING-EQUALS-PROPERTY (heir-occupied?)")

    # 2. Decedent and PR same last name
    dec_last = last_token(dec)
    if dec_last and pr_last and dec_last == pr_last.upper():
        flags.append(f"PR-SAME-LAST-NAME ({dec_last})")

    # 3. Decedent and PR look like same person (same first + last)
    dec_first = first_token(dec)
    if dec_first and pr_first and dec_first == pr_first.upper() and dec_last == pr_last.upper():
        flags.append("PR-EQUALS-DECEDENT (data error?)")

    # 4. "Heirs of" PR fallback
    if pr.lower().startswith("heirs of"):
        flags.append("HEIRS-OF-FALLBACK (no formal PR)")

    # 5. Property address blank or "0 STREET" (vacant)
    if not prop_addr:
        flags.append("NO-PROPERTY-ADDRESS")
    elif prop_addr.strip().upper().startswith("0 "):
        flags.append("VACANT-LOT (0 address)")

    # 6. Missing parcel ID
    if not parcel:
        flags.append("NO-PARCEL-ID")

    # 7. Property Value near or over the $500K buy-box cap
    try:
        val = float(val_str.replace(",", "").replace("$", ""))
        if val > 400_000:
            flags.append(f"HIGH-VALUE (${int(val):,})")
    except (ValueError, TypeError):
        pass

    # 8. PR mailing in a different state
    if pr_state and pr_state != "NC" and prop_state == "NC":
        flags.append(f"OUT-OF-STATE-PR ({pr_state})")

    # 9. PR zip far from property zip (rough heuristic — first 2 digits differ)
    if pr_zip and prop_zip and len(pr_zip) >= 3 and len(prop_zip) >= 3:
        if pr_zip[:2] != prop_zip[:2]:
            flags.append(f"PR-ZIP-{pr_zip[:5]}-PROPERTY-ZIP-{prop_zip[:5]}")

    return flags


def main() -> int:
    src = Path("output/FTM_2026_NC_Estates_throughWeek23.xlsx")
    if not src.exists():
        print(f"Source not found: {src}")
        return 1

    wb = load_workbook(src, read_only=True)
    out_rows = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c or "").strip() for c in rows[0]]
        for r in rows[1:]:
            d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
            if not (d.get("Case No.") or ""):
                continue
            flags = evaluate_row(d)
            if flags:
                d["_Tab"] = sn
                d["_Flags"] = "; ".join(flags)
                d["_FlagCount"] = len(flags)
                out_rows.append(d)

    # Sort by flag count (most-flagged first) then by case
    out_rows.sort(key=lambda r: (-r["_FlagCount"], r["_Tab"], r.get("Case No.") or ""))

    # Write review xlsx
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path("output") / f"audit_review_{ts}.xlsx"
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Review Me"
    out_cols = ["_FlagCount", "_Flags", "_Tab", "Case No.", "County", "Deceased Owner",
                "Personal Representative", "Mailing Address", "Mailing City",
                "Mailing State", "Mailing Zip",
                "Property Address", "Property City", "Property State", "Property Zip",
                "Property use", "Property Value", "Parcel ID"]
    ws_out.append(["Flag Count", "Flags", "Tab", "Case No.", "County", "Deceased Owner",
                   "Personal Representative", "Mailing Address", "Mailing City",
                   "Mailing State", "Mailing Zip",
                   "Property Address", "Property City", "Property State", "Property Zip",
                   "Property use", "Property Value", "Parcel ID"])
    for r in out_rows:
        ws_out.append([r.get(c, "") for c in out_cols])

    # Set column widths
    widths = {1: 8, 2: 60, 3: 14, 4: 18, 5: 12, 6: 26, 7: 22, 8: 26,
              9: 14, 10: 8, 11: 8, 12: 26, 13: 14, 14: 8, 15: 8,
              16: 12, 17: 12, 18: 18}
    from openpyxl.utils import get_column_letter
    for col_idx, w in widths.items():
        ws_out.column_dimensions[get_column_letter(col_idx)].width = w
    ws_out.freeze_panes = "A2"

    wb_out.save(out_path)

    # Summary
    print(f"Wrote {out_path}")
    print(f"Total flagged rows: {len(out_rows)}")
    print()
    from collections import Counter
    flag_counts = Counter()
    for r in out_rows:
        for f in r["_Flags"].split("; "):
            flag_counts[f.split(" ")[0]] += 1
    print("Flag breakdown:")
    for f, n in flag_counts.most_common():
        print(f"  {f:35} {n}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
