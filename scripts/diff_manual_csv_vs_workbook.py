"""Compare Oren's manual Week 24 CSV vs the auto workbook (Week 24 tab).

Usage:
    python scripts/diff_manual_csv_vs_workbook.py \\
        --manual "C:/Users/omark/Downloads/FTM_2026_NC Estates - Week 24.csv"
"""
from __future__ import annotations
import argparse
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


def load_csv_cases(path: Path) -> dict[tuple[str, str], dict]:
    out: dict[tuple[str, str], dict] = {}
    with path.open(encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            cn = (r.get("Case No.") or "").strip().upper()
            cty = (r.get("County") or "").strip().title()
            if cn and cty:
                out[(cty, cn)] = r
    return out


def load_wb_tab(wb_path: Path, tab: str) -> dict[tuple[str, str], dict]:
    wb = load_workbook(wb_path, read_only=True)
    if tab not in wb.sheetnames:
        return {}
    ws = wb[tab]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(c or "").strip() for c in rows[0]]
    out: dict[tuple[str, str], dict] = {}
    for r in rows[1:]:
        d = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        cn = str(d.get("Case No.") or "").strip().upper()
        cty = str(d.get("County") or "").strip().title()
        if cn and cty:
            out[(cty, cn)] = d
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--manual", required=True)
    ap.add_argument("--workbook", default="output/FTM_2026_NC_Estates_throughWeek24.xlsx")
    ap.add_argument("--tab", default="Week 24 2026")
    args = ap.parse_args()

    manual = load_csv_cases(Path(args.manual))
    wb = load_wb_tab(Path(args.workbook), args.tab)

    print(f"\n{'=' * 70}")
    print(f"WEEK 24 DIFF")
    print(f"{'=' * 70}")
    print(f"  Manual:   {len(manual)} cases")
    print(f"  Workbook: {len(wb)} cases")

    in_both = sorted(set(manual) & set(wb))
    manual_only = sorted(set(manual) - set(wb))
    wb_only = sorted(set(wb) - set(manual))

    print(f"\n  In both: {len(in_both)} | Manual-only: {len(manual_only)} | Workbook-only: {len(wb_only)}")

    print(f"\n  {'County':14} {'Manual':>7} {'Wkbk':>7} {'Both':>5} {'M-only':>7} {'W-only':>7}")
    all_counties = sorted({k[0] for k in set(manual) | set(wb)})
    for c in all_counties:
        m = sum(1 for k in manual if k[0] == c)
        w = sum(1 for k in wb if k[0] == c)
        b = sum(1 for k in in_both if k[0] == c)
        mo = sum(1 for k in manual_only if k[0] == c)
        wo = sum(1 for k in wb_only if k[0] == c)
        print(f"  {c:14} {m:>7} {w:>7} {b:>5} {mo:>7} {wo:>7}")

    if manual_only:
        print(f"\n  --- MANUAL-ONLY (scraper MISSED these — high priority) ---")
        for k in manual_only:
            r = manual[k]
            print(f"    {k[1]:18}  {k[0]:12}  {(r.get('Deceased Owner') or '')[:35]:35}  prop={r.get('Property Address') or '(blank)'}")

    if wb_only:
        print(f"\n  --- WORKBOOK-ONLY (scraper extras — verify they're real) ---")
        for k in wb_only:
            r = wb[k]
            print(f"    {k[1]:18}  {k[0]:12}  {str(r.get('Deceased Owner') or '')[:35]:35}  prop={r.get('Property Address') or '(blank)'}")

    # Property/parcel disagreements on matched rows
    print(f"\n  --- FIELD DISAGREEMENTS ON MATCHED CASES ---")
    diff_count = 0
    for k in in_both:
        m = manual[k]
        w = wb[k]
        diffs = []
        for field in ("Deceased Owner", "Property Address", "Parcel ID"):
            mv = str(m.get(field) or "").strip().lower()
            wv = str(w.get(field) or "").strip().lower()
            if mv and wv and mv != wv:
                diffs.append(f"{field}: M={m.get(field)!r} | W={w.get(field)!r}")
        if diffs:
            diff_count += 1
            print(f"\n    {k[1]} ({k[0]})")
            for d in diffs:
                print(f"      {d[:180]}")
    if diff_count == 0:
        print("    (none)")

    return 0


if __name__ == "__main__":
    sys.exit(main())
