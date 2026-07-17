"""One-shot: re-lookup property addresses for rows with parcel-but-no-address
(fixes the 5-digit-street-number-misparsed-as-ZIP bug now patched in
nc_gis_lookup._candidate_to_address_parts), then run the DataSift prep
(drop no-parcel + heirs-of rewrite). Writes fresh timestamped output.
"""

from __future__ import annotations

import csv
import os
import re
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / "src"))

from nc_gis_lookup import (  # noqa: E402
    _candidate_to_address_parts,
    _name_match_score,
    lookup_properties,
    simplify_use_code,
    split_decedent_name,
)
from reenrich_ftm_executors import write_csv, write_xlsx  # noqa: E402


_PARCEL_PROPERTY_FIELDS = (
    "Parcel ID", "Property Address", "Property City",
    "Property State", "Property Zip", "Property use",
)


def tag_reason(row: dict, code: str) -> None:
    """Append a short audit code to row['Match Reason'].

    Empty Match Reason after polish = parcel + PR came directly from the
    scrape (the default happy path — high confidence). Anything tagged
    here means a polish step mutated the row, and the code names which
    step / what kind of fallback. Idempotent: skips duplicate codes so
    re-running polish over the same rows doesn't grow the cell.

    Convention: short kebab-case codes. Sort the column to triage low-
    confidence rows fast.
    """
    if not code:
        return
    existing = (row.get("Match Reason") or "").strip()
    if not existing:
        row["Match Reason"] = code
        return
    parts = [p.strip() for p in existing.split("|") if p.strip()]
    if code in parts:
        return
    parts.append(code)
    row["Match Reason"] = " | ".join(parts)


# Counties whose GIS is so slow that we trim variations to just two
# (the precise full-middle form + the middle-initial form). Cabarrus's
# polaris3g endpoint averages 4-6 min per call; cutting from 4 to 2
# variations saves ~50% of weekly polish time.
_SLOW_GIS_COUNTIES = {"cabarrus"}


def _name_variations(decedent: str, county: str = "") -> list[str]:
    """Variations to try when re-searching for the correct parcel.

    Order matters: the first variation that returns matches wins, and
    every variation we try costs another GIS round-trip. User rules
    (matches Oren's manual search workflow):
      1. "LAST FIRST MIDDLE" — county GIS indices return matches most
         reliably in this form.
      2. "LAST FIRST M" — many indices store owners with just the
         middle initial.
      3. decedent as-passed (usually "Last, First Middle" with comma)
         — fast counties only.

    Dropped 2026-05-30 per user: "FIRST MIDDLE LAST" and bare "LAST"
    variations — user never uses them in manual pulls and bare-LAST
    returns too many false hits to be worth the round-trip.

    Slow-GIS counties stop after the first two; fast-GIS counties
    also try the as-passed form.
    """
    first, mid, last = split_decedent_name(decedent)
    mid_initial = mid[0] if mid else ""
    # "Drop first" variation: deed sometimes records the decedent under
    # her MIDDLE name as the first name, with the court-record first
    # suppressed. Peacock Edith Kathryn Campbell (Rowan 26E000684-790
    # Week 26): court=Edith but deed owner='PEACOCK KATHRYN CAMPBELL'.
    # Trying "Peacock, Kathryn Campbell" would match. Only emit when we
    # have a multi-word middle (so "Smith, John A" doesn't trigger and
    # match every Smith A in the county).
    mid_words = mid.split() if mid else []
    drop_first_variation = ""
    if last and len(mid_words) >= 2:
        drop_first_variation = f"{last}, {mid}"

    if county.lower() in _SLOW_GIS_COUNTIES:
        if last and first and mid:
            raw = [
                f"{last} {first} {mid}".strip(),
                f"{last} {first} {mid_initial}".strip(),
                decedent,
                drop_first_variation,
            ]
        elif last and first:
            # No middle name — try LAST FIRST then the as-passed comma form
            # as fallback. The space-separated form maps to a strict LIKE
            # query in some county GIS layers (e.g. Cabarrus) and misses
            # deed-spelling variants; the as-passed form takes a different
            # search path that's more spelling-tolerant. Cost is one extra
            # GIS call when LAST FIRST returns 0, which is the case we
            # actually want to recover (e.g. Sega Paulene -> deed PAULINE).
            raw = [f"{last} {first}".strip(), decedent]
        else:
            raw = [decedent]
    else:
        raw = [
            f"{last} {first} {mid}".strip() if (last and first and mid) else None,
            f"{last} {first} {mid_initial}".strip() if (last and first and mid) else None,
            decedent,
            drop_first_variation,
        ]
    seen: set[str] = set()
    out: list[str] = []
    for v in raw:
        v_norm = (v or "").strip()
        if v_norm and v_norm.upper() not in seen:
            seen.add(v_norm.upper())
            out.append(v_norm)
    return out


def research_blank_parcels(
    rows: list[dict],
    min_score: float = 0.7,
    audit_rejected_pids: set[tuple[str, str]] | None = None,
) -> int:
    """For rows where Parcel ID is blank but we have a decedent name + county,
    re-search the county GIS with name variations and try to find the
    correct parcel using the middle-name-aware matcher. Recovers cases
    where the original (buggy-matcher) scrape picked the wrong parcel
    and the audit then blanked it — e.g. Osborne, James Lee → correct
    parcel exists but original search picked the wrong James D Osborne.

    `audit_rejected_pids` (set of (county_lower, pid)) is a blacklist:
    when the audit step just rejected a PID for this decedent because
    the GIS owner doesn't match (e.g. property held by a family trust,
    not the decedent personally), the re-search must NOT pick up that
    same PID under a broader name variation — that's how
    "Walker, Betty Louise" kept getting re-bound to the same Walker
    Family Trust parcel after the audit had explicitly rejected it.
    """
    from nc_gis_lookup import filter_for_lead_quality, pick_best_candidate, _name_match_score
    rejected = audit_rejected_pids or set()
    recovered = 0
    for r in rows:
        if (r.get("Parcel ID") or "").strip():
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or not county or "IN THE MATTER" in dec.upper():
            continue
        found = None
        used_variation = ""
        # Compute the drop-first variation form once so we can detect when
        # a hit came via that path (and treat its score differently).
        _first_dec, _mid_dec, _last_dec = split_decedent_name(dec)
        _mid_words = _mid_dec.split() if _mid_dec else []
        _drop_first_form = (
            f"{_last_dec}, {_mid_dec}".strip()
            if (_last_dec and len(_mid_words) >= 2) else ""
        )
        for v in _name_variations(dec, county):
            try:
                results = lookup_properties(v, county, min_score=min_score)
            except Exception:
                continue
            if not results:
                continue
            # Re-score each candidate against the CANONICAL decedent name,
            # not the search variation. The variation "JOHNSTON GERALDINE H"
            # gets misparsed by split_decedent_name as first=JOHNSTON, last=H
            # — which then matches "HARRILL JOHN H HEIRS" at 0.7 via the
            # prefix-escape. Re-scoring against the original "Johnston,
            # Geraldine Hagan" parses correctly and drops that false
            # positive to 0.0. (Week 26 Johnston 26E000837-350.)
            #
            # EXCEPTION: the "drop-first" variation (e.g. "Peacock, Kathryn
            # Campbell" when court has "Edith Kathryn Campbell") ALWAYS
            # re-scores below threshold because Edith isn't in the deed
            # owner — but the variation was DESIGNED for that case. When
            # the variation match scored 1.0 AND we're on the drop-first
            # variation, keep the variation-based score (Peacock 26E000684-790).
            is_drop_first_match = bool(_drop_first_form) and v.strip().upper() == _drop_first_form.upper()
            for c in results:
                canon_score = _name_match_score(dec, c.owner_name or "")
                var_score = c.match_score
                if is_drop_first_match and var_score >= 1.0:
                    c.match_score = var_score
                else:
                    c.match_score = canon_score
            results = [c for c in results if c.match_score >= min_score]
            if not results:
                continue
            kept = filter_for_lead_quality(results)
            # Drop any candidate whose PID was just rejected by the audit
            kept = [c for c in kept if (county.lower(), c.pid or "") not in rejected]
            if not kept:
                continue
            best = pick_best_candidate(kept, dec, r.get("Decedent Address") or "")
            if not best:
                continue
            found = best
            used_variation = v
            break
        if not found:
            continue
        street, city, zipc = _candidate_to_address_parts(found)
        r["Parcel ID"] = found.pid or ""
        _set_row_acres_from_candidate(r, found)
        r["Property Address"] = street
        r["Property City"] = city
        r["Property State"] = "NC"
        r["Property Zip"] = zipc
        if found.use_code:
            new_use = simplify_use_code(found.use_code, found.use_description, found.county)
            if new_use:
                r["Property use"] = new_use
        tag_reason(r, "name-research")
        if getattr(found, "is_heir_transferred", False):
            # The deed already moved to the next-generation heir
            # (decedent Jr -> deed III). Still a real lead but flag it
            # so the operator knows post-probate transfer already
            # happened — the heir is the legal owner now.
            tag_reason(r, "heir-transferred-deed")
            existing_notes = (r.get("Notes") or "").strip()
            xfer_note = f"[HEIR-TRANSFERRED-DEED owner={found.owner_name}]"
            r["Notes"] = (existing_notes + ("\n" if existing_notes else "") + xfer_note).strip()
        print(f"  Re-found {county}/{dec} via {used_variation!r}: {found.pid} {street}, {city} NC {zipc}")
        recovered += 1
    return recovered


def _candidate_to_address_parts(c) -> tuple[str, str, str]:
    """Extract (street, city, zip) from a PropertyCandidate for Notes display."""
    street = c.situs_address or ""
    city = c.situs_city_override or ""
    zipc = c.situs_zip_override or ""
    return street, city, zipc


# Legacy horizontal "PLUS N PARCELS" lines look like:
#   PLUS 2 PARCELS
#     56210485720000 | 830 Florence St Nw Concord 28027 | [SFR]
#     56118667130000 | 1015 Central Dr Nw Concord 28027
# Captures: pid, address, optional use bracket. Indent + pipe separators
# are diagnostic — the vertical format never produces this shape.
_LEGACY_HORIZONTAL_LINE_RE = re.compile(
    r"^\s+(?P<pid>\S+)\s*\|\s*(?P<addr>.+?)(?:\s*\|\s*\[(?P<use>[^\]]+)\])?\s*$"
)
_LEGACY_HEADER_RE = re.compile(r"^PLUS\s+\d+\s+PARCELS?(\s*\([^)]+\))?\s*$", re.IGNORECASE)


def reformat_legacy_horizontal_notes(rows: list[dict]) -> int:
    """Rewrite pre-2026-06-20 horizontal 'PLUS N PARCELS' Notes blocks
    into the current vertical format. Preserves any other Notes content
    (swap-on-DQ markers, 2nd-pass-obit annotations, already-vertical
    blocks added by polish backfill).

    Background: commit 22cea35 (2026-06-20) switched the scrape-time
    Notes format from a single-line pipe-separated layout to a vertical
    multi-line layout that's easier to scan. But rows scraped BEFORE
    that date kept their horizontal Notes through merge + polish, since
    the polish only APPENDS new vertical blocks for newly-discovered
    siblings. This step parses the legacy lines and rewrites them.
    """
    from nc_ftm_writer import format_extra_parcels_vertical
    updated = 0
    for r in rows:
        notes = (r.get("Notes") or "").strip()
        if not notes or "PLUS " not in notes.upper() or "|" not in notes:
            continue
        lines = notes.split("\n")
        # Find each legacy header + the contiguous parcel-line block under
        # it. Replace each such block with a vertical rendering.
        out_lines: list[str] = []
        i = 0
        changed = False
        while i < len(lines):
            line = lines[i]
            if _LEGACY_HEADER_RE.match(line.strip()):
                # Collect contiguous legacy parcel lines until a non-legacy line
                items: list[dict] = []
                j = i + 1
                while j < len(lines):
                    m = _LEGACY_HORIZONTAL_LINE_RE.match(lines[j])
                    if not m:
                        # Allow a single blank line between header and items
                        if not lines[j].strip() and j == i + 1:
                            j += 1
                            continue
                        break
                    items.append({
                        "pid": m.group("pid"),
                        "address": m.group("addr").strip(),
                        "use": (m.group("use") or "").strip(),
                    })
                    j += 1
                if items:
                    vertical_block = format_extra_parcels_vertical(items)
                    out_lines.append(vertical_block)
                    changed = True
                    i = j
                    continue
            out_lines.append(line)
            i += 1
        if changed:
            r["Notes"] = "\n".join(out_lines).strip()
            updated += 1
    return updated


def apply_fetched_case_docs(rows: list[dict]) -> int:
    """For each row, look up its Case ID (hex) in the fetched_case_docs.json
    store and populate the will_* / application_* columns if structured data
    is available. Mutates rows in place. Returns count of rows updated.

    This is the late-arrival apply path: when a Will or Application PDF
    wasn't yet uploaded at scrape time, the scraper queues the case;
    subsequent daily runs drain the queue and store any newly-arrived
    parsed docs in fetched_case_docs.json. This step picks that up and
    updates the corresponding merged-CSV row by case_id_hex lookup.

    Idempotent — re-running with the same fetched cache produces the same
    row state.
    """
    try:
        import case_doc_queue as cdq
    except Exception as e:
        print(f"  case_doc_queue unavailable ({e}) — skipping late-doc apply")
        return 0
    import json as _json
    fetched = cdq.load_fetched()
    if not fetched:
        return 0
    updated = 0
    for r in rows:
        case_hex = (r.get("Case ID (hex)") or "").strip()
        if not case_hex:
            continue
        cache_entry = fetched.get(case_hex)
        if not cache_entry:
            continue
        row_changed = False
        # Will data — ALWAYS re-derive (not just when blank). The scrape's
        # case-doc enrichment runs BEFORE Parties (cookie-freshness reason)
        # so it picks the primary executor by default. Once polish runs,
        # First/Last Name are populated from Parties — re-derive here so
        # the acting-executor pick is correct (catches the case where the
        # alternate is acting because the primary predeceased).
        will = cache_entry.get("will")
        if will:
            people = will.get("people") or []
            primary = next((p for p in people if p.get("role") == "primary_executor"), None)
            alternate = next((p for p in people if p.get("role") == "alternate_executor"), None)
            acting = primary
            if alternate:
                last_in_row = (r.get("Last Name") or "").strip().upper()
                first_in_row = (r.get("First Name") or "").strip().upper()
                alt_name = (alternate.get("full_name") or "").upper()
                if last_in_row and last_in_row in alt_name and (not first_in_row or first_in_row in alt_name):
                    acting = alternate
            new_pr = (acting.get("full_name") or "").strip() if acting else ""
            new_rel = (acting.get("relationship") or "").strip() if acting else ""
            if new_pr and (new_pr != (r.get("PR Full Name (Will)") or "").strip()
                           or new_rel != (r.get("PR Relationship (Will)") or "").strip()):
                r["PR Full Name (Will)"] = new_pr
                r["PR Relationship (Will)"] = new_rel
                row_changed = True
        # Application data
        app = cache_entry.get("application")
        if app and not (r.get("PR Full Name (App)") or "").strip():
            applicant = app.get("applicant") or {}
            r["PR Full Name (App)"] = (applicant.get("full_name") or "").strip()
            r["PR Relationship (App)"] = (applicant.get("relationship_to_decedent") or "").strip()
            r["Date of Death (App)"] = (app.get("date_of_death") or "").strip()
            r["Attorney (App)"] = (app.get("attorney_name") or "").strip()
            val = app.get("preliminary_estate_value_usd")
            if val is not None:
                try:
                    r["Estate Value (App)"] = f"{int(round(float(val))):,}"
                except (TypeError, ValueError):
                    pass
            # When App data has a confirmed applicant AND the current main
            # PR was inferred (DM-promoted from obit, beneficiary-promoted,
            # or generic "Heirs of"), OVERWRITE the main PR fields with
            # the court-confirmed applicant. The App PDF is more
            # authoritative than obituary-derived DM picks.
            # Kiser 26E002388-590 Week 26: obit gave us "Linda Kiser" as DM,
            # she got promoted to PR. App PDF says actual PR is "Robert
            # Dustin Kiser" at 111 Lookout Ridge Cedar Point — different
            # person. App data should win.
            current_reason = (r.get("Match Reason") or "").lower()
            current_pr = (r.get("Personal Representative") or "").strip().lower()
            inferred_pr = (
                "dm-promoted-pr" in current_reason
                or "beneficiary-promoted-pr" in current_reason
                or current_pr.startswith("heirs of")
                or current_pr.startswith("estate of")
            )
            app_pr_name = (applicant.get("full_name") or "").strip()
            if inferred_pr and app_pr_name:
                # Split full name into First/Last for the search-friendly columns
                parts = app_pr_name.split()
                if len(parts) >= 2:
                    r["Personal Representative"] = app_pr_name
                    r["First Name"] = parts[0]
                    r["Last Name"] = parts[-1]
                    # Apply applicant's mailing too — court-confirmed address
                    if applicant.get("street"):
                        r["Mailing Address"] = applicant.get("street", "")
                        r["Mailing City"] = applicant.get("city", "")
                        r["Mailing State"] = applicant.get("state") or "NC"
                        r["Mailing Zip"] = applicant.get("zip", "")
                    tag_reason(r, "pr-from-app-override")
            heirs = app.get("heirs") or []
            if heirs:
                try:
                    r["Heirs (App)"] = _json.dumps(heirs, separators=(",", ":"))
                except Exception:
                    pass
            row_changed = True
        if row_changed:
            tag_reason(r, "late-doc-apply")
            updated += 1
    return updated


def backfill_sibling_parcels_to_notes(rows: list[dict], min_score: float = 0.7) -> int:
    """For each row with a Parcel ID, check if there are OTHER high-score
    GIS candidates for the same decedent that aren't already mentioned in
    Notes. Add them as 'PLUS N PARCELS' so the user sees the full estate.

    The scrape's collapse_by_case (nc_ftm_writer) already does this at scrape
    time, but only when the scrape's GIS lookup returned multiple candidates.
    If the scrape only returned one (or the polish's matcher fix in
    commit 331adda found extras), siblings get lost. This step is the
    polish-time backstop.

    Returns the number of rows where siblings were added.
    """
    from nc_gis_lookup import lookup_properties, filter_for_lead_quality
    updated = 0
    for r in rows:
        current_pid = (r.get("Parcel ID") or "").strip()
        if not current_pid:
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or not county or "IN THE MATTER" in dec.upper():
            continue
        try:
            cands = lookup_properties(dec, county, min_score=min_score)
        except Exception:
            continue
        if len(cands) <= 1:
            continue
        kept = filter_for_lead_quality(
            cands,
            beneficiaries_json=r.get("Beneficiaries", "") or "",
            decedent_name=dec,
        )
        siblings = [c for c in kept if (c.pid or "") and c.pid != current_pid]
        if not siblings:
            continue
        existing_notes = (r.get("Notes") or "").strip()
        # Skip if any sibling PID is already mentioned in Notes
        siblings = [c for c in siblings if c.pid not in existing_notes]
        if not siblings:
            continue

        # Score + sort: best parcels first so scanning Notes is meaningful
        from nc_gis_lookup import simplify_use_code, parcel_quality_score
        scored: list[tuple[int, str, object]] = []
        for s in siblings:
            use = simplify_use_code(s.use_code, s.use_description, s.county) or ""
            if not use:
                if s.is_vacant_land:
                    use = "Vacant Land"
                elif s.is_residential:
                    use = "SFR"
            score, tier = parcel_quality_score(s, simplified_use=use)
            scored.append((score, tier, s))
        scored.sort(key=lambda t: -t[0])

        # Build vertical sibling block via shared formatter
        from nc_ftm_writer import format_extra_parcels_vertical
        items: list[dict] = []
        for i, (score, tier, s) in enumerate(scored):
            street, city, zipc = _candidate_to_address_parts(s)
            addr = " ".join(filter(None, [street, city, zipc])).strip()
            use = simplify_use_code(s.use_code, s.use_description, s.county) or ""
            if not use:
                if s.is_vacant_land:
                    use = "Vacant Land"
                elif s.is_residential:
                    use = "SFR"
            item = {
                "address": addr,
                "use": use,
                "lot": s.lot_area,
                "value": s.market_value,
                "pid": s.pid,
                "tier": tier,
                "score": score,
            }
            if i == 0:
                item["_header_suffix"] = "(auto-ranked)"
            items.append(item)
        new_block = format_extra_parcels_vertical(items)
        if existing_notes and "PLUS " in existing_notes.upper() and "PARCEL" in existing_notes.upper():
            # Already has PLUS-N-PARCELS — append fresh block separated by blank line
            r["Notes"] = (existing_notes + "\n\n" + new_block).strip()
        elif existing_notes:
            r["Notes"] = (existing_notes + "\n\n" + new_block).strip()
        else:
            r["Notes"] = new_block
        updated += 1
    return updated


_LOT_CLUSTER_SUFFIXES = {
    "ST", "STREET", "RD", "ROAD", "AVE", "AVENUE", "DR", "DRIVE",
    "LN", "LANE", "CT", "COURT", "PL", "PLACE", "HWY", "HIGHWAY",
    "WAY", "BLVD", "BOULEVARD", "CIR", "CIRCLE", "TRL", "TRAIL",
    "PKWY", "PARKWAY", "TER", "TERRACE", "CR",
}
_LOT_CLUSTER_DIRECTIONALS = {"N", "S", "E", "W", "NE", "NW", "SE", "SW"}


def _street_signature(addr: str) -> str:
    """Canonical street name for cluster comparison.

    Strips leading house number, trailing street-type suffix, and trailing
    directional (so "1015 GOLD MINE DR" and "1083 GOLD MINE DR" both
    reduce to "GOLD MINE"; "4054 WANDERING LN NE" reduces to "WANDERING").
    Returns "" when no recognizable street name remains.
    """
    if not addr:
        return ""
    tokens = re.sub(r"[^\w\s\-]", " ", addr.upper()).split()
    if not tokens:
        return ""
    # Drop leading numeric house-number token (allowing "1015A" / "1015-2" etc.)
    if re.match(r"^\d", tokens[0]):
        tokens = tokens[1:]
    # Drop trailing directional then suffix (or suffix then directional)
    for _ in range(2):
        if tokens and tokens[-1] in _LOT_CLUSTER_DIRECTIONALS:
            tokens = tokens[:-1]
            continue
        if tokens and tokens[-1] in _LOT_CLUSTER_SUFFIXES:
            tokens = tokens[:-1]
            continue
        break
    return " ".join(tokens)


def tag_lot_clusters(rows: list[dict], min_score: float = 0.7) -> int:
    """Flag rows whose decedent owns 2+ parcels on the same street.

    Per Oren's Week 26 feedback ([[feedback_consecutive_vacant_lots]]):
    estates with multiple parcels on a single street are HIGH-value leads
    — mobile-home-on-land buyers prize adjacent parcels for siting MHs,
    combining lots, or building a small park. Without an explicit tag the
    user has to read the Notes column to spot the pattern; this surfaces
    it in Tags + Match Reason so it's visible at a glance.

    Detection: 2+ kept candidates for the same decedent whose street
    signature (street name minus house number + suffix + directional)
    matches the main parcel's. Vacant-vs-improved is NOT required —
    same-street clustering alone is the signal. Tag stays generic so it
    also catches residential cluster cases (family compound, etc).
    """
    from nc_gis_lookup import lookup_properties, filter_for_lead_quality
    tagged = 0
    for r in rows:
        current_pid = (r.get("Parcel ID") or "").strip()
        if not current_pid:
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or not county or "IN THE MATTER" in dec.upper():
            continue
        try:
            cands = lookup_properties(dec, county, min_score=min_score)
        except Exception:
            continue
        if len(cands) < 2:
            continue
        kept = filter_for_lead_quality(
            cands,
            beneficiaries_json=r.get("Beneficiaries", "") or "",
            decedent_name=dec,
        )
        # Anchor street = current row's main parcel street
        main = next((c for c in kept if c.pid == current_pid), None)
        if main:
            anchor = _street_signature(main.situs_address or r.get("Property Address", ""))
        else:
            anchor = _street_signature(r.get("Property Address", ""))
        if not anchor:
            continue
        same_street = [
            c for c in kept
            if c.pid and _street_signature(c.situs_address or "") == anchor
        ]
        if len(same_street) < 2:
            continue
        # Add tag + match reason
        cluster_tag = "Lot Cluster"
        existing_tags = (r.get("Tags") or "").strip()
        if cluster_tag.lower() not in existing_tags.lower():
            r["Tags"] = (existing_tags + ", " if existing_tags else "") + cluster_tag
        tag_reason(r, f"lot-cluster-{len(same_street)}")
        tagged += 1
    return tagged


def collect_multi_parcel_estates(rows: list[dict], threshold: int = 5) -> list[dict]:
    """For each row whose decedent has `threshold`+ scoring parcels in GIS,
    collect a per-parcel detail list. Returns:
        [{row: <source>, parcels: [{pid, score, tier, use, situs, acres, owner}]}, ...]
    Caller writes this to a side XLSX so Oren can audit complex estates.

    Re-uses the same lookup that backfill_sibling_parcels_to_notes uses;
    they're sequential steps and share the per-process cache.
    """
    from nc_gis_lookup import lookup_properties, filter_for_lead_quality, simplify_use_code, parcel_quality_score
    out: list[dict] = []
    for r in rows:
        if not (r.get("Parcel ID") or "").strip():
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or not county or "IN THE MATTER" in dec.upper():
            continue
        try:
            cands = lookup_properties(dec, county, min_score=0.7)
        except Exception:
            continue
        if len(cands) < threshold:
            continue
        kept = filter_for_lead_quality(
            cands,
            beneficiaries_json=r.get("Beneficiaries", "") or "",
            decedent_name=dec,
        )
        if len(kept) < threshold:
            continue
        scored = []
        for c in kept:
            use = simplify_use_code(c.use_code, c.use_description, c.county) or ""
            if not use:
                if c.is_vacant_land:
                    use = "Vacant Land"
                elif c.is_residential:
                    use = "SFR"
            score, tier = parcel_quality_score(c, simplified_use=use)
            scored.append({
                "pid": c.pid,
                "score": score,
                "tier": tier,
                "use": use,
                "owner": c.owner_name,
                "situs": c.situs_address or "",
                "city": c.situs_city_override or "",
                "zip": c.situs_zip_override or "",
                "acres": c.lot_area,
                "value": c.market_value,
            })
        scored.sort(key=lambda x: -x["score"])
        out.append({"row": r, "parcels": scored})
    return out


def write_multi_parcel_estates_review(entries: list[dict], out_path: Path) -> None:
    """One row per (decedent x parcel), tier-labeled and sortable in Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Multi-Parcel Estates"

    headers = [
        "Case No.", "County", "Deceased Owner", "Personal Representative",
        "Parcel Count", "Tier", "Score",
        "Parcel ID", "Property Use", "Situs", "City", "Zip", "Acres", "Market Value",
        "Owner Name (GIS)",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    for c, name in enumerate(headers, start=1):
        cell = ws.cell(1, c, name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # Tier color hints
    tier_fills = {
        "T1": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),  # light green
        "T2": PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid"),  # light yellow
        "T3": PatternFill(start_color="FFCC80", end_color="FFCC80", fill_type="solid"),  # light orange
        "T4": PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"),  # light gray
    }

    row_idx = 2
    for entry in entries:
        src = entry["row"]
        parcels = entry["parcels"]
        for p in parcels:
            ws.cell(row_idx, 1, src.get("Case No.", ""))
            ws.cell(row_idx, 2, src.get("County", ""))
            ws.cell(row_idx, 3, src.get("Deceased Owner", ""))
            ws.cell(row_idx, 4, src.get("Personal Representative", ""))
            ws.cell(row_idx, 5, len(parcels))
            tier_cell = ws.cell(row_idx, 6, p["tier"])
            if p["tier"] in tier_fills:
                tier_cell.fill = tier_fills[p["tier"]]
            ws.cell(row_idx, 7, p["score"])
            ws.cell(row_idx, 8, p["pid"])
            ws.cell(row_idx, 9, p["use"])
            ws.cell(row_idx, 10, p["situs"])
            ws.cell(row_idx, 11, p["city"])
            ws.cell(row_idx, 12, p["zip"])
            ws.cell(row_idx, 13, p["acres"])
            ws.cell(row_idx, 14, p["value"])
            ws.cell(row_idx, 15, p["owner"])
            row_idx += 1

    widths = {1: 18, 2: 12, 3: 28, 4: 22, 5: 6, 6: 5, 7: 6, 8: 16,
              9: 14, 10: 28, 11: 14, 12: 8, 13: 8, 14: 12, 15: 32}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


_BENEFICIARY_ADDR_RE = re.compile(
    # Pulls "<housenum> <street name> ..." lines from the Beneficiaries
    # blob the scraper writes. The column is multi-line — each beneficiary
    # is name + address + city/state/zip. We just want the street line.
    r"\b(\d{1,6}\s+[A-Z][A-Z\s\-\.]+(?:RD|ROAD|ST|STREET|AVE|AVENUE|DR|DRIVE|LN|LANE|"
    r"CT|COURT|PL|PLACE|HWY|HIGHWAY|WAY|BLVD|BOULEVARD|CIR|CIRCLE|TRL|TRAIL|"
    r"PKWY|TER|TERRACE)\b)",
    re.IGNORECASE,
)


def address_fallback_from_beneficiaries(rows: list[dict], min_score: float = 0.7) -> int:
    """For rows with blank Parcel ID, scan the Beneficiaries column for
    addresses + look them up in county GIS. If we find a parcel where the
    owner's surname matches the decedent's, populate the row.

    Catches cases like Young Carl Sr. (Gaston) where the actual property
    address appears in the Beneficiaries column (an heir lives at the
    inherited property) and our regular name search missed the parcel
    because of a name-variation gap or stale endpoint.
    """
    from nc_gis_lookup import (
        lookup_by_address, split_decedent_name, _ARCGIS_CONFIG, _compose_address,
        simplify_use_code,
    )
    recovered = 0
    for r in rows:
        if (r.get("Parcel ID") or "").strip():
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        beneficiaries = (r.get("Beneficiaries") or "").strip()
        if not dec or not county or not beneficiaries:
            continue
        if "IN THE MATTER" in dec.upper():
            continue
        # Extract candidate addresses from the beneficiaries blob
        addresses = list({m.strip() for m in _BENEFICIARY_ADDR_RE.findall(beneficiaries)})
        if not addresses:
            continue
        _, _, dec_last = split_decedent_name(dec)
        if not dec_last:
            continue
        cfg = _ARCGIS_CONFIG.get(county.lower())
        if not cfg:
            continue
        owner_fields = cfg.get("owner_fields") or []
        for addr in addresses:
            try:
                hits = lookup_by_address(addr, county)
            except Exception:
                continue
            for attrs in hits:
                # Verify owner's surname matches the decedent's lastname token
                owner_str = " ".join(
                    str(attrs.get(of) or "") for of in owner_fields
                ).upper()
                if dec_last.upper() not in re.split(r"[^A-Z]+", owner_str):
                    continue
                # Populate the row
                pid = str(attrs.get(cfg.get("parcel_field") or "PIN") or "")
                if not pid:
                    continue
                situs = _compose_address(attrs, cfg.get("situs_fields") or [])
                r["Parcel ID"] = pid
                _set_row_acres_from_attrs(r, attrs)
                if situs:
                    r["Property Address"] = situs
                # Use code
                uc = (cfg.get("use_field") or "")
                ud = (cfg.get("use_desc_field") or "")
                use = simplify_use_code(
                    str(attrs.get(uc) or "") if uc else "",
                    str(attrs.get(ud) or "") if ud else "",
                    county.lower(),
                )
                if use:
                    r["Property use"] = use
                existing_notes = (r.get("Notes") or "").strip()
                tag = f"[ADDR-FALLBACK from beneficiary address {addr}]"
                r["Notes"] = (existing_notes + ("\n" if existing_notes else "") + tag).strip()
                tag_reason(r, "beneficiary-address")
                print(f"  ADDR-FALLBACK {county}/{dec}: matched {pid} at {situs!r} "
                      f"via beneficiary address {addr!r}")
                recovered += 1
                break  # one address per row is enough
            if (r.get("Parcel ID") or "").strip():
                break
    return recovered


# NC town -> candidate county(ies) for the 7 counties we cover. Used to route an
# Application-listed property address to the RIGHT county GIS — the whole point
# of the Application-address recovery is catching estates filed in one county
# with property in another (Barnette 26E002615-590: filed Mecklenburg, property
# 156 Manor Circle in Mooresville = Iredell). Towns that straddle two counties
# list both; we try each.
_NC_TOWN_COUNTY: dict[str, list[str]] = {
    "CONCORD": ["Cabarrus"], "HARRISBURG": ["Cabarrus"], "MOUNT PLEASANT": ["Cabarrus"],
    "MIDLAND": ["Cabarrus"], "KANNAPOLIS": ["Cabarrus", "Rowan"],
    "HICKORY": ["Catawba"], "NEWTON": ["Catawba"], "CONOVER": ["Catawba"],
    "CLAREMONT": ["Catawba"], "MAIDEN": ["Catawba", "Lincoln"], "CATAWBA": ["Catawba"],
    "LONG VIEW": ["Catawba"], "TERRELL": ["Catawba"], "SHERRILLS FORD": ["Catawba"],
    "GASTONIA": ["Gaston"], "BELMONT": ["Gaston"], "MOUNT HOLLY": ["Gaston"],
    "BESSEMER CITY": ["Gaston"], "CHERRYVILLE": ["Gaston"], "DALLAS": ["Gaston"],
    "STANLEY": ["Gaston"], "LOWELL": ["Gaston"], "MCADENVILLE": ["Gaston"],
    "CRAMERTON": ["Gaston"], "RANLO": ["Gaston"],
    "STATESVILLE": ["Iredell"], "MOORESVILLE": ["Iredell"], "TROUTMAN": ["Iredell"],
    "HARMONY": ["Iredell"], "UNION GROVE": ["Iredell"], "OLIN": ["Iredell"],
    "TURNERSBURG": ["Iredell"],
    "LINCOLNTON": ["Lincoln"], "DENVER": ["Lincoln"], "IRON STATION": ["Lincoln"],
    "VALE": ["Lincoln", "Catawba"], "CROUSE": ["Lincoln"],
    "CHARLOTTE": ["Mecklenburg"], "MATTHEWS": ["Mecklenburg"], "MINT HILL": ["Mecklenburg"],
    "HUNTERSVILLE": ["Mecklenburg"], "CORNELIUS": ["Mecklenburg"], "DAVIDSON": ["Mecklenburg"],
    "PINEVILLE": ["Mecklenburg"],
    "SALISBURY": ["Rowan"], "CHINA GROVE": ["Rowan"], "LANDIS": ["Rowan"],
    "ROCKWELL": ["Rowan"], "CLEVELAND": ["Rowan"], "GRANITE QUARRY": ["Rowan"],
    "SPENCER": ["Rowan"], "EAST SPENCER": ["Rowan"], "FAITH": ["Rowan"],
    "GOLD HILL": ["Rowan"], "MOUNT ULLA": ["Rowan"], "WOODLEAF": ["Rowan"],
}


def _town_from_address(addr: str) -> str:
    """Pull the town out of 'street, TOWN, NC 28115' (or 'street TOWN NC zip')."""
    s = (addr or "").strip()
    m = re.search(r",\s*([A-Za-z .'-]+?)\s*,?\s*NC\b", s)
    if m:
        return m.group(1).strip()
    m = re.search(r"\b([A-Za-z][A-Za-z .'-]+?)\s+NC\s+\d{5}", s)
    return m.group(1).strip() if m else ""


def recover_parcel_from_app_realestate(rows: list[dict]) -> int:
    """For rows STILL lacking a parcel, use the property address the Application
    itself lists ("real estate owned by decedent", Preliminary Inventory) to
    find the parcel — routing to the county inferred from the address's town, so
    an estate filed in one county with property in another is caught. This field
    is usually blank on the form, so hit rate is low — but when present it's the
    court stating the decedent's property outright, no name-matching needed.
    Reads the Application from the fetched-case-docs store (needs the doc to have
    been drained). Keeps the row's County (the filing county) per Oren's convention.
    """
    try:
        import case_doc_queue as cdq
        from nc_gis_lookup import (
            lookup_by_address, split_decedent_name, _ARCGIS_CONFIG, _compose_address,
            simplify_use_code,
        )
    except Exception as e:
        print(f"  app-address recovery unavailable ({e})")
        return 0
    fetched = cdq.load_fetched()
    if not fetched:
        return 0
    recovered = 0
    for r in rows:
        if (r.get("Parcel ID") or "").strip():
            continue
        case_hex = (r.get("Case ID (hex)") or "").strip()
        entry = fetched.get(case_hex) if case_hex else None
        app = (entry or {}).get("application") or {}
        addrs = app.get("real_estate_owned") or []
        if not addrs:
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        _f, _m, dec_last = split_decedent_name(dec)
        case_county = (r.get("County") or "").strip()
        done = False
        for addr in addrs:
            if not addr or done:
                continue
            town = _town_from_address(addr)
            counties = list(_NC_TOWN_COUNTY.get(town.upper(), []))
            if case_county and case_county not in counties:
                counties.append(case_county)
            for cty in counties:
                cfg = _ARCGIS_CONFIG.get(cty.lower())
                if not cfg or not cfg.get("url"):
                    continue  # ArcGIS counties only (Meck address path differs)
                try:
                    hits = lookup_by_address(addr, cty)
                except Exception:
                    hits = []
                owner_fields = cfg.get("owner_fields") or []
                for attrs in hits:
                    owner_str = " ".join(str(attrs.get(of) or "") for of in owner_fields).upper()
                    # Surname sanity check (deed may misspell — allow a 1-char
                    # slip via startswith on the 5-char stem). Barnett vs Barnette.
                    stem = (dec_last or "").upper()[:5]
                    if stem and stem not in owner_str and not any(
                        tok.startswith(stem) for tok in re.split(r"[^A-Z]+", owner_str)
                    ):
                        continue
                    pid = str(attrs.get(cfg.get("parcel_field") or "PIN") or "")
                    if not pid:
                        continue
                    situs = _compose_address(attrs, cfg.get("situs_fields") or [])
                    r["Parcel ID"] = pid
                    _set_row_acres_from_attrs(r, attrs)
                    if situs:
                        r["Property Address"] = situs
                    uc, ud = cfg.get("use_field") or "", cfg.get("use_desc_field") or ""
                    use = simplify_use_code(
                        str(attrs.get(uc) or "") if uc else "",
                        str(attrs.get(ud) or "") if ud else "", cty.lower())
                    if use:
                        r["Property use"] = use
                    notes = (r.get("Notes") or "").strip()
                    tag = f"[APP-REALESTATE {addr}" + (f" — {cty} county]" if cty != case_county else "]")
                    r["Notes"] = (notes + ("\n" if notes else "") + tag).strip()
                    tag_reason(r, "app-realestate-parcel")
                    print(f"  APP-REALESTATE {case_county}/{dec}: matched {pid} at "
                          f"{situs!r} in {cty} via Application address {addr!r}")
                    recovered += 1
                    done = True
                    break
                if done:
                    break
    return recovered


def crosscheck_parcel_vs_decedent_address(rows: list[dict]) -> tuple[int, int]:
    """Validate a NAME-matched parcel against the decedent's OWN address.

    Targets the common-name false positive (Week 28: Morris, Michael and
    Adams, James). On a common surname the name matcher picks one of several
    same-named people's parcels essentially at random, and the decedent's real
    home is not among the candidates, so the address tiebreaker can't help —
    it only ranks WITHIN the name-match set. This runs a second, independent
    check against the decedent's listed address.

    Two outcomes:

    SWAP ("addr-corrected") — the decedent's address resolves to a parcel whose
      owner carries the decedent's surname, and it differs from the picked
      parcel. The decedent literally owned the home they lived in; that beats
      any namesake guess. Repoint the row (value cleared so Step 1.6/1.7
      re-derive; flows through the rest of the pipeline normally).

    FLAG ("addr-uncorroborated") — the decedent's address is NOT surname-owned,
      the picked parcel is not on that street, AND the name search is ambiguous
      (2+ parcels tied at the top score). We can't corroborate the pick, so
      flag it for review rather than silently ship a maybe-wrong address. This
      is the Morris case (lived at 485 Cornwall, owned by Lenkiewicz; picked a
      different Morris) and Adams (lived at 204 Buckskin, owned by Rollinson;
      picked one of four James Adamses).

    Skips rows already sourced from the decedent's address (Step 0.64). Runs
    before sibling backfill / value / filters so a swap flows through cleanly.
    """
    from nc_gis_lookup import (
        lookup_by_address, lookup_properties, split_decedent_name,
        _ARCGIS_CONFIG, _compose_address, _name_match_score, _middle_match_strength,
    )

    def owner_str(hit: dict, owner_fields: list[str]) -> str:
        return " ".join(str(hit.get(f) or "") for f in owner_fields).upper()

    def pid_of(hit: dict, pid_field: str) -> str:
        v = str(hit.get(pid_field) or "").strip()
        return v.split(".", 1)[0] if re.fullmatch(r"\d+\.0+", v) else v

    def _distinctive(addr: str) -> set[str]:
        # Street-name words only — drop the house number, suffix, and BOTH
        # spelled-out and abbreviated directionals, so "901 North Oakland St"
        # and "N Oakland St" compare equal on {oakland}.
        return {t for t in _street_core_tokens(addr)
                if t.isalpha() and len(t) > 1 and t not in _DIRECTIONAL_WORDS}

    def same_street(a: str, b: str) -> bool:
        da, db = _distinctive(a), _distinctive(b)
        return bool(da and db and (da & db))

    swapped = flagged = 0
    for r in rows:
        pid = (r.get("Parcel ID") or "").strip()
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        dec_addr = (r.get("Decedent Address") or "").strip()
        if not (pid and dec and county and dec_addr):
            continue
        # Already sourced from the decedent's address — nothing to corroborate.
        if "decedent-address" in (r.get("Match Reason") or ""):
            continue
        cfg = _ARCGIS_CONFIG.get(county.lower())
        if not cfg:
            continue
        _f, _m, last = split_decedent_name(dec)
        if not last:
            continue
        try:
            addr_hits = lookup_by_address(dec_addr, county)
        except Exception:  # noqa: BLE001
            continue
        owner_fields = cfg.get("owner_fields") or []
        pid_field = cfg.get("parcel_field") or "PIN"
        last_u = last.upper()
        surname_hits = [h for h in addr_hits
                        if last_u in re.split(r"[^A-Z]+", owner_str(h, owner_fields))]
        # A swap must match the decedent's FULL name, not just the surname.
        # Families cluster: Samuel Morrison (26E000672-480) lived at 1911 Old
        # Wilkesboro, a parcel owned by MORRISON RUBY B HEIRS — a DIFFERENT
        # decedent (Ruby, 26E000673-480, next door). The surname gate matched
        # the wrong Morrison; a full-name gate rejects it.
        fullname_hits = [h for h in surname_hits
                         if _name_match_score(dec, owner_str(h, owner_fields)) >= 0.9]

        # ── SWAP: the decedent owned the home they lived in ──
        if fullname_hits:
            home = fullname_hits[0]
            home_pid = pid_of(home, pid_field)
            if home_pid and home_pid != pid:
                situs = _compose_address(home, cfg.get("situs_fields") or [])
                note = f"[ADDR-CORRECTED to decedent's home {dec_addr}; was parcel {pid}]"
                r["Parcel ID"] = home_pid
                _set_row_acres_from_attrs(r, home)
                if situs:
                    r["Property Address"] = situs
                r["Property Value"] = ""   # re-derived by Step 1.6/1.7
                existing = (r.get("Notes") or "").strip()
                r["Notes"] = (existing + ("\n" if existing else "") + note).strip()
                tag_reason(r, "addr-corrected")
                print(f"  ADDR-CORRECT {county}/{dec}: {pid} -> {home_pid} "
                      f"at {situs!r} (decedent's own address {dec_addr!r})")
                swapped += 1
            continue

        # ── FLAG: uncorroborated namesake pick ──
        if same_street(r.get("Property Address", ""), dec_addr):
            continue  # picked the street the decedent lived on — good enough
        try:
            cands = lookup_properties(dec, county)
        except Exception:  # noqa: BLE001
            cands = []
        top = [c for c in cands if c.match_score >= 0.99]
        if len(top) < 2:
            continue  # unique-ish name — not a coin-flip, don't cry wolf
        # Don't flag when the PICKED parcel's owner carries the decedent's FULL
        # middle name — that's a confident identity match no matter where they
        # lived, so the address-mismatch is just "owns a property they don't
        # occupy" (normal in probate). Week 28: Ferguson 26E000924-350, deed
        # "FERGUSON SUSAN CARTER" carries the decedent's middle "Carter" — a real
        # lead Oren kept. Jones (initial "L" only) and Adams (no middle) score
        # below 2 and stay flagged — both genuinely needed a look.
        picked = next((c for c in cands if (c.pid or "").strip() == pid), None)
        if picked and _middle_match_strength(picked.owner_name, dec) >= 2:
            continue
        who = owner_str(addr_hits[0], owner_fields).title().strip() if addr_hits else ""
        note = (f"[VERIFY LOW-CONFIDENCE: matched on common name among {len(top)} "
                f"{last.title()} parcels; decedent's address {dec_addr} "
                + (f"is owned by {who}" if who else "was not found in GIS")
                + f", not a {last.title()}]")
        existing = (r.get("Notes") or "").strip()
        if "VERIFY LOW-CONFIDENCE" not in existing.upper():
            r["Notes"] = (existing + ("\n" if existing else "") + note).strip()
            tag_reason(r, "addr-uncorroborated")
            print(f"  ADDR-FLAG {county}/{dec}: pick {pid} at "
                  f"{r.get('Property Address')!r} not corroborated by decedent "
                  f"address {dec_addr!r} ({len(top)} namesakes)")
            flagged += 1
    return swapped, flagged


def parcel_fallback_from_decedent_address(rows: list[dict]) -> int:
    """For rows with a blank Parcel ID, look up the DECEDENT'S OWN address —
    the one Odyssey already returns on the Parties API — in county GIS.

    This is the strongest signal we have and it was going unused. Name search
    fails whenever the deed spells the middle name differently from the court
    record. Iredell 26E000660-480: court says "Pierce, Gail H" (Hope), the deed
    says "PIERCE GAIL P HEIRS" (P = Pennell, her maiden name used as a middle
    initial — confirmed by her obituary, "Gail Hope Pennell Pierce"). The name
    matcher scores that 0.60 and auto-drops it. But the decedent's address on
    the case, 680 Lippard Farm Rd, resolves straight to both of her parcels.

    Distinct from address_fallback_from_beneficiaries (Step 0.65), which reads
    heir addresses out of the Beneficiaries blob. That cannot help here: every
    Pierce beneficiary lives at the executor's house, not at the property.

    Guarded by the same surname check as the beneficiary fallback, because a
    decedent's last address may be a rental or a nursing home. Requires the
    owner-of-record's surname to contain the decedent's last name.

    Needs a cached WAF cookie; silently no-ops without one.
    """
    import json as _json
    from pathlib import Path as _Path
    from nc_gis_lookup import (
        lookup_by_address, split_decedent_name, _ARCGIS_CONFIG, _compose_address,
        simplify_use_code, _name_match_score,
    )

    targets = [r for r in rows
               if not (r.get("Parcel ID") or "").strip()
               and (r.get("Case ID (hex)") or "").strip()
               and (r.get("Deceased Owner") or "").strip()
               and (r.get("County") or "").strip()]
    if not targets:
        return 0

    waf_path = _Path("ecourts_waf_cookies.json")
    if not waf_path.exists():
        print("  (no cached WAF cookie — skipping decedent-address fallback)")
        return 0
    try:
        waf = _json.loads(waf_path.read_text())
        from ecourts_case_api import CaseDetailClient
        client = CaseDetailClient(waf_token=waf["aws_waf_token"],
                                  user_agent=waf.get("user_agent") or "Mozilla/5.0")
    except Exception as e:  # noqa: BLE001
        print(f"  (decedent-address fallback unavailable: {e})")
        return 0

    recovered = 0
    for r in targets:
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        cfg = _ARCGIS_CONFIG.get(county.lower())
        if not cfg:
            continue
        _, _, dec_last = split_decedent_name(dec)
        if not dec_last:
            continue
        try:
            detail = client.fetch_detail((r.get("Case ID (hex)") or "").strip())
        except Exception as e:  # noqa: BLE001
            print(f"  decedent-addr: Parties fetch failed for {r.get('Case No.')}: {e}")
            continue
        dec_party = detail.decedent
        if not dec_party:
            continue
        addr = dec_party.first_address
        if addr.is_blank() or not addr.line1:
            continue
        try:
            hits = lookup_by_address(addr.line1, county)
        except Exception as e:  # noqa: BLE001
            print(f"  decedent-addr: GIS lookup failed for {addr.line1!r}: {e}")
            continue

        owner_fields = cfg.get("owner_fields") or []
        pid_field = cfg.get("parcel_field") or "PIN"

        # Keep every parcel at that address whose owner-of-record NAME-matches
        # the decedent. A single address routinely resolves to the house AND its
        # adjacent lot (Pierce: 680 Lippard Farm Rd + the lot behind it).
        #
        # Score against the full name, not just the surname: families cluster,
        # and a surname-only gate would assign a same-surname RELATIVE's parcel.
        # Samuel Morrison (26E000672-480) lived at 1911 Old Wilkesboro, owned by
        # MORRISON RUBY B HEIRS — a different decedent (Ruby, 26E000673-480). The
        # 0.5 floor accepts a decedent's own deed even when the middle initial
        # differs (Pierce "Gail H" vs deed "GAIL P" = 0.60) while rejecting a
        # wrong-first-name relative (Samuel vs "RUBY" = 0.40).
        matched: list[tuple[str, str, dict]] = []
        for attrs in hits:
            owner_str = " ".join(str(attrs.get(of) or "") for of in owner_fields).upper()
            if _name_match_score(dec, owner_str) < 0.5:
                continue
            pid = str(attrs.get(pid_field) or "").strip()
            # ArcGIS hands back float-ish ids ("4705686222.000"); Oren records
            # the bare number. Mirrors nc_gis_lookup._arcgis_to_candidate.
            if re.fullmatch(r"\d+\.0+", pid):
                pid = pid.split(".", 1)[0]
            if not pid:
                continue
            matched.append((pid, _compose_address(attrs, cfg.get("situs_fields") or []), attrs))
        if not matched:
            continue

        # Main parcel = the one with a street number (the house). Siblings —
        # typically vacant lots — go to Notes, per Oren's PLUS convention.
        matched.sort(key=lambda m: (not re.match(r"\s*\d", m[1] or ""), m[0]))
        pid, situs, attrs = matched[0]
        owner_str = " ".join(str(attrs.get(of) or "") for of in owner_fields).upper().strip()

        r["Parcel ID"] = pid
        _set_row_acres_from_attrs(r, attrs)
        if situs:
            r["Property Address"] = situs
        # The matched parcel IS the decedent's address, and Odyssey gives that
        # address's city + zip cleanly. Carry them — authoritative, no absentee
        # risk (unlike the county GIS, which has no reliable situs-city field:
        # Iredell's CityLocationDescription is often just "00", and CITY tracks
        # owner mailing). Fills Pierce Gail 26E000660-480 -> Statesville 28625.
        sf_city = cfg.get("situs_city_field")
        situs_city = str(attrs.get(sf_city) or "").strip() if sf_city else ""
        city = (addr.city or situs_city).strip()
        if not (r.get("Property City") or "").strip() and city:
            r["Property City"] = city.title()
        if not (r.get("Property Zip") or "").strip() and addr.zip:
            r["Property Zip"] = str(addr.zip)[:5]
        uc, ud = cfg.get("use_field") or "", cfg.get("use_desc_field") or ""
        use = simplify_use_code(
            str(attrs.get(uc) or "") if uc else "",
            str(attrs.get(ud) or "") if ud else "",
            county.lower(),
        )
        if use:
            r["Property use"] = use

        note_lines = [f"[ADDR-FALLBACK from decedent address {addr.line1}]"]
        for sib_pid, sib_situs, _a in matched[1:]:
            note_lines.append(f"PLUS {sib_pid}\n{sib_situs or '(no situs)'}")
        existing = (r.get("Notes") or "").strip()
        r["Notes"] = (existing + ("\n" if existing else "") + "\n".join(note_lines)).strip()
        tag_reason(r, "decedent-address")
        extra = f" (+{len(matched)-1} sibling parcel(s))" if len(matched) > 1 else ""
        print(f"  DECEDENT-ADDR {county}/{dec}: matched {pid} at {situs!r} "
              f"via decedent address {addr.line1!r} (owner {owner_str!r}){extra}")
        recovered += 1
    return recovered


def address_lookup_for_small_estate_disposed(rows: list[dict]) -> int:
    """For Small Estate Disposed-recent rows with blank Parcel ID, look up
    the row's Mailing Address (populated from the Interested Person on
    the OData Parties response) in county GIS and accept the first hit.

    No surname-on-owner check (unlike address_fallback_from_beneficiaries)
    because Small Estate Affidavits often transfer title to the heir
    immediately — the deed is in their name, not the decedent's. The
    Disposed-recent + Small-Estate gate is the signal that this row is
    a Small Estate Affidavit lead.
    """
    from nc_gis_lookup import (
        lookup_by_address, _ARCGIS_CONFIG, _compose_address, simplify_use_code,
    )
    recovered = 0
    for r in rows:
        if (r.get("Parcel ID") or "").strip():
            continue
        if not _is_small_estate_disposed_recent(r):
            continue
        mailing = (r.get("Mailing Address") or "").strip()
        county = (r.get("County") or "").strip()
        if not mailing or not county:
            continue
        cfg = _ARCGIS_CONFIG.get(county.lower())
        if not cfg:
            continue
        try:
            hits = lookup_by_address(mailing, county)
        except Exception:
            continue
        if not hits:
            continue
        attrs = hits[0]
        pid = str(attrs.get(cfg.get("parcel_field") or "PIN") or "")
        if not pid:
            continue
        situs = _compose_address(attrs, cfg.get("situs_fields") or [])
        r["Parcel ID"] = pid
        _set_row_acres_from_attrs(r, attrs)
        if situs:
            r["Property Address"] = situs
        # Use code
        uc = (cfg.get("use_field") or "")
        ud = (cfg.get("use_desc_field") or "")
        use = simplify_use_code(
            str(attrs.get(uc) or "") if uc else "",
            str(attrs.get(ud) or "") if ud else "",
            county.lower(),
        )
        if use:
            r["Property use"] = use
        existing_notes = (r.get("Notes") or "").strip()
        note_tag = f"[SMALL-ESTATE matched {pid} via Interested Person address]"
        r["Notes"] = (existing_notes + ("\n" if existing_notes else "") + note_tag).strip()
        tag_reason(r, "small-estate-address")
        print(f"  SMALL-ESTATE {county}/{r.get('Deceased Owner','')}: matched {pid} at {situs!r} "
              f"via mailing {mailing!r}")
        recovered += 1
    return recovered


def collect_heir_transfer_candidates(rows: list[dict]) -> list[dict]:
    """For each row whose Parcel ID is still blank after re-search, query the
    county GIS for parcels whose NAME1/NAME2 ends in the decedent's surname
    (married-name pattern). Returns a list of dicts:
        {row: <source row>, candidates: [{pid, name1, name2, situs}, ...]}
    Empty list if no candidates anywhere.

    Only supports ArcGIS counties — Catawba PHP and Mecklenburg polaris3g
    don't expose the same WHERE-clause API and would need separate paths.
    """
    from nc_gis_lookup import find_heir_transfer_candidates
    out: list[dict] = []
    for r in rows:
        if (r.get("Parcel ID") or "").strip():
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or not county or "IN THE MATTER" in dec.upper():
            continue
        try:
            cands = find_heir_transfer_candidates(dec, county)
        except Exception as e:
            print(f"  heir-transfer query failed for {county}/{dec}: {e}")
            continue
        if not cands:
            continue
        out.append({"row": r, "candidates": cands})
    return out


def write_heir_transfer_review(entries: list[dict], out_path: Path) -> None:
    """Write a review-me XLSX flagging decedents with possible heir-transferred
    parcels. One row per (decedent x candidate) so the user can sort/filter.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Heir Transfer Review"

    headers = [
        "Case No.", "County", "Deceased Owner", "Personal Representative",
        "Beneficiaries",
        "Candidate PIN", "Candidate NAME1", "Candidate NAME2", "Candidate Situs",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    for c, name in enumerate(headers, start=1):
        cell = ws.cell(1, c, name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    row_idx = 2
    for entry in entries:
        src = entry["row"]
        cands = entry["candidates"]
        for cand in cands:
            ws.cell(row_idx, 1, src.get("Case No.", ""))
            ws.cell(row_idx, 2, src.get("County", ""))
            ws.cell(row_idx, 3, src.get("Deceased Owner", ""))
            ws.cell(row_idx, 4, src.get("Personal Representative", ""))
            ws.cell(row_idx, 5, (src.get("Beneficiaries") or "")[:300])
            ws.cell(row_idx, 6, cand.get("pid", ""))
            ws.cell(row_idx, 7, cand.get("name1", ""))
            ws.cell(row_idx, 8, cand.get("name2", ""))
            ws.cell(row_idx, 9, cand.get("situs", ""))
            row_idx += 1

    widths = {1: 18, 2: 12, 3: 28, 4: 22, 5: 40, 6: 16, 7: 28, 8: 28, 9: 32}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# A single decedent's name search should return a handful of parcels. These
# ceilings flag a corrupt/oversized GIS response (Cabarrus garbage-200) so it
# can't overwrite a correct existing parcel. Real large landowners top out well
# under these; the garbage response returned 146.
_VALIDATE_ANOMALY_TOTAL = 40   # total candidates (>= 0.4)
_VALIDATE_ANOMALY_HIGH = 20    # high-scoring candidates (>= min_score)


def validate_existing_matches(
    rows: list[dict], min_score: float = 0.7,
) -> tuple[int, set[tuple[str, str]]]:
    """Re-score each parceled row's GIS match with the current matcher.
    Blank the parcel + property fields on rows where the match no longer
    passes — these were false-positive homonym matches under the OLD
    matcher (e.g. "James Lee Osborne" vs "JAMES D OSBORNE").

    Returns (blanked_count, rejected_pids) where rejected_pids is a set
    of (county_lower, pid) tuples that the re-search step should treat
    as a blacklist — see research_blank_parcels.
    """
    # Group rows by (county, decedent) so we only call lookup_properties
    # once per unique decedent (most have a single parcel anyway).
    from collections import defaultdict
    by_decedent: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for r in rows:
        if not (r.get("Parcel ID") or "").strip():
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or not county or "IN THE MATTER" in dec.upper():
            continue
        by_decedent[(county, dec)].append(r)

    blanked = 0
    rejected_pids: set[tuple[str, str]] = set()
    for (county, dec), parcel_rows in by_decedent.items():
        try:
            # Use a wide min_score so we get ALL candidates including ones
            # that scored above 0.4 (lastname-only) under the old matcher
            results = lookup_properties(dec, county, min_score=0.4)
        except Exception as e:
            print(f"  ERROR validating {dec}: {e}")
            continue

        # Anomaly guard: Cabarrus (and any flaky ArcGIS server) intermittently
        # returns a bogus oversized result set — a name search for one decedent
        # coming back with dozens/hundreds of "matches" is a corrupt response,
        # not real ownership. Trusting it here BLANKS a correct existing parcel
        # (its real owner name isn't in the garbage set, or a garbage parcel
        # out-ranks it on repick). Barbee 26E000709-120 kept dropping this way
        # (146 matches one minute, the correct 1 the next). When the result set
        # is implausibly large, keep the existing parcels untouched and let a
        # later run (with a clean response) validate them.
        high_scoring = [c for c in results if c.match_score >= min_score]
        if len(results) > _VALIDATE_ANOMALY_TOTAL or len(high_scoring) > _VALIDATE_ANOMALY_HIGH:
            print(f"  ANOMALOUS GIS result {county}/{dec}: {len(results)} candidates "
                  f"({len(high_scoring)} >= {min_score}) — flaky server, keeping existing "
                  f"parcel(s) un-validated this run")
            continue

        # Index candidates by pid
        by_pid = {c.pid: c for c in results}
        for r in parcel_rows:
            pid = r["Parcel ID"]
            cand = by_pid.get(pid)
            if cand is None:
                # Parcel ID no longer in result set — keep as-is (could be
                # data drift); print warning so user can spot-check
                print(f"  STALE {county}/{dec} pid={pid}: not in current GIS results")
                continue
            new_score = _name_match_score(dec, cand.owner_name)
            if new_score < min_score:
                print(f"  REJECT {county}/{dec} pid={pid}: owner={cand.owner_name!r} "
                      f"score={new_score:.2f} (was accepted under old matcher)")
                # Blank the parcel + property fields so prepare_for_datasift
                # treats this as a no-parcel row and drops it
                for col in _PARCEL_PROPERTY_FIELDS:
                    r[col] = ""
                tag_reason(r, "audit-blanked")
                blanked += 1
                rejected_pids.add((county.lower(), pid))
                continue
            # Re-pick: if a strictly-better candidate exists, swap to it.
            # Catches the case where the original pick passes scoring but a
            # better parcel exists (e.g. Kinney current = LEONARD HEIRS no
            # suffix, better = LEONARD SR HEIRS with suffix; Keller current
            # = 0 Mt Hope vacant, better = 3820 Mt Hope SFR).
            from nc_gis_lookup import filter_for_lead_quality, pick_best_candidate
            kept = filter_for_lead_quality(
                [c for c in results if c.match_score >= min_score],
                beneficiaries_json=r.get("Beneficiaries", "") or "",
                decedent_name=dec,
            )
            best = pick_best_candidate(kept, dec, r.get("Decedent Address") or "")
            if best is None or best.pid == pid:
                continue
            # Buy-box guard: never repick from an in-box parcel to an over-cap
            # one. Marinakos, Peter 26E002614-590 Week 29: decedent owns 826
            # Ashmore ($254K SFR, co-owned w/ a business partner) and 11440
            # Bloomfield ($500K+, co-owned w/ the surviving spouse). Both score
            # equally on name, so the value tiebreaker repicked to the over-cap
            # Bloomfield — which Step 1.8 then swapped to a THIRD parcel
            # (Summerlin $458K), burying Oren's actual in-box lead. If the
            # current main is under its cap and the "better" candidate is over
            # its cap, keep the current one.
            def _cand_over_cap(c) -> bool:
                use = simplify_use_code(c.use_code, c.use_description, c.county) or ""
                is_vac = ("VACANT" in use.upper() or "LAND" in use.upper()
                          or bool(getattr(c, "is_vacant_land", False)))
                cap = _cap_for_use("VACANT LAND" if is_vac else use, _cand_acres(c))
                if c.market_value is None:
                    return False
                return float(c.market_value) > cap
            if _cand_over_cap(best) and not _cand_over_cap(cand):
                print(f"  REPICK-SKIP {county}/{dec}: better parcel {best.pid} "
                      f"is over buy-box cap; keeping in-box {pid}")
                continue
            # Only swap if the rank really moved (don't churn on equal tuples).
            # Use the 3-way _suffix_match_score (matches pick_best_candidate
            # from commit c1db989). The old boolean _owner_has_suffix returned
            # False for BOTH "Smith Thomas E" and "Smith Thomas Jr" when the
            # decedent had no suffix, leaving them tied — then audit-repick
            # fell through to value tiebreaker and picked the higher-value JR
            # (the wrong person). Smith Thomas Edward 26E000638-480 Week 26.
            # Key must mirror nc_gis_lookup.pick_best_candidate.sort_key exactly,
            # including the middle-name term. Without it, repick could swap away
            # from a parcel whose deed carries the decedent's middle initial to a
            # higher-value parcel that merely shares first+last (Week 28: decedent
            # "Miller, David Allen", correct deed "MILLER DAVID A", competing deed
            # "MILLER DAVID & GINGER REVOCABLE TRUST" — both score 1.00 because a
            # deed that omits the middle name is not penalized).
            from nc_gis_lookup import (
                _suffix_match_score, _use_tier, _extract_suffix, _middle_match_strength,
            )
            suffix = _extract_suffix(dec)
            current_key = (
                _suffix_match_score(cand.owner_name, suffix),
                cand.match_score,
                _middle_match_strength(cand.owner_name, dec),
                _use_tier(cand),
                float(cand.market_value or 0),
                float(cand.lot_area or 0),
            )
            best_key = (
                _suffix_match_score(best.owner_name, suffix),
                best.match_score,
                _middle_match_strength(best.owner_name, dec),
                _use_tier(best),
                float(best.market_value or 0),
                float(best.lot_area or 0),
            )
            if best_key <= current_key:
                continue
            print(f"  REPICK {county}/{dec}: {pid} -> {best.pid} "
                  f"(was {cand.owner_name!r} use_tier={_use_tier(cand)}; "
                  f"now {best.owner_name!r} use_tier={_use_tier(best)})")
            tag_reason(r, "audit-repick")
            street, city, zipc = _candidate_to_address_parts(best)
            r["Parcel ID"] = best.pid or ""
            _set_row_acres_from_candidate(r, best)
            r["Property Address"] = street
            r["Property City"] = city
            r["Property State"] = "NC"
            r["Property Zip"] = zipc
            if best.use_code:
                new_use = simplify_use_code(best.use_code, best.use_description, best.county) or ""
                if new_use:
                    r["Property use"] = new_use
            elif best.is_vacant_land:
                r["Property use"] = "Vacant Land"
            elif best.is_residential:
                r["Property use"] = "SFR"
            # Same fix as re_collapse_multi_parcel: refresh value from
            # the new parcel or blank so Step 1.7 re-derives. Without
            # this the stale value from the swapped-out parcel survives.
            # (Walton 26E002339-590 leaked $7,900 on a $260,400 house.)
            if best.market_value:
                r["Property Value"] = f"{int(round(float(best.market_value))):,}"
            else:
                r["Property Value"] = ""
            # Notes dedup: same pattern as swap-on-DQ. When audit-repick
            # promotes a parcel that was previously listed in Notes as a
            # sibling (via Step 0.7 sibling-backfill), strip that block so
            # the new main doesn't appear duplicated in Notes.
            # Edwards 26E002336-590 Week 26: audit-repick promoted Crown
            # Colony Dr to main but Notes still showed it as "PLUS 1 PARCEL".
            existing_notes = (r.get("Notes") or "").strip()
            new_pid = (best.pid or "").strip()
            if existing_notes and new_pid:
                r["Notes"] = _strip_parcel_from_notes(existing_notes, new_pid)
    return blanked, rejected_pids


# Property uses that we treat as "verify" because the original Cabarrus
# 'I' code (improved-but-no-detail) used to map to Commercial before the
# per-county short-circuit landed. Re-fetching them re-runs the now-fixed
# classifier.
_SUSPECT_USES = {"COMMERCIAL", "INDUSTRIAL", "OFFICE"}


def repair_addresses(rows: list[dict]) -> int:
    """Re-lookup property fields (street/city/zip + use) for rows where
    the original GIS data was incomplete, mangled by the old situs-parser
    bug, or pre-fix-era misclassified as Commercial.
    """
    repaired = 0
    for r in rows:
        pid = (r.get("Parcel ID") or "").strip()
        if not pid:
            continue
        prop_addr = (r.get("Property Address") or "").strip()
        prop_zip = (r.get("Property Zip") or "").strip()
        prop_use = (r.get("Property use") or "").strip()
        suspect_use = prop_use.upper() in _SUSPECT_USES
        # Re-fetch if anything is missing OR if Property use is suspect
        if prop_addr and prop_zip and prop_use and not suspect_use:
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or "IN THE MATTER" in dec.upper():
            continue
        try:
            results = lookup_properties(dec, county, min_score=0.5)
        except Exception as e:
            print(f"  ERROR for {dec}: {e}")
            continue
        match = next((c for c in results if c.pid == pid), None)
        if not match:
            print(f"  No matching parcel {pid} in fresh lookup for {dec}")
            continue
        street, city, zipc = _candidate_to_address_parts(match)
        changed: list[str] = []
        if street and not prop_addr:
            r["Property Address"] = street
            r["Property City"] = city
            r["Property State"] = "NC"
            changed.append("addr")
        if zipc and not prop_zip:
            r["Property Zip"] = zipc
            changed.append("zip")
        if match.use_code:
            new_use = simplify_use_code(match.use_code, match.use_description, match.county)
            if new_use and (not prop_use or (suspect_use and new_use.upper() not in _SUSPECT_USES)):
                old = prop_use
                r["Property use"] = new_use
                changed.append(f"use({old!r}->{new_use!r})" if old else "use")
        if changed:
            print(f"  Fixed {r['County']} {pid}: {','.join(changed)} -> "
                  f"{r['Property Address']}, {r['Property City']} NC {r['Property Zip']} "
                  f"[{r.get('Property use', '')}]")
            repaired += 1
    return repaired


_BAD_CITY_TOKENS = {
    "DR","ST","RD","LN","CT","AVE","BLVD","WAY","CIR","PL","TC","TER","TR","TRL","PKWY","HWY",
    "N","S","E","W","NE","NW","SE","SW",
    # Common abbreviated street suffixes seen in mangled city fields
    # (Sellers Irma Elizabeth 26E000406-540: city="Av" leftover from "Sherrill Avenue")
    "AV","CR","BL","HW","PK","PY","TL",
}
_US_STATE_CODES = {
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA","KS","KY","LA",
    "ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ","NM","NY","NC","ND","OH","OK",
    "OR","PA","RI","SC","SD","TN","TX","UT","VT","VA","WA","WV","WI","WY","DC",
}


_CENTROID_GEOCODE_CACHE = Path("output") / ".nc_centroid_geocode_cache.json"


def _nominatim_forward(query: str) -> dict | None:
    """Forward-geocode a full address string to {city, postcode} via Nominatim.

    Fallback for parcels with no map geometry (Mecklenburg polaris3g). Needs a
    city in the query to be unambiguous. 1 req/sec — caller paces.
    """
    import requests
    try:
        r = requests.get("https://nominatim.openstreetmap.org/search", params={
            "q": query, "countrycodes": "us", "format": "json",
            "addressdetails": "1", "limit": "1",
        }, headers={"User-Agent": "SiftStack-NC/1.0"}, timeout=10)
        if r.status_code != 200:
            return None
        hits = r.json()
    except Exception:  # noqa: BLE001
        return None
    if not hits:
        return None
    a = hits[0].get("address", {})
    city = a.get("city") or a.get("town") or a.get("village") or a.get("hamlet") or ""
    return {"city": city, "postcode": a.get("postcode", "")}


def fill_property_location_via_centroid(rows: list[dict]) -> int:
    """Fill blank Property City/Zip by reverse-geocoding the parcel centroid.

    DataSift marks a property record "Incomplete" without a zip, which keeps it
    out of marketing — so a blank Property Zip is not cosmetic. The county GIS
    has no reliable situs city/zip (Iredell's is "00"; CITY is owner mailing),
    but the parcel's map location is authoritative and absentee-proof: take the
    polygon centroid (WGS84) and reverse-geocode it via Nominatim. Verified
    2026-07-11: 1016 Sunset Dr -> Salisbury 28147, 342 Glenwood Dr ->
    Mooresville 28115, both exact.

    ArcGIS counties only (Cabarrus/Catawba have no situs geometry path here;
    Mecklenburg's city comes from polaris3g). Rate-limited to Nominatim's
    1 req/sec and cached across runs so the same parcel isn't re-hit.
    """
    import json as _json
    import time as _time
    from nc_gis_lookup import parcel_centroid
    try:
        from address_standardizer import _reverse_geocode
    except Exception as e:  # noqa: BLE001
        print(f"  (centroid geocode unavailable: {e})")
        return 0

    targets = [r for r in rows
               if (r.get("Parcel ID") or "").strip()
               and (r.get("Property Address") or "").strip()
               and not (r.get("Property Zip") or "").strip()]
    if not targets:
        return 0

    cache: dict[str, dict] = {}
    if _CENTROID_GEOCODE_CACHE.exists():
        try:
            cache = _json.loads(_CENTROID_GEOCODE_CACHE.read_text())
        except (ValueError, OSError):
            cache = {}

    filled = 0
    dirty = False
    for r in targets:
        county = (r.get("County") or "").strip()
        addr = (r.get("Property Address") or "").strip()
        key = f"{county.upper()}||{addr.upper()}"
        geo = cache.get(key)
        if geo is None:
            ll = parcel_centroid(county, addr, pid=(r.get("Parcel ID") or "").strip())
            if ll:
                _time.sleep(1.1)         # Nominatim: 1 req/sec
                geo = _reverse_geocode(str(ll[0]), str(ll[1])) or {}
            else:
                # No parcel geometry (Mecklenburg polaris3g, or a centroid miss).
                # Forward-geocode the address instead — needs a city to be
                # unambiguous, which Meck rows already have (Charlotte).
                have_city = (r.get("Property City") or "").strip()
                if have_city:
                    _time.sleep(1.1)
                    geo = _nominatim_forward(f"{addr}, {have_city}, NC") or {}
                else:
                    geo = {}
            cache[key] = geo
            dirty = True
        city = (geo.get("city") or "").strip()
        zipc = (geo.get("postcode") or "").strip()[:5]
        # Nominatim sometimes returns a zip but no city for rural points —
        # derive the city from the zip via the local lookup.
        if zipc and not city:
            city = _NC_ZIP_TO_CITY.get(zipc, "")
        if not city and not zipc:
            continue
        if city and not (r.get("Property City") or "").strip():
            r["Property City"] = city
        if zipc and not (r.get("Property Zip") or "").strip():
            r["Property Zip"] = zipc
        tag_reason(r, "centroid-geocode")
        print(f"  CENTROID-GEOCODE {county}/{r.get('Deceased Owner')}: "
              f"{addr!r} -> {city} {zipc}")
        filled += 1

    if dirty:
        try:
            _CENTROID_GEOCODE_CACHE.write_text(_json.dumps(cache, separators=(",", ":")))
        except OSError:
            pass
    return filled


def clean_bad_city_zip_in_place(rows: list[dict]) -> tuple[int, int]:
    """Final cleanup: when Property City contains a street-type suffix
    (Dr/St/Rd/Ln/etc.), a state code (Ga/Ny/etc.), or a numeric value,
    it's leftover noise from an upstream parser bug — merge the street
    suffix back into Property Address (when applicable) and clear the
    bad value. Wrong city data is worse than blank — DataSift's Smarty
    will populate blanks on upload, but can't undo wrong values.

    Also handles:
      - 9-digit-no-dash ZIPs (e.g., 281640000) -> '28164-0000' or '28164'
      - Mecklenburg pattern "ADDR CITYNAME NC" (and "ADDR CITYNAME UNINC NC")
        jammed into Property Address — strip out trailing city/state and
        move to Property City. Locklear/Edwards/Walton Week 26.
      - Gaston pattern: zip captured, city blank — populate city from a
        local zip->city table for our 7 NC counties.

    Returns (cities_cleaned, zips_cleaned).
    """
    cleaned_city = 0
    cleaned_zip = 0
    for r in rows:
        # ── 1. Bad-city tokens (street suffix / state code / numeric) ──
        city = (r.get("Property City") or "").strip()
        city_upper = city.upper()
        is_bad_city = (
            city_upper in _BAD_CITY_TOKENS
            or city_upper in _US_STATE_CODES
            or (city and city.replace("-", "").isdigit())
        )
        if is_bad_city:
            addr = (r.get("Property Address") or "").strip()
            if city_upper in _BAD_CITY_TOKENS and addr and not addr.upper().endswith(" " + city_upper):
                r["Property Address"] = f"{addr} {city}"
            r["Property City"] = ""
            cleaned_city += 1

        # ── 2. Mecklenburg-style jammed city: "ADDRESS CITYNAME NC" ──
        # Polaris3g sometimes returns situs with city+state appended:
        # "7918 CORDER DR CHARLOTTE NC" or "6813 MT HOLLY-HUNTERSVILLE
        # RD UNINC NC". Anchor split on the LAST street suffix — city is
        # whatever comes between that suffix and " NC" at the end.
        addr = (r.get("Property Address") or "").strip()
        if addr and addr.upper().endswith(" NC") and not (r.get("Property City") or "").strip():
            tokens = addr.split()
            if len(tokens) >= 3 and tokens[-1].upper() == "NC":
                tokens = tokens[:-1]  # drop trailing "NC"
                # Find LAST street suffix
                last_suffix_idx = -1
                for i in range(len(tokens) - 1, -1, -1):
                    if tokens[i].upper().rstrip(".") in _STREET_SUFFIXES_FOR_SPLIT:
                        last_suffix_idx = i
                        break
                if last_suffix_idx >= 0 and last_suffix_idx < len(tokens) - 1:
                    # Tokens AFTER the street suffix are the jammed city
                    city_tokens = tokens[last_suffix_idx + 1:]
                    city_str = " ".join(city_tokens)
                    if city_str.upper() not in {"UNINC", "UNINCORPORATED"}:
                        r["Property City"] = city_str.title()
                    r["Property Address"] = " ".join(tokens[: last_suffix_idx + 1])
                    r["Property State"] = "NC"
                    cleaned_city += 1
                elif last_suffix_idx == len(tokens) - 1:
                    # Suffix is the last token (no city before NC) — just
                    # drop the trailing " NC" and leave city blank
                    r["Property Address"] = " ".join(tokens)
                    r["Property State"] = "NC"

        # ── 3. ZIP normalization ──
        z = (r.get("Property Zip") or "").strip()
        if z and len(z) == 9 and z.isdigit():
            if z[5:] == "0000":
                r["Property Zip"] = z[:5]
            else:
                r["Property Zip"] = f"{z[:5]}-{z[5:]}"
            cleaned_zip += 1
        # Other bad-format ZIPs (not 5-digit or 5-4): clear
        z2 = (r.get("Property Zip") or "").strip()
        valid_zip = (len(z2) == 5 and z2.isdigit()) or (
            len(z2) == 10 and z2[5] == "-" and z2[:5].isdigit() and z2[6:].isdigit()
        )
        if z2 and not valid_zip:
            r["Property Zip"] = ""
            cleaned_zip += 1

        # ── 4. ZIP -> city lookup when city is blank ──
        # Gaston ArcGIS returns ZIP per parcel but no city; same for some
        # parcels in other counties. Populate city from a local lookup.
        if not (r.get("Property City") or "").strip():
            z3 = (r.get("Property Zip") or "").strip()[:5]
            city_from_zip = _NC_ZIP_TO_CITY.get(z3)
            if city_from_zip:
                r["Property City"] = city_from_zip
                cleaned_city += 1
    return cleaned_city, cleaned_zip


# Street-suffix tokens for the jammed-city splitter. Includes both
# abbreviated and spelled-out forms.
_STREET_SUFFIXES_FOR_SPLIT = {
    "ST", "STREET", "RD", "ROAD", "AVE", "AVENUE", "AV", "DR", "DRIVE",
    "LN", "LANE", "CT", "COURT", "PL", "PLACE", "BLVD", "BOULEVARD",
    "WAY", "WY", "CIR", "CIRCLE", "CR", "HWY", "HIGHWAY", "PKWY",
    "PARKWAY", "TER", "TERRACE", "TR", "TRL", "TRAIL", "TC", "PIKE",
    "RUN", "ROW", "LOOP", "PATH", "ALY", "ALLEY", "EXT", "EXTENSION",
    "EXPY", "EXPRESSWAY", "CRT",
    # Compass directions that often follow a street suffix (e.g. "5503 GLENVIEW DR NE")
    # are NOT included here because they're treated as part of the street.
}

# Local zip -> city for the 7 NC counties we scrape. Covers the cases
# where county ArcGIS returns zip but not city. Not exhaustive — extend
# as new patterns surface.
_NC_ZIP_TO_CITY = {
    # Cabarrus
    "28025": "Concord", "28026": "Concord", "28027": "Concord",
    "28075": "Harrisburg",
    "28081": "Kannapolis", "28082": "Kannapolis", "28083": "Kannapolis",
    "28107": "Midland",
    # Catawba
    "28601": "Hickory", "28602": "Hickory", "28603": "Hickory",
    "28609": "Catawba",
    "28610": "Claremont",
    "28613": "Conover",
    "28625": "Statesville",  # Iredell overlap
    "28678": "Stony Point",  # Iredell/Alexander (Taylorsville Hwy)
    "28134": "Pineville",    # south Mecklenburg (unincorporated situs like Sawtry Ct)
    "28658": "Newton",
    "28673": "Sherrills Ford",
    "28681": "Taylorsville",
    # Gaston
    "28006": "Alexis",
    "28012": "Belmont",
    "28016": "Bessemer City",
    "28021": "Cherryville",
    "28032": "Cramerton",
    "28034": "Dallas",
    "28052": "Gastonia", "28054": "Gastonia", "28056": "Gastonia",
    "28098": "Lowell",
    "28101": "Mc Adenville",
    "28120": "Mount Holly",
    "28164": "Stanley",
    # Iredell
    "28115": "Mooresville", "28117": "Mooresville",
    "28166": "Troutman",
    "28625": "Statesville", "28677": "Statesville",
    "28634": "Harmony",
    "28660": "Olin",
    "28673": "Sherrills Ford",
    # Lincoln
    "28033": "Crouse",
    "28037": "Denver",
    "28080": "Iron Station",
    "28092": "Lincolnton", "28093": "Lincolnton",
    "28168": "Vale",
    # Mecklenburg (Charlotte main + suburbs)
    "28078": "Huntersville",
    "28031": "Cornelius",
    "28036": "Davidson",
    "28104": "Matthews", "28105": "Matthews",
    "28134": "Pineville",
    "28202": "Charlotte", "28203": "Charlotte", "28204": "Charlotte",
    "28205": "Charlotte", "28206": "Charlotte", "28207": "Charlotte",
    "28208": "Charlotte", "28209": "Charlotte", "28210": "Charlotte",
    "28211": "Charlotte", "28212": "Charlotte", "28213": "Charlotte",
    "28214": "Charlotte", "28215": "Charlotte", "28216": "Charlotte",
    "28217": "Charlotte", "28226": "Charlotte", "28227": "Charlotte",
    "28262": "Charlotte", "28269": "Charlotte", "28270": "Charlotte",
    "28273": "Charlotte", "28277": "Charlotte", "28278": "Charlotte",
    "28280": "Charlotte",
    # Rowan
    "28023": "China Grove",
    "28071": "Gold Hill",
    "28072": "Granite Quarry",
    "28138": "Rockwell",
    "28144": "Salisbury", "28145": "Salisbury", "28146": "Salisbury", "28147": "Salisbury",
    "28159": "Spencer",
    "28023": "China Grove",
    "28041": "Cleveland",
    "28039": "Cleveland",
    "28023": "China Grove",
}


def _money(v) -> float:
    """Parse a market-value string to float ($ signs, commas tolerated). 0 if blank."""
    if v is None:
        return 0.0
    s = str(v).replace("$", "").replace(",", "").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def flag_initial_only_middle_matches(rows: list[dict]) -> int:
    """Tag rows where the matched deed owner has middle-initial-ONLY but
    the decedent has a FULL middle name. The matcher accepts these at
    score 1.0 (initial-of-Edward matches "E") but the deed "E" could
    actually be Eugene, Eric, Ethan, etc — same-name homonym risk.

    Surfaced Week 26 audit (Smith Thomas Edward 26E000638-480 Iredell):
    deed showed "SMITH THOMAS E", matcher picked it confidently, but
    Oren verified via BeenVerified that the real owner is Thomas
    Eugene Smith — different person.

    We can't disambiguate automatically (Iredell deed + GIS both lose
    the full middle). Best we can do is FLAG the risk so Oren scans
    those rows on BeenVerified manually. Adds 'verify-middle-initial'
    to Match Reason — doesn't drop the row.

    Returns count of rows flagged.
    """
    from nc_gis_lookup import lookup_properties, split_decedent_name
    import re as _re
    # Tokens that are NOT middle-name candidates (suffixes, noise, etc.)
    _SKIP_TOKENS = {"JR", "SR", "II", "III", "IV", "V", "WF", "HSB",
                    "HEIRS", "ESTATE", "TRUSTEE", "TRUST", "LIVING",
                    "REVOC", "REVOCABLE", "LFI", "AKA", "C/O"}
    flagged = 0
    for r in rows:
        pid = (r.get("Parcel ID") or "").strip()
        if not pid:
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or not county or "IN THE MATTER" in dec.upper():
            continue
        # Skip if Match Reason already has the flag (idempotent)
        if "verify-middle-initial" in (r.get("Match Reason") or "").lower():
            continue
        # Decedent must have a FULL middle name (>=2 chars). Single-char
        # middle = no information to verify against; skip silently.
        _first, dec_middle, dec_last = split_decedent_name(dec)
        if not dec_middle:
            continue
        dec_middle_words = [w for w in dec_middle.split() if w]
        if not any(len(w) >= 2 for w in dec_middle_words):
            continue  # decedent middle is itself initial-only

        # Look up the matched candidate to inspect deed owner middle
        try:
            cands = lookup_properties(dec, county, min_score=0.5)
        except Exception:
            continue
        match = next((c for c in cands if c.pid == pid), None)
        if not match:
            continue
        owner = (match.owner_name or "").upper()
        if not owner:
            continue
        # Take the segment containing the decedent's last name (handles
        # joint owners like "SMITH JOHN | SMITH JANE")
        segments = _re.split(r"\s*\|\s*|\s+&\s+|\s*;\s*", owner)
        target_seg = next(
            (s for s in segments if dec_last and dec_last.upper() in s),
            owner,
        )
        tokens = [t.strip(",.") for t in target_seg.split() if t.strip(",.")]
        tokens = [t for t in tokens if t not in _SKIP_TOKENS]
        if dec_last and dec_last.upper() in tokens:
            idx = tokens.index(dec_last.upper())
            tokens = tokens[:idx] + tokens[idx + 1:]
        # After dropping last+first, what's left is middle-position tokens.
        # If ALL remaining are single chars, owner has initial-only middle.
        if len(tokens) < 2:
            continue
        middle_tokens = tokens[1:]
        if not middle_tokens or not all(len(t) == 1 for t in middle_tokens):
            continue
        # Initial-only middle AND decedent has full middle -> flag as
        # informational. The deed-E could be Edward, Eugene, Eric, etc;
        # most cases ARE the right person (Tadlock, Earney) but Smith
        # Thomas Edward / Eugene class isn't distinguishable from our data.
        # Treat as a manual-verification flag, not a drop.
        tag_reason(r, "verify-middle-initial")
        flagged += 1
    return flagged


def flag_no_middle_ambiguous_matches(rows: list[dict]) -> int:
    """Tag rows where the deed owner has NO middle name at all, the decedent
    DOES have a middle name, AND the county GIS holds two-or-more same-name
    parcels — so the match rests on first+last only and can't be trusted.

    Distinct from flag_initial_only_middle_matches (which needs a middle
    INITIAL on the deed): here the deed is bare "First Last", giving nothing
    to disambiguate a common name against. Only flags when a competing
    homonym actually exists in the county, so unique names stay clean.

    Surfaced Week 29 audit: Wallace, Mary Louise 26E002588-590 — deed "MARY
    WALLACE" (no middle) matched 223 Halsey St, but a "MARY L WALLACE" parcel
    (the L = Louise) sits at a different address. Dudley, Robert Lewis
    26E002605-590 — deed "ROBERT DUDLEY" with two Robert Dudley households in
    Mecklenburg. Adds 'verify-name-ambiguous' to Match Reason — doesn't drop.

    Returns count of rows flagged.
    """
    from nc_gis_lookup import lookup_properties, split_decedent_name, _name_match_score
    import re as _re
    _SKIP_TOKENS = {"JR", "SR", "II", "III", "IV", "V", "WF", "HSB",
                    "HEIRS", "ESTATE", "TRUSTEE", "TRUST", "LIVING",
                    "REVOC", "REVOCABLE", "LFI", "AKA", "C/O", "LIFE"}
    flagged = 0
    for r in rows:
        pid = (r.get("Parcel ID") or "").strip()
        if not pid:
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or not county or "IN THE MATTER" in dec.upper():
            continue
        reason = (r.get("Match Reason") or "").lower()
        if "verify-name-ambiguous" in reason or "verify-middle-initial" in reason:
            continue  # already carries a name-verification flag
        _first, dec_middle, dec_last = split_decedent_name(dec)
        if not dec_middle or not dec_last:
            continue  # deed's missing middle can't disambiguate what isn't there
        try:
            cands = lookup_properties(dec, county, min_score=0.5)
        except Exception:
            continue
        match = next((c for c in cands if c.pid == pid), None)
        if not match:
            continue
        # Matched owner's decedent-segment must be bare first+last (no middle).
        owner = (match.owner_name or "").upper()
        segments = _re.split(r"\s*\|\s*|\s+&\s+|\s*;\s*", owner)
        target_seg = next((s for s in segments if dec_last.upper() in s), owner)
        toks = [t.strip(",.") for t in target_seg.split() if t.strip(",.")]
        toks = [t for t in toks if t not in _SKIP_TOKENS]
        if len(toks) != 2:
            continue  # has a middle token (verify-middle-initial covers that) or malformed
        # The deed is bare first+last while the court record HAS a middle, so
        # nothing corroborates the match. Per Oren (Week 29): flag every one of
        # these for an eyeball, not just the ones with an in-county homonym —
        # Gibson 26E000948-350 (deed "GIBSON JEFFREY", court "Jeffrey Scott")
        # had no competing Jeffrey Gibson and so passed silently at score 1.0.
        # Two tiers so the audit shows which is riskier:
        #   verify-name-ambiguous — a competing same-name parcel exists (2+
        #     strong first+last matches). Could be the WRONG person.
        #   verify-name-nomiddle  — unique name, just no middle to confirm.
        strong = sum(1 for c in cands if _name_match_score(dec, c.owner_name or "") >= 0.9)
        tag_reason(r, "verify-name-ambiguous" if strong >= 2 else "verify-name-nomiddle")
        flagged += 1
    return flagged


def refresh_property_use_from_gis(rows: list[dict]) -> int:
    """Re-derive Property use from the current cached GIS data for each
    row with a parcel. Catches cases where simplify_use_code was upgraded
    AFTER the row's scrape (Reid Townhouse Week 26: scraped 6/24 with the
    old SFR mapping; my Townhouse fix shipped 6/25; today's polish would
    otherwise keep the stale SFR because dedup/seen_ids prevents re-scrape).

    Only OVERWRITES when the new derived value is non-empty AND differs
    from the current. Doesn't blank an existing use just because GIS lost
    a candidate (defensive — keep what we have if GIS goes empty).

    Returns count of rows whose Property use was updated.
    """
    from nc_gis_lookup import lookup_properties, simplify_use_code
    updated = 0
    for r in rows:
        pid = (r.get("Parcel ID") or "").strip()
        if not pid:
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or not county or "IN THE MATTER" in dec.upper():
            continue
        try:
            cands = lookup_properties(dec, county, min_score=0.5)
        except Exception:
            continue
        match = next((c for c in cands if c.pid == pid), None)
        if not match:
            continue
        new_use = simplify_use_code(match.use_code, match.use_description, match.county) or ""
        # Apply the same vacant-override the scrape path uses
        if match.is_vacant_land and "VACANT" not in new_use.upper():
            new_use = "Vacant Land"
        if new_use and new_use != (r.get("Property use") or "").strip():
            old = r.get("Property use") or ""
            r["Property use"] = new_use
            updated += 1
            print(f"    USE-REFRESH {county}/{dec[:30]} pid={pid}: {old!r} -> {new_use!r}")
    return updated


def populate_property_values(rows: list[dict]) -> int:
    """Backfill Property Value + Property Acres from a fresh GIS lookup for rows
    that have a Parcel ID but are missing either. Required before the
    drop_over_500k filter can run — it reads BOTH (acreage drives the >2-acre
    subdivide exemption to the $500K cap).

    Acreage is backfilled even when Property Value is already set: rows whose
    parcel was attached by an earlier polish step (repick / swap / recovery)
    carry a value but never passed through the scrape-time writer that fills
    acres. Without this, those rows look acreage-unknown and keep the $500K cap.

    Catawba fallback: the county's PHP layer used for name search exposes
    no value field at all (see project_catawba_value_field_missing.md).
    For Catawba rows still blank after the standard lookup, hit the
    bitek_parcel_report_view endpoint (which backs the "Parcel Report"
    page on gis.catawbacountync.gov) to pull total_value directly.
    """
    from nc_gis_lookup import _catawba_parcel_report
    filled = 0
    for r in rows:
        pid = (r.get("Parcel ID") or "").strip()
        if not pid:
            continue
        needs_value = not (r.get("Property Value") or "").strip()
        needs_acres = not (r.get("Property Acres") or "").strip()
        if not needs_value and not needs_acres:
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or not county or "IN THE MATTER" in dec.upper():
            continue
        try:
            # Cached per (decedent, county) — the extra acres-only lookups here
            # ride the same cache entry the value lookup already warmed.
            results = lookup_properties(dec, county, min_score=0.5)
        except Exception:
            results = []
        match = next((c for c in results if c.pid == pid), None)
        if match and needs_acres and match.lot_area:
            r["Property Acres"] = f"{float(match.lot_area):.2f}"
        if not needs_value:
            continue
        if match and match.market_value:
            r["Property Value"] = f"{int(round(float(match.market_value))):,}"
            filled += 1
            continue
        # Catawba PHP endpoint fallback — ArcGIS layer has no value field.
        if county.lower() == "catawba":
            rec = _catawba_parcel_report(pid)
            if rec:
                v = rec.get("total_value")
                try:
                    v_f = float(v) if v not in (None, "", 0, "0") else None
                except (TypeError, ValueError):
                    v_f = None
                if v_f:
                    r["Property Value"] = f"{int(round(v_f)):,}"
                    filled += 1
    return filled


# Per-use-type buy-box value caps. Vacant land has a higher cap because
# raw land is harder to price + heirs more willing to sell at higher $.
# Per Oren 2026-06-20: "Residential SFR parcels would still keep the 500k
# limit, but vacant parcels up to 1M."
_VALUE_CAP_BY_USE = {
    "VACANT LAND": 1_000_000,
    "VACANT":      1_000_000,
    "LAND":        1_000_000,
    "SFR":         500_000,
    "MH":          500_000,
    "RESIDENTIAL": 500_000,
    "MULTI-FAMILY": 500_000,
    "DUPLEX":      500_000,
}
_DEFAULT_VALUE_CAP = 500_000

# Subdivide exemption. Per Oren 2026-07-16: a house on a large parcel carries
# subdivision potential, so its value alone shouldn't DQ it — the land is the
# play, not the structure. A house or mobile home on more than this many acres
# gets the vacant-land cap ($1M) instead of the $500K structure cap.
#
# Acreage-unknown rows keep the $500K cap (Oren's call): a pricey house with no
# acreage on file is far more likely a small in-town lot than a subdividable
# tract, and assuming otherwise would pull real DQs back into the workbook.
# ~14% of Iredell and ~6% of Lincoln parcels have no acreage field populated;
# every other county returns it on every parcel.
_SUBDIVIDE_MIN_ACRES = 2.0

# Explicit allowlist, NOT "anything that isn't vacant". Oren asked for houses,
# and confirmed mobile homes (the land is the play either way). An allowlist
# also keeps the exemption off use types that merely fall through to the
# default cap — a $900K commercial building on 50 acres is not a subdivide
# lead, and Step 2 drops it regardless.
_SUBDIVIDE_ELIGIBLE_USES = {"SFR", "MH", "RESIDENTIAL"}


def _cap_for_use(use: str, acres: float | None = None) -> float:
    """Buy-box value cap for a parcel, given its simplified use and acreage.

    `acres` None/0 means unknown — the caller couldn't determine lot size, and
    the row keeps the conservative structure cap.
    """
    u = (use or "").upper().strip()
    if not u:
        return _DEFAULT_VALUE_CAP
    if "VACANT" in u or u == "LAND":
        return _VALUE_CAP_BY_USE["VACANT LAND"]
    cap = _VALUE_CAP_BY_USE.get(u, _DEFAULT_VALUE_CAP)
    # Subdivide exemption: only ever lifts a cap, never lowers one.
    if (acres and acres > _SUBDIVIDE_MIN_ACRES
            and u in _SUBDIVIDE_ELIGIBLE_USES):
        cap = max(cap, _VALUE_CAP_BY_USE["VACANT LAND"])
    return cap


def _row_acres(row: dict) -> float | None:
    """Parsed acreage off a CSV row, or None when blank/unparseable."""
    raw = (row.get("Property Acres") or "").strip().replace(",", "")
    if not raw:
        return None
    try:
        v = float(raw)
    except ValueError:
        return None
    return v if v > 0 else None


def _cand_acres(c) -> float | None:
    """Parsed acreage off a PropertyCandidate, or None when unset."""
    try:
        v = float(getattr(c, "lot_area", None) or 0)
    except (TypeError, ValueError):
        return None
    return v if v > 0 else None


def _set_row_acres(row: dict, acres: float | None) -> None:
    """Write Property Acres, CLEARING it when acreage is unknown.

    Must be called by every step that reassigns Parcel ID. Clearing on unknown
    is the whole point: a swapped-in parcel that inherits the swapped-out
    parcel's acreage gets the wrong buy-box cap. Same trap Property Value hit
    in Week 26 (Walton 26E002339-590) — Step 1.7 only refills BLANK fields, so
    a stale non-blank value survives a swap forever.
    """
    row["Property Acres"] = f"{acres:.2f}" if acres and acres > 0 else ""


def _set_row_acres_from_candidate(row: dict, c) -> None:
    _set_row_acres(row, _cand_acres(c))


def _set_row_acres_from_attrs(row: dict, attrs: dict) -> None:
    from nc_gis_lookup import acres_from_arcgis_attrs
    _set_row_acres(row, acres_from_arcgis_attrs(attrs or {}))


# Floor for a STANDALONE main parcel. Per Oren (Week 28 audit): a lone scrap
# of land isn't a lead — Hutchins 26E000897-350 was a $400 landlocked strip on
# Wellington Dr. But the floor must NOT touch multi-parcel estates: a cheap lot
# beside two more cheap lots is the mobile-home-on-land play, and there the
# individual value is irrelevant. So: only drop when the row has no siblings.
_MIN_STANDALONE_VALUE = 10_000


def _has_sibling_parcels(row: dict) -> bool:
    """True when the row's Notes list additional parcels for the same estate.

    Both notations appear: the legacy horizontal 'PLUS 2 PARCELS' block and the
    vertical 'PLUS <pid>' lines written by the multi-parcel collapse and the
    decedent-address fallback.
    """
    return "PLUS " in (row.get("Notes") or "").upper()


def _landportal_revalue_vacant(r: dict, floor: float, prev_val: str) -> bool:
    """For a vacant parcel about to be scrapped for sub-floor value, ask
    LandPortal for its market estimate (tlp_estimate). County tax values badly
    underprice raw land — Lowman 26E000824-170 Week 29: county $5,300 vs
    LandPortal $56,454. When the market estimate clears the floor, overwrite the
    value + annotate and return True (keep the row). Free/no-op without a key.
    """
    try:
        import landportal_lookup as lp
    except Exception:
        return False
    if not lp.available():
        return False
    pin = (r.get("Parcel ID") or "").strip()
    county = (r.get("County") or "").strip()
    if not pin or not county:
        return False
    info = lp.get_vacant_market_value(pin, county)
    est = (info or {}).get("tlp_estimate")
    if est is None or est < floor:
        return False
    r["Property Value"] = f"{int(round(est)):,}"
    cv = (info or {}).get("county_value")
    note = ("[LANDPORTAL re-valued: market est "
            f"${int(round(est)):,}"
            + (f" vs county tax value ${int(round(cv)):,}" if cv else "")
            + "]")
    existing = (r.get("Notes") or "").strip()
    r["Notes"] = (existing + ("\n" if existing else "") + note).strip()
    tag_reason(r, "landportal-revalued")
    print(f"  LANDPORTAL RESCUE {county}/{r.get('Deceased Owner')}: "
          f"{r.get('Property Address')!r} county said {prev_val} -> "
          f"market est ${int(round(est)):,} — keeping")
    return True


def drop_under_min_value(rows: list[dict], floor: float = _MIN_STANDALONE_VALUE) -> tuple[list[dict], int]:
    """Drop rows whose ONLY parcel is worth less than `floor`.

    Skips rows with no priced value (never drop what we couldn't price) and
    skips any multi-parcel estate — see _MIN_STANDALONE_VALUE. Before dropping a
    VACANT lot, tries a LandPortal market re-valuation (county tax value grossly
    underprices raw land); keeps + re-prices the row when the estimate clears.
    """
    kept: list[dict] = []
    dropped = 0
    for r in rows:
        v_str = (r.get("Property Value") or "").strip()
        if not v_str or _has_sibling_parcels(r):
            kept.append(r)
            continue
        if _money(v_str) >= floor:
            kept.append(r)
            continue
        use = (r.get("Property use") or "").upper()
        if ("VACANT" in use or use == "LAND") and _landportal_revalue_vacant(r, floor, v_str):
            kept.append(r)
            continue
        print(f"  MIN-VALUE DROP {r.get('County')}/{r.get('Deceased Owner')}: "
              f"{r.get('Property Address')!r} valued {v_str} < {floor:,.0f} (standalone parcel)")
        dropped += 1
    return kept, dropped


_MIN_VACANT_ACRES = 0.25


def drop_tiny_vacant_lots(rows: list[dict],
                          floor: float = _MIN_VACANT_ACRES) -> tuple[list[dict], int]:
    """Drop a STANDALONE VACANT lot smaller than `floor` acres — scrap too small
    to build on or resell, regardless of its dollar value.

    The dollar floor (drop_under_min_value, Step 1.82) misses these two ways:
    a sliver can carry a plausible tax value, or LandPortal can re-value it above
    the $10K floor. Long, Rebecca M 26E000944-350 Week 29: 0.11 ac, county value
    $1,010, LandPortal re-valued it to $57,141 and it survived — but Oren:
    "obviously a scrap parcel, should not even go to LandPortal." A hard size
    floor catches it before value ever enters the picture.

    Guards mirror the dollar floor exactly, and matter as much as the rule:
      * VACANT LAND ONLY. A 0.03-ac townhome (Anderson 26E000820-170) or a
        0.07-ac SFR on a city lot (Massey) is a real building on a small parcel,
        not scrap — size says nothing about those. Never size-floor a structure.
      * Multi-parcel estates EXEMPT. 2+ adjacent vacant lots are the
        mobile-home-on-land play Oren prizes; individual lot size is irrelevant
        there (see the consecutive-vacant-lots rule).
      * Never drop UNKNOWN acreage. Blank means we couldn't measure it — often a
        county GIS outage (half of Week 29's Gaston rows had blank acres the
        night Gaston was down). Dropping on blank would silently delete real
        leads whenever a server hiccuped. Only a positive number under the floor
        is scrap.
    """
    kept: list[dict] = []
    dropped = 0
    for r in rows:
        use = (r.get("Property use") or "").upper()
        acres = _row_acres(r)
        is_vacant = "VACANT" in use or use == "LAND"
        if (not is_vacant or acres is None or acres >= floor
                or _has_sibling_parcels(r)):
            kept.append(r)
            continue
        print(f"  TINY-LOT DROP {r.get('County')}/{r.get('Deceased Owner')}: "
              f"{r.get('Property Address')!r} {acres:.2f}ac < {floor}ac "
              f"(standalone vacant scrap)")
        tag_reason(r, "tiny-vacant-lot")
        dropped += 1
    return kept, dropped


def drop_life_estate_parcels(rows: list[dict]) -> tuple[list[dict], int]:
    """Drop rows whose GIS-matched parcel is titled as the DECEDENT's LIFE
    ESTATE. A life estate ends at the life tenant's death and the property
    passes to the remaindermen (the heirs) automatically, OUTSIDE probate —
    so there is no estate sale to pursue, and the remaindermen commonly already
    occupy it. Per Oren (Week 29 audit), hard-DQ these.

    Parker, Geraldine 26E000935-350 — GIS owner "PARKER GERALDINE SNEED LIFE
    ESTATE" at 205 S Pink St, owner-occupied; Oren confirmed via BeenVerified
    the remaindermen (Cathy Parker Safley / Grover Clifton Parker) now hold and
    occupy it. County GIS still shows the life estate, so the "LIFE ESTATE"
    string on the decedent's own parcel is the detectable tell.

    Guards: skips vacant land (keep-land rule stands — a lot still sells) and
    skips multi-parcel estates (a "PLUS" sibling may be the real lead; don't
    nuke the whole row over one life-estate parcel).

    Returns (kept_rows, dropped_count).
    """
    from nc_gis_lookup import lookup_properties, split_decedent_name
    kept: list[dict] = []
    dropped = 0
    for r in rows:
        pid = (r.get("Parcel ID") or "").strip()
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        use = (r.get("Property use") or "").upper()
        if (not pid or not dec or not county or "IN THE MATTER" in dec.upper()
                or "VACANT" in use or use == "LAND" or _has_sibling_parcels(r)):
            kept.append(r)
            continue
        try:
            cands = lookup_properties(dec, county, min_score=0.5)
            match = next((c for c in cands if (c.pid or "") == pid), None)
        except Exception:
            match = None
        if not match:
            kept.append(r)
            continue
        owner = (match.owner_name or "").upper()
        _f, _m, dec_last = split_decedent_name(dec)
        dec_last = (dec_last or "").upper()
        if "LIFE ESTATE" in owner and dec_last and dec_last in owner:
            print(f"  LIFE-ESTATE DROP {county}/{dec}: parcel {pid} owner={owner!r} "
                  f"(auto-transfers to remaindermen — not a probate sale)")
            dropped += 1
            continue
        kept.append(r)
    return kept, dropped


# Rowan publishes no use code at all — simplify_use_code() defaults every Rowan
# parcel to SFR. Its LEG_DESC field is the only type hint: a condo carries a unit
# number ("U102"), while houses/land carry lot numbers ("L23-24"), acreage
# ("15.23AC"), or "-". Verified on 5 known parcels 2026-07-09.
#
# Per Oren: FLAG, do not drop. One confirmed example isn't enough to start
# silently deleting rows from his third-largest county; a wrong guess would be
# invisible. Revisit once the flag has proven itself over a few weeks.
_ROWAN_UNIT_RE = re.compile(r"\bU\d{1,4}\b")


def flag_rowan_possible_condos(rows: list[dict]) -> int:
    """Mark Rowan rows whose legal description carries a unit number as Condo.

    Rowan publishes no use code, so a condo otherwise defaults to SFR. The unit
    number in LEG_DESC ("U102") is the tell — verified against known parcels
    (condo "U102" vs houses "L23-24"/"15.23AC") and confirmed on Riley
    26E000710-790 (Oren: "definitely a Condo"). Promoted from flag to auto-drop
    2026-07-12: sets Property use = "Condo" so the existing Condo/Townhouse drop
    (Step 2.1) removes it and counts it — one uniform path. Oren drops condos.
    """
    from nc_gis_lookup import lookup_by_address

    marked = 0
    for r in rows:
        if (r.get("County") or "").strip().lower() != "rowan":
            continue
        if (r.get("Property use") or "").strip().upper() in ("CONDO", "TOWNHOUSE"):
            continue  # already classified
        addr = (r.get("Property Address") or "").strip()
        if not addr:
            continue
        try:
            hits = lookup_by_address(addr, "Rowan")
        except Exception:  # noqa: BLE001
            continue
        pid = (r.get("Parcel ID") or "").strip()
        match = next((h for h in hits if pid and str(h.get("PARCEL_ID") or "").strip() == pid),
                     hits[0] if hits else None)
        if not match:
            continue
        leg = str(match.get("LEG_DESC") or "").upper()
        if not _ROWAN_UNIT_RE.search(leg):
            continue
        r["Property use"] = "Condo"   # -> dropped + counted by Step 2.1
        tag_reason(r, "rowan-condo")
        print(f"  ROWAN CONDO Rowan/{r.get('Deceased Owner')}: {addr!r} "
              f"(LEG_DESC={leg!r}) -> Condo, will drop")
        marked += 1
    return marked


def drop_recently_sold(rows: list[dict], months: int = 24, min_price: float = 50_000) -> tuple[list[dict], int]:
    """Drop rows whose parcel sold within the last `months` for at
    least `min_price`. Catches the Conners 26E000691-170 class: deed
    transferred to a new owner January 2026, no equity for an heir to
    sell. Per Oren's pre-mail workflow ("I typically check the property
    to see if it has been listed, sold, under contract, etc."), these
    are dead leads -- the property is already in market or just changed
    hands, mail to the dead decedent's heir is wasted.

    Window: 24 months (configurable). Threshold: $50K (filters out
    intra-family deed transfers stamped as a sale for $0/$10/$100;
    those aren't real sales and shouldn't trigger the filter).

    Looks up each row's parcel in GIS (cache-hot after Step 1.7
    Property Value backfill) and reads PropertyCandidate.sale_date.
    Returns (kept_rows, dropped_count).
    """
    from datetime import datetime
    from nc_gis_lookup import lookup_properties
    kept: list[dict] = []
    dropped = 0
    today = datetime.now().date()
    cutoff_days = months * 30  # approximate; 24mo ≈ 720 days
    for r in rows:
        pid = (r.get("Parcel ID") or "").strip()
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not (pid and dec and county) or "IN THE MATTER" in dec.upper():
            kept.append(r)
            continue
        try:
            results = lookup_properties(dec, county, min_score=0.5)
        except Exception:
            kept.append(r)
            continue
        match = next((c for c in results if c.pid == pid), None)
        if not match or not match.sale_date:
            kept.append(r)
            continue
        try:
            sold = datetime.strptime(match.sale_date[:10], "%Y-%m-%d").date()
        except ValueError:
            kept.append(r)
            continue
        age_days = (today - sold).days
        if age_days < 0 or age_days > cutoff_days:
            kept.append(r)
            continue
        # Within the window — verify it was a real arms-length sale, not
        # an intra-family $0/$10/$100 deed stamp or unrecorded price.
        # Unknown sale_price (None) is treated as "not a real sale" and
        # KEPT — without this guard, intra-family deed transfers were
        # over-dropping live cases (Lingerfelt 26E000849-350 Week 26:
        # Larry Keith Lingerfelt's residence had a Jan 2026 deed event
        # with price=None, very likely the deed-to-life-estate that
        # opened the probate — not a market sale).
        if not match.sale_price or float(match.sale_price) < min_price:
            kept.append(r)
            continue
        tag_reason(r, "dq-recently-sold")
        price_str = f"${int(float(match.sale_price)):,}" if match.sale_price else "?"
        print(f"  RECENTLY-SOLD {county}/{dec}: pid={pid} sold {match.sale_date} for {price_str} ({age_days//30}mo ago)")
        dropped += 1
    return kept, dropped


def drop_over_500k(rows: list[dict], cap: float = 500_000) -> tuple[list[dict], int]:
    """Drop rows whose Property Value exceeds the per-use buy-box cap.
    SFR/MH/Residential get the default $500K cap; Vacant Land gets $1M
    (per Oren's buy-box: vacant land sells at higher $ — wider net).
    Skips rows without a populated value (we don't want to drop rows we
    couldn't price). `cap` arg kept for back-compat but per-use map wins.
    """
    kept: list[dict] = []
    dropped = 0
    for r in rows:
        v_str = (r.get("Property Value") or "").strip()
        if not v_str:
            kept.append(r)
            continue
        acres = _row_acres(r)
        use_cap = _cap_for_use(r.get("Property use", ""), acres)
        if _money(v_str) <= use_cap:
            # Log the subdivide exemption — a >$500K house staying in the
            # workbook is surprising unless you know the acreage is why.
            if acres and acres > _SUBDIVIDE_MIN_ACRES and _money(v_str) > _DEFAULT_VALUE_CAP:
                tag_reason(r, "subdivide-exempt")
                print(f"  SUBDIVIDE-EXEMPT {r.get('County','')}/{r.get('Deceased Owner','')}: "
                      f"{r.get('Property use','')} at {v_str} on {acres:.2f}ac "
                      f"(case {r.get('Case No.','')})")
            kept.append(r)
            continue
        # Over cap. Before dropping, try to repoint a multi-parcel estate to an
        # under-cap sibling (vacant lot preferred) — keeps Shuford-class estates
        # whose real lead is a cheap vacant lot next to an over-cap residence.
        if _try_swap_to_under_cap_sibling(r):
            kept.append(r)
            continue
        dropped += 1
    return kept, dropped


# Statuses that mark a probate case as finished — no mail opportunity left.
# Belt-and-suspenders with the scrape-time filter in ecourts_scraper.
_DROP_CASE_STATUSES_POLISH = {"DISPOSED", "CLOSED", "INACTIVE", "TRANSFERRED"}


_ENTITY_DECEDENT_PATTERNS = (
    # Whole-word matches in the Deceased Owner string. We're looking for
    # trust / corporate-structure markers that indicate the "decedent"
    # is actually an entity, not a person — so there's no probate lead
    # and the existing matcher pipeline would just chase noise (see the
    # 220-candidate Starnes Trust example in the 2026-06-11 audit).
    "TRUST",
    "TRUSTEE",
    "LLC",
    "INC",
    "CORP",
    "FBO",
    "F/B/O",
    "RETIREMENT",
    "BENEFIT TRUST",
    "FAMILY TRUST",
    "REVOCABLE TRUST",
    "IRREVOCABLE TRUST",
)


def _is_entity_decedent(name: str) -> bool:
    """True when Deceased Owner looks like a trust / corporate entity
    rather than a person. Conservative — we'd rather keep a quirky
    person-name than chase a trust ghost."""
    if not name:
        return False
    upper = name.upper()
    # Whole-word check for the short tokens (TRUST/LLC/etc); substring is
    # fine for multi-word phrases (F/B/O, BENEFIT TRUST).
    tokens = set(re.findall(r"[A-Z]+", upper))
    short_markers = {"TRUST", "TRUSTEE", "LLC", "INC", "CORP", "FBO", "RETIREMENT"}
    if tokens & short_markers:
        return True
    return any(pat in upper for pat in ("F/B/O", "BENEFIT TRUST", "FAMILY TRUST",
                                       "REVOCABLE TRUST", "IRREVOCABLE TRUST"))


def drop_entity_decedents(rows: list[dict]) -> tuple[list[dict], int, list[str]]:
    """Drop rows whose Deceased Owner is a trust or corporate entity rather
    than a person. Returns (kept, dropped_count, sample_names) — caller
    logs the first few dropped names so user can spot a false positive.
    """
    kept: list[dict] = []
    dropped_names: list[str] = []
    for r in rows:
        dec = (r.get("Deceased Owner") or "").strip()
        if _is_entity_decedent(dec):
            dropped_names.append(dec)
            continue
        kept.append(r)
    return kept, len(dropped_names), dropped_names


def backfill_case_status_from_odata(rows: list[dict]) -> int:
    """For rows where Case Status is blank but Case ID (hex) is known,
    fetch the canonical status from Tyler Tech's CaseSummariesSlim
    OData endpoint (no auth required) and populate the column.

    Some search-result rows don't render a status cell, so the scrape's
    cell-scan misses them — leaving Case Status blank. Week 26 saw 8
    such rows (Hoopingarner, Honeycutt, Gaston Wilbert, Lambert,
    Sellers, etc). Backfilling here fixes display + downstream filters
    that key on status.

    Returns count of rows whose status was backfilled.
    """
    import requests as _requests
    BASE = "https://portal-nc.tylertech.cloud/app/RegisterOfActionsService/CaseSummariesSlim"
    HEADERS = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "application/json",
        "Referer": "https://portal-nc.tylertech.cloud/",
    }
    backfilled = 0
    for r in rows:
        if (r.get("Case Status") or "").strip():
            continue
        case_hex = (r.get("Case ID (hex)") or "").strip()
        if not case_hex:
            continue
        try:
            resp = _requests.get(BASE, params={"key": case_hex}, headers=HEADERS, timeout=15)
            if resp.status_code != 200:
                continue
            data = resp.json()
        except Exception:
            continue
        # Status lives at CaseInformation.CaseStatuses[0].CaseStatusId.Description
        try:
            statuses = ((data.get("CaseInformation") or {}).get("CaseStatuses")) or []
            if statuses and isinstance(statuses, list):
                desc = (((statuses[0] or {}).get("CaseStatusId")) or {}).get("Description", "")
                if desc:
                    r["Case Status"] = desc.strip()
                    backfilled += 1
        except (AttributeError, IndexError, TypeError):
            continue
    return backfilled


def drop_non_pending(rows: list[dict]) -> tuple[list[dict], int, dict[str, int]]:
    """Tally Case Status distribution + keep all rows.

    Per Oren 2026-06-27: Disposed cases ARE real probate leads — Disposed
    in NC probate typically means Letters issued and estate is being
    administered, often with real estate just transferred to heirs. Those
    heirs may want to sell. The heir-occupied filter (Step 1.9) already
    handles the "heirs live there → dead lead" case; Disposed status alone
    isn't a DQ signal.

    Previously dropped Closed/Disposed/etc. Now passes everything and just
    tracks status distribution for telemetry. Returns the same shape so
    callers don't need to change.
    """
    histo: dict[str, int] = {}
    for r in rows:
        status = (r.get("Case Status") or "").strip()
        histo[status or "(blank)"] = histo.get(status or "(blank)", 0) + 1
        # Tag any Disposed-recent rows for visibility (still kept)
        if status.upper() in _DROP_CASE_STATUSES_POLISH and _is_small_estate_disposed_recent(r):
            tag_reason(r, "small-estate-disposed-recent")
    return rows, 0, histo


# Polish-side mirror of ecourts_scraper.SMALL_ESTATE_RECENT_DAYS — kept
# in sync so both layers honor the same carve-out window.
SMALL_ESTATE_RECENT_DAYS = 14


_SMALL_ESTATE_NOTE = "SMALL ESTATE (disposed by affidavit)"


def mark_small_estate_in_notes(rows: list[dict]) -> int:
    """Put a visible SMALL ESTATE marker in the Notes column for surviving
    small-estate-disposed cases.

    Oren keeps these (a disposed estate doesn't mean the house is spoken for)
    but wants to know which ones they are so he can treat them differently on
    the call. Match Reason already carries `small-estate-disposed-recent`, but
    that's an audit column he doesn't scan while dialing — Notes is the field
    he reads. Keyed off the durable Match Reason tag rather than re-checking the
    14-day File Date window, so a later re-polish (or an archived-week apply)
    marks the same rows the scrape identified, not whatever is <14 days old now.

    Idempotent: skips a row whose Notes already carries the marker.
    """
    marked = 0
    for r in rows:
        if "small-estate-disposed-recent" not in (r.get("Match Reason") or ""):
            continue
        notes = (r.get("Notes") or "").strip()
        if _SMALL_ESTATE_NOTE in notes:
            continue
        r["Notes"] = (_SMALL_ESTATE_NOTE + ("\n" + notes if notes else "")).strip()
        marked += 1
    return marked


def drop_zillow_listed_or_sold(rows: list[dict]) -> tuple[list[dict], int]:
    """Drop rows whose property Zillow shows actively listed / under contract /
    pending / recently sold — the live-MLS class county GIS can't see (it only
    knows a sale once the deed records post-close). Always writes the status
    into Notes, even when keeping, so Oren no longer clicks every Zillow link.

    Per Oren (Week 29): auto-drop the obvious ones. Guardrails:
      * FAIL TOWARD KEEPING. zillow_status returns "unknown" (→ keep) on any
        fetch/parse failure, a blocked page, or an unsure LLM. Only a HIGH-
        confidence for_sale/pending/under_contract/sold_recently drops.
      * Survivors only (this runs at Step 4.93, after every cheaper filter),
        because each check is a Firecrawl + LLM call — the pipeline's costliest
        per-row step. Disk-cached ~4 days so nightly re-polishes are ~free.
      * Skips rows with no real street property address (blank, or a vacant
        "0 <street>" with no house number — Zillow can't resolve those).
      * Multi-parcel estates: only the MAIN property is checked. A cluster whose
        main is a vacant lot simply reads "unknown" and is kept.
      * Off switch: ZILLOW_DISABLE=1.

    Beam 26E000829-170 Week 29: sibling parcel 1392 21St Ave Ne actively listed
    $285k — Oren's manual DQ, now automatic.
    """
    try:
        import zillow_status as zs
    except Exception as e:  # noqa: BLE001
        print(f"  (zillow_status unavailable: {e}) — skipping")
        return rows, 0
    if not zs.available():
        print("  (Firecrawl not configured or ZILLOW_DISABLE=1) — skipping")
        return rows, 0

    kept: list[dict] = []
    dropped = 0
    for r in rows:
        street = (r.get("Property Address") or "").strip()
        # Need a house number for Zillow to resolve a single property. A bare
        # street or a vacant "0 Debbie Ln" won't; keep and move on.
        if not street or not re.match(r"^\d", street) or street.startswith("0 "):
            kept.append(r)
            continue
        st = zs.get_status(street,
                           (r.get("Property City") or "").strip(),
                           (r.get("Property State") or "NC").strip(),
                           (r.get("Property Zip") or "").strip())
        if st.status != "unknown" and st.detail:
            note = f"[Zillow: {st.detail}]"
            existing = (r.get("Notes") or "").strip()
            if note not in existing:
                r["Notes"] = (existing + ("\n" if existing else "") + note).strip()
        if st.should_drop():
            print(f"  ZILLOW DROP {r.get('County')}/{r.get('Deceased Owner')}: "
                  f"{street!r} — {st.status} ({st.detail})")
            tag_reason(r, f"zillow-{st.status.replace('_', '-')}")
            dropped += 1
            continue
        kept.append(r)
    return kept, dropped


def _is_small_estate_disposed_recent(r: dict) -> bool:
    """True when the row is a Disposed case with a recent File Date —
    the Small Estate Affidavit pattern (Filed and Disposed same day).
    Same heuristic as ecourts_scraper._row_to_notice but applied to
    polish-stage dict rows (File Date may be 'M/D/YYYY' or 'YYYY-MM-DD').
    Requires Case Status to actually be Disposed (or a synonym) —
    otherwise Pending rows would spuriously bypass downstream filters
    that this helper gates.
    """
    status = (r.get("Case Status") or "").strip().upper()
    if not any(s in status for s in ("DISPOSED", "CLOSED", "INACTIVE", "TRANSFERRED")):
        return False
    file_date = (r.get("File Date") or "").strip()
    if not file_date:
        return False
    parsed = None
    for fmt in ("%m/%d/%Y", "%-m/%-d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            parsed = datetime.strptime(file_date, fmt).date()
            break
        except ValueError:
            continue
    if parsed is None:
        return False
    age = (datetime.now().date() - parsed).days
    return 0 <= age <= SMALL_ESTATE_RECENT_DAYS


MANUAL_INDEX_PATH = Path("output") / ".manual_archive_index.json"


def _name_token_key(name: str) -> str:
    """Order-independent name key — covers 'Last, First Middle' /
    'First Middle Last' / 'AKA <alt>' variations. Must match
    build_manual_archive_index.py:name_token_key.
    """
    s = (name or "").upper()
    s = re.sub(r"\bAKA\b.*", "", s)
    s = re.sub(r"\b(JR|SR|II|III|IV|MR|MRS|MS|DR)\.?\b", "", s)
    tokens = sorted(t for t in re.findall(r"[A-Z]+", s) if len(t) >= 3)
    return " ".join(tokens)


def _parse_file_date_str(s: str) -> datetime | None:
    """Parse a 'File Date' cell value to datetime (MM/DD/YYYY or YYYY-MM-DD)."""
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _current_iso_week(rows: list[dict]) -> str | None:
    """Derive the ISO week tag (e.g. 'Week 21') from the most common File
    Date in this batch. Returns None when no parseable dates exist.
    """
    from collections import Counter
    counts: Counter[int] = Counter()
    for r in rows:
        d = _parse_file_date_str(r.get("File Date", ""))
        if d:
            counts[d.isocalendar().week] += 1
    if not counts:
        return None
    return f"Week {counts.most_common(1)[0][0]}"


def _week_number(label: str) -> int | None:
    """Extract the ISO week number from any week label. Archive entries store
    the week in several shapes depending on source — 'Week 27', 'Week 27 2026'
    (legacy XLSX tab), 'Week 27 (csv)' (per-week CSV) — so cross-week dedup
    must compare NUMBERS, not raw strings, or the same week reads as different."""
    m = re.search(r"week\s*0*(\d+)", label or "", re.IGNORECASE)
    return int(m.group(1)) if m else None


def backfill_from_manual_archive(rows: list[dict]) -> tuple[int, int]:
    """Compare each row against the manual XLSX archive of prior pulls.

    - Blank-case-no row + archive match → backfill case# + flag as duplicate
    - Has-case-no row + archive match in a PRIOR week → flag as duplicate
    - No match → no-op

    Rows flagged as duplicates carry `_archive_duplicate = True` for the
    downstream `drop_archive_duplicates` step to remove.

    Returns (backfilled_or_flagged, no_match_count).
    """
    import json
    if not MANUAL_INDEX_PATH.exists():
        print(f"  Manual archive index not found ({MANUAL_INDEX_PATH}). "
              f"Run build_manual_archive_index.py first.")
        return (0, 0)
    with MANUAL_INDEX_PATH.open(encoding="utf-8") as f:
        archive = json.load(f).get("_entries", {})

    current_week = _current_iso_week(rows)
    if current_week:
        print(f"  Current scrape ISO week (from file dates): {current_week}")

    backfilled = 0
    no_match = 0
    for r in rows:
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or not county:
            continue
        key = f"{county.upper()}||{_name_token_key(dec)}"
        entry = archive.get(key)
        if not entry:
            if not (r.get("Case No.") or "").strip():
                no_match += 1
            continue
        # Match found — decide whether it's a cross-week duplicate. Compare by
        # ISO WEEK NUMBER (not raw label): the archive stores weeks as
        # "Week 27 (csv)" / "Week 27 2026" / "Week 27" depending on source, so a
        # string compare treats the SAME week as different and wrongly drops
        # current-week cases the user pulled by hand this week (Shuford/Deal/
        # Ramsey 26E000779/780-170, 26E000647-480 — Week 27).
        archive_week = entry.get("week", "")
        cur_wk = _week_number(current_week or "")
        arc_wk = _week_number(archive_week)
        is_blank_case = not (r.get("Case No.") or "").strip()
        is_prior_week = cur_wk is not None and arc_wk is not None and arc_wk < cur_wk
        # Note: we do NOT skip same-week matches. The manual pull is
        # authoritative, so we still backfill BLANK fields from it below (fills
        # only gaps, never overrides scraped data) — this heals scrape-time
        # misses like a parcel a GIS outage dropped (Ramsey/Iredell, Week 27).
        # Only the DROP is gated on a genuinely prior week (see below).

        if is_blank_case:
            # Backfill the case number from archive
            r["Case No."] = entry.get("case_no", "")
        # Backfill missing PR/mailing fields when archive has them and ours are blank
        for archive_key, csv_col in [
            ("first_name", "First Name"),
            ("last_name", "Last Name"),
            ("mailing_address", "Mailing Address"),
            ("mailing_city", "Mailing City"),
            ("mailing_state", "Mailing State"),
            ("mailing_zip", "Mailing Zip"),
            ("parcel_id", "Parcel ID"),
            ("property_address", "Property Address"),
            ("property_city", "Property City"),
            ("property_zip", "Property Zip"),
        ]:
            if not (r.get(csv_col) or "").strip() and entry.get(archive_key):
                r[csv_col] = entry[archive_key]
        # Tag the row so user can see it came from archive
        existing_tags = (r.get("Tags") or "").strip()
        marker = f"manual-archive:{entry.get('week', '?')}"
        if marker not in existing_tags:
            r["Tags"] = f"{existing_tags} | {marker}" if existing_tags else marker
        # Drop ONLY for a genuinely PRIOR week: the case was filed weeks ago and
        # the user already has it in an earlier week's pipeline, so this row is
        # just the later newspaper Notice-to-Creditors. A same-week match is the
        # user's own current pull of the same case — keep it (same-week dupes are
        # handled by the soft-dedup and collapse steps, not here).
        if is_prior_week:
            r["_archive_duplicate"] = True
            print(f"  Archive match (drop as cross-week dupe): {county}/{dec!r} -> "
                  f"case {entry['case_no']} (from {entry.get('week', '?')})")
        backfilled += 1
    return backfilled, no_match


def backfill_pr_from_parties(rows: list[dict]) -> int:
    """Fill blank / 'Heirs of' Personal Representative from the eCourts Parties API.

    The scrape sets PR from CaseDetail.executor; cases scraped before an
    executor-recognition fix (e.g. the 2026-07-12 Co-Executor change) keep a
    stale blank. The nightly merge rebuilds from those raw scrapes and skips
    re-scraping seen cases, so a blank PR never self-heals on its own. This
    re-fetches Parties for blank-PR rows every run, so such fixes take effect on
    already-scraped cases without a full re-scrape. Cheap: only blank-PR rows,
    OData Parties endpoint (not the rate-limited DisplayDoc). No-ops without a
    cached WAF cookie.
    """
    import json as _json
    from pathlib import Path as _Path

    def _blank_pr(r: dict) -> bool:
        pr = (r.get("Personal Representative") or "").strip().lower()
        return (not pr) or pr.startswith("heirs of")

    targets = [r for r in rows
               if _blank_pr(r) and (r.get("Case ID (hex)") or "").strip()]
    if not targets:
        return 0

    waf_path = _Path("ecourts_waf_cookies.json")
    if not waf_path.exists():
        print("  (no cached WAF cookie — skipping PR backfill)")
        return 0
    try:
        waf = _json.loads(waf_path.read_text())
        from ecourts_case_api import CaseDetailClient
        from reenrich_ftm_executors import detail_to_fill_dict, apply_fill_to_row
        client = CaseDetailClient(waf_token=waf["aws_waf_token"],
                                  user_agent=waf.get("user_agent") or "Mozilla/5.0")
    except Exception as e:  # noqa: BLE001
        print(f"  (PR backfill unavailable: {e})")
        return 0

    filled = 0
    for r in targets:
        try:
            detail = client.fetch_detail((r.get("Case ID (hex)") or "").strip())
        except Exception as e:  # noqa: BLE001
            print(f"  PR backfill: Parties fetch failed for {r.get('Case No.')}: {e}")
            continue
        fill = detail_to_fill_dict(detail)   # None for guardianships, {} if no PR
        if fill and fill.get("Personal Representative"):
            apply_fill_to_row(r, fill)
            tag_reason(r, "pr-backfill-parties")
            filled += 1
            print(f"  PR backfill {r.get('Case No.')} {r.get('Deceased Owner')}: "
                  f"-> {fill['Personal Representative']}")
    return filled


_MANUAL_DROPS_PATH = Path("manual_drops.txt")


def _load_manual_drops(path: Path = _MANUAL_DROPS_PATH) -> dict[str, str]:
    """Load the user-maintained 'always drop these case numbers' list.

    Plain text, one entry per line. `#` starts a comment. The first
    whitespace/comma token on a line is the Case No.; any trailing text is a
    human-readable reason (ignored by code). Returns {CASE_NO_UPPER: reason}.

    Used for rows the county GIS can't classify on its own — e.g. condos in
    zoning-only counties (Lincoln/Catawba) that read as SFR. The user adds the
    Case No. and the row stays dropped across every future re-scrape.
    """
    drops: dict[str, str] = {}
    if not path.exists():
        return drops
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = line.replace(",", " ", 1).split(None, 1)
        case_no = parts[0].strip().upper()
        reason = parts[1].strip() if len(parts) > 1 else ""
        if case_no:
            drops[case_no] = reason
    return drops


def drop_manual_exclusions(rows: list[dict]) -> tuple[list[dict], int, list[str]]:
    """Drop rows whose Case No. is in the user's manual_drops.txt list."""
    drops = _load_manual_drops()
    if not drops:
        return rows, 0, []
    kept: list[dict] = []
    removed: list[str] = []
    for r in rows:
        cn = (r.get("Case No.") or "").strip().upper()
        if cn and cn in drops:
            reason = drops[cn]
            removed.append(f"{cn} {r.get('Deceased Owner', '')}".strip()
                           + (f" — {reason}" if reason else ""))
        else:
            kept.append(r)
    return kept, len(removed), removed


def drop_archive_duplicates(rows: list[dict]) -> tuple[list[dict], int]:
    """Drop rows that were backfilled from the manual archive.
    These represent cases already in the user's prior-week pipeline —
    keeping them here would create duplicate uploads to DataSift.
    """
    kept = [r for r in rows if not r.get("_archive_duplicate")]
    dropped = len(rows) - len(kept)
    # Strip the private marker from kept rows for cleanliness
    for r in kept:
        r.pop("_archive_duplicate", None)
    return kept, dropped


def _norm_dec_name(name: str) -> tuple[str, str]:
    """Return (last_name, first_name) normalized for cross-source matching.
    Handles both 'Last, First Middle' and 'First Middle Last' formats, and
    common newspaper-publishing parser quirks like 'Jr, Floyd David Long'.
    """
    name = (name or "").strip()
    if not name:
        return ("", "")
    # Strip 'Aka X Y Z' aliases — keep the primary name
    if " AKA " in name.upper():
        name = re.split(r"\s+AKA\s+", name, flags=re.IGNORECASE, maxsplit=1)[0].strip()
    # Strip noise tokens (JR / SR / II / III) that the newspaper parser sometimes
    # treats as a separate "name" element
    noise = {"JR", "JR.", "SR", "SR.", "II", "III", "IV"}
    tokens = re.split(r"[,\s]+", name)
    tokens = [t.strip().upper() for t in tokens if t.strip() and t.strip().upper() not in noise]
    if not tokens:
        return ("", "")
    # Heuristic: in our scraper data, last name is usually:
    #  - First token if format was "Last, First Middle" (comma split)
    #  - Last token if format was "First Middle Last"
    # We try both and let the caller match on either order.
    return (tokens[0], tokens[-1])


def _dec_name_tokens(name: str) -> set[str]:
    """Return the full set of meaningful name tokens for subset/superset
    matching. Strips JR/SR/aliases like _norm_dec_name but keeps middle
    names — useful for catching newspaper-truncated rows like
    'Walker, Betty' that match the eCourts row 'Walker, Betty Louise'.
    """
    name = (name or "").strip()
    if not name:
        return set()
    if " AKA " in name.upper():
        name = re.split(r"\s+AKA\s+", name, flags=re.IGNORECASE, maxsplit=1)[0].strip()
    noise = {"JR", "JR.", "SR", "SR.", "II", "III", "IV"}
    return {
        t.strip().upper() for t in re.split(r"[,\s]+", name)
        if t.strip() and t.strip().upper() not in noise
    }


def soft_dedup_blank_case_no(rows: list[dict]) -> tuple[list[dict], int, int]:
    """For rows with blank Case No. (newspaper-published notices), check if
    the same decedent already exists as a named-case row in the same county.
    If so, MERGE: drop the blank row, but copy any fields the named row was
    missing (e.g., PR mailing that newspaper had but eCourts didn't).

    Returns (rows, merged_count, kept_unique_count).
    """
    import re as _re  # local alias because fn defined inside file
    named = [r for r in rows if (r.get("Case No.") or "").strip()]
    blanks = [r for r in rows if not (r.get("Case No.") or "").strip()]

    # Build lookup: by (county, last_name_tokens)
    named_by_key: dict[tuple[str, str], list[dict]] = {}
    for r in named:
        last, first = _norm_dec_name(r.get("Deceased Owner", ""))
        for tok in (last, first):  # try both orderings
            if tok:
                named_by_key.setdefault((r["County"], tok), []).append(r)

    merged = 0
    to_drop_ids: set[int] = set()
    for b in blanks:
        last_b, first_b = _norm_dec_name(b.get("Deceased Owner", ""))
        if not last_b:
            continue
        tokens_b = _dec_name_tokens(b.get("Deceased Owner", ""))
        # Look up candidates using ANY of the tokens
        candidates = []
        for tok in (last_b, first_b):
            for r in named_by_key.get((b["County"], tok), []):
                if r not in candidates:
                    candidates.append(r)
        # Match strategies:
        #   1. Strict 2-token match (Last, First ↔ First Last orderings)
        #   2. Subset match (truncated newspaper name is a subset of the
        #      full eCourts name — catches 'Walker, Betty' ↔
        #      'Walker, Betty Louise')
        for r in candidates:
            last_r, first_r = _norm_dec_name(r.get("Deceased Owner", ""))
            tokens_r = _dec_name_tokens(r.get("Deceased Owner", ""))
            strict = (
                (last_b == last_r or last_b == first_r)
                and (first_b == first_r or first_b == last_r)
            )
            # Subset = blank name's tokens are all contained in named name's
            # tokens, AND there's at least 2 tokens of overlap (so single
            # 'WHITE' doesn't merge into 'WHITE, JANE DOE').
            subset = (
                tokens_b
                and tokens_b.issubset(tokens_r)
                and len(tokens_b & tokens_r) >= 2
            )
            if strict or subset:
                # Match found — merge
                for col in ("Mailing Address", "Mailing City", "Mailing State", "Mailing Zip",
                            "Property Address", "Property City", "Property Zip", "Property use",
                            "Parcel ID", "Beneficiaries", "Notes", "First Name", "Last Name",
                            "Personal Representative"):
                    if not (r.get(col) or "").strip() and (b.get(col) or "").strip():
                        r[col] = b[col]
                to_drop_ids.add(id(b))
                merged += 1
                break

    kept = [r for r in rows if id(r) not in to_drop_ids]
    remaining_blanks = sum(1 for r in kept if not (r.get("Case No.") or "").strip())
    return kept, merged, remaining_blanks


def fill_pr_mailing_via_people_search(rows: list[dict], state: str = "NC") -> tuple[int, int]:
    """For rows with a named PR/IP but no mailing address from the court
    record, look up the PR's CURRENT residence via the people-search
    waterfall (Serper + Firecrawl + LLM, then Tracerfy). Runs BEFORE the
    property-address fallback so direct mail targets the PR's real address
    when we can find one.

    Builds the search name from the row's separate First Name / Last Name
    columns — those are correctly ordered "First Last" (from Odyssey's
    NameFirst/NameLast). Do NOT use "Personal Representative" raw — that may
    carry Odyssey's "Last, First" comma form, which the people-search URL
    builders mis-parse (they take token[0]=first, token[-1]=last).

    Returns (found, attempted).
    """
    try:
        import config as cfg
        from obituary_enricher import _lookup_dm_address, _lookup_dm_address_tracerfy
        import tracerfy_budget
    except Exception as e:  # pragma: no cover
        print(f"  people-search unavailable ({e}) — skipping, property fallback will apply")
        return (0, 0)

    api_key = cfg.ANTHROPIC_API_KEY
    tracerfy_ok = bool(getattr(cfg, "TRACERFY_API_KEY", ""))
    cache: dict[str, dict] = {}
    found = attempted = 0
    for r in rows:
        if r.get("First Name") == "Heirs":
            continue
        first = (r.get("First Name") or "").strip()
        last = (r.get("Last Name") or "").strip()
        if not first or not last:
            continue  # need a real PR name to search
        if (r.get("Mailing Address") or "").strip():
            continue  # already have a court-supplied mailing
        # Prefer the will-extracted full legal name when available — it
        # carries middle names that the eCourts Parties API truncates
        # ("Daniel Cox" -> "DANIEL CLINTON COX"). Full middle names are
        # dramatically more disambiguating in both Tier 0 county GIS
        # search AND Tier 2 people-search. See [[project_county_gis_as_pr_address_verification]].
        will_full = (r.get("PR Full Name (Will)") or "").strip()
        if will_full:
            name = will_full
        else:
            name = f"{first} {last}"
        # Property city is a soft locality hint — PRs are often local to the
        # decedent. The lookup tolerates a wrong/empty city (national search).
        city = (r.get("Property City") or "").strip()
        # County is the STRICT anchor (Week 26 audit fix): without it, common-
        # name PRs like "Daniel Cox" matched a stranger 4 hrs away in Oak City
        # NC. The LLM now treats same-county-or-adjacent as "high" confidence
        # only; other matches return blank and fall through to property mailing.
        county = (r.get("County") or "").strip()
        attempted += 1
        # Cache key includes county so different counties don't share results
        key = f"{name}|{city}|{county}".upper()
        res = cache.get(key)
        if res is None:
            try:
                # Tier 1: free Serper + Firecrawl + LLM via CyberBackgroundChecks.
                res = _lookup_dm_address(name, city, api_key, state=state, county=county)
            except Exception as e:
                print(f"    people-search error for {name!r}: {e}")
                res = {}
            # Tier 2: Tracerfy paid skip-trace (5 credits / $0.10 per hit, 0
            # on miss). Anchors on the property address to disambiguate the
            # PR — most PRs are heirs whose former address is the property.
            if (not res or not res.get("street")) and tracerfy_ok and tracerfy_budget.can_spend():
                try:
                    tf = _lookup_dm_address_tracerfy(
                        name, city,
                        address=r.get("Property Address", "") or "",
                        zip_code=r.get("Property Zip", "") or "",
                        state=state,
                    )
                except Exception:
                    tf = None
                if tf and tf.get("street"):
                    tf["source"] = "tracerfy"
                    res = tf
                    warn_now, msg = tracerfy_budget.record_hit()
                    if warn_now:
                        print(f"  {msg}")
            cache[key] = res
        if res and res.get("street"):
            r["Mailing Address"] = res["street"]
            r["Mailing City"] = res.get("city") or city
            r["Mailing State"] = res.get("state") or "NC"
            r["Mailing Zip"] = res.get("zip") or ""
            source = (res.get("source") or "").lower()
            if "tracerfy" in source:
                tag_reason(r, "pr-tracerfy")
            else:
                tag_reason(r, "pr-people-search")
            found += 1
            print(f"    PR address: {name} -> {res['street']}, "
                  f"{r['Mailing City']} {r['Mailing Zip']} [{res.get('source', '')}]")
    return (found, attempted)


# NC county GIS owner-name formats: Mecklenburg (Polaris) lists an owner as
# "FIRST [MIDDLE] LAST"; the six ArcGIS counties list "LAST FIRST [MIDDLE]".
_COOWNER_NOISE = {
    "WF", "HB", "HUS", "HUSB", "ETUX", "ETAL", "ET", "AL", "TRUSTEE", "TR",
    "TRUST", "LIFE", "ESTATE", "HEIRS", "AND", "LLC", "INC",
    "JR", "SR", "II", "III", "IV", "V",
}


def _gis_owner_name_parts(segment: str, county: str) -> tuple[str, str, str]:
    """Parse one GIS owner segment into (First, Last, Middle), county-format-
    aware. 'CAMPBELL CARSON CORRELL III' (Rowan) -> ('Carson','Campbell','Correll').
    The middle name matters for people-search precision on common names."""
    toks = [t.strip(",.") for t in re.split(r"\s+", (segment or "").upper()) if t.strip(",.")]
    core = [t for t in toks if t not in _COOWNER_NOISE and len(t) > 1]
    if len(core) < 2:
        return "", "", ""
    if county.strip().lower() == "mecklenburg":
        first, last, mids = core[0], core[-1], core[1:-1]
    else:
        last, first, mids = core[0], core[1], core[2:]
    return first.title(), last.title(), " ".join(mids).title()


def promote_deed_coowner_to_dm(
    rows: list[dict], *, do_address_lookup: bool = True, state: str = "NC",
) -> tuple[int, int]:
    """For no-PR ('Heirs of') rows, name the deed CO-OWNER as the decision
    maker. When a decedent jointly owns the property with someone the court
    never named as PR — and that co-owner has a DIFFERENT surname, so they're a
    distinct person, not same-family noise — that co-owner is the person to
    contact: they own the property too. The county deed already carries the
    name, so this needs no obituary (more reliable — county records are always
    present and authoritative).

    Hunt 26E000726-790 Rowan: parcel 2230 Amity Hill Rd owner
    "CAMPBELL CARSON CORRELL III | HUNT ARCHIE F" — the pipeline used the HUNT
    half to match the parcel and ignored the CAMPBELL half, falling to "Heirs
    of." Now it names Carson Campbell (co-owner) as DM and, when do_address_lookup
    is on, looks up his CURRENT mailing address via the people-search waterfall
    (Oren found 650 Peach Orchard Rd by hand).

    Sets DM columns only — the row stays "Heirs of <Decedent>" so it still flows
    through Oren's deep-prospecting workflow, now with a concrete contact.
    Returns (dm_set, addr_found).
    """
    from nc_gis_lookup import lookup_properties, extract_co_owner_names, split_decedent_name
    addr_fn = None
    api_key = ""
    if do_address_lookup:
        try:
            import config as cfg
            from obituary_enricher import _lookup_dm_address
            addr_fn = _lookup_dm_address
            api_key = getattr(cfg, "ANTHROPIC_API_KEY", "")
        except Exception as e:
            print(f"  co-owner address lookup unavailable ({e}) — DM name only")
            addr_fn = None
    dm_set = addr_found = 0
    addr_cache: dict[str, dict] = {}
    for r in rows:
        first = (r.get("First Name") or "").strip()
        if first and first != "Heirs":
            continue  # already has a court-named PR
        dm = (r.get("DM Name") or "").strip().lower()
        if dm and dm != "estate of":
            continue  # already have a decision maker (obituary, etc.)
        pid = (r.get("Parcel ID") or "").strip()
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not pid or not dec or not county or "IN THE MATTER" in dec.upper():
            continue
        try:
            cands = lookup_properties(dec, county, min_score=0.3)
            match = next((c for c in cands if (c.pid or "") == pid), None)
        except Exception:
            match = None
        if not match or not match.owner_name:
            continue
        co_owners = extract_co_owner_names(match.owner_name, dec)
        if not co_owners:
            continue
        # Prefer a co-owner whose surname differs from the decedent's — a
        # definitively distinct person, not a same-name relative listing.
        _f, _m, dec_last = split_decedent_name(dec)
        dec_last_up = (dec_last or "").upper()
        picked = None
        for co in co_owners:
            fn, ln, mid = _gis_owner_name_parts(co, county)
            if fn and ln:
                if ln.upper() != dec_last_up:
                    picked = (fn, ln, mid)
                    break
                if picked is None:
                    picked = (fn, ln, mid)  # same-surname co-owner as a fallback
        if not picked:
            continue
        fn, ln, mid = picked
        r["DM Name"] = f"{fn} {ln}"
        r["DM Relationship"] = "co-owner (deed)"
        tag_reason(r, "coowner-dm")
        dm_set += 1
        print(f"  CO-OWNER DM {county}/{dec}: {fn} {ln} (deed co-owner of "
              f"{match.situs_address or pid})")

        # Tier 0 (free, deed-authoritative): the co-owned parcel's OWN tax-bill
        # mailing address — that's literally where the county mails the owners.
        # Only use it when it differs from the property situs (an owner-occupied
        # parcel mails to itself, which tells us nothing new). Hunt: Amity Hill's
        # tax mailing is 901 N Jackson St, Salisbury — Carson's mailing on file.
        street = city = zipc = ""
        situs_norm = _norm_addr_simple(match.situs_address or "")
        if match.mailing_address and _norm_addr_simple(match.mailing_address) != situs_norm:
            street, city, zipc = _parse_mailing_blob(match.mailing_address)
        # Tier 1 (fallback): people-search the co-owner's current residence.
        if not street and addr_fn:
            search_name = f"{fn} {mid} {ln}".replace("  ", " ").strip() if mid else f"{fn} {ln}"
            key = f"{search_name}|{county}".upper()
            res = addr_cache.get(key)
            if res is None:
                try:
                    res = addr_fn(search_name, (r.get("Property City") or ""),
                                  api_key, state=state, county=county) or {}
                except Exception:
                    res = {}
                addr_cache[key] = res
            street, city, zipc = res.get("street", ""), res.get("city", ""), res.get("zip", "")
        if street:
            r["Mailing Address"] = street
            if city:
                r["Mailing City"] = city
            if zipc:
                r["Mailing Zip"] = zipc
            tag_reason(r, "coowner-address")
            addr_found += 1
            print(f"    -> mailing: {street} {city} {zipc}")
    return dm_set, addr_found


_MAIL_SUFFIXES = {
    "ST", "STREET", "RD", "ROAD", "DR", "DRIVE", "AVE", "AVENUE", "LN", "LANE",
    "CT", "COURT", "BLVD", "PL", "PLACE", "CIR", "CIRCLE", "WAY", "TRL", "TRAIL",
    "PKWY", "HWY", "LOOP", "RUN", "PT", "POINT", "TER", "TERRACE", "CV", "COVE",
    "SQ", "XING", "BND", "BLF", "HWY", "EXT",
}


def _parse_mailing_blob(blob: str) -> tuple[str, str, str]:
    """Split '901 N JACKSON ST SALISBURY NC 28144-3411' into
    (street, city, zip5) by finding the state+zip tail, then the last street
    suffix (city is whatever follows it). Best-effort."""
    s = (blob or "").strip()
    m = re.search(r"\b([A-Z]{2})\s+(\d{5})(?:-\d{4})?\s*$", s)
    if not m:
        return (s, "", "")
    zipc = m.group(2)
    head = s[:m.start()].strip()          # "901 N JACKSON ST SALISBURY"
    toks = head.split()
    last_suf = -1
    for i, t in enumerate(toks):
        if t.strip(",.").upper() in _MAIL_SUFFIXES:
            last_suf = i
    if 0 <= last_suf < len(toks) - 1:
        street = " ".join(toks[:last_suf + 1])
        city = " ".join(toks[last_suf + 1:])
    else:
        street, city = head, ""
    return (street.strip(), city.strip(), zipc)


def _norm_addr_simple(s: str) -> str:
    return "".join(ch for ch in (s or "").upper() if ch.isalnum())


def fill_missing_pr_mailing_from_property(rows: list[dict]) -> int:
    """For rows where a Personal Rep / Executor / Interested Person is named
    but their mailing address was not in the court record, fall back to
    using the property address as the mailing. Direct mail still gets
    delivered (to the property), addressed to the PR by name.

    This matches the user's manual workflow. Phase 2 (future): replace
    with the PR's actual mailing or the heir's mailing once deep
    prospecting (obituary + skip trace) finds them.

    Returns the number of rows updated. Skips 'Heirs of' rows (their
    mailing was already set this way by the Heirs-of transform).
    """
    filled = 0
    for r in rows:
        if r.get("First Name") == "Heirs":
            continue
        if not (r.get("Personal Representative") or "").strip():
            continue  # no PR named — falls through to Heirs-of treatment elsewhere
        if (r.get("Mailing Address") or "").strip():
            continue  # already has a mailing
        prop_addr = (r.get("Property Address") or "").strip()
        if not prop_addr:
            continue  # no property to fall back to
        r["Mailing Address"] = prop_addr
        if (r.get("Property City") or "").strip():
            r["Mailing City"] = r["Property City"]
        r["Mailing State"] = "NC"
        if (r.get("Property Zip") or "").strip():
            r["Mailing Zip"] = r["Property Zip"]
        tag_reason(r, "mailing-from-property")
        # Keep-but-flag (Oren 2026-07-01): we had NO independent address for this
        # PR / interested person, so we're mailing to the property itself. That
        # often means they live there (heir-occupied) — but we can't confirm
        # without the court Application. Surface it for manual review rather than
        # dropping or silently mailing. Austin 26E000883-350 (Cantler @ 119 Cane
        # Forest Dr) is the motivating case.
        tags = (r.get("Tags") or "").strip()
        if "Verify: No PR Address" not in tags:
            r["Tags"] = (tags + ", " if tags else "") + "Verify: No PR Address"
        filled += 1
    return filled


_STREET_SUFFIX_NORMALIZE = {
    # Map long form -> short form so "Drive" and "Dr" compare equal.
    # (Court records use Drive/Avenue/Street/etc.; county GIS often
    # abbreviates. Without this map, heir-occupancy comparison missed
    # cases like Young 26E002125-590: mailing="8110 Gera Emma Drive"
    # vs property="8110 Gera Emma Dr".)
    "drive": "dr",
    "avenue": "ave",
    "av": "ave",      # Sellers 26E000406-540: "319 Sherrill Av" vs "319 Sherrill Avenue"
    "street": "st",
    "road": "rd",
    "boulevard": "blvd",
    "bl": "blvd",
    "court": "ct",
    "place": "pl",
    "lane": "ln",
    "circle": "cir",
    # Some counties (incl. Mecklenburg) abbreviate Circle as "Cr". Collapse
    # both Cr and Circle to the same canonical "cir" so they compare equal.
    # Without this, VEIT 26E002255-590 leaked through heir-occupied:
    # mailing="8631 KNOLLWOOD CIRCLE" vs property="8631 Knollwood Cr".
    "cr": "cir",
    "highway": "hwy",
    "parkway": "pkwy",
    "terrace": "ter",
    "trail": "trl",
    "way": "wy",
    # Compass directions — county GIS abbreviates ("19827 N Ferry St"),
    # court records spell out ("19827 NORTH FERRY STREET"). Without these,
    # BURTON 26E002256-590 leaked through heir-occupied.
    "north": "n",
    "south": "s",
    "east": "e",
    "west": "w",
    "northeast": "ne",
    "northwest": "nw",
    "southeast": "se",
    "southwest": "sw",
}

# Street-suffix / directional tokens to ignore when comparing the
# "distinctive" part of two street names (see _same_property_by_number_zip).
_STREET_SUFFIX_TOKENS = set(_STREET_SUFFIX_NORMALIZE.values()) | {
    "n", "s", "e", "w", "ne", "nw", "se", "sw",
}


def _leading_house_number(addr: str | None) -> str:
    """Return the leading house number of an address, or '' if none."""
    m = re.match(r"\s*(\d+)", addr or "")
    return m.group(1) if m else ""


# Directionals in both forms, so "North Oakland" and "N Oakland" match.
_DIRECTIONAL_WORDS = {
    "n", "s", "e", "w", "ne", "nw", "se", "sw",
    "north", "south", "east", "west",
    "northeast", "northwest", "southeast", "southwest",
}


def _street_core_tokens(addr: str | None) -> set[str]:
    """Normalized street tokens minus the house number and street
    suffixes/directionals — the 'distinctive' part of a street name."""
    toks = (addr or "").lower().split()
    core: set[str] = set()
    for i, t in enumerate(toks):
        t = _STREET_SUFFIX_NORMALIZE.get(t.rstrip(".").strip(","), t.rstrip(".").strip(","))
        if i == 0 and t.isdigit():
            continue  # house number
        if t and t not in _STREET_SUFFIX_TOKENS:
            core.add(t)
    return core


def _streets_near_identical(a_addr: str | None, b_addr: str | None) -> bool:
    """True when two street names are the SAME modulo a typo/abbreviation —
    every distinctive core token in the shorter name has an exact or fuzzy
    (>=0.80 similarity) match in the longer one. Catches 'Swallow Trail Lane'
    vs 'Swallow Tail Ln' (Walsh 26E002459-590: heir Cook's mailing == property,
    just misspelled) while rejecting genuinely different streets that merely
    share a house number ('Main' vs 'Oak'). Caller must also confirm the house
    number + ZIP match before treating this as heir-occupied.
    """
    import difflib
    ca, cb = _street_core_tokens(a_addr), _street_core_tokens(b_addr)
    if not ca or not cb:
        return False
    small, big = (ca, cb) if len(ca) <= len(cb) else (cb, ca)
    for t in small:
        if t in big:
            continue
        if any(difflib.SequenceMatcher(None, t, u).ratio() >= 0.80 for u in big):
            continue
        return False  # a distinctive token with no near-match — different street
    return True


def _same_property_by_number_zip(
    a_addr: str | None, a_zip: str | None, b_addr: str | None, b_zip: str | None
) -> bool:
    """True when two addresses share the same leading house number AND ZIP
    AND a distinctive street token (a numeric route or a >=4-char name word).

    Catches route-style renderings that exact-string matching misses, e.g.
    mailing "2535 Old Highway 27" vs property "2535 Old NC 27 Hwy" (same
    place) — while NOT false-matching two different streets that merely share
    a house number in one ZIP ("2535 Main St" vs "2535 Oak Ave" share no
    distinctive token, so this returns False).
    """
    az = (a_zip or "").strip()[:5]
    bz = (b_zip or "").strip()[:5]
    if not az or az != bz:
        return False
    an = _leading_house_number(a_addr)
    if not an or an != _leading_house_number(b_addr):
        return False
    shared = _street_core_tokens(a_addr) & _street_core_tokens(b_addr)
    return any(t.isdigit() or len(t) >= 4 for t in shared)


def _house_number(norm: str) -> str:
    """Leading digit run of a norm_addr() string (the house number), or ''."""
    m = re.match(r"^(\d+)", norm or "")
    return m.group(1) if m else ""


def _within_one_edit(a: str, b: str) -> bool:
    """True when a and b differ by at most one insert/delete/substitute."""
    if a == b:
        return True
    la, lb = len(a), len(b)
    if abs(la - lb) > 1:
        return False
    i = 0
    while i < la and i < lb and a[i] == b[i]:
        i += 1
    if la == lb:            # substitution — remainder must match
        return a[i + 1:] == b[i + 1:]
    if la < lb:             # single insertion in b
        return a[i:] == b[i + 1:]
    return a[i + 1:] == b[i:]  # single deletion in b


def _heir_addr_match(a: str, b: str) -> bool:
    """Heir-occupancy address equality with a tight fuzz for deed-vs-court
    spelling drift. Exact match, OR: identical house number AND the street
    portion is within a single edit. The identical-house-number guard keeps
    this from ever collapsing two different addresses on the same street —
    it only forgives a one-character typo/OCR variance in the street name.

    James 26E002599-590 Week 29: beneficiary Elizabeth James lives at the
    property, but her court address spelled it "5126 Glenbriar Dr" while the
    deed spelled it "5126 Glenbrier Dr" — one letter (a/e) defeated the exact
    match and the heir-occupied DQ never fired.
    """
    if not a or not b:
        return False
    if a == b:
        return True
    ha, hb = _house_number(a), _house_number(b)
    if not ha or ha != hb:
        return False
    return _within_one_edit(a, b)


def drop_executor_at_property(rows: list[dict]) -> tuple[list[dict], int]:
    """Drop rows where the executor's mailing address matches the property
    address — meaning the executor LIVES at the property. They almost
    certainly inherit and stay (heir-occupied — bad probate lead).

    ALSO drops rows where ANY beneficiary's address (from the
    Beneficiaries column) matches the property address — same signal,
    just from a different person in the court file. Catches Lima Heidi,
    Mathis Lillie, DAVIDGE Edward Charles class.

    Only EXEMPTION: 'Heirs of ...' rows where we deliberately set
    mailing := property as the no-PR-found backstop. Even Small Estate
    Affidavit cases are subject to this filter — if the applicant
    lives at the property, the lead is dead by Oren's policy regardless
    of case type (per user feedback Week 25 audit).
    """
    def norm_addr(s: str | None) -> str:
        # Lowercase, normalize Drive/Dr / Avenue/Ave / etc., then strip to
        # alphanumeric for the equality check. The token rewrite has to
        # happen BEFORE the alphanumeric strip — otherwise "drive" and "dr"
        # both collapse to "dr" the wrong way around when only one side has
        # the full form.
        tokens = (s or "").lower().split()
        norm_tokens = [_STREET_SUFFIX_NORMALIZE.get(t.rstrip("."), t.rstrip(".")) for t in tokens]
        joined = " ".join(norm_tokens)
        return ''.join(c for c in joined if c.isalnum())

    def _row_dq_signals(r: dict, prop_norm: str) -> tuple[bool, str]:
        """Return (is_dq, reason_code). Heir-occupied DQ fires when either
        the executor's mailing or any listed beneficiary's address matches
        the property address (all normalized via norm_addr).

        Vacant land exemption: a vacant lot cannot be "occupied" by anyone.
        When situs falls back to the owner's mailing (Cabarrus/etc. for
        parcels without NG911 addresses), a vacant lot LOOKS heir-occupied
        because situs == mailing. Skip the DQ for these. Surfaced Week 26
        via Bostian 26E000662-120: residence at 4699 Rainbow Dr correctly
        heir-occupied, but the decedent's adjacent vacant lot (separate
        parcel, no NG911 address, situs fell back to 4699 Rainbow Dr too)
        was incorrectly dropped along with it.
        """
        use = (r.get("Property use") or "").upper()
        if "VACANT" in use or use == "LAND":
            return (False, "")
        mail = norm_addr(r.get("Mailing Address"))
        if mail and _heir_addr_match(prop_norm, mail):
            return (True, "dq-executor-at-property")
        ben_block = (r.get("Beneficiaries") or "")
        if ben_block:
            for ben_addr in _BENEFICIARY_ADDR_RE.findall(ben_block):
                if _heir_addr_match(norm_addr(ben_addr), prop_norm):
                    return (True, "dq-beneficiary-at-property")
        # Application-PDF heirs (when available) — the OData Parties API
        # often misses an heir/applicant address that the Application PDF
        # records. Once the case-doc retry queue fetches the App PDF,
        # Heirs (App) column carries JSON [{full_name, street, city, ...}, ...].
        # Surfaced Week 26 audit: Ramsey 26E002367-590 — PR Bryan Scott
        # Ramsey lives at the property (348 Touch Mc Not Lane) per his
        # own Application but his address didn't appear in the Parties API
        # response, so heir-occupied check missed it and the case stayed
        # in the workbook. With App data on the row, this catches it.
        app_heirs_json = r.get("Heirs (App)") or ""
        if app_heirs_json:
            try:
                import json as _json
                heirs = _json.loads(app_heirs_json)
                if isinstance(heirs, list):
                    for h in heirs:
                        if not isinstance(h, dict):
                            continue
                        street = (h.get("street") or "").strip()
                        if street and _heir_addr_match(norm_addr(street), prop_norm):
                            return (True, "dq-app-heir-at-property")
            except (ValueError, TypeError):
                pass
        # PR-is-spouse-co-owner check (Tenancy-by-Entirety auto-transfer).
        # Only fires when the deed segment containing the DECEDENT has an
        # explicit marriage marker (WF/HSB) AND the PR's first+last appears
        # as a separate co-owner segment. In NC, married couples on a deed
        # default to tenancy by the entirety (auto-survivorship) — definite
        # JTWROS, no probate needed for that property.
        #
        # Without a marriage marker (e.g. "MORRISON JEANETTE C | MORRISON
        # VICTOR HARVEY"), joint ownership could be either tenants-in-common
        # (decedent's share IS in probate — keep) or JTWROS (drop). Ambiguous
        # — DON'T auto-drop. User can judge from the workbook.
        #
        # Surfaced Week 26: Gaston Wilbert 26E000673-120 — owner "GASTON
        # BETTY K | GASTON WILBERT HSB" has HSB marker on decedent Wilbert,
        # PR Betty is the surviving spouse-co-owner -> DQ. Morrison
        # 26E000725-170 — owner "MORRISON JEANETTE C | MORRISON VICTOR
        # HARVEY" has no marker, kept.
        first = (r.get("First Name") or "").strip().upper()
        last = (r.get("Last Name") or "").strip().upper()
        if first and last:
            pid_check = (r.get("Parcel ID") or "").strip()
            dec_check = (r.get("Deceased Owner") or "").strip()
            county_check = (r.get("County") or "").strip()
            if pid_check and dec_check and county_check and "IN THE MATTER" not in dec_check.upper():
                try:
                    from nc_gis_lookup import lookup_properties as _lp, split_decedent_name as _split
                    cands_check = _lp(dec_check, county_check, min_score=0.5)
                    match_check = next((c for c in cands_check if c.pid == pid_check), None)
                except Exception:
                    match_check = None
                if match_check and match_check.is_jointly_owned:
                    owner_upper = (match_check.owner_name or "").upper()
                    segments = [s.strip() for s in
                                __import__("re").split(r"\s*\|\s*|\s+&\s+|\s*;\s*", owner_upper)]
                    _MARRIAGE_MARKERS = {"WF", "HSB", "WIFE", "HUSBAND"}
                    # Find the segment containing the DECEDENT's last name AND a
                    # marriage marker — that's the spouse-on-the-deed pattern.
                    _, _, _dec_last = _split(dec_check)
                    dec_last_up = (_dec_last or "").upper()
                    decedent_segment_has_marker = False
                    for seg in segments:
                        seg_tokens = set(t.strip(",.") for t in seg.split())
                        if dec_last_up and dec_last_up in seg_tokens and (seg_tokens & _MARRIAGE_MARKERS):
                            decedent_segment_has_marker = True
                            break
                    if decedent_segment_has_marker:
                        # Now confirm PR is a co-owner (separate segment)
                        for seg in segments:
                            seg_tokens = set(t.strip(",.") for t in seg.split())
                            if first in seg_tokens and last in seg_tokens:
                                return (True, "dq-pr-spouse-coowner")
        # Fuzzy same-property: executor mailing shares house number + ZIP +
        # a distinctive street token with the property, but the street text
        # renders differently (route-style addresses). This is a SWAP-ONLY
        # signal — the caller will re-point a multi-parcel estate to a non-DQ
        # sibling, but will NOT drop a single-parcel lead on this alone
        # (two streets can share a number within one ZIP). Catches Hoover
        # 26E000875-350: mailing "2535 Old Highway 27" == property "2535 Old
        # NC 27 Hwy" (son lives in the decedent's house).
        m_addr, p_addr = r.get("Mailing Address"), r.get("Property Address")
        if _same_property_by_number_zip(
            m_addr, r.get("Mailing Zip"), p_addr, r.get("Property Zip"),
        ):
            # Same house# + ZIP. If the street names are also near-identical
            # (typo/abbrev, e.g. "Swallow Trail Lane" vs "Swallow Tail Ln"),
            # it's confidently the same property — DROP (heir lives there). If
            # only one distinctive token is shared, it's weaker — swap-only, so
            # a single-parcel lead survives (two streets can share a number).
            if _streets_near_identical(m_addr, p_addr):
                return (True, "dq-executor-at-property-street-typo")
            return (True, "dq-executor-at-property-hn")
        return (False, "")

    kept = []
    dropped = 0
    for r in rows:
        if r.get("First Name") == "Heirs":
            kept.append(r)
            continue
        prop = norm_addr(r.get("Property Address"))
        if not prop:
            kept.append(r)
            continue
        is_dq, _reason = _row_dq_signals(r, prop)
        if is_dq:
            # Before dropping, try to swap to a non-DQ sibling parcel from
            # the same decedent's estate. Catches multi-parcel cases like
            # Queen 26E000714-170 where decedent owns residence (heir-
            # occupied) + vacant lot next door — vacant lot is the real
            # lead the residence DQ would otherwise kill.
            if _try_swap_to_non_dq_sibling(r, norm_addr, _row_dq_signals):
                # Re-check DQ on the now-swapped row. The swap's synth-row
                # dq check uses a stripped-down row dict (no First/Last),
                # so PR-as-co-owner can't fire during swap selection. Re-
                # run the full dq_check against the actual row now that
                # all fields are populated. Catches Gaston Wilbert
                # 26E000673-120 Week 26: swapped to a sibling where PR
                # Betty Gaston is on the deed.
                new_prop = norm_addr(r.get("Property Address"))
                post_dq, post_reason = _row_dq_signals(r, new_prop) if new_prop else (False, "")
                if post_dq:
                    tag_reason(r, post_reason)
                    dropped += 1
                    continue
                kept.append(r)
                continue
            # Swap failed. A fuzzy house#+ZIP match alone is not strong enough
            # to DROP a single-parcel lead (two streets can share a number in
            # one ZIP) — keep it. Only the exact-address / beneficiary /
            # app-heir / co-owner signals drop.
            if _reason == "dq-executor-at-property-hn":
                kept.append(r)
                continue
            tag_reason(r, _reason)
            dropped += 1
            continue
        kept.append(r)
    return kept, dropped


def _try_swap_to_non_dq_sibling(
    row: dict,
    norm_addr,
    dq_check,
) -> bool:
    """For a row whose currently-selected main parcel is heir-occupied,
    look up the decedent's full estate via GIS and try to swap in a
    non-DQ sibling (vacant lot, rental, or other property whose situs
    doesn't match the executor mailing OR any beneficiary address).

    Preference order for swap candidate: non-DQ vacant land first (most
    likely to sell), then non-DQ residential/other. Mutates `row` in
    place and returns True on swap; returns False (no swap) when no
    sibling is acceptable.

    `norm_addr` and `dq_check` are passed in to keep the heir-occupied
    semantics identical to the caller (same address-normalization map,
    same DQ predicate that considers executor AND beneficiaries).
    """
    from nc_gis_lookup import lookup_properties, filter_for_lead_quality
    dec = (row.get("Deceased Owner") or "").strip()
    county = (row.get("County") or "").strip()
    if not dec or not county or "IN THE MATTER" in dec.upper():
        return False
    try:
        cands = lookup_properties(dec, county, min_score=0.7)
    except Exception:
        return False
    if not cands:
        return False
    kept_cands = filter_for_lead_quality(
        cands,
        beneficiaries_json=row.get("Beneficiaries", "") or "",
        decedent_name=dec,
    )
    current_pid = (row.get("Parcel ID") or "").strip()
    siblings = [c for c in kept_cands if (c.pid or "") and c.pid != current_pid]
    if not siblings:
        return False

    def rank(c) -> tuple[int, int]:
        """Lower tuple = better. First tier: 0 = non-DQ vacant, 1 = non-DQ
        other, 99 = DQ (skip)."""
        # Build a temp row-shape so dq_check can use the same predicate
        temp_situs = c.situs_address or ""
        if not temp_situs:
            # Vacant parcels often have no situs in GIS — fall back to
            # the owner mailing as a proxy, same as scrape-time logic.
            temp_situs = c.mailing_address or ""
        temp_norm = norm_addr(temp_situs)
        if not temp_norm:
            # No comparable address — risky to swap to; treat as DQ.
            return (99, 0)
        # Build a synthetic row preserving the original row's mailing +
        # beneficiaries (those are what we're comparing against) but with
        # the sibling's situs as the property. Include Property use so
        # the vacant-land exemption in _row_dq_signals also applies to
        # candidate siblings — without this, vacant siblings whose situs
        # falls back to owner mailing would still be rejected.
        sibling_use = (c.use_description or "").upper()
        if not ("VACANT" in sibling_use or "LAND" in sibling_use) and c.is_vacant_land:
            sibling_use = "VACANT LAND"
        synth = {
            "Mailing Address": row.get("Mailing Address", ""),
            "Beneficiaries": row.get("Beneficiaries", ""),
            "Property Address": temp_situs,
            "Property use": sibling_use,
        }
        is_dq, _ = dq_check(synth, temp_norm)
        if is_dq:
            return (99, 0)
        use_desc = (c.use_description or "").upper()
        is_vacant = "VACANT" in use_desc or "LAND" in use_desc or bool(c.is_vacant_land)
        return (0 if is_vacant else 1, -int((c.market_value or 0)))

    siblings.sort(key=rank)
    best = siblings[0]
    if rank(best)[0] >= 99:
        return False  # every sibling also heir-occupied

    return _apply_sibling_swap(
        row, best, current_pid,
        marker=f"[SWAPPED-ON-HEIR-OCCUPIED from {current_pid}: prior main DQ'd]",
        reason_tag="swap-on-dq", log_label="SWAP-ON-DQ",
    )


def _apply_sibling_swap(row: dict, best, old_pid: str, *,
                        marker: str, reason_tag: str, log_label: str) -> bool:
    """Repoint `row`'s main parcel to sibling candidate `best` (mutate in
    place). Shared by the heir-occupied swap (_try_swap_to_non_dq_sibling) and
    the over-cap swap (_try_swap_to_under_cap_sibling). Updates
    address/use/value, keeps Notes coherent (strips the new main from the
    sibling block, records a marker for the old main), tags the reason, logs,
    and returns True.
    """
    dec = (row.get("Deceased Owner") or "").strip()
    county = (row.get("County") or "").strip()
    street, city, zipc = _candidate_to_address_parts(best)
    # Vacant/landlocked parcels often have no situs in GIS — fall back to
    # the user's manual convention rather than dumping the mailing-style
    # "City State Zip" string into the street field (Hefner 26E000742-170
    # Week 26: swap-on-DQ landed on a vacant parcel with situs="",
    # mailing="Hickory NC 28602" -> Property Address showed the mailing
    # which is wrong and confusing). When street is unknown:
    #   * If we have a real situs street, use "0 <street>" (Oren's
    #     vacant-lot convention)
    #   * Otherwise use "No Address" (landlocked or NG911-unassigned)
    if not street:
        # situs_address may exist but be a no-number street name only
        situs_text = (best.situs_address or "").strip()
        if situs_text and not situs_text[0].isdigit():
            street = "0 " + situs_text
        else:
            street = "No Address"
    row["Parcel ID"] = best.pid or ""
    _set_row_acres_from_candidate(row, best)
    row["Property Address"] = street
    row["Property City"] = city
    row["Property State"] = "NC"
    row["Property Zip"] = zipc
    new_use = simplify_use_code(best.use_code, best.use_description, best.county) or ""
    if not new_use:
        if best.is_vacant_land:
            new_use = "Vacant Land"
        elif best.is_residential:
            new_use = "SFR"
    if new_use:
        row["Property use"] = new_use
    # Property Value must track the new main — the old value belonged to the
    # parcel we just swapped away from (e.g. a $720K residence we're leaving
    # for a cheap vacant lot). Set from the sibling when GIS gave us a value,
    # else clear so the stale figure doesn't linger (Step 1.7 refills next run;
    # a blank value is never re-dropped by drop_over_500k).
    row["Property Value"] = str(int(best.market_value)) if best.market_value else ""
    # Notes coherence: when the sibling-backfill (Step 0.7) already added
    # the new main parcel to Notes as a sibling, we now have it duplicated
    # (Main + Notes both show the same parcel). Strip the new main's PID
    # from Notes if present, then append a marker for the old main so
    # the user can still see which parcel got dropped and why.
    # Taylor 26E000853-350 Week 26: main was Jonathan Dr (heir-occupied),
    # swap moved to 404 S 6th St, but old Notes had 404 S 6th St as PLUS
    # 1 PARCEL while Jonathan Dr disappeared entirely. Both wrong.
    existing_notes = (row.get("Notes") or "").strip()
    new_main_pid = (best.pid or "").strip()
    if existing_notes and new_main_pid:
        existing_notes = _strip_parcel_from_notes(existing_notes, new_main_pid)
    row["Notes"] = (existing_notes + ("\n" if existing_notes else "") + marker).strip()
    tag_reason(row, reason_tag)
    print(f"  {log_label} {county}/{dec}: {old_pid} -> {best.pid} ({new_use or '?'})")
    return True


def _try_swap_to_under_cap_sibling(row: dict) -> bool:
    """For a row whose main parcel exceeds its buy-box value cap, look up the
    decedent's full estate and swap in a sibling that's under its own cap —
    preferring a vacant lot. Per Oren's rule, a multi-parcel estate with cheap
    vacant lots is a live mobile-home-on-land lead even when the residence is
    over-cap. Mutates `row` and returns True on swap; False when no sibling
    qualifies.

    Runs at Step 1.8, BEFORE the heir-occupied filter (Step 1.9), which then
    applies its own vacant-lot exemption to the swapped main.

    Surfaced Week 27: Shuford 26E000779-170 — Step 0 repicked the $720,900
    residence (1627 Cauble Dairy Rd) as main; over-cap dropped the whole
    estate, discarding three vacant lots ($308K/$193K/$185K) on Cauble Dairy
    Rd that are the actual lead.
    """
    from nc_gis_lookup import (
        lookup_properties, filter_for_lead_quality, _middle_match_strength,
    )
    dec = (row.get("Deceased Owner") or "").strip()
    county = (row.get("County") or "").strip()
    if not dec or not county or "IN THE MATTER" in dec.upper():
        return False
    try:
        cands = lookup_properties(dec, county, min_score=0.7)
    except Exception:
        return False
    kept_cands = filter_for_lead_quality(
        cands, beneficiaries_json=row.get("Beneficiaries", "") or "", decedent_name=dec,
    )
    current_pid = (row.get("Parcel ID") or "").strip()
    # A sibling must identify the decedent AT LEAST as confidently as the
    # over-cap parcel — otherwise the "swap" trades the decedent's real property
    # for a cheaper NAMESAKE just to duck the cap. Week 28: Jones Michael Gregory
    # 26E000800-170 — over-cap parcel "JONES MICHAEL GREGORY" (3946 Granite St,
    # $1.07M, full-middle match) got swapped to "JONES MICHAEL G" (1866 Fairway
    # Dr, $360K, initial only), a different Michael Jones. The $1M house is his
    # and is legitimately over the buy-box; the row should DROP, not swap.
    current = next((c for c in cands if (c.pid or "").strip() == current_pid), None)
    cur_strength = _middle_match_strength(current.owner_name, dec) if current else 0
    siblings = [c for c in kept_cands
                if (c.pid or "") and c.pid != current_pid
                and _middle_match_strength(c.owner_name, dec) >= cur_strength]
    if not siblings:
        return False

    def _is_vacant(c) -> bool:
        use = (c.use_description or c.use_code or "").upper()
        return "VACANT" in use or "LAND" in use or bool(c.is_vacant_land)

    def _under_cap(c) -> bool:
        is_vac = _is_vacant(c)
        cap = _cap_for_use("VACANT LAND" if is_vac
                           else (simplify_use_code(c.use_code, c.use_description, c.county) or ""),
                           _cand_acres(c))
        # Unknown value: keep vacant lots (assumed under the $1M vacant cap),
        # skip unknown-value non-vacant (can't confirm under the $500K cap).
        if c.market_value is None:
            return is_vac
        return float(c.market_value) <= cap

    def rank(c):
        # vacant first, then situs-present (better main address), then value ASC.
        # Value ascending (cheapest-first) because when a decedent owns several
        # under-cap parcels, the more affordable one is the better buy-box lead.
        # Marinakos, Peter 26E002614-590 Week 29: over-cap main (11440 Bloomfield,
        # the family home) had two under-cap siblings — 826 Ashmore ($254K, Oren's
        # actual lead) and 6215 Summerlin ($458K). Value-DESC picked Summerlin;
        # value-ASC picks Ashmore.
        has_situs = 1 if (c.situs_address or "").strip() else 0
        return (0 if _is_vacant(c) else 1, 0 if has_situs else 1, int(c.market_value or 0))

    eligible = [c for c in siblings if _under_cap(c)]
    if not eligible:
        return False
    eligible.sort(key=rank)
    best = eligible[0]
    return _apply_sibling_swap(
        row, best, current_pid,
        marker=f"[SWAPPED-ON-OVER-CAP from {current_pid}: prior main over buy-box cap]",
        reason_tag="swap-on-over-cap", log_label="SWAP-ON-CAP",
    )


def _strip_parcel_from_notes(notes: str, pid_to_strip: str) -> str:
    """Remove a vertical PLUS-N-PARCELS block that references the given
    PID — used after swap-on-DQ to deduplicate the new main from Notes.

    Recognizes both vertical format ("PLUS N PARCELS" header followed by
    address / meta / "  PID xxx" trios separated by blank lines) and the
    legacy horizontal format. Best-effort — leaves Notes unchanged when
    structure doesn't match the expected shape.
    """
    if not notes or not pid_to_strip:
        return notes
    lines = notes.split("\n")
    out: list[str] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        # Look ahead 5 lines for the PID marker to confirm a block to skip
        ahead = "\n".join(lines[i:i + 6])
        if pid_to_strip in ahead and (line.strip().startswith("PLUS ") or
                                       (line.strip() and not line[0].isspace())):
            # Skip lines until we hit a blank-line gap OR end OR a non-related block
            j = i
            block_end = len(lines)
            in_target_block = False
            while j < len(lines):
                if pid_to_strip in lines[j]:
                    in_target_block = True
                if in_target_block and (j > i and lines[j].strip() == ""):
                    block_end = j + 1
                    break
                j += 1
            if in_target_block:
                i = block_end
                continue
        out.append(line)
        i += 1
    # Also collapse the now-orphan "PLUS 1 PARCEL" header if its only item was stripped
    cleaned = "\n".join(out).strip()
    return cleaned


def _parcel_use_tier(use: str) -> int:
    """Same residential-first tiering used by nc_ftm_writer.collapse_by_case."""
    u = (use or "").upper()
    if "COMMERCIAL" in u or "INDUSTRIAL" in u or "OFFICE" in u:
        return 0
    if "VACANT" in u or "LAND" in u:
        return 1
    if u in {"SFR", "RESIDENTIAL", "TOWNHOUSE", "CONDO", "MH",
             "MULTI-FAMILY", "DUPLEX"}:
        return 3
    # Unknown use (blank/uncategorized) is no better than vacant — without
    # this guard, Step 1.5 re_collapse_multi_parcel swapped vacant-land
    # leads to blank-use parcels because blank ranked higher (was tier 2).
    # Bostian 26E000662-120 Week 26: my CO mapping revert left both Cabarrus
    # parcels at use="". The vacant-land $53K parcel (correct lead) got
    # swapped to the $316K residence at 4699 Rainbow Dr (heir-occupied,
    # then dropped). Tying blank to tier 1 keeps vacant from being
    # outranked by uncategorized targets.
    return 1


def re_collapse_multi_parcel(rows: list[dict]) -> int:
    """For rows whose Notes contain 'PLUS N PARCELS' (multi-parcel decedents
    where the original scrape collapsed by market_value and may have
    picked a vacant lot or commercial parcel as main), re-fetch the
    decedent's parcels from GIS and pick the residential one as main.
    Updates the row in place and rewrites the PLUS-N-PARCELS note with
    the remaining parcels.

    Returns the number of rows whose main parcel was swapped.
    """
    from nc_gis_lookup import filter_for_lead_quality
    swapped = 0
    for r in rows:
        notes = (r.get("Notes") or "").strip()
        if "PLUS " not in notes.upper() or "PARCEL" not in notes.upper():
            continue
        current_use = (r.get("Property use") or "").strip()
        # Skip if main is already residential — no swap needed
        if _parcel_use_tier(current_use) == 3:
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or not county or "IN THE MATTER" in dec.upper():
            continue
        try:
            results = lookup_properties(dec, county, min_score=0.7)
        except Exception:
            continue
        kept = filter_for_lead_quality(results)
        if len(kept) < 2:
            continue
        # Apply residential-first priority
        def sort_key(c):
            new_use = simplify_use_code(c.use_code, c.use_description, c.county) or ""
            return (_parcel_use_tier(new_use),
                    float(c.market_value or 0))
        sorted_kept = sorted(kept, key=sort_key, reverse=True)
        new_main = sorted_kept[0]
        new_use = simplify_use_code(new_main.use_code, new_main.use_description, new_main.county) or ""
        # Only swap if the new main is a higher tier than current
        if _parcel_use_tier(new_use) <= _parcel_use_tier(current_use):
            continue
        # Apply the new main's data
        street, city, zipc = _candidate_to_address_parts(new_main)
        r["Parcel ID"] = new_main.pid or ""
        _set_row_acres_from_candidate(r, new_main)
        r["Property Address"] = street
        r["Property City"] = city
        r["Property State"] = "NC"
        r["Property Zip"] = zipc
        r["Property use"] = new_use
        # Refresh Property Value from the new main parcel — without this
        # the stale value from the swapped-out parcel survives because
        # Step 1.7 only fires on blank values. Walton 26E002339-590
        # surfaced this Week 26: scraper picked $7,900 vacant strip as
        # main, polish swapped to $260K house at 7918 Corder Dr but
        # value stayed $7,900. Always overwrite when we have a value
        # from the new main; otherwise blank so Step 1.7 re-derives.
        if new_main.market_value:
            r["Property Value"] = f"{int(round(float(new_main.market_value))):,}"
        else:
            r["Property Value"] = ""
        # Rebuild the PLUS-N-PARCELS note (vertical format) with the
        # remaining parcels
        extras = sorted_kept[1:]
        if extras:
            from nc_ftm_writer import format_extra_parcels_vertical
            items: list[dict] = []
            for e in extras:
                es, ec, ez = _candidate_to_address_parts(e)
                addr = " ".join(filter(None, [es, ec, ez])).strip()
                eu = simplify_use_code(e.use_code, e.use_description, e.county) or ""
                items.append({
                    "address": addr,
                    "use": eu,
                    "lot": e.lot_area,
                    "value": e.market_value,
                    "pid": e.pid,
                })
            r["Notes"] = format_extra_parcels_vertical(items)
        else:
            r["Notes"] = ""
        print(f"  Re-collapsed {county}/{dec}: {current_use!r} -> {new_use!r}, "
              f"main now {street}, {city}")
        swapped += 1
    return swapped


def drop_commercial(rows: list[dict]) -> tuple[list[dict], int]:
    """Drop rows whose Property use is genuinely Commercial/Industrial/Office.

    Run AFTER repair_addresses (which re-fetches suspect Cabarrus 'I'
    codes back to SFR). What's left tagged Commercial after that is the
    real commercial — drop it from probate lead lists.
    """
    kept = [r for r in rows if (r.get("Property use") or "").strip().upper() not in _SUSPECT_USES]
    dropped = len(rows) - len(kept)
    return kept, dropped


_DROP_CONDO_TOWNHOUSE_USES = {"CONDO", "TOWNHOUSE", "CONDOMINIUM", "TOWN HOUSE"}


def refine_lincoln_structure_type(rows: list[dict]) -> int:
    """Upgrade Lincoln rows' Property use from the real dwelling style.

    Lincoln's parcel query only exposes ZONING, so condos/townhouses read as
    SFR. The Property Record Card's backing table has the true style (AHDESC).
    For each Lincoln row with a Parcel ID, look it up; if it's a Condo or
    Townhouse, set Property use so Step 2.1 drops it. Only overrides TO
    Condo/Townhouse — never downgrades an existing classification.
    """
    from nc_gis_lookup import lincoln_structure_type
    marked = 0
    for r in rows:
        if (r.get("County") or "").strip().lower() != "lincoln":
            continue
        pid = (r.get("Parcel ID") or "").strip()
        if not pid:
            continue
        cur = (r.get("Property use") or "").strip()
        if cur in ("Condo", "Townhouse"):
            continue
        bucket = lincoln_structure_type(pid)
        if bucket in ("Condo", "Townhouse"):
            r["Property use"] = bucket
            tag_reason(r, "lincoln-ahdesc-structure")
            marked += 1
            print(f"  LINCOLN {bucket.upper()} {r.get('Deceased Owner')}: "
                  f"{r.get('Property Address')!r} (PIN {pid}) -> {bucket}, will drop")
    return marked


def drop_condos_and_townhouses(rows: list[dict]) -> tuple[list[dict], int]:
    """Drop rows whose Property use is Condo or Townhouse — per Oren's
    investor buy-box, these don't pencil out (HOA constraints, thin margins).
    """
    kept: list[dict] = []
    dropped = 0
    for r in rows:
        use = (r.get("Property use") or "").strip().upper()
        if use in _DROP_CONDO_TOWNHOUSE_USES:
            dropped += 1
            continue
        kept.append(r)
    return kept, dropped


def collapse_duplicate_decedents(rows: list[dict]) -> tuple[list[dict], int]:
    """For rows with blank Case No. that share (County, Deceased Owner),
    collapse to one row + a 'PLUS N PARCELS' note. Mirrors the case-based
    collapse done at scrape time but works on the FTM CSV by-decedent
    when case_no never got populated.
    """
    from collections import defaultdict
    groups: dict[tuple[str, str], list[dict]] = defaultdict(list)
    keep_individually: list[dict] = []
    for r in rows:
        if (r.get("Case No.") or "").strip():
            keep_individually.append(r)  # has case# — already collapsed by scrape
            continue
        key = (r.get("County", ""), (r.get("Deceased Owner") or "").strip().upper())
        if not key[1]:
            keep_individually.append(r)
            continue
        groups[key].append(r)
    collapsed = 0
    for key, items in groups.items():
        if len(items) == 1:
            keep_individually.append(items[0])
            continue
        # Pick main = first (already sorted by scrape order)
        main = items[0]
        extras = items[1:]
        # Build a vertical-format PLUS-N-PARCELS note appended to existing Notes
        from nc_ftm_writer import format_extra_parcels_vertical
        items: list[dict] = []
        for e in extras:
            addr = " ".join(filter(None, [e.get("Property Address", ""),
                                          e.get("Property City", ""),
                                          e.get("Property Zip", "")])).strip()
            items.append({
                "address": addr,
                "use": (e.get("Property use") or "").strip(),
                "pid": (e.get("Parcel ID") or "").strip(),
                # collapse_duplicate_decedents works at the polish-row layer
                # and doesn't carry market_value -- left blank
            })
        extra_note = format_extra_parcels_vertical(items)
        existing_notes = (main.get("Notes") or "").strip()
        if existing_notes:
            main["Notes"] = existing_notes + "\n\n" + extra_note
        else:
            main["Notes"] = extra_note
        keep_individually.append(main)
        collapsed += len(extras)
    return keep_individually, collapsed


# Beneficiary name tokens that mean "not a person" — skip these as PR fallback.
_NON_PERSON_TOKENS = (
    "TRUST", "LLC", "CORP", "ESTATE OF", "INC", "FOUNDATION",
    "TRUSTEE", "FAMILY TRUST", "REVOCABLE", "IRREVOCABLE",
    "MINISTRIES", "CHURCH", "BAPTIST", "METHODIST", "CATHOLIC",
)


def _is_person_name(name: str) -> bool:
    """True when the beneficiary name looks like an individual (not a trust/LLC/etc)."""
    if not name:
        return False
    up = name.upper()
    if any(tok in up for tok in _NON_PERSON_TOKENS):
        return False
    return True


def _parse_beneficiary_line(line: str) -> dict | None:
    """Parse one Beneficiaries-cell line in our format:
        "Last, First Middle - street, city, NC zip"
    Returns dict with name/first/last/street/city/state/zip, or None on failure.
    """
    if " - " not in line:
        return None
    name_part, _, addr_part = line.partition(" - ")
    name_part = name_part.strip()
    addr_part = addr_part.strip()
    if not name_part or not _is_person_name(name_part):
        return None
    # Parse name: "Last, First Middle" preferred; fallback to space-separated
    first = middle = last = ""
    if "," in name_part:
        last_chunk, _, fm_chunk = name_part.partition(",")
        last = last_chunk.strip()
        fm_tokens = fm_chunk.strip().split()
        if fm_tokens:
            first = fm_tokens[0]
            middle = " ".join(fm_tokens[1:])
    else:
        tokens = name_part.split()
        if len(tokens) >= 2:
            first, last = tokens[0], tokens[-1]
            middle = " ".join(tokens[1:-1])
    if not first or not last:
        return None
    # Parse address: "street, city, NC zip"
    parts = [p.strip() for p in addr_part.split(",") if p.strip()]
    street = city = state = zipc = ""
    if len(parts) >= 1:
        street = parts[0]
    if len(parts) >= 2:
        city = parts[1]
    if len(parts) >= 3:
        # Last element: "NC 12345" or "NC 12345-6789"
        tail = parts[-1].strip().split()
        if tail:
            if len(tail[0]) == 2 and tail[0].isalpha():
                state = tail[0].upper()
            for t in tail:
                if t and (t[0].isdigit() or (len(t) >= 5 and t[:5].isdigit())):
                    zipc = t
                    break
    # Need at least street + (city OR zip) to be a usable mailing
    if not street or not (city or zipc):
        return None
    return {
        "name": name_part,
        "first": first,
        "middle": middle,
        "last": last,
        "street": street,
        "city": city,
        "state": state or "NC",
        "zip": zipc,
    }


def promote_beneficiary_to_pr(row: dict) -> dict | None:
    """For rows where no court-named executor exists, look at the Beneficiaries
    column for a person with a usable mailing address. If found, return the
    parsed beneficiary dict — the caller wires it into the Executor / Mailing
    fields. Returns None when no suitable beneficiary exists (caller falls
    back to generic 'Heirs of [Decedent]' treatment).
    """
    bens_raw = (row.get("Beneficiaries") or "").strip()
    if not bens_raw:
        return None
    # The CSV stores beneficiaries as newline-separated; the XLSX collapses
    # to ' | '. Handle both delimiters.
    if "\n" in bens_raw:
        lines = [ln.strip() for ln in bens_raw.split("\n") if ln.strip()]
    else:
        lines = [ln.strip() for ln in bens_raw.split(" | ") if ln.strip()]
    for ln in lines:
        parsed = _parse_beneficiary_line(ln)
        if parsed:
            return parsed
    return None


def _split_person_name(full: str) -> tuple[str, str]:
    """Split a 'First Middle Last' string into (first, last). Best-effort."""
    parts = (full or "").strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[-1]


def promote_dm_to_pr(row: dict) -> dict | None:
    """If DM Name is populated (obituary enricher found a verified-living
    heir) and the row has no court PR, promote the DM to PR contact.

    Prefers the DM's mailing address if available; otherwise falls back to
    property address so direct mail still reaches the lead.

    Returns dict with promoted fields or None if no usable DM.
    """
    dm_name = (row.get("DM Name") or "").strip()
    if not dm_name:
        return None
    # Skip degenerate placeholders — let the row fall through to the
    # regular "Heirs of [Decedent]" path instead:
    #   - "Heirs of ..." (already a generic placeholder)
    #   - "Estate of ..." (obituary enricher's Path 5 estate-fallback —
    #     fires when no survivors are named in the obituary)
    #   - Pure relationship words like "Wife"/"Husband" (parser pulled
    #     the relation label instead of the actual person's name)
    dm_lower = dm_name.lower()
    # "estate of" / "heirs of" ANYWHERE, not just at the start: the enricher's
    # estate-fallback inherits the decedent's "Last, First" name order, so it
    # arrives comma-inverted as "Farley, Estate of Robert Matthew"
    # (26E000827-170). A leading-only check let that promote to PR and split
    # into First="Estate of Robert" — a mailer to no one. No real person's name
    # contains either phrase, so substring matching is safe.
    if "estate of" in dm_lower or "heirs of" in dm_lower:
        return None
    if dm_lower in {"wife", "husband", "spouse", "son", "daughter",
                    "brother", "sister", "mother", "father", "child",
                    "children", "family"}:
        return None
    first, last = _split_person_name(dm_name)
    return {
        "name": dm_name,
        "first": first,
        "last": last,
        "relationship": (row.get("DM Relationship") or "").strip(),
    }


def apply_manual_corrections(
    rows: list[dict], path: str = "manual_corrections.csv",
) -> tuple[int, int]:
    """Apply durable per-record field overrides from manual_corrections.csv
    (columns: Case No., Field, Value). Runs LAST — right before write — so a
    hand-verified value always wins over anything the pipeline derived, and
    survives the nightly rebuild. Mirrors manual_drops.txt, but corrects fields
    instead of dropping cases. Use for one-offs the pipeline can't get right:
    Stegall 26E002606-590 (Oren's phone beats the skip-trace number), Sisk
    26E000688-480 (executor mailing the App PDF didn't carry).

    Returns (fields_applied, cases_touched).
    """
    p = Path(path)
    if not p.exists():
        return (0, 0)
    corr: dict[str, dict[str, str]] = {}
    with p.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            case = (row.get("Case No.") or "").strip()
            field = (row.get("Field") or "").strip()
            if not case or not field or case.startswith("#"):
                continue
            corr.setdefault(case, {})[field] = (row.get("Value") or "").strip()
    if not corr:
        return (0, 0)
    fields_applied = 0
    cases_touched: set[str] = set()
    for r in rows:
        c = (r.get("Case No.") or "").strip()
        overrides = corr.get(c)
        if not overrides:
            continue
        for field, value in overrides.items():
            if field in r:
                r[field] = value
                fields_applied += 1
                cases_touched.add(c)
            else:
                print(f"  MANUAL-CORRECTION warn: {c} unknown field {field!r} — skipped")
    return (fields_applied, len(cases_touched))


def _clean_person_name(name: str | None) -> str:
    """Title-case an OCR'd applicant name and repair the common suffix misread
    'IR' -> 'Jr' (e.g. 'BILL J. BAITY IR'). Returns '' for blank input."""
    s = (name or "").strip()
    if not s:
        return ""
    s = re.sub(r"\bIR\b\.?$", "Jr", s)  # OCR: JR often reads as IR
    out = []
    for tok in s.split():
        t = tok.strip(".")
        up = t.upper()
        if up in {"JR", "SR", "II", "III", "IV"}:
            out.append({"JR": "Jr", "SR": "Sr"}.get(up, up))
        elif len(t) == 1:
            out.append(t.upper() + ".")   # middle initial
        else:
            out.append(t.capitalize())
    return " ".join(out)


def _split_app_pr_name(app_pr: str) -> tuple[str, str, str]:
    """Split a cleaned applicant name into (full, first, last) for the search-
    friendly columns, dropping a single-initial middle token so the Last field
    stays clean. 'Bill J. Baity Jr' -> ('Bill Baity Jr', 'Bill', 'Baity Jr')."""
    parts = app_pr.split()
    first = parts[0]
    last_parts = [p for p in parts[1:] if len(p.rstrip(".")) > 1] or [parts[-1]]
    last = " ".join(last_parts)
    return (f"{first} {last}", first, last)


def second_pass_obituary_for_heirs_of(rows: list[dict]) -> tuple[int, int, int]:
    """For rows that fell through to "Heirs of [Decedent]" (no PR found,
    no usable beneficiary), do an aggressive second-pass obituary search.
    If we find a survivor we can name AND find their current mailing
    address, promote them to the PR slot — turning a dead-end
    heirs-of-fallback into a real person we can mail and skip-trace.

    Why this matters: DataSift skip-trace operates on real-person
    first/last names + mailing addresses. "Heirs of John Smith" at the
    property address cannot be skip-traced and burns an upload slot.

    Differences from the first-pass obituary enricher (which ran at
    scrape time): broader keyword variations, multiple URL candidates
    per row, NC state forced, ignores DOD recency gate.

    Returns (full_hits, name_only_hits, attempted) — full = name + address,
    name_only = heir name found but no address.
    """
    try:
        import config as cfg
        from obituary_enricher import (
            _search_obituary, _fetch_cached_text,
            _parse_obituary_with_llm, identify_decision_maker,
            _lookup_dm_address, _lookup_dm_address_tracerfy,
        )
        import tracerfy_budget
    except Exception as e:  # pragma: no cover - import-time guards only
        print(f"  Second-pass obituary unavailable ({e}) — skipping")
        return (0, 0, 0)

    api_key = getattr(cfg, "ANTHROPIC_API_KEY", "")
    tracerfy_ok = bool(getattr(cfg, "TRACERFY_API_KEY", ""))
    full_hits = name_only_hits = attempted = 0

    for r in rows:
        if r.get("First Name") != "Heirs":
            continue
        decedent = (r.get("Deceased Owner") or "").strip()
        if not decedent or "IN THE MATTER" in decedent.upper():
            continue

        # Court-named applicant (Application PDF) beats an obituary heir. When
        # the App PDF gave us a real PR name, promote it and skip the obit — the
        # court record outranks a scraped obituary, which otherwise fills the
        # slot with whatever survivor it finds (often the heir living in the
        # house). Sisk, Nathan 26E000688-480 Week 29: App PDF "BILL J. BAITY JR"
        # but the obit named the heir-occupant wife "Dian Sisk". (The exact
        # applicant mailing, when it isn't on the row, comes from a later App-
        # mailing pass or manual_corrections.csv.)
        app_pr = _clean_person_name(r.get("PR Full Name (App)"))
        if (app_pr and len(app_pr.split()) >= 2 and not _is_entity_decedent(app_pr)
                and not app_pr.lower().startswith(("estate of", "heirs of"))):
            full, first, last = _split_app_pr_name(app_pr)
            r["Personal Representative"] = full
            r["First Name"] = first
            r["Last Name"] = last
            r["DM Name"] = full
            tag_reason(r, "pr-from-app-over-obit")
            print(f"  APP-PR-OVER-OBIT {r.get('County')}/{decedent}: "
                  f"{full} (Application PDF) — skipping obit")
            continue

        attempted += 1
        property_city = (r.get("Property City") or "").strip()

        # Broader search than first-pass: multiple keyword variants.
        seen_urls: set[str] = set()
        obit_results: list[dict] = []
        for terms in ("obituary", '"death notice"', '"passed away"', "funeral memorial"):
            try:
                rs = _search_obituary(decedent, property_city, extra_terms=terms, state="NC")
            except Exception:
                continue
            for o in rs:
                u = (o.get("url") or "").strip()
                if not u or u in seen_urls:
                    continue
                seen_urls.add(u)
                obit_results.append(o)
        if not obit_results:
            continue

        # Try up to 6 URLs, stop on first that yields survivors.
        parsed = None
        for obit in obit_results[:6]:
            url = obit["url"]
            try:
                text = _fetch_cached_text(url)
            except Exception:
                continue
            if not text:
                continue
            try:
                cand = _parse_obituary_with_llm(
                    obituary_text=text,
                    owner_name=decedent,
                    city=property_city,
                    address=r.get("Property Address", ""),
                    api_key=api_key,
                    state="NC",
                )
            except Exception:
                continue
            if cand and cand.get("survivors"):
                parsed = cand
                parsed["_url"] = url
                break

        if not parsed:
            continue

        dm_name, dm_rel = identify_decision_maker(parsed["survivors"])
        if not dm_name:
            continue

        # Tier 1: free people-search waterfall (Serper + Firecrawl + LLM).
        addr = None
        try:
            addr = _lookup_dm_address(dm_name, property_city, api_key, state="NC")
        except Exception:
            addr = None
        # Tier 2: Tracerfy paid skip-trace, only if Tier 1 whiffed AND we
        # have budget headroom. 5 credits ($0.10) per hit, 0 on miss.
        # Tracerfy needs an address anchor — use the property address.
        if (not addr or not addr.get("street")) and tracerfy_ok and tracerfy_budget.can_spend():
            try:
                addr = _lookup_dm_address_tracerfy(
                    dm_name,
                    property_city,
                    address=r.get("Property Address", "") or "",
                    zip_code=r.get("Property Zip", "") or "",
                    state="NC",
                )
            except Exception:
                addr = None
            if addr and addr.get("street"):
                warn_now, msg = tracerfy_budget.record_hit()
                if warn_now:
                    print(f"  {msg}")
                tag_reason(r, "tracerfy")

        tokens = dm_name.split()
        first = tokens[0] if tokens else "Heir"
        last = tokens[-1] if len(tokens) > 1 else ""

        if addr and addr.get("street"):
            r["Personal Representative"] = dm_name
            r["First Name"] = first
            r["Last Name"] = last
            r["DM Name"] = dm_name
            r["DM Relationship"] = dm_rel or "heir"
            r["Mailing Address"] = addr["street"]
            r["Mailing City"] = addr.get("city") or property_city
            r["Mailing State"] = addr.get("state") or "NC"
            r["Mailing Zip"] = addr.get("zip") or ""
            existing_notes = (r.get("Notes") or "").strip()
            tag = f"[2ND-PASS-OBIT from {parsed.get('_url','')}]"
            r["Notes"] = (existing_notes + ("\n" if existing_notes else "") + tag).strip()
            tag_reason(r, "second-pass-obit-full")
            full_hits += 1
            print(f"  2ND-PASS-OBIT FULL {r.get('County')}/{decedent}: heir={dm_name} ({dm_rel}) at {addr['street']}")
        else:
            # Name-only: keep property as mailing (already set by heirs-of-fallback)
            # but at least replace the "Heirs of X" with a real name. DataSift
            # skip-trace post-upload may still surface phone/email for this person.
            r["Personal Representative"] = dm_name
            r["First Name"] = first
            r["Last Name"] = last
            r["DM Name"] = dm_name
            r["DM Relationship"] = dm_rel or "heir"
            tag_reason(r, "second-pass-obit-name-only")
            name_only_hits += 1
            print(f"  2ND-PASS-OBIT NAME-ONLY {r.get('County')}/{decedent}: heir={dm_name} ({dm_rel})")

    return (full_hits, name_only_hits, attempted)


def populate_zillow_urls(rows: list[dict]) -> int:
    """Build a Zillow search URL for each row's Property Address and
    write it to the 'Zillow URL' column. Skips rows with no usable
    address (blank, or vacant lots prefixed with "0 ").

    User clicks the URL from the workbook (or from DataSift after
    upload — the column rides along) before calling/texting/mailing
    to confirm the property isn't already listed / under contract.
    Catches the MLS-only cases the recently-sold filter (Step 1.85)
    can't see in county GIS.
    """
    from nc_ftm_writer import build_zillow_url
    n = 0
    for r in rows:
        url = build_zillow_url(
            r.get("Property Address", ""),
            r.get("Property City", ""),
            r.get("Property State", "") or "NC",
            r.get("Property Zip", ""),
        )
        if url:
            r["Zillow URL"] = url
            n += 1
    return n


def prep_for_datasift(rows: list[dict]) -> tuple[list[dict], int, int, int, int]:
    """For rows with a parcel + no court-named executor:
      1. If an obituary-verified DM exists (DM Name populated by the
         enricher), promote the DM to PR contact — real person, court-
         independent confirmation.
      2. Otherwise, try to promote a usable BENEFICIARY (person, not
         trust/LLC, with a mailing address) — far better than a generic
         'Heirs of [Decedent]' mailer.
      3. If neither exists, fall back to 'Heirs of [Decedent]' with the
         property as mailing address.

    Returns (kept_rows, dropped_no_parcel, dm_promoted, beneficiary_promoted, generic_heirs).
    """
    kept = [r for r in rows if (r.get("Parcel ID") or "").strip()]
    dropped = len(rows) - len(kept)
    dm_promoted = 0
    promoted = 0
    heirs = 0
    for r in kept:
        if (r.get("Personal Representative") or "").strip():
            continue
        decedent = (r.get("Deceased Owner") or "").strip()
        if not decedent or "IN THE MATTER" in decedent.upper():
            continue

        # Try DM promotion first (obituary-verified living heir)
        dm = promote_dm_to_pr(r)
        if dm:
            r["Personal Representative"] = dm["name"]
            r["First Name"] = dm["first"]
            r["Last Name"] = dm["last"]
            # DM mailing address isn't preserved through the polish CSV
            # round-trip — fall back to property address so direct mail
            # still reaches the property (lead can be skip-traced in
            # DataSift post-upload).
            if (r.get("Property Address") or "").strip():
                r["Mailing Address"] = r["Property Address"]
            if (r.get("Property City") or "").strip():
                r["Mailing City"] = r["Property City"]
            r["Mailing State"] = "NC"
            if (r.get("Property Zip") or "").strip():
                r["Mailing Zip"] = r["Property Zip"]
            tag_reason(r, "dm-promoted-pr")
            dm_promoted += 1
            continue

        # Then beneficiary promotion
        ben = promote_beneficiary_to_pr(r)
        if ben:
            # Build a nice "Last, First Middle" display name preserving original casing
            r["Personal Representative"] = ben["name"]
            r["First Name"] = ben["first"]
            r["Last Name"] = ben["last"]
            r["Mailing Address"] = ben["street"]
            r["Mailing City"] = ben["city"]
            r["Mailing State"] = ben["state"]
            r["Mailing Zip"] = ben["zip"]
            tag_reason(r, "beneficiary-promoted-pr")
            promoted += 1
            continue

        # Fall back to generic Heirs-of with property as mailing
        _f, _m, last = split_decedent_name(decedent)
        r["Personal Representative"] = f"Heirs of {decedent}"
        r["First Name"] = "Heirs"
        r["Last Name"] = last.title() if last.isupper() else last
        if (r.get("Property Address") or "").strip():
            r["Mailing Address"] = r["Property Address"]
        if (r.get("Property City") or "").strip():
            r["Mailing City"] = r["Property City"]
        r["Mailing State"] = "NC"
        if (r.get("Property Zip") or "").strip():
            r["Mailing Zip"] = r["Property Zip"]
        tag_reason(r, "heirs-of-fallback")
        heirs += 1
    return kept, dropped, dm_promoted, promoted, heirs


# Low-confidence parcel review band (2B). The scrape/enrichment path accepts a
# match at >= 0.70. Below that, a blank parcel is normally dropped at Step 4.
# Instead, when a still-blank row has a DISTINCT front-runner candidate in the
# review band, keep it with the parcel filled + a loud "Low-Confidence Parcel"
# flag so Oren can eyeball it rather than silently losing the lead.
_LOWCONF_MIN = 0.55          # floor: below this is too weak to bother surfacing
_LOWCONF_MAX = 0.70          # ceiling: the confident accept threshold
_LOWCONF_SEPARATION = 0.15   # top must be this far ahead of the 2nd candidate

def flag_low_confidence_parcels(rows: list[dict]) -> int:
    """Keep-and-flag blank-parcel rows whose best GIS candidate sits just under
    the accept threshold AND is a clear front-runner (a distinct top score in
    [0.55, 0.70), at least 0.15 ahead of the runner-up). Fills the parcel from
    that candidate and marks the row 'Low-Confidence Parcel' (Tags + Notes +
    Match Reason) so it survives Step 4 as a review lead instead of vanishing.

    Deliberately does NOT surface flat ties (e.g. a decedent whose surname
    yields many equal 0.40 candidates — Owensby-class): with no front-runner
    there's no signal to pick the right parcel, so those stay blank and drop as
    before. Oren works those by hand."""
    from nc_gis_lookup import lookup_properties
    flagged = 0
    for r in rows:
        if (r.get("Parcel ID") or "").strip():
            continue
        dec = (r.get("Deceased Owner") or "").strip()
        county = (r.get("County") or "").strip()
        if not dec or not county or "IN THE MATTER" in dec.upper():
            continue
        try:
            cands = lookup_properties(dec, county, min_score=0.4)
        except Exception:
            continue
        if not cands:
            continue
        cands = sorted(cands, key=lambda c: (c.match_score or 0), reverse=True)
        top_score = cands[0].match_score or 0
        second = (cands[1].match_score or 0) if len(cands) > 1 else 0
        if not (_LOWCONF_MIN <= top_score < _LOWCONF_MAX):
            continue
        if (top_score - second) < _LOWCONF_SEPARATION:
            continue  # ambiguous — no clear front-runner, don't guess
        top = cands[0]
        street, city, zipc = _candidate_to_address_parts(top)
        if not street:
            situs = (top.situs_address or "").strip()
            street = "0 " + situs if (situs and not situs[0].isdigit()) else "No Address"
        use = simplify_use_code(top.use_code, top.use_description, top.county) or ""
        if not use:
            use = "Vacant Land" if top.is_vacant_land else ("SFR" if top.is_residential else "")

        # Buy-box guard: a low-confidence attach must clear the SAME filters a
        # real match does. This step runs at the very end (Step 3.7), long after
        # the over-cap drop (1.8) and heir-occupied drop (1.9) — so a parcel
        # attached here bypasses both. Leave the row blank-parcel (it then drops
        # at Step 4) when the candidate is over-cap or heir-occupied.
        # Potter, Sharon Davis 26E000718-120 Week 29: a 0.60 match to
        # "POTTER PAUL | POTTER SHARON E" ($530K SFR, mailing == property)
        # slipped into the workbook over both filters.
        eff_use = use or ("VACANT LAND" if getattr(top, "is_vacant_land", False) else "")
        if top.market_value is not None and float(top.market_value) > _cap_for_use(eff_use):
            print(f"  LOW-CONF-SKIP {county}/{dec}: pid={top.pid} "
                  f"${int(top.market_value):,} over buy-box cap — leaving blank")
            continue
        if "VACANT" not in eff_use.upper() and eff_use.upper() != "LAND":
            def _na(s: str | None) -> str:
                toks = (s or "").lower().split()
                toks = [_STREET_SUFFIX_NORMALIZE.get(t.rstrip("."), t.rstrip(".")) for t in toks]
                return "".join(c for c in " ".join(toks) if c.isalnum())
            mail_n, prop_n = _na(r.get("Mailing Address")), _na(street)
            if mail_n and prop_n and _heir_addr_match(prop_n, mail_n):
                print(f"  LOW-CONF-SKIP {county}/{dec}: pid={top.pid} "
                      f"heir-occupied (mailing == property) — leaving blank")
                continue

        r["Parcel ID"] = top.pid or ""
        _set_row_acres_from_candidate(r, top)
        r["Property Address"] = street
        r["Property City"] = city
        r["Property State"] = "NC"
        r["Property Zip"] = zipc
        if use:
            r["Property use"] = use
        if top.market_value:
            r["Property Value"] = str(int(top.market_value))
        pct = int(round(top_score * 100))
        owner = (top.owner_name or "").strip()
        tag_reason(r, f"low-confidence-parcel({pct}%)")
        tags = (r.get("Tags") or "").strip()
        if "Low-Confidence Parcel" not in tags:
            r["Tags"] = (tags + ", " if tags else "") + "Low-Confidence Parcel"
        note = f"[LOW-CONFIDENCE PARCEL {pct}% — VERIFY owner: {owner}]"
        existing = (r.get("Notes") or "").strip()
        r["Notes"] = (existing + ("\n" if existing else "") + note).strip()
        print(f"  LOW-CONF {county}/{dec}: pid={top.pid} score={top_score:.2f} owner={owner!r}")
        flagged += 1
    return flagged


def run(src_path: Path, tag: str, ts: str) -> None:
    print(f"\n=== {tag.upper()}: {src_path.name} ===")
    with src_path.open(newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    # Back-compat: pre-rename CSVs used "Executor Full Name". Normalize to
    # the current "Personal Representative" so downstream logic only sees
    # one column name. Idempotent — no-op when column is already renamed.
    for r in rows:
        if "Executor Full Name" in r and "Personal Representative" not in r:
            r["Personal Representative"] = r.pop("Executor Full Name")
    print(f"Loaded {len(rows)} rows")

    print("Step -1.5: apply late-arriving case-doc data (wills/applications fetched on prior runs)")
    n_doc_applied = apply_fetched_case_docs(rows)
    print(f"  Rows updated from fetched case-doc cache: {n_doc_applied}")

    print("Step -1: backfill blank Case No. from user's manual XLSX archive")
    n_archive_hit, n_archive_miss = backfill_from_manual_archive(rows)
    print(f"  Archive-backfilled: {n_archive_hit}  No-match (still blank-case): {n_archive_miss}")

    print("Step -0.93: backfill blank Personal Representative from Parties API (recognizes co-executors)")
    n_pr_backfill = backfill_pr_from_parties(rows)
    print(f"  PR backfilled from Parties: {n_pr_backfill}")

    print("Step -0.97: drop manually-excluded case numbers (manual_drops.txt)")
    rows, n_manual_drop, manual_samples = drop_manual_exclusions(rows)
    print(f"  Manually dropped: {n_manual_drop}  Remaining: {len(rows)}")
    for s in manual_samples[:10]:
        print(f"    - {s}")

    print("Step -0.95: backfill blank Case Status from Tyler OData (CaseSummariesSlim)")
    n_status = backfill_case_status_from_odata(rows)
    print(f"  Case Status backfilled: {n_status}")

    print("Step -0.9: tally Case Status (Disposed/Closed kept per Oren 2026-06-27 — heir-occupied filter handles dead leads)")
    rows, n_non_pending, status_histo = drop_non_pending(rows)
    histo_pretty = ", ".join(f"{k}={v}" for k, v in sorted(status_histo.items(), key=lambda x: -x[1]))
    print(f"  Status distribution: {histo_pretty}  Remaining: {len(rows)}")

    print("Step -0.85: drop trust / corporate-entity decedents (no probate lead)")
    rows, n_entity, entity_samples = drop_entity_decedents(rows)
    print(f"  Dropped entity decedents: {n_entity}  Remaining: {len(rows)}")
    for sample in entity_samples[:5]:
        print(f"    - {sample[:90]}")
    if len(entity_samples) > 5:
        print(f"    ... and {len(entity_samples) - 5} more")

    print("Step -0.8: drop archive duplicates (already in prior-week manual pipeline)")
    rows, n_dropped_archive = drop_archive_duplicates(rows)
    print(f"  Dropped archive-duplicate rows: {n_dropped_archive}  Remaining: {len(rows)}")

    print("Step -0.5: soft-dedup blank-case-no rows against same-decedent named rows in this week")
    rows, n_merged, n_still_blank = soft_dedup_blank_case_no(rows)
    print(f"  Merged dupes: {n_merged}  Remaining blank-case-no: {n_still_blank}")

    print("Step 0: validate existing parcel matches against new middle-name-aware matcher")
    n_blanked, audit_rejected_pids = validate_existing_matches(rows)
    print(f"  Blanked false-positive matches: {n_blanked}  "
          f"(rejected PIDs blacklisted from re-search: {len(audit_rejected_pids)})")

    print("Step 0.5: re-search for correct parcel where audit blanked a wrong match")
    n_refound = research_blank_parcels(rows, audit_rejected_pids=audit_rejected_pids)
    print(f"  Re-found correct parcels: {n_refound}")

    # Runs BEFORE the beneficiary fallback: the decedent's own address is a
    # far stronger signal than an heir's, and it rescues the deed-spells-the-
    # middle-name-differently class that name search can never match.
    print("Step 0.64: parcel-fallback lookup using the decedent's address (Odyssey Parties)")
    n_dec_recovered = parcel_fallback_from_decedent_address(rows)
    print(f"  Recovered via decedent-address GIS lookup: {n_dec_recovered}")

    print("Step 0.65: address-fallback lookup using Beneficiaries column")
    n_addr_recovered = address_fallback_from_beneficiaries(rows)
    print(f"  Recovered via beneficiary-address GIS lookup: {n_addr_recovered}")

    print("Step 0.66: Small Estate Disposed-recent — match parcel via Interested Person mailing")
    n_se_recovered = address_lookup_for_small_estate_disposed(rows)
    print(f"  Small Estate parcels recovered via Interested-Person address: {n_se_recovered}")

    print("Step 0.67: recover parcel from the Application's listed real-estate address (cross-county aware)")
    n_app_re = recover_parcel_from_app_realestate(rows)
    print(f"  Parcels recovered via Application real-estate address: {n_app_re}")

    print("Step 0.68: cross-check name-matched parcels against the decedent's own address")
    n_addr_swap, n_addr_flag = crosscheck_parcel_vs_decedent_address(rows)
    print(f"  Corrected to decedent's home: {n_addr_swap}  Flagged uncorroborated: {n_addr_flag}")

    print("Step 0.6: scan blank-parcel rows for heir-transfer candidates (embedded surname)")
    heir_transfer_rows = collect_heir_transfer_candidates(rows)
    if heir_transfer_rows:
        review_path = Path("output") / f"heir_transfer_review_{tag}_{ts}.xlsx"
        write_heir_transfer_review(heir_transfer_rows, review_path)
        print(f"  Wrote review-me: {review_path}  ({len(heir_transfer_rows)} decedents with "
              f"{sum(len(r['candidates']) for r in heir_transfer_rows)} candidates)")
    else:
        print(f"  No heir-transfer candidates found.")

    print("Step 0.69: reformat pre-2026-06-20 horizontal 'PLUS N PARCELS' Notes to vertical")
    n_reformatted = reformat_legacy_horizontal_notes(rows)
    print(f"  Rows with legacy Notes rewritten: {n_reformatted}")

    print("Step 0.7: backfill sibling parcels into Notes (multi-parcel estates)")
    n_siblings = backfill_sibling_parcels_to_notes(rows)
    print(f"  Rows updated with sibling-parcel notes: {n_siblings}")

    print("Step 0.75: tag Lot Cluster rows (2+ parcels on same street — high-value MH niche)")
    n_clusters = tag_lot_clusters(rows)
    print(f"  Rows tagged 'Lot Cluster': {n_clusters}")

    print("Step 0.8: write Multi-Parcel Estates review-me XLSX (decedents with 5+ parcels)")
    mpe_entries = collect_multi_parcel_estates(rows, threshold=5)
    if mpe_entries:
        mpe_path = Path("output") / f"multi_parcel_estates_{tag}_{ts}.xlsx"
        write_multi_parcel_estates_review(mpe_entries, mpe_path)
        print(f"  Wrote review-me: {mpe_path}  ({len(mpe_entries)} decedents with "
              f"{sum(len(e['parcels']) for e in mpe_entries)} parcels)")
    else:
        print(f"  No decedents with 5+ parcels to review.")

    print("Step 1: repair property addresses + re-classify suspect Commercial rows")
    n_repaired = repair_addresses(rows)
    print(f"  Repaired: {n_repaired}")

    print("Step 1.5: re-collapse multi-parcel decedents (prefer residential as main)")
    n_swapped = re_collapse_multi_parcel(rows)
    print(f"  Swapped vacant/commercial-main -> residential-main: {n_swapped}")

    print("Step 1.6: re-derive Property use from current GIS (catches stale classifications)")
    n_use_refreshed = refresh_property_use_from_gis(rows)
    print(f"  Property use refreshed: {n_use_refreshed}")

    print("Step 1.65: flag rows where deed has middle-initial only but decedent has full middle (verify-middle-initial)")
    n_flagged = flag_initial_only_middle_matches(rows)
    print(f"  Rows flagged for manual middle-initial verification: {n_flagged}")

    print("Step 1.66: flag bare first+last matches with a competing same-name parcel (verify-name-ambiguous)")
    n_flagged_amb = flag_no_middle_ambiguous_matches(rows)
    print(f"  Rows flagged for manual name-ambiguity verification: {n_flagged_amb}")

    print("Step 1.7: backfill Property Value from GIS where missing")
    n_priced = populate_property_values(rows)
    print(f"  Filled Property Value: {n_priced}")

    print("Step 1.8: drop properties over buy-box cap (SFR/MH/Residential $500K, Vacant Land $1M)")
    rows, n_over_500k = drop_over_500k(rows, cap=500_000)
    print(f"  Dropped over-cap: {n_over_500k}  Remaining: {len(rows)}")

    # BEFORE the dollar floor: a tiny vacant lot is scrap on size alone, and Oren
    # doesn't want it reaching LandPortal for a re-valuation (Long 26E000944-350:
    # 0.11ac, LandPortal lifted $1,010 -> $57,141 and it wrongly survived).
    print(f"Step 1.81: drop standalone vacant lots under {_MIN_VACANT_ACRES}ac "
          "(scrap size — multi-parcel estates exempt, structures exempt)")
    rows, n_tiny_lot = drop_tiny_vacant_lots(rows)
    print(f"  Dropped tiny-vacant-lot: {n_tiny_lot}  Remaining: {len(rows)}")

    print("Step 1.82: drop standalone parcels under $10K (scrap land — multi-parcel estates exempt)")
    rows, n_under_min = drop_under_min_value(rows)
    print(f"  Dropped under-min-value: {n_under_min}  Remaining: {len(rows)}")

    print("Step 1.85: drop properties sold within last 24 months for $50K+ (already-in-market filter)")
    rows, n_recently_sold = drop_recently_sold(rows, months=24, min_price=50_000)
    print(f"  Dropped recently-sold: {n_recently_sold}  Remaining: {len(rows)}")

    print("Step 1.87: mark Rowan condos (unit number in legal description) as Condo -> dropped at Step 2.1")
    n_condo_flagged = flag_rowan_possible_condos(rows)
    print(f"  Rowan condos marked (drop at 2.1): {n_condo_flagged}")

    print("Step 1.88: refine Lincoln structure type from Property Record Card (AHDESC) -> mark Condo/Townhouse")
    n_lincoln_struct = refine_lincoln_structure_type(rows)
    print(f"  Lincoln condos/townhouses marked (drop at 2.1): {n_lincoln_struct}")

    # Step order rationale: people search FIRST (fills legit PR mailing),
    # THEN heir-occupied drop (now mailing is populated for the check),
    # THEN property-as-mailing fallback last (the "no mailing found anywhere,
    # mail to the property" backstop — this intentionally creates mail==property
    # so it has to run AFTER the heir-occupancy drop to avoid false-dropping).
    print("Step 1.93: look up PR mailing via people search (Serper+Firecrawl, before property fallback)")
    n_ps_found, n_ps_tried = fill_pr_mailing_via_people_search(rows, state="NC")
    print(f"  PR addresses found via people search: {n_ps_found}/{n_ps_tried}")

    # Run the city/zip cleanup BEFORE heir-occupied check so mangled
    # Property Address strings (e.g. "319 Sherrill" with city="Av")
    # get repaired before address-equality comparison. Sellers Irma
    # Elizabeth 26E000406-540 Week 26: mangled address foiled the
    # heir-occupied DQ even though PR clearly lived at the property.
    print("Step 1.89: clean bad-city / bad-zip leftovers (before heir-occupied check)")
    n_clean_city, n_clean_zip = clean_bad_city_zip_in_place(rows)
    print(f"  Cleaned cities: {n_clean_city}  Reformatted/cleared zips: {n_clean_zip}")

    print("Step 1.9: drop heir-occupied (executor mailing == property address)")
    rows, n_heir_occupied = drop_executor_at_property(rows)
    print(f"  Dropped heir-occupied: {n_heir_occupied}  Remaining: {len(rows)}")

    # Step 1.92: re-apply the <$10K standalone floor. Step 1.9's swap-on-DQ can
    # demote a multi-parcel estate whose main house is heir-occupied down to its
    # only remaining sibling — and that sibling may be a sub-floor scrap lot the
    # first floor pass (Step 1.82) skipped because the row was still multi-parcel
    # then (Kilgo 26E002578-590 Week 29: $293K heir-occupied house swapped to a
    # $3,800 Creekwood Ct lot, leaked into the workbook). Rows that still carry a
    # "PLUS" sibling note stay exempt — that's the consecutive-vacant-lots play.
    # Step 1.91: re-apply the over-cap drop. Step 1.8 checked the cap BEFORE
    # the heir-occupied swap (Step 1.9), so a swap that installs a new, more
    # expensive main can push a row back over the buy-box cap unchecked.
    # James, David Ray 26E002599-590 Week 29: main 5126 Glenbrier (heir lives
    # there) DQ'd, swapped to 10108 Woodview ($518K) — over the $500K SFR cap
    # but it slipped through because the cap check had already run. Re-drop
    # here (which itself tries an under-cap sibling before dropping).
    print("Step 1.91: re-apply over-cap drop after heir-occupied swaps")
    rows, n_over_cap_postswap = drop_over_500k(rows, cap=500_000)
    print(f"  Dropped over-cap after swap: {n_over_cap_postswap}  Remaining: {len(rows)}")

    print("Step 1.92: re-apply <$10K standalone floor after heir-occupied swaps")
    rows, n_under_min_postswap = drop_under_min_value(rows)
    print(f"  Dropped sub-floor after swap: {n_under_min_postswap}  Remaining: {len(rows)}")

    print("Step 1.925: drop decedent life-estate parcels (auto-transfer to remaindermen, not a probate sale)")
    rows, n_life_estate = drop_life_estate_parcels(rows)
    print(f"  Dropped life-estate: {n_life_estate}  Remaining: {len(rows)}")

    # Step 1.927: for no-PR rows, name the deed co-owner (different surname) as
    # the decision maker + look up their current address. Off-switch NC_COOWNER_DM=0.
    if os.environ.get("NC_COOWNER_DM") == "0":
        print("Step 1.927: co-owner DM promotion  — skipped (NC_COOWNER_DM=0)")
    else:
        print("Step 1.927: name deed co-owner as DM for no-PR rows (+ current-address lookup)")
        n_co_dm, n_co_addr = promote_deed_coowner_to_dm(rows, state="NC")
        print(f"  Co-owner DMs named: {n_co_dm}  (current address found: {n_co_addr})")

    print("Step 1.95: fill missing PR mailing from property (so direct mail still goes out)")
    n_filled_pr = fill_missing_pr_mailing_from_property(rows)
    print(f"  PR mailing property fallback applied: {n_filled_pr}")

    print("Step 2: drop genuinely commercial rows")
    rows, n_commercial = drop_commercial(rows)
    print(f"  Dropped commercial: {n_commercial}  Remaining: {len(rows)}")

    print("Step 2.1: drop Condo / Townhouse rows (not investor-friendly per buy-box)")
    rows, n_condo_th = drop_condos_and_townhouses(rows)
    print(f"  Dropped condos/townhouses: {n_condo_th}  Remaining: {len(rows)}")

    print("Step 3: collapse duplicate decedent rows (blank Case No.)")
    rows, n_collapsed = collapse_duplicate_decedents(rows)
    print(f"  Collapsed: {n_collapsed} parcels into PLUS-N-PARCELS notes  Remaining: {len(rows)}")

    print("Step 3.5: clean bad-city / bad-zip leftovers (suffix/state/numeric noise)")
    n_clean_city, n_clean_zip = clean_bad_city_zip_in_place(rows)
    print(f"  Cleaned cities: {n_clean_city}  Reformatted/cleared zips: {n_clean_zip}")

    print("Step 3.6: fill blank Property City/Zip via parcel-centroid reverse-geocode "
          "(DataSift needs the zip)")
    n_centroid = fill_property_location_via_centroid(rows)
    print(f"  Filled via centroid geocode: {n_centroid}")

    print("Step 3.7: keep+flag blank-parcel rows with a distinct low-confidence match (review band)")
    n_lowconf = flag_low_confidence_parcels(rows)
    print(f"  Low-confidence parcels flagged (kept for review): {n_lowconf}")

    print("Step 4: filter to has-parcel; promote DM/beneficiary or apply 'Heirs of' fallback")
    kept, dropped, dm_promoted, promoted, heirs = prep_for_datasift(rows)
    print(f"  Rows in: {len(rows)}  Dropped (no parcel): {dropped}  "
          f"DM-promoted: {dm_promoted}  Beneficiary-promoted: {promoted}  "
          f"Generic Heirs-of: {heirs}  Out: {len(kept)}")

    # Step 4.05: the heir-occupancy DQ, re-run for PRs that only came into
    # existence at Step 4.
    #
    # Step 1.9 can only judge rows that HAVE a PR. A case where the court named
    # nobody has no name to people-search at 1.93 and no mailing to compare at
    # 1.9, so it sails through both; only at Step 4 does an obituary-derived DM
    # get promoted in — carrying the address the enricher found for them. When
    # that address IS the property, the row is heir-occupied and by Oren's rule
    # it's a dead lead, but the check that would say so has already run.
    #
    # VanDriesen 26E000738-120 Week 29: no court PR, DM "Gary VanDriesen" (son)
    # promoted at Step 4 with mailing 14 Search Dr == the property. Oren DQ'd it
    # by hand off BeenVerified; the evidence was already sitting in the row.
    #
    # Scoped to promoted rows only. drop_executor_at_property also performs
    # swap-on-DQ, and re-running it across rows that already cleared Step 1.9
    # re-swaps their parcels (seen on James 26E002599-590).
    print("Step 4.05: heir-occupancy re-check for DM/beneficiary-promoted PRs")
    _promoted_ids = {id(r) for r in kept
                     if any(t in (r.get("Match Reason") or "")
                            for t in ("dm-promoted-pr", "beneficiary-promoted-pr"))}
    if _promoted_ids:
        _subset = [r for r in kept if id(r) in _promoted_ids]
        _survivors, _n_dq = drop_executor_at_property(_subset)
        _survivor_ids = {id(r) for r in _survivors}
        kept = [r for r in kept
                if id(r) not in _promoted_ids or id(r) in _survivor_ids]
        print(f"  Promoted PRs checked: {len(_subset)}  "
              f"Dropped heir-occupied: {_n_dq}  Remaining: {len(kept)}")
    else:
        print("  Promoted PRs checked: 0  Dropped heir-occupied: 0  "
              f"Remaining: {len(kept)}")

    print("Step 4.5: second-pass aggressive obituary search for 'Heirs of' rows "
          "(replace with real heir name + address when found)")
    n_full, n_name, n_tried = second_pass_obituary_for_heirs_of(kept)
    print(f"  Tried: {n_tried}  Promoted with address: {n_full}  Promoted name-only: {n_name}")

    print("Step 4.6: split jammed mailing addresses (city/state/zip stuck in the street field)")
    n_split = split_jammed_mailing_address(kept)
    print(f"  Mailing addresses split into components: {n_split}")

    print("Step 4.7: finalize blank State (=NC) + City-from-ZIP (DataSift Incomplete guard)")
    n_final = finalize_state_and_city(kept)
    print(f"  State/City fields filled: {n_final}")

    print("Step 4.8: collapse same-property spouse pairs (one house, one row)")
    kept, n_pairs = collapse_same_property_pairs(kept)
    print(f"  Same-property rows merged: {n_pairs}  Out: {len(kept)}")

    print("Step 4.9: default blank Property Type to SFR for valued residential parcels")
    n_sfr = default_property_type_residential(kept)
    print(f"  Property Type defaulted to SFR: {n_sfr}")

    print("Step 4.92: surface SMALL ESTATE marker into Notes (kept but flagged)")
    n_se_marked = mark_small_estate_in_notes(kept)
    print(f"  Small-estate rows marked in Notes: {n_se_marked}")

    # LAST filter, on survivors only: a Firecrawl + LLM call per row is the
    # costliest per-row step, so spend it only on rows that would actually ship.
    # Catches the live-MLS class county GIS is blind to (active listing, under
    # contract, sold-but-not-yet-recorded). Disk-cached; fails toward keeping.
    print("Step 4.93: Zillow listing-status check (drop actively-listed/sold; note status)")
    kept, n_zillow = drop_zillow_listed_or_sold(kept)
    print(f"  Dropped via Zillow status: {n_zillow}  Remaining: {len(kept)}")

    print("Step 4.95: apply durable manual field corrections (manual_corrections.csv)")
    n_corr, n_cases = apply_manual_corrections(kept)
    print(f"  Manual field corrections applied: {n_corr} field(s) across {n_cases} case(s)")

    # Step 4.7 (populate_zillow_urls) removed 2026-07-11 per Oren — DataSift
    # provides the Zillow/listing link in the property record after upload, so
    # generating one per row here is redundant. The function is retained (unused)
    # in case the standalone workbook ever wants it again.

    out_csv = Path("output") / f"nc_estates_ftm_{ts}_{tag}_datasift.csv"
    out_xlsx = Path("output") / f"nc_estates_ftm_{ts}_{tag}_datasift.xlsx"
    write_csv(kept, out_csv)
    write_xlsx(kept, out_xlsx)
    print(f"  Wrote: {out_csv}")
    print(f"  Wrote: {out_xlsx}")

    # DataSift-native upload file: headers match DataSift's field names so the
    # upload wizard auto-maps every column (no manual dragging), and Tags / List
    # / Zillow are dropped (Oren tags by week in DataSift). The FTM csv+xlsx above
    # stay as-is for the review workbook.
    from nc_datasift_export import write_datasift_upload_csv
    _wk_m = re.search(r"week(\d+)", tag)
    _wk = int(_wk_m.group(1)) if _wk_m else None
    upload_csv = Path("output") / f"nc_estates_ftm_{ts}_{tag}_datasift_upload.csv"
    n_up = write_datasift_upload_csv(kept, upload_csv, week=_wk)
    print(f"  Wrote: {upload_csv}  ({n_up} rows, DataSift-native headers, "
          f"Tags=Courthouse Data + Week {_wk})")


def _parse_money(v) -> float:
    s = re.sub(r"[^\d.]", "", str(v or ""))
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def split_jammed_mailing_address(rows: list[dict]) -> int:
    """Split a full 'street city NC zip' jammed into Mailing Address into parts.

    Some Odyssey/eCourts party addresses arrive as one line — e.g.
    'Po Box 310 Stanley Nc 281640310' or '211 Todd St Belmont, NC 28012' —
    leaving Mailing City/ZIP blank so DataSift maps them to nothing. Only acts
    when Mailing City is blank AND the street ends in a NC + 5-digit-zip tail.
    """
    fixed = 0
    tail = re.compile(r"[,\s]+(NC|N\.?C\.?|NORTH CAROLINA)\s+(\d{5})(?:-?\d{0,4})?\s*$",
                      re.IGNORECASE)
    for r in rows:
        if (r.get("Mailing City") or "").strip():
            continue
        street = (r.get("Mailing Address") or "").strip()
        m = tail.search(street)
        if not m:
            continue
        zip5 = m.group(2)
        head = street[:m.start()].strip().rstrip(",").strip()
        city = _NC_ZIP_TO_CITY.get(zip5, "")
        if city and head.upper().endswith(city.upper()):
            head = head[:len(head) - len(city)].strip().rstrip(",").strip()
        elif not city:
            parts = head.rsplit(None, 1)
            if len(parts) == 2:
                head, city = parts[0], parts[1]
        if not head:
            continue
        r["Mailing Address"] = head
        r["Mailing State"] = "NC"
        r["Mailing Zip"] = zip5
        if city:
            r["Mailing City"] = city
        tag_reason(r, "mailing-addr-split")
        fixed += 1
    return fixed


def finalize_state_and_city(rows: list[dict]) -> int:
    """Fill blank Property/Mailing State (always NC here) + City-from-ZIP.

    DataSift flags records with a blank State/City as Incomplete. State is
    always NC for this pipeline; City is authoritative from the ZIP.
    """
    filled = 0
    for r in rows:
        if not (r.get("Property State") or "").strip() and (r.get("Property Address") or "").strip():
            r["Property State"] = "NC"; filled += 1
        if not (r.get("Mailing State") or "").strip() and (r.get("Mailing Address") or "").strip():
            r["Mailing State"] = "NC"; filled += 1
        pz = (r.get("Property Zip") or "").strip()[:5]
        if not (r.get("Property City") or "").strip() and pz in _NC_ZIP_TO_CITY:
            r["Property City"] = _NC_ZIP_TO_CITY[pz]; filled += 1
        mz = (r.get("Mailing Zip") or "").strip()[:5]
        if not (r.get("Mailing City") or "").strip() and mz in _NC_ZIP_TO_CITY:
            r["Mailing City"] = _NC_ZIP_TO_CITY[mz]; filled += 1
    return filled


def collapse_same_property_pairs(rows: list[dict]) -> tuple[list[dict], int]:
    """Merge rows for the SAME property + SAME PR into one (spouse pairs).

    Two probate cases on one house — e.g. Wesley & Carolyn Titterington at
    6899 Forrest Creek Dr, both with PR Dorothy Riggleman — upload as two rows
    that DataSift then collides by address (one silently overwrites the other,
    and the house gets mailed twice). Collapse to the highest-value row and note
    the other decedent(s). Requires same PR so unrelated estates never merge.
    """
    from collections import defaultdict
    groups: dict[tuple, list[dict]] = defaultdict(list)
    passthrough: list[dict] = []
    for r in rows:
        pid = (r.get("Parcel ID") or "").strip()
        addr = (r.get("Property Address") or "").strip().upper()
        pr = (r.get("Personal Representative") or "").strip().upper()
        if (pid or addr) and pr:
            groups[(pid or addr, pr)].append(r)
        else:
            passthrough.append(r)
    keep: list[dict] = []
    collapsed = 0
    for grp in groups.values():
        if len(grp) == 1:
            keep.append(grp[0]); continue
        grp.sort(key=lambda x: -_parse_money(x.get("Property Value")))
        main = grp[0]
        others = [g.get("Deceased Owner", "").strip() for g in grp[1:]
                  if g.get("Deceased Owner", "").strip()
                  and g.get("Deceased Owner", "").strip() != main.get("Deceased Owner", "").strip()]
        if others:
            note = (main.get("Notes") or "").strip()
            add = "Also deceased owner(s) at this property: " + "; ".join(others)
            main["Notes"] = (note + " | " + add) if note else add
        tag_reason(main, "collapsed-same-property")
        keep.append(main)
        collapsed += len(grp) - 1
    keep.extend(passthrough)
    return keep, collapsed


def default_property_type_residential(rows: list[dict]) -> int:
    """Default a blank Property use to SFR when the parcel has a real value.

    Cabarrus/Catawba county codes don't always classify, leaving Property use
    (-> DataSift 'Property Type') blank. An improved parcel with a market value
    that wasn't tagged Vacant/Condo/Commercial is a house — default to SFR so
    the custom field isn't empty. Rows with no value stay blank (can't confirm).
    """
    n = 0
    for r in rows:
        if (r.get("Property use") or "").strip():
            continue
        if _parse_money(r.get("Property Value")) > 0:
            r["Property use"] = "SFR"
            tag_reason(r, "default-sfr")
            n += 1
    return n


def main() -> None:
    """Process every week that has a merged file. Auto-picks the latest
    *_weekN_merged.csv per ISO week — week-agnostic so new weeks (22, 23,
    ...) are handled without code changes.
    """
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")

    # Discover every week with a merged file; keep the most recent file per
    # week. glob() sorted ascending puts older timestamps first, so the last
    # assignment per week wins (newest scrape).
    by_week: dict[int, Path] = {}
    for fp in sorted(Path("output").glob("nc_estates_ftm_*_week*_merged.csv")):
        m = re.search(r"_week(\d+)_merged\.csv$", fp.name)
        if not m:
            continue
        by_week[int(m.group(1))] = fp

    if not by_week:
        print("No *_weekN_merged.csv files in output/. Run prepare_weekly_input.py first.")
        return

    # Skip archived weeks (output/archive_week<N>_done/ exists)
    sys.path.insert(0, str(Path(__file__).parent))
    from iso_week_archive import get_archived_weeks
    archived = get_archived_weeks()
    if archived:
        print(f"Archived ISO weeks (skipping): {sorted(archived)}")

    for wk in sorted(by_week):
        if wk in archived:
            print(f"\n=== Week {wk}: archived, skipping {by_week[wk].name} ===")
            continue
        print(f"\n{'=' * 70}")
        print(f"=== Week {wk}: {by_week[wk].name} ===")
        print(f"{'=' * 70}")
        run(by_week[wk], f"week{wk}", ts)

    record_gis_outage_state(sorted(w for w in by_week if w not in archived))


# Where a run leaves word that a county GIS was down. auto_archive_weeks.py
# reads this and refuses to freeze a week whose data is known-incomplete;
# scripts/daily_report.py surfaces it.
GIS_OUTAGE_PATH = Path("output") / ".gis_outage_last_run.json"


def record_gis_outage_state(weeks: list[int]) -> None:
    """Record which county GIS servers went down during this polish.

    A downed county returns zero rows, so its rows lose their parcel and get
    dropped at Step 4. That drop is normally harmless -- the merged input still
    holds the row, so the next night's polish restores it once the county is
    back. It stops being harmless the moment the week is ARCHIVED, because then
    no further polish ever runs. So a run that hit an outage must leave a note,
    and archiving must honour it. Always written (empty counties on a clean
    run) so a stale note can never defer archiving forever.
    """
    import json as _json
    try:
        from nc_gis_lookup import downed_counties
        down = sorted(downed_counties())
        GIS_OUTAGE_PATH.parent.mkdir(parents=True, exist_ok=True)
        GIS_OUTAGE_PATH.write_text(_json.dumps({
            "ts": datetime.now().isoformat(timespec="seconds"),
            "counties": down,
            "weeks": weeks,
        }), encoding="utf-8")
        if down:
            print(f"\n{'*' * 70}")
            print(f"*** GIS OUTAGE this run: {', '.join(down)}")
            print("*** Rows for these counties are INCOMPLETE — they lost their")
            print("*** parcel to a dead server, not to 'owns nothing'. Weeks "
                  f"{weeks} will NOT be archived until a clean run.")
            print(f"{'*' * 70}")
    except Exception as e:  # never let bookkeeping break the pipeline
        print(f"  (could not record GIS outage state: {e})")


if __name__ == "__main__":
    main()
