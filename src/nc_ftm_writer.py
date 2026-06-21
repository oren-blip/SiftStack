"""NC Estates CSV writer in the user's manual FTM-style format.

Matches the column layout in `FTM_2026_NC Estates - Week N YYYY` files:

    File Date, County, Case No., Deceased Owner, First Name, Last Name,
    Mailing Address, Mailing City, Mailing State, Mailing Zip,
    Parcel ID, Property Address, Property City, Property State,
    Property Zip, Property use, Notes, Phone 1, Tags, List

- "First Name" / "Last Name" are the EXECUTOR's (Affiant/Applicant/etc).
- "Mailing *" are the EXECUTOR's mailing address.
- "Property *" + "Parcel ID" come from the county GIS lookup.
- "Notes" is freeform — beneficiary list block, extra parcel notes, etc.
- "Tags" = "NC Estates Week {N} {YYYY}" (ISO week of the file date).
- "List" = "PROBATE".

Records are grouped by case so the multi-parcel rows from one decedent
(e.g. 9 parcels for the Thrower estate) share the executor data.
"""

from __future__ import annotations

import csv
import json
import logging
from datetime import datetime
from pathlib import Path

from notice_parser import NoticeData

logger = logging.getLogger(__name__)


FTM_COLUMNS = [
    "File Date",
    "County",
    "Case No.",
    "Case Status",
    "Deceased Owner",
    "Personal Representative",
    "First Name",
    "Last Name",
    "Mailing Address",
    "Mailing City",
    "Mailing State",
    "Mailing Zip",
    "Parcel ID",
    "Property Address",
    "Property City",
    "Property State",
    "Property Zip",
    "Property use",
    "Property Value",
    "Notes",
    "Beneficiaries",
    "Phone 1",
    "Tags",
    "List",
    # Audit aid: how each row landed in its current state. Empty = parcel +
    # PR came directly from the scrape (the default happy path). Polish-step
    # tags are appended " | "-separated as the row mutates. Sort by this
    # column to triage low-confidence rows fast.
    "Match Reason",
    # Clickable Zillow search URL for the property — operator checks
    # listing status (active / under contract / sold) before calling/
    # texting/mailing. Built from Property Address+City+State+Zip via
    # build_zillow_url(). Catches the cases Step 1.85 recently-sold
    # filter can't (MLS-only listing data isn't in county GIS).
    "Zillow URL",
    # Heir / Decision Maker columns — populated for rows that lack a
    # court-named executor, via heir_prospect_no_executor.py (obituary
    # survivors + multi-tier skip trace). Stays blank for executor rows.
    "DM Name",
    "DM Relationship",
    "DM Phone",
    "DM Email",
    "DM 2 Name",
    "DM 2 Relationship",
    "DM 3 Name",
    "DM 3 Relationship",
]

# Columns that exist in the data but are hidden from the xlsx display.
# Case Status is a silent guardrail — every kept row should be "Pending"
# (polish drops Disposed/Closed). Hiding it removes visual noise but
# preserves the field for audit / debugging / future use.
HIDDEN_FROM_WORKBOOK = {"Case Status"}

# County values for the Sheets dropdown validation
NC_COUNTY_OPTIONS = [
    "Cabarrus", "Catawba", "Gaston", "Iredell", "Lincoln", "Mecklenburg", "Rowan",
]

# Per-county row tints — subtle pastel backgrounds so rows visually group
# by county when the sheet is sorted by County. Picked low-saturation
# shades that pair with the dark green header + don't clash.
NC_COUNTY_COLORS: dict[str, str] = {
    "Cabarrus":    "D6EAF8",  # light blue
    "Catawba":     "D5F5E3",  # light green
    "Gaston":      "FCF3CF",  # light yellow (close to FTM band)
    "Iredell":     "FADBD8",  # light pink
    "Lincoln":     "E8DAEF",  # light lavender
    "Mecklenburg": "FDEBD0",  # light peach
    "Rowan":       "D1F2EB",  # light teal
}
# Fallback for unknown counties (white = no fill)
NC_COUNTY_DEFAULT_COLOR = ""


def _format_date(yyyymmdd: str) -> str:
    """Convert 'YYYY-MM-DD' → 'M/D/YYYY' to match the FTM file."""
    if not yyyymmdd:
        return ""
    try:
        d = datetime.strptime(yyyymmdd, "%Y-%m-%d")
        return f"{d.month}/{d.day}/{d.year}"
    except ValueError:
        return yyyymmdd


def _format_decedent(name: str) -> str:
    """Convert 'First Middle Last' → 'Last, First Middle' (FTM convention).

    Already-comma-formatted names pass through. Trust/business names pass through.
    """
    if not name or "," in name or " trust" in name.lower() or " llc" in name.lower():
        return name
    tokens = name.strip().split()
    if len(tokens) < 2:
        return name
    last = tokens[-1]
    rest = " ".join(tokens[:-1])
    return f"{last}, {rest}"


def build_zillow_url(address: str, city: str = "", state: str = "NC", zip_code: str = "") -> str:
    """Build a Zillow search URL for the given property address.

    Zillow's stable URL pattern is `/homes/{slug}_rb/` which lands on
    a search page filtered to that address. The slug is the address
    components joined by hyphens, with spaces and most punctuation
    converted to hyphens, and runs of hyphens collapsed.

    Example: 830 Florence St NW + Concord + NC + 28027 ->
      https://www.zillow.com/homes/830-Florence-St-NW-Concord-NC-28027_rb/

    Returns empty string when address is blank — vacant lots with no
    house number ("0 Carriage Rd") aren't meaningfully searchable on
    Zillow so we don't generate a URL for them.
    """
    if not address or not address.strip():
        return ""
    # Vacant lots get "0 <street>" prefix in our convention; Zillow
    # can't resolve a search for "0 street" cleanly. Skip those.
    a = address.strip()
    if a.startswith("0 ") or a == "0":
        return ""
    bits = [a, city.strip(), (state or "NC").strip(), zip_code.strip()]
    bits = [b for b in bits if b]
    if not bits:
        return ""
    raw = " ".join(bits)
    # Slug: lowercase NOT required, but normalize whitespace + remove
    # most punct except '.' (e.g. "St." stays meaningful).
    import re
    slug = re.sub(r"[^\w\s.-]", "", raw)
    slug = re.sub(r"\s+", "-", slug.strip())
    slug = re.sub(r"-+", "-", slug)
    return f"https://www.zillow.com/homes/{slug}_rb/"


def _fmt_money(v) -> str:
    """Compact dollar formatter: $1,698,570 -> '$1.7M', $77,170 -> '$77K'."""
    if v in (None, "", 0, 0.0):
        return ""
    try:
        n = float(v)
    except (TypeError, ValueError):
        return ""
    if n <= 0:
        return ""
    if n >= 1_000_000:
        return f"${n/1_000_000:.1f}M"
    if n >= 1_000:
        return f"${n/1_000:.0f}K"
    return f"${int(n)}"


def format_extra_parcels_vertical(parcels: list[dict]) -> str:
    """Build a vertically-laid-out 'PLUS N PARCELS' note from a list of
    parcel dicts. Each parcel dict accepts these keys (all optional):
      address, value, use, lot, pid, tier, score, suffix

    Output shape (one blank line between parcels):

      PLUS 14 PARCELS (auto-ranked)

      830 Florence St NW Concord 28027
        $1.7M  Commercial  2.1ac
        PID 56210485720000

      1015 Central Dr NW Concord 28027
        $136K  SFR
        PID 56118667130000

    The vertical format reads much faster than the prior pipe-separated
    single-line format when scanning a multi-parcel estate for
    commercial-portfolio patterns (Watts Mitchell W class).
    """
    if not parcels:
        return ""
    suffix = ""
    # Allow caller to pass a header suffix like "(auto-ranked)"
    if parcels and isinstance(parcels[0], dict) and parcels[0].get("_header_suffix"):
        suffix = " " + parcels[0]["_header_suffix"]
    header = f"PLUS {len(parcels)} PARCEL{'S' if len(parcels) > 1 else ''}{suffix}"
    blocks: list[str] = [header]
    for p in parcels:
        addr = (p.get("address") or "").strip()
        if not addr:
            addr = "(no address)"
        lines = [addr]

        # Line 2: value + use + lot, space-separated; skip blanks
        meta_bits = []
        v = _fmt_money(p.get("value"))
        if v:
            meta_bits.append(v)
        use = (p.get("use") or "").strip()
        if use:
            meta_bits.append(use)
        lot = p.get("lot")
        if lot is not None:
            try:
                f = float(lot)
                if f > 0:
                    meta_bits.append(f"{f:.2f}ac")
            except (TypeError, ValueError):
                pass
        # Optional ranking-tier prefix (e.g. "T1=85") for the smart-picker
        tier = p.get("tier")
        score = p.get("score")
        if tier and score is not None:
            meta_bits.insert(0, f"{tier}={score}")
        if meta_bits:
            lines.append("  " + "  ".join(meta_bits))

        # Line 3: PID
        pid = (p.get("pid") or "").strip()
        if pid:
            lines.append(f"  PID {pid}")

        blocks.append("\n".join(lines))
    # Join blocks with a blank line separator for visual scanning
    return "\n\n".join(blocks)


def _format_extra_parcels(extras: list[NoticeData]) -> str:
    """Scrape-time wrapper: converts NoticeData extras into the dict
    shape that format_extra_parcels_vertical expects, then delegates.
    """
    if not extras:
        return ""
    items: list[dict] = []
    for e in extras:
        addr = " ".join(filter(None, [e.address, e.city, e.zip])).strip()
        items.append({
            "address": addr,
            "use": e.property_use_simple,
            "pid": e.parcel_id,
            # NoticeData doesn't carry market_value -- left blank.
        })
    return format_extra_parcels_vertical(items)


def _build_notes(extra_parcels: list[NoticeData] | None = None) -> str:
    """Build the Notes block — extra-parcel notes only (beneficiaries moved
    to their own column in build 1.0.30+).
    """
    return _format_extra_parcels(extra_parcels or [])


def _build_beneficiaries(notice: NoticeData) -> str:
    """Format the beneficiary list for its own column.

    One beneficiary per line (multi-line cell content). Each line:
        "Last, First Middle — street, city ST zip"
    """
    if not notice.beneficiaries_json:
        return ""
    try:
        bens = json.loads(notice.beneficiaries_json)
    except (ValueError, TypeError):
        return ""
    if not bens:
        return ""
    lines: list[str] = []
    for b in bens:
        name = (b.get("name") or "").strip()
        street = (b.get("street") or "").strip()
        city = (b.get("city") or "").strip()
        state = (b.get("state") or "").strip()
        zipc = (b.get("zip") or "").strip()
        addr_bits: list[str] = []
        if street:
            addr_bits.append(street)
        csz = " ".join(filter(None, [city + "," if city else "", state, zipc])).strip()
        if csz:
            addr_bits.append(csz)
        addr = ", ".join(addr_bits)
        if name and addr:
            lines.append(f"{name} - {addr}")
        elif name:
            lines.append(name)
        elif addr:
            lines.append(addr)
    return "\n".join(lines)


def _iso_week_tag(date_str: str) -> str:
    """Build 'NC Estates Week N YYYY' from a YYYY-MM-DD date string."""
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d") if date_str else datetime.now()
    except ValueError:
        d = datetime.now()
    week = d.isocalendar()[1]
    return f"NC Estates Week {week} {d.year}"


def notice_to_ftm_row(
    notice: NoticeData,
    *,
    tag_override: str = "",
    extra_parcels: list[NoticeData] | None = None,
) -> dict[str, str]:
    """Convert one NoticeData record to a single FTM-format dict.

    `extra_parcels` (if any) get listed in Notes as 'PLUS N PARCELS: ...'
    so a decedent with multiple properties collapses to one row.

    DM Name / DM Relationship / DM 2 / DM 3 are populated from
    `notice.decision_maker_*` (set by the obituary enricher when
    `--nc-obituary` runs). Phone/Email columns are left blank — DataSift's
    post-upload skip-trace fills those.
    """
    tag = tag_override or _iso_week_tag(notice.date_added)
    exec_full = " ".join(filter(None, [notice.executor_first_name, notice.executor_last_name])).strip()
    return {
        "File Date":          _format_date(notice.date_added),
        "County":             notice.county,
        "Case No.":           notice.case_number,
        "Case Status":        notice.case_status,
        "Deceased Owner":     _format_decedent(notice.decedent_name),
        "Personal Representative": exec_full,
        "First Name":         notice.executor_first_name,
        "Last Name":          notice.executor_last_name,
        "Mailing Address":  notice.owner_street,
        "Mailing City":     notice.owner_city,
        "Mailing State":    notice.owner_state,
        "Mailing Zip":      notice.owner_zip,
        "Parcel ID":        notice.parcel_id,
        "Property Address": notice.address,
        "Property City":    notice.city,
        "Property State":   notice.state if notice.state == "NC" else "NC",
        "Property Zip":     notice.zip,
        "Property use":     notice.property_use_simple,
        "Property Value":   notice.estimated_value or "",
        "Notes":            _build_notes(extra_parcels=extra_parcels),
        "Beneficiaries":    _build_beneficiaries(notice),
        "Phone 1":          notice.primary_phone,
        "Tags":             tag,
        "List":             "PROBATE",
        "DM Name":          notice.decision_maker_name or "",
        "DM Relationship":  notice.decision_maker_relationship or "",
        "DM Phone":         "",
        "DM Email":         "",
        "DM 2 Name":        notice.decision_maker_2_name or "",
        "DM 2 Relationship": notice.decision_maker_2_relationship or "",
        "DM 3 Name":        notice.decision_maker_3_name or "",
        "DM 3 Relationship": notice.decision_maker_3_relationship or "",
    }


def _market_value_key(n: NoticeData) -> float:
    """Read estimated_value as a float; 0 when missing."""
    try:
        return float(n.estimated_value or 0)
    except (TypeError, ValueError):
        return 0.0


def _main_parcel_priority(n: NoticeData) -> tuple[int, int, float]:
    """Sort key for picking the 'main' parcel per decedent.

    Priority (sorted DESCENDING, so highest tuple wins):
      1. IN-PROBATE beats transfers-by-survivorship. A parcel is "in
         probate" when it's either solely owned OR jointly owned with
         someone who ISN'T a beneficiary in the estate (likely Tenants-
         in-Common, decedent's share goes through probate). A parcel
         transfers by survivorship when it's jointly owned with someone
         who IS a beneficiary (JTWROS, decedent's share auto-transfers).
      2. Use-class preference depends on ownership profile:
         - IN-PROBATE: residential > unknown > vacant > commercial
         - SURVIVORSHIP: vacant > unknown > residential > commercial
           (residential is likely surviving spouse's home)
      3. Highest market value wins within same class.

    "In-probate" tier (descending):
      1 = in probate (sole, or joint-with-non-beneficiary aka likely TIC)
      0 = transfers by survivorship (joint with a beneficiary)
    """
    is_joint = bool(getattr(n, "is_jointly_owned", False))
    is_survivorship = bool(getattr(n, "is_likely_survivorship", True))
    # In-probate when sole (not joint) OR joint with non-beneficiary
    in_probate = (not is_joint) or (is_joint and not is_survivorship)
    probate_tier = 1 if in_probate else 0
    use = (getattr(n, "property_use_simple", "") or "").upper()
    if "COMMERCIAL" in use or "INDUSTRIAL" in use or "OFFICE" in use:
        use_class = "COMMERCIAL"
    elif "VACANT" in use or "LAND" in use:
        use_class = "VACANT"
    elif use in {"SFR", "RESIDENTIAL", "TOWNHOUSE", "CONDO", "MH",
                 "MULTI-FAMILY", "DUPLEX"}:
        use_class = "RESIDENTIAL"
    else:
        use_class = "UNKNOWN"
    if in_probate:
        # The probate asset is most likely the residence (where the
        # decedent lived); vacant lots are secondary.
        use_tier = {"RESIDENTIAL": 3, "UNKNOWN": 2, "VACANT": 1, "COMMERCIAL": 0}[use_class]
    else:
        # Survivorship — residential is likely surviving spouse's home,
        # demote it; vacant might still be a TIC fragment.
        use_tier = {"VACANT": 3, "UNKNOWN": 2, "RESIDENTIAL": 1, "COMMERCIAL": 0}[use_class]
    return (probate_tier, use_tier, _market_value_key(n))


def collapse_by_case(notices: list[NoticeData]) -> list[tuple[NoticeData, list[NoticeData]]]:
    """Group notices by case_number; return [(main_parcel, [extra_parcels])].

    Main parcel selection (in order of preference):
      1. Residential / SFR / Condo / Townhouse (the actual house)
      2. Unknown use-type (when no classification is available)
      3. Vacant land
      4. Commercial / industrial
    Within same use-tier, highest estimated_value wins.

    Notices without a case_number get returned individually.
    """
    groups: dict[str, list[NoticeData]] = {}
    out: list[tuple[NoticeData, list[NoticeData]]] = []
    for n in notices:
        key = n.case_number or f"_no_case_{id(n)}"
        groups.setdefault(key, []).append(n)
    for key, items in groups.items():
        if len(items) == 1:
            out.append((items[0], []))
            continue
        items_sorted = sorted(items, key=_main_parcel_priority, reverse=True)
        main, extras = items_sorted[0], items_sorted[1:]
        out.append((main, extras))
    return out


# FTM-style color palette
_HEADER_FILL_COLOR = "1B5E20"        # dark green
_HEADER_TEXT_COLOR = "FFFFFF"        # white
_BAND_FILL_COLOR = "FFFDE7"          # very light yellow (banded rows)
_DEFAULT_ROW_HEIGHT = 16             # single-line height


def write_ftm_xlsx(
    notices: list[NoticeData],
    out_path: Path,
    *,
    tag_override: str = "",
    collapse_multi_parcel: bool = True,
) -> int:
    """Write the FTM Estates output as XLSX with a County dropdown.

    Styling: dark-green header with white bold text + alternating white /
    pale-yellow banded rows. Single-line row height (Notes shown truncated
    in cell but full content preserved in the underlying value — hover or
    expand row to see).

    Imports openpyxl lazily because not all callers need XLSX.
    Returns the number of data rows written.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment

    out_path.parent.mkdir(parents=True, exist_ok=True)
    if collapse_multi_parcel:
        groups = collapse_by_case(notices)
        rows = [notice_to_ftm_row(main, tag_override=tag_override, extra_parcels=extras)
                for main, extras in groups]
    else:
        rows = [notice_to_ftm_row(n, tag_override=tag_override) for n in notices]

    # Sort rows by County (alphabetical) then by Case No. — keeps each
    # county's leads grouped together for visual scan-ability.
    rows.sort(key=lambda r: ((r.get("County") or "ZZZ"), r.get("Case No.", "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "NC Estates"

    # Header row — dark green fill, bold white text
    header_font = Font(bold=True, color=_HEADER_TEXT_COLOR)
    header_fill = PatternFill(start_color=_HEADER_FILL_COLOR,
                              end_color=_HEADER_FILL_COLOR, fill_type="solid")
    for col_idx, col_name in enumerate(FTM_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # Data rows — single-line height, per-county color tint (replaces the
    # old yellow-alternating-band so county groups are visually distinct).
    county_fills = {
        c: PatternFill(start_color=h, end_color=h, fill_type="solid")
        for c, h in NC_COUNTY_COLORS.items()
    }
    multiline_cols = {"Notes", "Beneficiaries"}
    for r_idx, r in enumerate(rows, start=2):
        row_fill = county_fills.get(r.get("County", ""))
        for c_idx, col_name in enumerate(FTM_COLUMNS, start=1):
            val = r.get(col_name, "")
            if col_name in multiline_cols and val:
                val = " | ".join(s.strip() for s in str(val).split("\n") if s.strip())
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[r_idx].height = _DEFAULT_ROW_HEIGHT

    # County dropdown — applied to the full County column
    county_col_idx = FTM_COLUMNS.index("County") + 1
    county_col_letter = get_column_letter(county_col_idx)
    county_formula = '"' + ",".join(NC_COUNTY_OPTIONS) + '"'
    dv = DataValidation(type="list", formula1=county_formula, allow_blank=True)
    dv.error = "Pick one of the 7 NC counties"
    dv.errorTitle = "Invalid county"
    dv.prompt = "Select a NC county"
    dv.promptTitle = "County"
    ws.add_data_validation(dv)
    dv.add(f"{county_col_letter}2:{county_col_letter}1048576")

    # Column widths — make the important fields legible
    col_widths = {
        "File Date": 11, "County": 14, "Case No.": 18, "Deceased Owner": 32,
        "Personal Representative": 25, "First Name": 16, "Last Name": 18,
        "Mailing Address": 28, "Mailing City": 16, "Mailing State": 7, "Mailing Zip": 8,
        "Parcel ID": 16, "Property Address": 28, "Property City": 16,
        "Property State": 8, "Property Zip": 8, "Property use": 14,
        "Property Value": 14,
        "Notes": 40, "Beneficiaries": 80, "Phone 1": 14, "Tags": 26, "List": 10,
        "Match Reason": 24,
        "Zillow URL": 50,
        # Heir / Decision Maker columns
        "DM Name": 22, "DM Relationship": 14, "DM Phone": 14, "DM Email": 26,
        "DM 2 Name": 22, "DM 2 Relationship": 14,
        "DM 3 Name": 22, "DM 3 Relationship": 14,
    }
    for c_idx, col_name in enumerate(FTM_COLUMNS, start=1):
        dim = ws.column_dimensions[get_column_letter(c_idx)]
        dim.width = col_widths.get(col_name, 14)
        if col_name in HIDDEN_FROM_WORKBOOK:
            dim.hidden = True

    # Freeze the header row so it stays visible on scroll
    ws.freeze_panes = "A2"

    wb.save(out_path)
    logger.info("Wrote %d NC Estates rows to %s", len(rows), out_path)
    return len(rows)


def write_ftm_csv(
    notices: list[NoticeData],
    out_path: Path,
    *,
    tag_override: str = "",
    collapse_multi_parcel: bool = True,
) -> int:
    """Write NoticeData list to a CSV matching the FTM Estates format.

    With `collapse_multi_parcel=True` (default), decedents with multiple
    parcels collapse to one row (highest-value parcel as main; the rest
    listed in Notes as 'PLUS N PARCELS: ...'). Set False to emit one
    row per parcel.

    Returns the number of rows written.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if collapse_multi_parcel:
        groups = collapse_by_case(notices)
        rows = [notice_to_ftm_row(main, tag_override=tag_override, extra_parcels=extras)
                for main, extras in groups]
    else:
        rows = [notice_to_ftm_row(n, tag_override=tag_override) for n in notices]
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FTM_COLUMNS, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
    logger.info("Wrote %d NC Estates rows to %s", len(rows), out_path)
    return len(rows)
