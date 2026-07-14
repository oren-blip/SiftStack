"""Re-enrich missing Executor + Beneficiaries data in an existing FTM CSV.

When the daily eCourts pipeline saturates Tyler's per-IP rate limit, the
Parties OData endpoint returns HTTP 202 ("queued, retry") for the last
batch of cases — most often Mecklenburg + Rowan. The CSV rows then have
blank `Personal Representative` and `Beneficiaries` even though the data
exists at Tyler.

Tyler's throttle has been observed to last for *hours*, not minutes —
so this script is designed for **incremental progress across multiple
invocations**:

  1. Hex-map cache (`output/.ecourts_hex_cache.json`) — case_no → 256-char
     case_id. Avoids re-running the per-county wildcard searches after
     the first invocation.
  2. Fills cache (`output/.reenrich_fills.json`) — case_no → {executor +
     beneficiary fields}. Every successful Parties call appends to this
     file IMMEDIATELY so progress survives kill / crash.
  3. Adaptive throttle exit — if we see 5 consecutive 202-exhausted
     failures, Tyler is in ban mode; we abort and tell the user to
     re-run in a few hours.
  4. Always writes a fresh CSV + XLSX at the end (or on abort), merging
     cached fills with whatever this invocation managed to fetch.

Run as many times as needed — each run picks up where the last left off.

Usage:
    python reenrich_ftm_executors.py
    python reenrich_ftm_executors.py --csv output/nc_estates_ftm_X.csv
    python reenrich_ftm_executors.py --counties Mecklenburg,Rowan
    python reenrich_ftm_executors.py --headed   # watch the browser
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import logging
import sys
import time
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / "src"))

from playwright.async_api import async_playwright  # noqa: E402

from ecourts_case_api import CaseDetail, CaseDetailClient  # noqa: E402
from ecourts_scraper import (  # noqa: E402
    CASE_TYPE_BY_NOTICE_TYPE,
    PORTAL_URL,
    SMART_SEARCH_URL,
    _DEFAULT_UA,
    _is_waf_gate,
    _load_cached_waf_cookie,
    _navigate_to_smart_search,
    _open_advanced_filters,
    _parse_results,
    _search_criteria_for,
    _select_only_county,
    _set_case_type,
    _set_date_range,
    _set_search_criteria,
    _solve_and_inject_waf,
    _submit_search,
)
from nc_ftm_writer import FTM_COLUMNS, NC_COUNTY_COLORS, NC_COUNTY_OPTIONS, workbook_header  # noqa: E402

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("reenrich")


HEX_CACHE_PATH = Path("output") / ".ecourts_hex_cache.json"
FILLS_CACHE_PATH = Path("output") / ".reenrich_fills.json"
# Cases whose Parties API returned a response (success or empty) but didn't
# yield an executor name. Skipped on subsequent runs within TRIED_EMPTY_TTL
# hours so we don't waste throttle budget re-attacking known-empty cases.
TRIED_EMPTY_CACHE_PATH = Path("output") / ".reenrich_tried_empty.json"
TRIED_EMPTY_TTL_HOURS = 24

# Fields we copy from the Parties API into the FTM row (and persist to fills cache)
FILL_FIELDS = (
    "Personal Representative", "First Name", "Last Name",
    "Mailing Address", "Mailing City", "Mailing State", "Mailing Zip",
    "Beneficiaries",
)


# ── CSV row helpers ───────────────────────────────────────────────────


def find_latest_ftm_csv() -> Path:
    candidates = sorted(Path("output").glob("nc_estates_ftm_*.csv"))
    if not candidates:
        raise FileNotFoundError("No nc_estates_ftm_*.csv in output/")
    return candidates[-1]


def load_csv(path: Path) -> list[dict]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def missing_executor_rows(rows: list[dict]) -> list[dict]:
    return [r for r in rows if not (r.get("Personal Representative") or "").strip()]


def parse_file_date(s: str) -> datetime | None:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def csv_date_range(rows: list[dict]) -> tuple[datetime, datetime]:
    dates = [d for d in (parse_file_date(r.get("File Date", "")) for r in rows) if d]
    if not dates:
        end = datetime.now()
        return end - timedelta(days=7), end
    return min(dates), max(dates)


# ── Cache I/O ─────────────────────────────────────────────────────────


def load_json_cache(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        with path.open(encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        logger.warning("Cache %s unreadable (%s) — starting empty", path, e)
        return {}


def save_json_cache(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(data, f, separators=(",", ":"))
    tmp.replace(path)


# ── eCourts re-search (only for counties not yet in hex cache) ───────


async def build_case_hex_map(
    counties_to_search: list[str],
    date_start: datetime,
    date_end: datetime,
    *,
    headless: bool,
) -> tuple[dict[str, str], str, str]:
    """Re-run the per-county wildcard search to recover case_no → case_hex.

    Returns:
        ({case_no: case_hex}, waf_token, user_agent)

    Note: case_no is unique enough across NC counties (county suffix is in
    the number itself like '26E001790-590') so we key by case_no alone.
    """
    mdy_start = date_start.strftime("%m/%d/%Y")
    mdy_end = date_end.strftime("%m/%d/%Y")
    logger.info("Re-search %s..%s for counties=%s", mdy_start, mdy_end, counties_to_search)

    case_hex: dict[str, str] = {}
    waf_token = ""
    user_agent = _DEFAULT_UA

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=headless)
        ctx = await browser.new_context(
            viewport={"width": 1440, "height": 900}, user_agent=_DEFAULT_UA
        )
        ctx.set_default_timeout(60_000)
        page = await ctx.new_page()

        cached = _load_cached_waf_cookie()
        if cached:
            logger.info("Reusing cached WAF cookie")
            await ctx.close()
            ctx = await browser.new_context(
                viewport={"width": 1440, "height": 900},
                user_agent=cached.get("user_agent") or _DEFAULT_UA,
            )
            await ctx.add_cookies([{
                "name": "aws-waf-token",
                "value": cached["aws_waf_token"],
                "domain": ".tylertech.cloud",
                "path": "/",
                "httpOnly": False,
                "secure": True,
                "sameSite": "Lax",
            }])
            ctx.set_default_timeout(60_000)
            page = await ctx.new_page()

        await page.goto(PORTAL_URL, wait_until="domcontentloaded", timeout=45_000)
        await page.wait_for_timeout(2500)

        if await _is_waf_gate(page):
            logger.info("WAF gate hit — solving via CapSolver")
            ctx, page = await _solve_and_inject_waf(browser, ctx, page)

        await _navigate_to_smart_search(page)
        await _open_advanced_filters(page)

        case_type_value = CASE_TYPE_BY_NOTICE_TYPE["probate"]
        criteria = _search_criteria_for("probate", date_end.year)

        for i, county in enumerate(counties_to_search):
            logger.info("Re-searching %s (%s)", county, case_type_value)
            if i > 0:
                await page.goto(SMART_SEARCH_URL, wait_until="domcontentloaded", timeout=45_000)
                await page.wait_for_timeout(2000)
                await _open_advanced_filters(page)
            try:
                await _set_search_criteria(page, criteria)
                await _set_case_type(page, case_type_value)
                await _set_date_range(page, mdy_start, mdy_end)
                await _select_only_county(page, county)
                if not await _submit_search(page):
                    logger.warning("Submit failed for %s — skipping", county)
                    continue
                results = await _parse_results(page, county, "probate")
            except Exception:
                logger.exception("Re-search failed for %s", county)
                continue
            for n in results:
                hex_id = getattr(n, "_roa_id", "")
                if hex_id and n.case_number:
                    case_hex[n.case_number] = hex_id
            logger.info("%s: mapped %d case → hex IDs", county,
                        sum(1 for k in results if getattr(k, "_roa_id", "")))

        ctx_cookies = await ctx.cookies("https://portal-nc.tylertech.cloud/")
        waf_cookie = next((c for c in ctx_cookies if c["name"] == "aws-waf-token"), None)
        waf_token = waf_cookie["value"] if waf_cookie else ""
        user_agent = (cached.get("user_agent") if cached else _DEFAULT_UA) or _DEFAULT_UA

        await ctx.close()
        await browser.close()

    return case_hex, waf_token, user_agent


# ── Fill extraction (Parties → row dict) ──────────────────────────────


def detail_to_fill_dict(detail: CaseDetail) -> dict | None:
    """Extract the executor + beneficiary fields from a CaseDetail into a
    plain dict suitable for persisting + applying to a row. Returns None
    for guardianships (caller should drop). Returns {} for cases with no
    executor (still a valid 'we tried' result so caller can skip retries).
    """
    if detail.is_guardianship:
        return None

    fill: dict[str, str] = {}
    ex = detail.executor
    if ex:
        full = " ".join(filter(None, [ex.first_name, ex.last_name])).strip() or ex.full_name
        fill["Personal Representative"] = full
        fill["First Name"] = ex.first_name
        fill["Last Name"] = ex.last_name
        addr = ex.first_address
        if not addr.is_blank():
            fill["Mailing Address"] = " ".join(filter(None, [addr.line1, addr.line2])).strip()
            fill["Mailing City"] = addr.city
            fill["Mailing State"] = addr.state
            fill["Mailing Zip"] = addr.zip

    if detail.beneficiaries:
        lines: list[str] = []
        for b in detail.beneficiaries:
            addr = b.first_address
            street = " ".join(filter(None, [addr.line1, addr.line2])).strip()
            csz = " ".join(filter(None,
                [addr.city + "," if addr.city else "", addr.state, addr.zip])).strip()
            addr_bits: list[str] = []
            if street:
                addr_bits.append(street)
            if csz:
                addr_bits.append(csz)
            addr_str = ", ".join(addr_bits)
            if b.full_name and addr_str:
                lines.append(f"{b.full_name} - {addr_str}")
            elif b.full_name:
                lines.append(b.full_name)
        if lines:
            fill["Beneficiaries"] = "\n".join(lines)

    return fill


def apply_fill_to_row(row: dict, fill: dict) -> None:
    """Merge a fill dict into a CSV row (only writes non-empty values)."""
    for k, v in fill.items():
        if v:
            row[k] = v


# ── Output writers ────────────────────────────────────────────────────


def write_csv(rows: list[dict], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FTM_COLUMNS,
                                quoting=csv.QUOTE_MINIMAL, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
    logger.info("Wrote CSV: %s", out_path)


def write_xlsx(rows: list[dict], out_path: Path) -> None:
    """Mirror nc_ftm_writer.write_ftm_xlsx styling — operates on row dicts."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment

    HEADER_FILL = "1B5E20"
    BAND_FILL = "FFFDE7"

    # Sort rows by County then Case No. for grouped visual layout
    rows = sorted(rows, key=lambda r: ((r.get("County") or "ZZZ"), r.get("Case No.", "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "NC Estates"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    for col_idx, col_name in enumerate(FTM_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=workbook_header(col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # Per-county color tint (replaces old yellow-alternating band)
    county_fills = {
        c: PatternFill(start_color=h, end_color=h, fill_type="solid")
        for c, h in NC_COUNTY_COLORS.items()
    }
    # Multi-line columns get wrap_text + top-align so the FULL content
    # is preserved in the cell (visible in the formula bar on click, and
    # rendered fully when the user expands the row). Rows default to
    # compact single-line height so the workbook stays scannable —
    # matches Oren's Google Sheets workflow. To see all multi-line
    # content at once: Excel Alt+H+O+A or Sheets Format > Row >
    # Auto-fit row height.
    multiline_cols = {"Notes", "Beneficiaries", "Heirs (App)"}
    band_fill = PatternFill(start_color=BAND_FILL, end_color=BAND_FILL, fill_type="solid")
    for r_idx, r in enumerate(rows, start=2):
        # Per-county tint applied ONLY to the County column cell — per Oren
        # 2026-06-26, full-row tinting was visually noisy.
        row_fill = county_fills.get(r.get("County", ""))
        # Alternating band: even data rows (3rd, 5th, ...) get a pale tint
        # for readability. Skip County so its per-county color stays clean.
        is_banded = (r_idx - 2) % 2 == 1
        for c_idx, col_name in enumerate(FTM_COLUMNS, start=1):
            val = r.get(col_name, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if col_name in multiline_cols:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if col_name == "County" and row_fill:
                cell.fill = row_fill
            elif is_banded and col_name != "County":
                cell.fill = band_fill
        ws.row_dimensions[r_idx].height = 16

    county_col_idx = FTM_COLUMNS.index("County") + 1
    county_col_letter = get_column_letter(county_col_idx)
    county_formula = '"' + ",".join(NC_COUNTY_OPTIONS) + '"'
    dv = DataValidation(type="list", formula1=county_formula, allow_blank=True)
    dv.error = "Pick one of the 7 NC counties"
    dv.errorTitle = "Invalid county"
    dv.prompt = "Select a NC county"
    dv.promptTitle = "County"
    ws.add_data_validation(dv)
    dv.add(f"{county_col_letter}2:{county_col_letter}1048576")

    col_widths = {
        "File Date": 11, "County": 14, "Case No.": 18, "Deceased Owner": 32,
        "Personal Representative": 25, "First Name": 16, "Last Name": 18,
        "Mailing Address": 28, "Mailing City": 16, "Mailing State": 7, "Mailing Zip": 8,
        "Parcel ID": 16, "Property Address": 28, "Property City": 16,
        "Property State": 8, "Property Zip": 8, "Property use": 14,
        "Property Value": 14,
        "Notes": 40, "Beneficiaries": 80, "Phone 1": 14, "Phone 1 Tier": 12, "Tags": 26, "List": 10,
    }
    for c_idx, col_name in enumerate(FTM_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col_name, 14)

    ws.freeze_panes = "A2"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    logger.info("Wrote XLSX: %s", out_path)


def write_outputs(rows: list[dict]) -> tuple[Path, Path]:
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    out_csv = Path("output") / f"nc_estates_ftm_{ts}.csv"
    out_xlsx = Path("output") / f"nc_estates_ftm_{ts}.xlsx"
    write_csv(rows, out_csv)
    write_xlsx(rows, out_xlsx)
    return out_csv, out_xlsx


# ── Main ──────────────────────────────────────────────────────────────


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", type=str, default=None,
                    help="FTM CSV path (default: latest in output/)")
    ap.add_argument("--counties", type=str, default=None,
                    help="Comma-separated subset (default: all counties with blanks)")
    ap.add_argument("--headed", action="store_true", help="Watch the browser")
    ap.add_argument("--inter-case-delay", type=float, default=8.0,
                    help="Seconds to sleep between Parties calls (default 8)")
    ap.add_argument("--max-consecutive-fails", type=int, default=5,
                    help="Abort if this many cases in a row come back empty (default 5)")
    ap.add_argument("--retries", type=int, default=3,
                    help="Parties API retry budget per case (default 3 → ~30s max wait)")
    args = ap.parse_args()

    src_csv = Path(args.csv) if args.csv else find_latest_ftm_csv()
    logger.info("Source CSV: %s", src_csv)
    rows = load_csv(src_csv)
    logger.info("Loaded %d rows", len(rows))

    # Apply any previously-cached fills (from earlier runs) BEFORE counting blanks
    fills_cache: dict[str, dict] = load_json_cache(FILLS_CACHE_PATH)
    if fills_cache:
        applied = 0
        for r in rows:
            cno = r.get("Case No.", "")
            if cno in fills_cache:
                apply_fill_to_row(r, fills_cache[cno])
                applied += 1
        logger.info("Applied %d previously-cached fills from %s", applied, FILLS_CACHE_PATH.name)

    missing = missing_executor_rows(rows)
    if not missing:
        logger.info("No blank-executor rows — nothing to do")
        out_csv, out_xlsx = write_outputs(rows)
        return

    by_county = defaultdict(list)
    for r in missing:
        by_county[r["County"]].append(r)
    logger.info("Blank-executor rows by county:")
    for c in sorted(by_county):
        logger.info("  %s: %d", c, len(by_county[c]))
    logger.info("Total: %d blank-executor rows remaining", len(missing))

    requested_counties = set(c.strip() for c in (args.counties or "").split(",") if c.strip())
    target_counties = sorted(
        c for c in by_county
        if not requested_counties or c in requested_counties
    )
    if not target_counties:
        logger.info("No target counties match request — nothing to do")
        return
    logger.info("Targeting counties: %s", target_counties)

    target_rows = [r for r in missing if r["County"] in target_counties]

    # ── Tried-empty cache: skip cases that returned no executor recently ──
    tried_empty: dict[str, str] = load_json_cache(TRIED_EMPTY_CACHE_PATH)
    if tried_empty:
        cutoff = datetime.now() - timedelta(hours=TRIED_EMPTY_TTL_HOURS)
        recently_tried: set[str] = set()
        for cno, iso in tried_empty.items():
            try:
                if datetime.fromisoformat(iso) > cutoff:
                    recently_tried.add(cno)
            except (TypeError, ValueError):
                continue
        if recently_tried:
            before = len(target_rows)
            target_rows = [r for r in target_rows if r["Case No."] not in recently_tried]
            skipped = before - len(target_rows)
            if skipped:
                logger.info("Skipping %d cases tried-empty within last %dh "
                            "(re-attempt after TTL expires)", skipped, TRIED_EMPTY_TTL_HOURS)

    if not target_rows:
        logger.info("Nothing left to attempt this run — all remaining blanks are "
                    "in tried-empty TTL window")
        out_csv, out_xlsx = write_outputs(rows)
        return

    # ── Hex map: check cache first; only re-search counties with gaps ─
    hex_cache: dict[str, str] = load_json_cache(HEX_CACHE_PATH)
    logger.info("Hex cache: %d entries", len(hex_cache))

    counties_needing_search = sorted({
        r["County"] for r in target_rows if r["Case No."] not in hex_cache
    })

    # Always go through build_case_hex_map — even with an empty county list,
    # it opens Playwright + validates/re-solves the WAF cookie. The WAF
    # cookie is bound to the requesting IP, so a cached cookie from a
    # previous VPN exit will fail on the new IP unless re-solved here.
    date_start, date_end = csv_date_range(rows)
    date_end = date_end + timedelta(days=1)
    if counties_needing_search:
        logger.info("Need to re-search %d counties (cache miss)", len(counties_needing_search))
    else:
        logger.info("All target case_nos already in hex cache — verifying WAF cookie matches current IP")
    new_hex, waf_token, user_agent = asyncio.run(
        build_case_hex_map(counties_needing_search, date_start, date_end,
                           headless=not args.headed)
    )
    if new_hex:
        hex_cache.update(new_hex)
        save_json_cache(HEX_CACHE_PATH, hex_cache)
        logger.info("Saved hex cache (%d total entries)", len(hex_cache))

    if not waf_token:
        logger.error("No WAF token available — aborting")
        sys.exit(1)

    # ── Patient single-pass with throttle detection ──────────────────
    client = CaseDetailClient(waf_token=waf_token, user_agent=user_agent)
    consecutive_fails = 0
    filled_this_run = 0
    skipped_no_hex = 0
    aborted = False

    logger.info("Starting fetch loop: %d cases, retries=%d, %.0fs inter-case, "
                "abort after %d consecutive throttles",
                len(target_rows), args.retries, args.inter_case_delay,
                args.max_consecutive_fails)

    for i, r in enumerate(target_rows):
        if i > 0:
            time.sleep(args.inter_case_delay)
        county = r["County"]
        case_no = r["Case No."]
        hex_id = hex_cache.get(case_no)
        if not hex_id:
            logger.warning("[%d/%d] %s/%s — no case_hex (skipping)",
                           i + 1, len(target_rows), county, case_no)
            skipped_no_hex += 1
            continue

        parties = client.fetch_parties(hex_id, retries=args.retries)
        if not parties:
            consecutive_fails += 1
            logger.info("[%d/%d] %s/%s — throttled (%d in a row)",
                        i + 1, len(target_rows), county, case_no, consecutive_fails)
            if consecutive_fails >= args.max_consecutive_fails:
                logger.warning("Aborting: %d consecutive throttle failures. "
                               "Tyler is rate-limiting. Re-run in 2-4 hours.",
                               consecutive_fails)
                aborted = True
                break
            continue

        detail = CaseDetail(case_id=hex_id, parties=parties)
        fill = detail_to_fill_dict(detail)
        consecutive_fails = 0

        if fill is None:
            logger.info("[%d/%d] %s/%s — guardianship case, skipping",
                        i + 1, len(target_rows), county, case_no)
            tried_empty[case_no] = datetime.now().isoformat(timespec="seconds")
            save_json_cache(TRIED_EMPTY_CACHE_PATH, tried_empty)
            continue
        if not fill or not fill.get("Personal Representative"):
            logger.info("[%d/%d] %s/%s — no executor on file yet",
                        i + 1, len(target_rows), county, case_no)
            # Still apply whatever we got (e.g. beneficiaries) and persist
            if fill:
                apply_fill_to_row(r, fill)
                fills_cache[case_no] = fill
                save_json_cache(FILLS_CACHE_PATH, fills_cache)
            tried_empty[case_no] = datetime.now().isoformat(timespec="seconds")
            save_json_cache(TRIED_EMPTY_CACHE_PATH, tried_empty)
            continue

        apply_fill_to_row(r, fill)
        # Persist the fill IMMEDIATELY so a kill / crash doesn't lose it
        fills_cache[case_no] = fill
        save_json_cache(FILLS_CACHE_PATH, fills_cache)
        filled_this_run += 1
        logger.info("[%d/%d] %s/%s OK — exec=%s",
                    i + 1, len(target_rows), county, case_no,
                    fill.get("Personal Representative", "")[:30])

    # ── Decedent name repair: rows where parser garbled the name ─────
    # Older runs (pre-fix) left some rows with "OF, IN THE MATTER" or
    # similar garbage as the decedent. Re-fetch parties to recover the
    # canonical decedent name. Skipped if the main loop already aborted.
    if not aborted:
        broken_decedent_rows = [
            r for r in rows
            if "IN THE MATTER" in (r.get("Deceased Owner") or "").upper()
        ]
        if broken_decedent_rows:
            logger.info("Repairing %d garbled decedent name(s)", len(broken_decedent_rows))
            decedent_repaired = 0
            for r in broken_decedent_rows:
                time.sleep(args.inter_case_delay)
                case_no = r["Case No."]
                hex_id = hex_cache.get(case_no)
                if not hex_id:
                    continue
                parties = client.fetch_parties(hex_id, retries=args.retries)
                if not parties:
                    logger.info("Decedent repair throttled at %s — stopping repair pass", case_no)
                    break
                detail = CaseDetail(case_id=hex_id, parties=parties)
                if detail.decedent and detail.decedent.full_name:
                    old = r["Deceased Owner"]
                    r["Deceased Owner"] = detail.decedent.full_name
                    logger.info("Decedent repair %s: %r -> %r",
                                case_no, old, detail.decedent.full_name)
                    decedent_repaired += 1
            logger.info("Repaired %d/%d decedent name(s)",
                        decedent_repaired, len(broken_decedent_rows))

    # ── Always write outputs, even on abort ──────────────────────────
    out_csv, out_xlsx = write_outputs(rows)

    still_blank = sum(1 for r in rows if not (r.get("Personal Representative") or "").strip())
    logger.info("=" * 60)
    logger.info("Filled this run: %d   No hex: %d   Aborted: %s",
                filled_this_run, skipped_no_hex, aborted)
    logger.info("Blank-executor rows: %d → %d", len(missing), still_blank)
    logger.info("Cumulative fills in cache: %d", len(fills_cache))
    if aborted or still_blank > 0:
        logger.info("Re-run this script in 2-4 hours to recover more — "
                    "cached hex IDs + fills mean it picks up where this left off.")


if __name__ == "__main__":
    main()
