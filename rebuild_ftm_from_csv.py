"""Rebuild the FTM CSV + XLSX from an existing FTM CSV — quick way to apply
column / formatting changes without re-running the full pipeline (which
takes ~50 min thanks to Tyler's throttle).

Reads the latest nc_estates_ftm_*.csv, adds the new 'Personal Representative'
column, and writes both CSV (new column added) + XLSX (with County dropdown).
"""

import csv
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / "src"))

from nc_ftm_writer import FTM_COLUMNS, NC_COUNTY_OPTIONS  # noqa: E402


def main() -> None:
    src_csv = Path("output/nc_estates_ftm_2026-05-17_211501.csv")
    if not src_csv.exists():
        print(f"ERROR: source file missing: {src_csv}")
        sys.exit(1)

    with src_csv.open(newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    print(f"Loaded {len(rows)} rows from {src_csv.name}")

    # Add the new Personal Representative column derived from First + Last
    # AND split the old combined Notes into Notes + Beneficiaries columns.
    for r in rows:
        first = (r.get("First Name") or "").strip()
        last = (r.get("Last Name") or "").strip()
        r["Personal Representative"] = " ".join(filter(None, [first, last])).strip()

        # Old Notes format: "PLUS N PARCELS\n  ...\n\nBeneficiary\n<name>\n<addr>..."
        old_notes = (r.get("Notes") or "").strip()
        if not old_notes:
            r["Notes"] = ""
            r["Beneficiaries"] = ""
            continue
        # Split on the 'Beneficiary' header line if present
        if "Beneficiary" in old_notes:
            before, _, after = old_notes.partition("Beneficiary")
            notes_part = before.strip()
            ben_block = after.strip()
            # Re-format beneficiary block: every 3 lines = name / street / city,state zip
            lines = [ln.strip() for ln in ben_block.split("\n") if ln.strip()]
            ben_entries: list[str] = []
            i = 0
            while i < len(lines):
                name = lines[i]
                street = lines[i + 1] if i + 1 < len(lines) else ""
                csz = lines[i + 2] if i + 2 < len(lines) else ""
                addr_parts = [street, csz] if street and csz else ([street] if street else [csz] if csz else [])
                if addr_parts:
                    ben_entries.append(f"{name} - {', '.join(addr_parts)}")
                else:
                    ben_entries.append(name)
                # Heuristic: try to detect groups by checking if next line looks like a name (Last, First pattern)
                # Default to 3-line groups for now
                i += 3
            r["Notes"] = notes_part
            r["Beneficiaries"] = "\n".join(ben_entries)
        else:
            r["Notes"] = old_notes
            r["Beneficiaries"] = ""

    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    csv_path = Path("output") / f"nc_estates_ftm_{ts}.csv"
    xlsx_path = Path("output") / f"nc_estates_ftm_{ts}.xlsx"

    # Write the CSV with the new column order
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FTM_COLUMNS, quoting=csv.QUOTE_MINIMAL,
                                extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
    print(f"Wrote CSV: {csv_path}")

    # Write the XLSX with County dropdown
    _write_xlsx(rows, xlsx_path)
    print(f"Wrote XLSX: {xlsx_path}")


def _write_xlsx(rows: list[dict], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment

    HEADER_FILL = "1B5E20"  # dark green
    BAND_FILL = "FFFDE7"     # very light yellow

    wb = Workbook()
    ws = wb.active
    ws.title = "NC Estates"

    # Header — dark green fill, bold white text
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    for col_idx, col_name in enumerate(FTM_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # Data rows — single-line height, alternating banded fill
    band_fill = PatternFill(start_color=BAND_FILL, end_color=BAND_FILL, fill_type="solid")
    multiline_cols = {"Notes", "Beneficiaries"}
    for r_idx, r in enumerate(rows, start=2):
        for c_idx, col_name in enumerate(FTM_COLUMNS, start=1):
            val = r.get(col_name, "")
            # Collapse multi-line cell content into ' | ' for single-line display.
            if col_name in multiline_cols and val:
                val = " | ".join(s.strip() for s in str(val).split("\n") if s.strip())
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if (r_idx % 2) == 0:
                cell.fill = band_fill
        ws.row_dimensions[r_idx].height = 16

    # County dropdown
    county_col_idx = FTM_COLUMNS.index("County") + 1
    county_col_letter = get_column_letter(county_col_idx)
    county_formula = '"' + ",".join(NC_COUNTY_OPTIONS) + '"'
    dv = DataValidation(type="list", formula1=county_formula, allow_blank=True)
    dv.error = "Pick one of the 7 NC counties"
    dv.errorTitle = "Invalid county"
    dv.prompt = "Select a NC county"
    dv.promptTitle = "County"
    ws.add_data_validation(dv)
    dv.add(f"{county_col_letter}2:{county_col_letter}1048576")

    col_widths = {
        "File Date": 11, "County": 14, "Case No.": 18, "Deceased Owner": 32,
        "Personal Representative": 25, "First Name": 16, "Last Name": 18,
        "Mailing Address": 28, "Mailing City": 16, "Mailing State": 7, "Mailing Zip": 8,
        "Parcel ID": 16, "Property Address": 28, "Property City": 16,
        "Property State": 8, "Property Zip": 8, "Property use": 14,
        "Notes": 40, "Beneficiaries": 80, "Phone 1": 14, "Tags": 26, "List": 10,
    }
    for c_idx, col_name in enumerate(FTM_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col_name, 14)

    ws.freeze_panes = "A2"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


if __name__ == "__main__":
    main()
